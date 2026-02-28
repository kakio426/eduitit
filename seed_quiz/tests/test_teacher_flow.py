import uuid
import re
import io
import zipfile
from datetime import timedelta
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from core.models import UserProfile
from happy_seed.models import HSClassroom
from seed_quiz.models import SQGenerationLog, SQRagDailyUsage, SQQuizBank, SQQuizBankItem, SQQuizSet
from seed_quiz.services.bank import parse_csv_upload

User = get_user_model()


def _make_teacher(username):
    teacher = User.objects.create_user(
        username=username, password="pw", email=f"{username}@test.com"
    )
    UserProfile.objects.update_or_create(
        user=teacher, defaults={"nickname": username, "role": "school"}
    )
    return teacher

VALID_AI_RESPONSE = {
    "items": [
        {
            "question_text": "우리나라 수도는?",
            "choices": ["부산", "서울", "대구", "인천"],
            "correct_index": 1,
            "explanation": "서울입니다.",
            "difficulty": "easy",
        },
        {
            "question_text": "1+1은?",
            "choices": ["1", "2", "3", "4"],
            "correct_index": 1,
            "explanation": "2입니다.",
            "difficulty": "easy",
        },
        {
            "question_text": "태양은 어디서 뜨나요?",
            "choices": ["서쪽", "남쪽", "동쪽", "북쪽"],
            "correct_index": 2,
            "explanation": "동쪽에서 뜹니다.",
            "difficulty": "easy",
        },
    ]
}


class TeacherFlowTest(TestCase):
    def setUp(self):
        self.teacher = _make_teacher("tf_teacher")
        self.other_teacher = _make_teacher("tf_other")
        self.classroom = HSClassroom.objects.create(
            name="흐름 테스트반",
            teacher=self.teacher,
            slug=f"flow-{uuid.uuid4().hex[:6]}",
        )
        self.client = Client()
        self.client.force_login(self.teacher)
        self.bank = SQQuizBank.objects.create(
            title="공식 상식 기본 세트",
            preset_type="orthography",
            grade=3,
            source="manual",
            is_official=True,
            is_public=False,
            is_active=True,
            created_by=self.teacher,
        )
        SQQuizBankItem.objects.create(
            bank=self.bank,
            order_no=1,
            question_text="대한민국 수도는?",
            choices=["부산", "서울", "대구", "광주"],
            correct_index=1,
            explanation="서울입니다.",
            difficulty="easy",
        )
        SQQuizBankItem.objects.create(
            bank=self.bank,
            order_no=2,
            question_text="1+1은?",
            choices=["1", "2", "3", "4"],
            correct_index=1,
            explanation="2입니다.",
            difficulty="easy",
        )
        SQQuizBankItem.objects.create(
            bank=self.bank,
            order_no=3,
            question_text="지구는 어느 행성계에 있나요?",
            choices=["은하수", "태양계", "안드로메다", "명왕성계"],
            correct_index=1,
            explanation="태양계입니다.",
            difficulty="easy",
        )

    def test_dashboard_returns_200(self):
        url = reverse(
            "seed_quiz:teacher_dashboard",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "문제 가져오기")
        self.assertContains(resp, "3단계로 끝나요")
        self.assertContains(resp, 'id="csv-client-check"')
        self.assertContains(resp, "문제는 1~200개")
        self.assertContains(resp, "제작 가이드 보기")
        self.assertContains(resp, "둘 다 넣을 필요가 없습니다")
        self.assertContains(resp, "방법 A. 붙여넣기 (권장)")
        self.assertContains(resp, "방법 B. 파일 올리기")
        self.assertNotContains(resp, "랜덤 1세트 선택")

    def test_landing_redirects_to_first_active_classroom_dashboard(self):
        url = reverse("seed_quiz:landing")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(
            resp["Location"],
            reverse("seed_quiz:teacher_dashboard", kwargs={"classroom_id": self.classroom.id}),
        )

    def test_landing_redirects_to_classroom_create_when_no_classroom(self):
        teacher_without_classroom = _make_teacher("tf_no_class")
        no_class_client = Client()
        no_class_client.force_login(teacher_without_classroom)
        url = reverse("seed_quiz:landing")
        resp = no_class_client.get(url)
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp["Location"], reverse("happy_seed:classroom_create"))

    def test_download_csv_guide(self):
        url = reverse(
            "seed_quiz:download_csv_guide",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/html; charset=utf-8")
        body = resp.content.decode("utf-8")
        self.assertIn("문제 파일 만들기 가이드", body)
        self.assertIn("주제 입력 가이드 (매우 중요)", body)
        self.assertIn("주제 코드(영문)", body)

    def test_download_csv_template(self):
        url = reverse(
            "seed_quiz:download_csv_template",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "text/csv; charset=utf-8")
        body = resp.content.decode("utf-8-sig")
        self.assertIn("주제,학년,문제", body)
        self.assertIn("#가이드:", body)
        self.assertIn("맞춤법", body)

    def test_download_xlsx_template(self):
        url = reverse(
            "seed_quiz:download_xlsx_template",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertIn("씨앗퀴즈 템플릿", wb.sheetnames)
        self.assertIn("필독_작성가이드", wb.sheetnames)

        ws = wb["씨앗퀴즈 템플릿"]
        self.assertEqual(ws["A1"].value, "주제")
        self.assertIn("필독_작성가이드", str(ws["A2"].value or ""))
        self.assertIn("4행부터", str(ws["A3"].value or ""))
        self.assertEqual(ws["A4"].value, "맞춤법")

    def test_download_csv_sample_pack(self):
        url = reverse(
            "seed_quiz:download_csv_sample_pack",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/zip")

        zf = zipfile.ZipFile(io.BytesIO(resp.content))
        names = zf.namelist()
        self.assertIn("README.txt", names)
        csv_names = [n for n in names if n.startswith("samples/") and n.endswith(".csv")]
        self.assertEqual(len(csv_names), len(SQQuizSet.PRESET_CHOICES))

        first_csv = zf.read(csv_names[0])
        parsed_sets, errors = parse_csv_upload(first_csv)
        self.assertFalse(errors)
        self.assertEqual(len(parsed_sets), 1)
        self.assertEqual(len(parsed_sets[0]["items"]), 4)

    def test_download_csv_error_report_requires_valid_token(self):
        url = reverse(
            "seed_quiz:download_csv_error_report",
            kwargs={"classroom_id": self.classroom.id, "token": "invalid-token"},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_topic_summary_returns_200(self):
        url = reverse(
            "seed_quiz:htmx_topic_summary",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "주제별 운영 요약")

    def test_csv_history_returns_200(self):
        SQGenerationLog.objects.create(
            level="info",
            code="CSV_UPLOAD_PREVIEW_READY",
            message="CSV 미리보기 성공: 1세트",
            payload={"classroom_id": str(self.classroom.id), "teacher_id": self.teacher.id, "set_count": 1},
        )
        url = reverse(
            "seed_quiz:htmx_csv_history",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "CSV 업로드 이력")
        self.assertContains(resp, "CSV_UPLOAD_PREVIEW_READY")

    def test_other_teacher_forbidden(self):
        other_client = Client()
        other_client.force_login(self.other_teacher)
        url = reverse(
            "seed_quiz:teacher_dashboard",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = other_client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_bank_browse_returns_official_list(self):
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url, {"preset_type": "orthography", "grade": "3", "scope": "official"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "공식 상식 기본 세트")

    def test_bank_browse_with_all_grade_includes_other_grades(self):
        bank2 = SQQuizBank.objects.create(
            title="공식 맞춤법 심화 세트",
            preset_type="orthography",
            grade=4,
            source="manual",
            is_official=True,
            is_public=False,
            is_active=True,
            created_by=self.teacher,
        )
        for idx in range(1, 4):
            SQQuizBankItem.objects.create(
                bank=bank2,
                order_no=idx,
                question_text=f"문제 {idx}",
                choices=["A", "B", "C", "D"],
                correct_index=0,
                explanation="",
                difficulty="easy",
            )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url, {"preset_type": "orthography", "grade": "all", "scope": "official"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "공식 상식 기본 세트")
        self.assertContains(resp, "공식 맞춤법 심화 세트")

    def test_bank_browse_applies_search_query(self):
        SQQuizBank.objects.create(
            title="공식 문법 연습 세트",
            preset_type="orthography",
            grade=3,
            source="manual",
            is_official=True,
            quality_status="approved",
            created_by=self.teacher,
        )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(
            url,
            {"preset_type": "orthography", "grade": "3", "scope": "official", "q": "상식"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "공식 상식 기본 세트")
        self.assertNotContains(resp, "공식 문법 연습 세트")
        self.assertContains(resp, "검색 결과")

    def test_bank_browse_paginates_results(self):
        for idx in range(1, 26):
            SQQuizBank.objects.create(
                title=f"공식 대량 세트 {idx:02d}",
                preset_type="orthography",
                grade=3,
                source="manual",
                is_official=True,
                quality_status="approved",
                created_by=self.teacher,
            )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )

        first_resp = self.client.get(
            url,
            {"preset_type": "orthography", "grade": "3", "scope": "official", "page": "1"},
        )
        self.assertEqual(first_resp.status_code, 200)
        first_body = first_resp.content.decode("utf-8")
        self.assertEqual(first_body.count("이 세트 적용"), 20)
        self.assertIn("1/2페이지", first_body)
        self.assertIn("page=2", first_body)

        second_resp = self.client.get(
            url,
            {"preset_type": "orthography", "grade": "3", "scope": "official", "page": "2"},
        )
        self.assertEqual(second_resp.status_code, 200)
        second_body = second_resp.content.decode("utf-8")
        self.assertEqual(second_body.count("이 세트 적용"), 6)
        self.assertIn("2/2페이지", second_body)
        self.assertIn("page=1", second_body)

    def test_bank_browse_hides_future_official_set(self):
        SQQuizBank.objects.create(
            title="내일 공개 공식 세트",
            preset_type="orthography",
            grade=3,
            source="ai",
            is_official=True,
            quality_status="approved",
            available_from=timezone.localdate() + timedelta(days=1),
            created_by=self.teacher,
        )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url, {"preset_type": "orthography", "grade": "3", "scope": "official"})
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "내일 공개 공식 세트")

    def test_bank_browse_public_hides_unapproved(self):
        SQQuizBank.objects.create(
            title="미승인 공개 세트",
            preset_type="orthography",
            grade=3,
            source="csv",
            is_public=True,
            share_opt_in=True,
            quality_status="review",
            created_by=self.other_teacher,
        )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url, {"preset_type": "orthography", "grade": "3", "scope": "public"})
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "미승인 공개 세트")

    def test_bank_browse_all_includes_my_review_set(self):
        SQQuizBank.objects.create(
            title="내 검토대기 세트",
            preset_type="orthography",
            grade=3,
            source="csv",
            is_public=False,
            share_opt_in=True,
            quality_status="review",
            created_by=self.teacher,
        )
        url = reverse(
            "seed_quiz:htmx_bank_browse",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.get(url, {"preset_type": "orthography", "grade": "3", "scope": "all"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "내 검토대기 세트")

    def test_bank_select_creates_draft_preview(self):
        url = reverse(
            "seed_quiz:htmx_bank_select",
            kwargs={"classroom_id": self.classroom.id, "bank_id": self.bank.id},
        )
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "2) 저장된 문제 확인")
        draft = SQQuizSet.objects.filter(classroom=self.classroom, status="draft").first()
        self.assertIsNotNone(draft)
        self.assertEqual(draft.source, "bank")
        self.assertEqual(draft.items.count(), 3)

    def test_bank_select_allows_more_than_three_items(self):
        SQQuizBankItem.objects.create(
            bank=self.bank,
            order_no=4,
            question_text="추가 문항",
            choices=["A", "B", "C", "D"],
            correct_index=0,
            explanation="",
            difficulty="easy",
        )
        url = reverse(
            "seed_quiz:htmx_bank_select",
            kwargs={"classroom_id": self.classroom.id, "bank_id": self.bank.id},
        )
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 200)
        draft = SQQuizSet.objects.filter(classroom=self.classroom, status="draft").first()
        self.assertIsNotNone(draft)
        self.assertEqual(draft.items.count(), 4)

    def test_bank_select_creates_new_draft_each_time(self):
        url = reverse(
            "seed_quiz:htmx_bank_select",
            kwargs={"classroom_id": self.classroom.id, "bank_id": self.bank.id},
        )
        first = self.client.post(url)
        second = self.client.post(url)
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        drafts = SQQuizSet.objects.filter(classroom=self.classroom, status="draft")
        self.assertEqual(drafts.count(), 2)

    @patch("seed_quiz.services.generation._call_ai")
    def test_generate_creates_new_draft_each_time(self, mock_ai):
        mock_ai.return_value = VALID_AI_RESPONSE
        url = reverse(
            "seed_quiz:htmx_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        self.client.post(url, {"preset_type": "orthography", "grade": "3"})
        self.client.post(url, {"preset_type": "orthography", "grade": "3"})
        drafts = SQQuizSet.objects.filter(classroom=self.classroom, status="draft")
        self.assertEqual(drafts.count(), 2)

    def test_csv_upload_creates_bank(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,대한민국 수도는?,서울,부산,대구,광주,1,서울입니다,easy\n"
            "orthography,3,1+1은?,2,1,3,4,1,2입니다,easy\n"
            "orthography,3,바다 색은?,파랑,빨강,검정,흰색,1,파랑이 일반적입니다,easy\n"
            "orthography,3,낮과 밤이 반복되는 이유는?,지구 자전,지구 공전,달 공전,태양 공전,1,지구가 자전하기 때문입니다,medium\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "CSV 미리보기 완료")

        match = re.search(r'name="preview_token" value="([^"]+)"', resp.content.decode("utf-8"))
        self.assertIsNotNone(match)
        token = match.group(1)

        confirm_url = reverse(
            "seed_quiz:htmx_csv_confirm",
            kwargs={"classroom_id": self.classroom.id},
        )
        confirm_resp = self.client.post(confirm_url, {"preview_token": token})
        self.assertEqual(confirm_resp.status_code, 200)
        self.assertContains(confirm_resp, "CSV 저장이 완료되었습니다")
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertTrue(bank.title.startswith("CSV-orthography-G3-"))
        self.assertEqual(bank.items.count(), 4)

    def test_csv_upload_rejects_mixed_topic_in_single_file(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,대한민국 수도는?,서울,부산,대구,광주,1,서울입니다,easy\n"
            "vocabulary,3,낱말 뜻은?,의미,뜻,내용,문맥,2,뜻과 의미를 확인합니다,easy\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "주제는 모든 문항에서 동일해야 합니다", status_code=400)
        body = resp.content.decode("utf-8")
        report_match = re.search(r'href="([^"]+/csv-error-report/[^"]+/)"', body)
        self.assertIsNotNone(report_match)
        report_url = report_match.group(1)
        report_resp = self.client.get(report_url)
        self.assertEqual(report_resp.status_code, 200)
        report_body = report_resp.content.decode("utf-8-sig")
        self.assertIn("no,error_message", report_body)
        self.assertIn("주제는 모든 문항에서 동일해야 합니다", report_body)
        self.assertTrue(
            SQGenerationLog.objects.filter(
                code="CSV_UPLOAD_PREVIEW_FAILED",
                payload__classroom_id=str(self.classroom.id),
                payload__teacher_id=self.teacher.id,
            ).exists()
        )

    def test_csv_upload_with_korean_headers_is_supported(self):
        csv_text = (
            "주제,학년,문제,보기1,보기2,보기3,보기4,정답번호,해설,난이도\n"
            "맞춤법,3,대한민국 수도는?,서울,부산,대구,광주,1,서울입니다,쉬움\n"
            "맞춤법,3,1+1은?,2,1,3,4,1,2입니다,쉬움\n"
            "맞춤법,3,바다 색은?,파랑,빨강,검정,흰색,1,파랑이 일반적입니다,쉬움\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz_ko.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "CSV 미리보기 완료")

    def test_text_upload_tsv_creates_bank(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "한글 맞춤법\t3\t대한민국 수도는?\t서울\t부산\t대구\t광주\t1\t서울입니다\t쉬움\n"
            "한글 맞춤법\t3\t1+1은?\t2\t1\t3\t4\t1\t2입니다\t쉬움\n"
            "한글 맞춤법\t3\t바다 색은?\t파랑\t빨강\t검정\t흰색\t1\t파랑이 일반적입니다\t보통\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        upload_resp = self.client.post(upload_url, {"pasted_text": pasted_text})
        self.assertEqual(upload_resp.status_code, 200)
        self.assertContains(upload_resp, "붙여넣기 미리보기 완료")

        match = re.search(r'name="preview_token" value="([^"]+)"', upload_resp.content.decode("utf-8"))
        self.assertIsNotNone(match)
        token = match.group(1)

        confirm_url = reverse(
            "seed_quiz:htmx_csv_confirm",
            kwargs={"classroom_id": self.classroom.id},
        )
        confirm_resp = self.client.post(confirm_url, {"preview_token": token})
        self.assertEqual(confirm_resp.status_code, 200)
        self.assertContains(confirm_resp, "CSV 저장이 완료되었습니다")
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.items.count(), 3)

    def test_text_upload_save_mode_creates_bank_immediately(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "한글 맞춤법\t3\t대한민국 수도는?\t서울\t부산\t대구\t광주\t1\t서울입니다\t쉬움\n"
            "한글 맞춤법\t3\t1+1은?\t2\t1\t3\t4\t1\t2입니다\t쉬움\n"
            "한글 맞춤법\t3\t바다 색은?\t파랑\t빨강\t검정\t흰색\t1\t파랑이 일반적입니다\t보통\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            upload_url,
            {"pasted_text": pasted_text, "submit_mode": "save"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "문제 만들기가 완료되었습니다")
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.items.count(), 3)

    def test_text_upload_accepts_common_topic_alias(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "독해\t3\t대한민국 수도는?\t서울\t부산\t대구\t광주\t1\t서울입니다\t쉬움\n"
            "독해\t3\t1+1은?\t2\t1\t3\t4\t1\t2입니다\t쉬움\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            upload_url,
            {"pasted_text": pasted_text, "submit_mode": "save"},
        )
        self.assertEqual(resp.status_code, 200)
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.preset_type, "main_sentence")
        self.assertEqual(bank.grade, 3)

    def test_text_upload_accepts_human_friendly_grade_label(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "맞춤법\t초3\t대한민국 수도는?\t서울\t부산\t대구\t광주\t1\t서울입니다\t쉬움\n"
            "맞춤법\t초3\t1+1은?\t2\t1\t3\t4\t1\t2입니다\t쉬움\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            upload_url,
            {"pasted_text": pasted_text, "submit_mode": "save"},
        )
        self.assertEqual(resp.status_code, 200)
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.preset_type, "orthography")
        self.assertEqual(bank.grade, 3)

    def test_text_upload_accepts_topic_label_with_spacing(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "주제 제목 고르기\t3\t이 글의 제목은?\t봄나들이\t빗소리\t바닷가\t겨울눈\t1\t글 전체 내용과 맞습니다\t보통\n"
            "주제 제목 고르기\t3\t가장 알맞은 제목은?\t우리 집 강아지\t바닷속 여행\t우주 탐험\t사막 횡단\t1\t중심 내용과 맞습니다\t보통\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            upload_url,
            {"pasted_text": pasted_text, "submit_mode": "save"},
        )
        self.assertEqual(resp.status_code, 200)
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.preset_type, "topic_title")
        self.assertEqual(bank.grade, 3)

    def test_text_upload_accepts_grade_with_parenthesis(self):
        pasted_text = (
            "주제\t학년\t문제\t보기1\t보기2\t보기3\t보기4\t정답번호\t해설\t난이도\n"
            "맞춤법\t3학년(초등)\t대한민국 수도는?\t서울\t부산\t대구\t광주\t1\t서울입니다\t쉬움\n"
            "맞춤법\t3학년(초등)\t1+1은?\t2\t1\t3\t4\t1\t2입니다\t쉬움\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            upload_url,
            {"pasted_text": pasted_text, "submit_mode": "save"},
        )
        self.assertEqual(resp.status_code, 200)
        bank = SQQuizBank.objects.get(source="csv", created_by=self.teacher)
        self.assertEqual(bank.preset_type, "orthography")
        self.assertEqual(bank.grade, 3)

    def test_text_upload_supports_stacked_lines_format(self):
        pasted_text = (
            "주제\n학년\n문제\n보기1\n보기2\n보기3\n보기4\n정답번호\n해설\n난이도\n"
            "맞춤법\n3\n대한민국 수도는?\n서울\n부산\n대구\n광주\n1\n서울입니다\n쉬움\n"
            "맞춤법\n3\n1+1은?\n2\n1\n3\n4\n1\n2입니다\n쉬움\n"
        )
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(upload_url, {"pasted_text": pasted_text})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "붙여넣기 미리보기 완료")
        self.assertContains(resp, "세로 나열")

    def test_text_upload_rejects_empty(self):
        upload_url = reverse(
            "seed_quiz:htmx_text_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(upload_url, {"pasted_text": "   "})
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "붙여넣기 내용이 비어 있습니다", status_code=400)

    @override_settings(SEED_QUIZ_CSV_MAX_ROWS=2)
    def test_csv_upload_rejects_when_row_limit_exceeded(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,대한민국 수도는?,서울,부산,대구,광주,1,서울입니다,easy\n"
            "orthography,3,1+1은?,2,1,3,4,1,2입니다,easy\n"
            "orthography,3,바다 색은?,파랑,빨강,검정,흰색,1,파랑이 일반적입니다,easy\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "CSV 행 수가 제한(2행)을 초과했습니다.", status_code=400)

    def test_csv_upload_rejects_when_grade_is_mixed_in_single_file(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,문제1,가,나,다,라,1,해설,easy\n"
            "orthography,4,문제2,가,나,다,라,2,해설,easy\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "학년은 모든 문항에서 동일해야 합니다", status_code=400)

    @override_settings(SEED_QUIZ_CSV_MAX_FILE_BYTES=120)
    def test_csv_upload_rejects_when_file_size_exceeded(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,대한민국 수도는 어디인가요?,서울,부산,대구,광주,1,서울입니다,easy\n"
            "orthography,3,1+1은?,2,1,3,4,1,2입니다,easy\n"
            "orthography,3,바다 색은?,파랑,빨강,검정,흰색,1,파랑이 일반적입니다,easy\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "CSV 파일 용량이 제한을 초과했습니다.", status_code=400)

    def test_csv_confirm_auto_shares_as_public_approved(self):
        csv_text = (
            "preset_type,grade,question_text,choice_1,choice_2,choice_3,choice_4,correct_index,explanation,difficulty\n"
            "orthography,3,대한민국 수도는?,서울,부산,대구,광주,1,서울입니다,easy\n"
            "orthography,3,1+1은?,2,1,3,4,1,2입니다,easy\n"
            "orthography,3,바다 색은?,파랑,빨강,검정,흰색,1,파랑이 일반적입니다,easy\n"
        )
        from django.core.files.uploadedfile import SimpleUploadedFile

        upload_url = reverse(
            "seed_quiz:htmx_csv_upload",
            kwargs={"classroom_id": self.classroom.id},
        )
        upload_resp = self.client.post(
            upload_url,
            {"csv_file": SimpleUploadedFile("quiz.csv", csv_text.encode("utf-8"), content_type="text/csv")},
        )
        self.assertEqual(upload_resp.status_code, 200)

        match = re.search(r'name="preview_token" value="([^"]+)"', upload_resp.content.decode("utf-8"))
        self.assertIsNotNone(match)
        token = match.group(1)

        confirm_url = reverse(
            "seed_quiz:htmx_csv_confirm",
            kwargs={"classroom_id": self.classroom.id},
        )
        confirm_resp = self.client.post(confirm_url, {"preview_token": token})
        self.assertEqual(confirm_resp.status_code, 200)
        self.assertContains(confirm_resp, "공유 완료")
        bank = SQQuizBank.objects.get(
            source="csv",
            created_by=self.teacher,
        )
        self.assertTrue(bank.title.startswith("CSV-orthography-G3-"))
        self.assertEqual(bank.quality_status, "approved")
        self.assertTrue(bank.is_public)
        self.assertTrue(bank.share_opt_in)
        self.assertTrue(
            SQGenerationLog.objects.filter(
                code="CSV_UPLOAD_CONFIRM_SUCCESS",
                payload__classroom_id=str(self.classroom.id),
                payload__teacher_id=self.teacher.id,
            ).exists()
        )

    def test_csv_confirm_fails_when_preview_expired(self):
        confirm_url = reverse(
            "seed_quiz:htmx_csv_confirm",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(confirm_url, {"preview_token": "expired-token"})
        self.assertEqual(resp.status_code, 400)
        self.assertContains(resp, "세션이 만료", status_code=400)

    @override_settings(SEED_QUIZ_RAG_DAILY_LIMIT=1, SEED_QUIZ_ALLOW_RAG=True)
    @patch("seed_quiz.views.generate_bank_from_context_ai")
    def test_rag_generate_returns_preview_and_consumes_quota(self, mock_generate):
        mock_generate.return_value = self.bank
        url = reverse(
            "seed_quiz:htmx_rag_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"preset_type": "orthography", "grade": "3", "source_text": "오늘 읽은 글 지문입니다."},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "지문 기반 맞춤 생성 완료")
        self.assertContains(resp, "대한민국 수도는?")
        usage = SQRagDailyUsage.objects.get(
            usage_date=timezone.localdate(),
            classroom=self.classroom,
            teacher=self.teacher,
        )
        self.assertEqual(usage.count, 1)

    @override_settings(SEED_QUIZ_RAG_DAILY_LIMIT=1, SEED_QUIZ_ALLOW_RAG=True)
    def test_rag_generate_blocks_when_daily_limit_exceeded(self):
        SQRagDailyUsage.objects.create(
            usage_date=timezone.localdate(),
            classroom=self.classroom,
            teacher=self.teacher,
            count=1,
        )
        url = reverse(
            "seed_quiz:htmx_rag_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"preset_type": "orthography", "grade": "3", "source_text": "오늘 읽은 글 지문입니다."},
        )
        self.assertEqual(resp.status_code, 429)
        self.assertIn("오늘의 맞춤 생성 횟수를 모두 사용", resp.content.decode("utf-8"))

    @override_settings(SEED_QUIZ_RAG_DAILY_LIMIT=1, SEED_QUIZ_ALLOW_RAG=True)
    @patch("seed_quiz.views.generate_bank_from_context_ai", side_effect=RuntimeError("생성 실패"))
    def test_rag_generate_refunds_quota_on_generation_error(self, _mock_generate):
        url = reverse(
            "seed_quiz:htmx_rag_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(
            url,
            {"preset_type": "orthography", "grade": "3", "source_text": "오늘 읽은 글 지문입니다."},
        )
        self.assertEqual(resp.status_code, 400)
        usage = SQRagDailyUsage.objects.get(
            usage_date=timezone.localdate(),
            classroom=self.classroom,
            teacher=self.teacher,
        )
        self.assertEqual(usage.count, 0)

    @patch("seed_quiz.services.generation._call_ai")
    def test_generate_returns_preview(self, mock_ai):
        mock_ai.return_value = VALID_AI_RESPONSE
        url = reverse(
            "seed_quiz:htmx_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        resp = self.client.post(url, {"preset_type": "orthography", "grade": "3"})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "우리나라 수도는?")

    @patch("seed_quiz.services.generation._call_ai")
    def test_publish_changes_status(self, mock_ai):
        mock_ai.return_value = VALID_AI_RESPONSE
        # 생성
        gen_url = reverse(
            "seed_quiz:htmx_generate",
            kwargs={"classroom_id": self.classroom.id},
        )
        self.client.post(gen_url, {"preset_type": "orthography", "grade": "3"})
        quiz_set = SQQuizSet.objects.filter(classroom=self.classroom).first()
        self.assertIsNotNone(quiz_set)

        # 배포
        pub_url = reverse(
            "seed_quiz:htmx_publish",
            kwargs={
                "classroom_id": self.classroom.id,
                "set_id": quiz_set.id,
            },
        )
        resp = self.client.post(pub_url)
        self.assertEqual(resp.status_code, 200)
        quiz_set.refresh_from_db()
        self.assertEqual(quiz_set.status, "published")

    @patch("seed_quiz.services.generation._call_ai")
    def test_publish_requires_force_and_closes_existing_published(self, mock_ai):
        mock_ai.return_value = VALID_AI_RESPONSE
        gen_url = reverse(
            "seed_quiz:htmx_generate",
            kwargs={"classroom_id": self.classroom.id},
        )

        # 첫 번째 생성
        self.client.post(gen_url, {"preset_type": "orthography", "grade": "3"})
        qs1 = SQQuizSet.objects.filter(classroom=self.classroom).order_by("created_at").first()
        self.assertIsNotNone(qs1)
        pub_url = reverse(
            "seed_quiz:htmx_publish",
            kwargs={"classroom_id": self.classroom.id, "set_id": qs1.id},
        )
        self.client.post(pub_url)
        qs1.refresh_from_db()
        self.assertEqual(qs1.status, "published")

        # 두 번째 생성 (다시 생성하면 같은 draft 세트 재사용)
        self.client.post(gen_url, {"preset_type": "orthography", "grade": "3"})
        # draft가 다시 생겼는지 확인 (qs1이 published → 새 draft)
        qs2 = SQQuizSet.objects.filter(classroom=self.classroom, status="draft").first()
        self.assertIsNotNone(qs2)
        pub_url2 = reverse(
            "seed_quiz:htmx_publish",
            kwargs={"classroom_id": self.classroom.id, "set_id": qs2.id},
        )
        first_resp = self.client.post(pub_url2)
        self.assertEqual(first_resp.status_code, 200)
        self.assertContains(first_resp, "이미 배포 중인 퀴즈가 있습니다.")
        qs1.refresh_from_db()
        qs2.refresh_from_db()
        self.assertEqual(qs1.status, "published")
        self.assertEqual(qs2.status, "draft")

        # 확인 후 강제 배포
        force_resp = self.client.post(pub_url2, {"force_replace": "1"})
        self.assertEqual(force_resp.status_code, 200)
        self.assertContains(force_resp, "직전 배포로 되돌리기")
        qs1.refresh_from_db()
        qs2.refresh_from_db()
        self.assertEqual(qs1.status, "closed")
        self.assertEqual(qs2.status, "published")

    @patch("seed_quiz.services.generation._call_ai")
    def test_publish_rollback_restores_previous_set(self, mock_ai):
        mock_ai.return_value = VALID_AI_RESPONSE
        gen_url = reverse(
            "seed_quiz:htmx_generate",
            kwargs={"classroom_id": self.classroom.id},
        )

        # 첫 배포
        self.client.post(gen_url, {"preset_type": "orthography", "grade": "3"})
        old_set = SQQuizSet.objects.filter(classroom=self.classroom).order_by("created_at").first()
        self.assertIsNotNone(old_set)
        old_publish_url = reverse(
            "seed_quiz:htmx_publish",
            kwargs={"classroom_id": self.classroom.id, "set_id": old_set.id},
        )
        self.client.post(old_publish_url)
        old_set.refresh_from_db()
        self.assertEqual(old_set.status, "published")

        # 새 배포(강제 덮어쓰기)
        self.client.post(gen_url, {"preset_type": "orthography", "grade": "3"})
        new_set = SQQuizSet.objects.filter(classroom=self.classroom, status="draft").first()
        self.assertIsNotNone(new_set)
        new_publish_url = reverse(
            "seed_quiz:htmx_publish",
            kwargs={"classroom_id": self.classroom.id, "set_id": new_set.id},
        )
        self.client.post(new_publish_url, {"force_replace": "1"})
        old_set.refresh_from_db()
        new_set.refresh_from_db()
        self.assertEqual(old_set.status, "closed")
        self.assertEqual(new_set.status, "published")

        rollback_url = reverse(
            "seed_quiz:htmx_publish_rollback",
            kwargs={"classroom_id": self.classroom.id},
        )
        rollback_resp = self.client.post(
            rollback_url,
            {"restore_set_id": str(old_set.id), "current_set_id": str(new_set.id)},
        )
        self.assertEqual(rollback_resp.status_code, 200)
        self.assertContains(rollback_resp, "직전 배포로 되돌렸습니다.")
        old_set.refresh_from_db()
        new_set.refresh_from_db()
        self.assertEqual(old_set.status, "published")
        self.assertEqual(new_set.status, "closed")
