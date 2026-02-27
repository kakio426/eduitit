import base64
import logging
import csv
from datetime import timedelta
from io import BytesIO, StringIO
from zipfile import ZIP_DEFLATED, ZipFile
from uuid import uuid4

import openpyxl
import qrcode
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, Paginator
from django.db import transaction
from django.db.models import Count, F, Max, Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from happy_seed.models import HSClassroom, HSStudent
from seed_quiz.models import SQAttempt, SQGenerationLog, SQQuizBank, SQQuizItem, SQQuizSet
from seed_quiz.services.bank import (
    copy_bank_to_draft,
    generate_bank_from_context_ai,
    parse_csv_upload,
    save_parsed_sets_to_bank,
)
from seed_quiz.services.gate import (
    clear_session,
    find_or_create_attempt,
    get_current_item,
    get_today_published_set,
    set_session,
)
from seed_quiz.services.generation import generate_and_save_draft
from seed_quiz.services.paste_parser import parse_pasted_text_to_csv_bytes
from seed_quiz.services.grading import submit_and_reward
from seed_quiz.services.rag import consume_rag_quota, refund_rag_quota
from seed_quiz.topics import DEFAULT_TOPIC, TOPIC_LABELS, normalize_topic

logger = logging.getLogger("seed_quiz.views")


def _build_qr_data_url(raw_text: str) -> str:
    if not raw_text:
        return ""

    qr_image = qrcode.make(raw_text)
    with BytesIO() as buffer:
        qr_image.save(buffer, format="PNG")
        encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"

CSV_PREVIEW_PAYLOAD_KEY = "sq_csv_preview_payload"
CSV_PREVIEW_TOKEN_KEY = "sq_csv_preview_token"
CSV_ERROR_REPORT_TOKEN_KEY = "sq_csv_error_report_token"
CSV_ERROR_REPORT_ROWS_KEY = "sq_csv_error_report_rows"

CSV_CANONICAL_HEADERS = [
    "preset_type",
    "grade",
    "question_text",
    "choice_1",
    "choice_2",
    "choice_3",
    "choice_4",
    "correct_index",
    "explanation",
    "difficulty",
]

CSV_KOREAN_HEADERS = [
    "주제",
    "학년",
    "문제",
    "보기1",
    "보기2",
    "보기3",
    "보기4",
    "정답번호",
    "해설",
    "난이도",
]

CSV_GUIDE_COLUMNS = [
    {
        "ko": "주제",
        "required": "필수",
        "desc": "한 파일에서는 모든 행이 같은 주제여야 합니다. 입력 예: 맞춤법, 띄어쓰기, 어휘 뜻, 속담, 관용어, 사자성어, 한자어 뜻, 중심문장 찾기, 문장 순서 배열, 주제/제목 고르기, 사실/의견 구분, 영어 단어 뜻, 영어 문장 의미, 영어 빈칸 채우기, 수학 연산, 규칙 찾기, 분수/소수 비교, 시간/달력 계산, 단위 변환, 생활 안전 상식",
    },
    {"ko": "학년", "required": "필수", "desc": "1~6 중 하나. 학년 구분이 없으면 0 또는 학년무관. 한 파일에서 학년은 동일해야 합니다."},
    {"ko": "문제", "required": "필수", "desc": "문항 내용"},
    {"ko": "보기1", "required": "필수", "desc": "선택지 1"},
    {"ko": "보기2", "required": "필수", "desc": "선택지 2"},
    {"ko": "보기3", "required": "필수", "desc": "선택지 3"},
    {"ko": "보기4", "required": "필수", "desc": "선택지 4"},
    {"ko": "정답번호", "required": "필수", "desc": "1~4 중 하나 (보기1이 정답이면 1, 보기2이면 2, 보기3이면 3, 보기4이면 4)"},
    {"ko": "해설", "required": "선택", "desc": "정답 설명 (없어도 됩니다)"},
    {"ko": "난이도", "required": "선택", "desc": "쉬움, 보통, 어려움 중 하나 (없으면 보통으로 처리)"},
]

CSV_XLSX_COLUMN_WIDTHS = [16, 8, 42, 22, 22, 22, 22, 12, 36, 12]

TOPIC_TEMPLATE_HELP = {
    "orthography": {
        "focus": "철자, 받침, 헷갈리는 맞춤법 표현을 묻는 문제",
        "examples": "되/돼, 안/않, 왠지/웬지, 어이없다",
        "tip": "짧은 문장 속에서 바른 표기를 고르게 하면 좋습니다.",
    },
    "spacing": {
        "focus": "띄어쓰기 원칙을 적용해 바른 문장을 고르는 문제",
        "examples": "할 수 있다, 함께 가다, 집에 가는 길",
        "tip": "비슷한 보기 4개를 넣고 한 칸 차이를 분명히 보여주세요.",
    },
    "vocabulary": {
        "focus": "낱말 뜻, 유의어/반의어, 문맥 속 어휘 의미를 묻는 문제",
        "examples": "유의어 찾기, 낱말 뜻 풀이, 문맥에 맞는 단어",
        "tip": "학년 수준보다 어려운 한자어는 해설에 뜻을 짧게 덧붙이세요.",
    },
    "proverb": {
        "focus": "속담의 의미와 상황 적용을 묻는 문제",
        "examples": "가는 말이 고와야 오는 말이 곱다, 티끌 모아 태산",
        "tip": "생활 장면을 먼저 제시하고 어울리는 속담을 고르게 하세요.",
    },
    "idiom": {
        "focus": "관용어의 실제 의미와 쓰임을 묻는 문제",
        "examples": "발이 넓다, 귀가 얇다, 손이 크다",
        "tip": "글자 그대로 뜻이 아닌 관용적 의미를 구분하도록 구성하세요.",
    },
    "sino_idiom": {
        "focus": "사자성어 뜻과 쓰임 상황을 묻는 문제",
        "examples": "유비무환, 동문서답, 자업자득",
        "tip": "뜻풀이 보기와 상황 보기를 섞어 난이도를 조절하세요.",
    },
    "hanja_word": {
        "focus": "교과서 수준 한자어의 의미를 이해하는 문제",
        "examples": "보호, 관찰, 책임, 협동",
        "tip": "낱말의 실제 문장 활용 예시를 문제에 함께 넣으면 좋습니다.",
    },
    "main_sentence": {
        "focus": "짧은 글에서 중심문장을 찾는 독해 문제",
        "examples": "설명문 핵심 문장 찾기, 글쓴이 주장 찾기",
        "tip": "지문은 2~4문장 정도로 간결하게 유지하세요.",
    },
    "sentence_order": {
        "focus": "문장 배열 순서와 글의 흐름을 판단하는 문제",
        "examples": "도입-전개-마무리 순서, 사건 순서 배열",
        "tip": "연결어(먼저, 다음에, 그래서)를 활용하면 풀이가 명확해집니다.",
    },
    "topic_title": {
        "focus": "글의 주제 또는 제목을 고르는 문제",
        "examples": "가장 알맞은 제목, 글의 중심 생각",
        "tip": "정답과 오답의 차이를 해설에서 짧게 설명해 주세요.",
    },
    "fact_opinion": {
        "focus": "사실과 의견을 구분하는 문해력 문제",
        "examples": "관찰 사실 vs 개인 생각, 객관 정보 vs 주장",
        "tip": "숫자/날짜가 들어간 문장을 사실 보기로 활용하면 구분이 쉽습니다.",
    },
    "eng_vocab": {
        "focus": "영어 단어의 기본 뜻을 묻는 문제",
        "examples": "apple, library, beautiful, quickly",
        "tip": "보기는 모두 같은 품사 수준으로 맞추면 학습 효과가 큽니다.",
    },
    "eng_sentence": {
        "focus": "짧은 영어 문장의 의미를 묻는 문제",
        "examples": "I have a dog. / She is reading a book.",
        "tip": "문장은 5~10단어 내외로 유지하고 핵심 어휘를 반복하세요.",
    },
    "eng_cloze": {
        "focus": "문맥에 맞는 영어 낱말/표현을 빈칸에 넣는 문제",
        "examples": "I ___ to school. (go/goes/went/going)",
        "tip": "시제나 수일치처럼 한 가지 규칙만 묻도록 설계하세요.",
    },
    "arithmetic": {
        "focus": "사칙연산과 기초 계산 정확도를 묻는 문제",
        "examples": "두 자리 수 덧셈/뺄셈, 곱셈표, 나눗셈",
        "tip": "보기는 계산 실수 패턴(자리올림 누락 등)을 반영하면 좋습니다.",
    },
    "pattern": {
        "focus": "수열/도형/규칙 찾기 문제",
        "examples": "2,4,6,?, 모양 반복 규칙",
        "tip": "규칙의 단서를 문제 본문에 명확히 드러내 주세요.",
    },
    "fraction_decimal": {
        "focus": "분수와 소수의 크기 비교·변환 문제",
        "examples": "1/2 vs 0.4, 0.75를 분수로",
        "tip": "같은 값의 다른 표현을 보기로 섞어 개념 연결을 돕습니다.",
    },
    "time_calendar": {
        "focus": "시각 읽기와 달력 계산 문제",
        "examples": "경과 시간, 요일 계산, 일정 날짜 찾기",
        "tip": "현실 상황(수업 시작/끝 시간)을 넣으면 이해가 쉬워집니다.",
    },
    "unit_conversion": {
        "focus": "길이·무게·들이·시간 단위 변환 문제",
        "examples": "cm-m, g-kg, mL-L, 분-시간",
        "tip": "숫자 단위가 바뀌는 기준(10, 100, 1000)을 해설에 써 주세요.",
    },
    "safety_common": {
        "focus": "생활 안전 지식과 상황 대처를 묻는 문제",
        "examples": "횡단보도 안전, 화재 대피, 감염 예방",
        "tip": "정답 문항은 실제 행동 지침과 일치하도록 작성하세요.",
    },
}


def _get_topic_template_rows() -> list[dict]:
    rows: list[dict] = []
    for topic_key, topic_label in SQQuizSet.PRESET_CHOICES:
        info = TOPIC_TEMPLATE_HELP.get(topic_key, {})
        rows.append(
            {
                "key": topic_key,
                "label": topic_label,
                "focus": info.get("focus", "해당 주제의 핵심 개념을 묻는 문제"),
                "examples": info.get("examples", "교과서 핵심 단원 중심"),
                "tip": info.get("tip", "문항 수준을 학년에 맞춰 주세요."),
            }
        )
    return rows


def _build_template_guide_rows() -> list[list[str]]:
    return [
        ["#가이드: 한 파일은 1세트로 저장되며 문항은 1~200개까지 자유롭게 작성할 수 있습니다."] + [""] * 9,
        ["#가이드: 한 파일 안에서는 주제와 학년을 섞지 마세요."] + [""] * 9,
        ["#가이드: 주제는 한글 이름 또는 영문 코드로 입력합니다."] + [""] * 9,
        ["#가이드: 학년은 0~6(0=학년무관), 정답번호는 1~4입니다."] + [""] * 9,
        ["#가이드: 난이도는 쉬움/보통/어려움이며 비우면 보통으로 처리됩니다."] + [""] * 9,
        ["#가이드: 상세 주제 예시는 XLSX의 '필독_작성가이드' 시트 또는 CSV 가이드를 참고하세요."] + [""] * 9,
        ["#가이드: 아래 샘플 문항은 그대로 업로드해도 통과됩니다."] + [""] * 9,
    ]


def _build_default_sample_rows() -> list[list[str]]:
    return [
        ["맞춤법", 3, "다음 중 띄어쓰기가 올바른 것은?", "할수있다", "할 수 있다", "할 수있다", "할수 있다", 2, "'할 수 있다'처럼 의존 명사는 띄어 씁니다.", "쉬움"],
        ["맞춤법", 3, "다음 중 맞춤법이 올바른 것은?", "왠지", "웬지", "왠 지", "웬 지", 1, "'왠지'가 맞는 표기입니다.", "보통"],
        ["맞춤법", 3, "다음 중 표기가 바른 것은?", "되요", "돼요", "되욤", "되여", 2, "'되어요'가 줄어 '돼요'가 됩니다.", "쉬움"],
        ["맞춤법", 3, "다음 중 맞춤법이 올바른 문장은?", "않 가요", "안 가요", "안가요", "않가요", 2, "부정 부사 '안'은 띄어 씁니다.", "보통"],
        ["맞춤법", 3, "다음 중 알맞은 표현은?", "어의없다", "어이없다", "어이 업다", "어의 업다", 2, "'어이없다'가 바른 표현입니다.", "보통"],
    ]


def _build_topic_sample_rows(topic_label: str, grade: int = 3) -> list[list[str]]:
    return [
        [topic_label, grade, f"[{topic_label}] 예시 문제 1", f"{topic_label} 정답 1", f"{topic_label} 오답 1-1", f"{topic_label} 오답 1-2", f"{topic_label} 오답 1-3", 1, f"{topic_label} 예시 해설 1", "쉬움"],
        [topic_label, grade, f"[{topic_label}] 예시 문제 2", f"{topic_label} 오답 2-1", f"{topic_label} 정답 2", f"{topic_label} 오답 2-2", f"{topic_label} 오답 2-3", 2, f"{topic_label} 예시 해설 2", "쉬움"],
        [topic_label, grade, f"[{topic_label}] 예시 문제 3", f"{topic_label} 오답 3-1", f"{topic_label} 오답 3-2", f"{topic_label} 정답 3", f"{topic_label} 오답 3-3", 3, f"{topic_label} 예시 해설 3", "보통"],
        [topic_label, grade, f"[{topic_label}] 예시 문제 4", f"{topic_label} 오답 4-1", f"{topic_label} 오답 4-2", f"{topic_label} 오답 4-3", f"{topic_label} 정답 4", 4, f"{topic_label} 예시 해설 4", "보통"],
    ]


def _get_positive_int_setting(name: str, default: int) -> int:
    try:
        value = int(getattr(settings, name, default))
    except (TypeError, ValueError):
        return default
    return value if value > 0 else default


def _get_csv_upload_limits() -> dict:
    return {
        "max_file_bytes": _get_positive_int_setting("SEED_QUIZ_CSV_MAX_FILE_BYTES", 2 * 1024 * 1024),
        "max_rows": _get_positive_int_setting("SEED_QUIZ_CSV_MAX_ROWS", 1200),
        "max_sets": _get_positive_int_setting("SEED_QUIZ_CSV_MAX_SETS", 400),
    }


def _get_text_upload_limits() -> dict:
    csv_limits = _get_csv_upload_limits()
    return {
        "max_chars": _get_positive_int_setting("SEED_QUIZ_TEXT_MAX_CHARS", 500_000),
        "max_rows": csv_limits["max_rows"],
        "max_sets": csv_limits["max_sets"],
    }


def _humanize_bytes(num: int) -> str:
    value = float(max(0, int(num)))
    units = ["B", "KB", "MB", "GB"]
    unit_idx = 0
    while value >= 1024 and unit_idx < len(units) - 1:
        value /= 1024
        unit_idx += 1
    if unit_idx == 0:
        return f"{int(value)}{units[unit_idx]}"
    return f"{value:.1f}{units[unit_idx]}"


def _trim_errors(errors: list[str], max_items: int = 5, max_len: int = 180) -> list[str]:
    trimmed = []
    for err in (errors or [])[:max_items]:
        msg = str(err or "").replace("\n", " ").replace("\r", " ").strip()
        if len(msg) > max_len:
            msg = msg[: max_len - 1] + "..."
        if msg:
            trimmed.append(msg)
    return trimmed


def _log_csv_event(
    *,
    classroom,
    teacher,
    code: str,
    level: str,
    message: str,
    payload: dict | None = None,
):
    data = dict(payload or {})
    data.setdefault("classroom_id", str(classroom.id))
    data.setdefault("teacher_id", int(getattr(teacher, "id", 0) or 0))
    SQGenerationLog.objects.create(
        level=level,
        code=code,
        message=message,
        payload=data,
    )


def _parse_bank_filters(raw_preset: str, raw_grade: str):
    preset_type = normalize_topic(raw_preset) or DEFAULT_TOPIC
    raw_grade_value = (raw_grade or "").strip().lower()
    if raw_grade_value in {"all", "*", ""}:
        return preset_type, None
    try:
        grade = int(raw_grade_value)
    except (TypeError, ValueError):
        grade = 3
    if grade not in range(0, 7):
        grade = 3
    return preset_type, grade


def _build_bank_queryset(classroom, user, preset_type: str, grade: int | None, scope: str):
    banks_qs = SQQuizBank.objects.filter(
        is_active=True,
        preset_type=preset_type,
    )
    if grade is not None:
        banks_qs = banks_qs.filter(grade=grade)

    approved_filter = Q(quality_status="approved")
    today = timezone.localdate()
    available_filter = (
        (Q(available_from__isnull=True) | Q(available_from__lte=today))
        & (Q(available_to__isnull=True) | Q(available_to__gte=today))
    )
    if scope == "mine":
        banks_qs = banks_qs.filter(created_by=user)
    elif scope == "official":
        banks_qs = banks_qs.filter(is_official=True).filter(approved_filter).filter(available_filter)
    elif scope == "public":
        banks_qs = banks_qs.filter(is_public=True).filter(approved_filter).filter(available_filter)
    else:
        banks_qs = banks_qs.filter(
            (Q(is_official=True) & approved_filter & available_filter)
            | (Q(is_public=True) & approved_filter & available_filter)
            | Q(created_by=user)
        ).distinct()
    return banks_qs


def _clear_csv_error_report(request):
    request.session.pop(CSV_ERROR_REPORT_TOKEN_KEY, None)
    request.session.pop(CSV_ERROR_REPORT_ROWS_KEY, None)
    request.session.modified = True


def _store_csv_error_report(request, errors: list[str]) -> str:
    token = uuid4().hex
    sanitized = []
    for raw in errors:
        msg = str(raw or "").replace("\r", " ").replace("\n", " ").strip()
        if msg:
            sanitized.append(msg)
    request.session[CSV_ERROR_REPORT_TOKEN_KEY] = token
    request.session[CSV_ERROR_REPORT_ROWS_KEY] = sanitized[:500]
    request.session.modified = True
    return token


# ---------------------------------------------------------------------------
# 랜딩
# ---------------------------------------------------------------------------

@login_required
def landing(request):
    """씨앗 퀴즈 교사 진입점.

    - 활성 학급이 있으면 첫 학급 대시보드로 이동
    - 없으면 학급 생성 화면으로 안내
    """
    classroom = (
        HSClassroom.objects.filter(teacher=request.user, is_active=True)
        .order_by("created_at")
        .first()
    )
    if classroom:
        return redirect("seed_quiz:teacher_dashboard", classroom_id=classroom.id)

    messages.info(
        request,
        "씨앗 퀴즈를 시작하려면 먼저 학급을 만들어 주세요.",
    )
    return redirect("happy_seed:classroom_create")


# ---------------------------------------------------------------------------
# 교사 뷰
# ---------------------------------------------------------------------------

@login_required
def teacher_dashboard(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    csv_limits = _get_csv_upload_limits()
    text_limits = _get_text_upload_limits()
    initial_preset, initial_grade = _parse_bank_filters(
        request.GET.get("preset_type", DEFAULT_TOPIC),
        request.GET.get("grade", "all"),
    )
    initial_scope = (request.GET.get("scope", "official") or "official").strip().lower()
    if initial_scope not in {"official", "public", "all", "mine"}:
        initial_scope = "official"
    initial_query = (request.GET.get("q") or "").strip()
    today_sets = SQQuizSet.objects.filter(
        classroom=classroom,
        target_date=timezone.localdate(),
    ).order_by("-created_at")
    return render(
        request,
        "seed_quiz/teacher_dashboard.html",
        {
            "classroom": classroom,
            "today_sets": today_sets,
            "preset_choices": SQQuizSet.PRESET_CHOICES,
            "initial_preset": initial_preset,
            "initial_grade": initial_grade,
            "initial_grade_str": "all" if initial_grade is None else str(initial_grade),
            "initial_scope": initial_scope,
            "initial_query": initial_query,
            "rag_daily_limit": max(0, int(getattr(settings, "SEED_QUIZ_RAG_DAILY_LIMIT", 1))),
            "allow_rag": bool(getattr(settings, "SEED_QUIZ_ALLOW_RAG", False)),
            "csv_limits": csv_limits,
            "csv_limits_human": {
                "max_file": _humanize_bytes(csv_limits["max_file_bytes"]),
                "max_rows": csv_limits["max_rows"],
                "max_sets": csv_limits["max_sets"],
            },
            "text_limits": text_limits,
            "csv_headers_ko_text": ",".join(CSV_KOREAN_HEADERS),
            "csv_headers_en_text": ",".join(CSV_CANONICAL_HEADERS),
            "csv_guide_columns": CSV_GUIDE_COLUMNS,
            "topic_guide_rows": _get_topic_template_rows(),
        },
    )


@login_required
def download_csv_guide(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    if not classroom:
        return HttpResponseForbidden("권한이 없습니다.")
    return render(
        request,
        "seed_quiz/csv_guide.html",
        {
            "classroom": classroom,
            "csv_headers_ko_text": ",".join(CSV_KOREAN_HEADERS),
            "csv_headers_en_text": ",".join(CSV_CANONICAL_HEADERS),
            "csv_guide_columns": CSV_GUIDE_COLUMNS,
            "topic_guide_rows": _get_topic_template_rows(),
        },
    )


@login_required
def download_csv_template(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    if not classroom:
        return HttpResponseForbidden("권한이 없습니다.")

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_KOREAN_HEADERS)
    for guide_row in _build_template_guide_rows():
        writer.writerow(guide_row)
    for sample_row in _build_default_sample_rows():
        writer.writerow(sample_row)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="seed_quiz_template.csv"'
    response.write("\ufeff")
    response.write(output.getvalue())
    return response


@login_required
def download_xlsx_template(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    if not classroom:
        return HttpResponseForbidden("권한이 없습니다.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "씨앗퀴즈 템플릿"
    ws.append(CSV_KOREAN_HEADERS)

    header_fill = PatternFill(fill_type="solid", start_color="E8F0FE", end_color="E8F0FE")
    for col_idx, header in enumerate(CSV_KOREAN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = CSV_XLSX_COLUMN_WIDTHS[col_idx - 1]

    # 메인 시트는 입력 중심으로 단순하게 유지하고, 상세 설명은 별도 가이드 시트로 분리한다.
    guide_sheet_name = "필독_작성가이드"

    ws.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=len(CSV_KOREAN_HEADERS),
    )
    guide_cell = ws.cell(
        row=2,
        column=1,
        value="#필독: 하단 시트 탭 '필독_작성가이드'를 먼저 확인하세요.",
    )
    guide_cell.font = Font(color="1E40AF", bold=True, underline="single")
    guide_cell.fill = PatternFill(fill_type="solid", start_color="FFF4D6", end_color="FFF4D6")
    guide_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    guide_cell.hyperlink = f"#'{guide_sheet_name}'!A1"
    ws.row_dimensions[2].height = 22

    ws.merge_cells(
        start_row=3,
        start_column=1,
        end_row=3,
        end_column=len(CSV_KOREAN_HEADERS),
    )
    input_hint_cell = ws.cell(
        row=3,
        column=1,
        value="#가이드: 샘플은 4행부터 시작합니다. 실제 입력도 4행부터 이어서 작성하세요.",
    )
    input_hint_cell.font = Font(color="7C4A03", bold=True)
    input_hint_cell.fill = PatternFill(fill_type="solid", start_color="FFF4D6", end_color="FFF4D6")
    input_hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[3].height = 22

    row_no = 4
    sample_fill = PatternFill(fill_type="solid", start_color="F9FBFF", end_color="F9FBFF")
    for sample_row in _build_default_sample_rows():
        ws.append(sample_row)
        for col_idx in range(1, len(CSV_KOREAN_HEADERS) + 1):
            cell = ws.cell(row=row_no, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = sample_fill
        row_no += 1

    ws.freeze_panes = "A4"

    # 입력 오류를 줄이기 위한 드롭다운(파서 규칙과 동일)
    dv_grade = DataValidation(type="list", formula1='"0,1,2,3,4,5,6,학년무관"', allow_blank=False)
    dv_correct = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=False)
    dv_difficulty = DataValidation(type="list", formula1='"쉬움,보통,어려움"', allow_blank=True)
    ws.add_data_validation(dv_grade)
    ws.add_data_validation(dv_correct)
    ws.add_data_validation(dv_difficulty)
    dv_grade.add("B4:B1003")
    dv_correct.add("H4:H1003")
    dv_difficulty.add("J4:J1003")

    topic_ws = wb.create_sheet(guide_sheet_name)
    topic_ws.sheet_properties.tabColor = "F59E0B"
    topic_ws.merge_cells("A1:E1")
    title_cell = topic_ws["A1"]
    title_cell.value = "씨앗 퀴즈 작성 가이드"
    title_cell.font = Font(bold=True, size=14, color="1E3A8A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    topic_ws["A2"] = "이 템플릿은 CSV 파서 규칙과 100% 동일합니다."
    topic_ws["A3"] = "메인 시트(씨앗퀴즈 템플릿)의 1행 헤더는 수정하지 말고, 3행부터 입력하세요."
    topic_ws["A4"] = "한 파일(한 번 업로드)은 같은 주제/같은 학년으로만 작성합니다."
    topic_ws["A5"] = "문항 수는 1~200개, 정답번호는 1~4, 난이도는 쉬움/보통/어려움(빈칸=보통)."
    for row_idx in range(2, 6):
        topic_ws[f"A{row_idx}"].alignment = Alignment(wrap_text=True, vertical="top")

    rule_header_row = 7
    rule_headers = ["항목", "허용 입력", "예시", "비고", ""]
    for col_idx, header in enumerate(rule_headers, start=1):
        cell = topic_ws.cell(row=rule_header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rules = [
        ["주제", "한글 또는 영문 코드", "맞춤법 / orthography", "한 파일 안에서는 같은 값 유지", ""],
        ["학년", "0~6", "0(학년무관), 3", "한 파일 안에서는 같은 값 유지", ""],
        ["정답번호", "1~4", "보기2가 정답이면 2", "필수 입력", ""],
        ["난이도", "쉬움/보통/어려움", "비우면 보통 처리", "선택 입력", ""],
    ]
    for offset, rule in enumerate(rules):
        row_idx = rule_header_row + 1 + offset
        for col_idx, value in enumerate(rule, start=1):
            topic_ws.cell(row=row_idx, column=col_idx, value=value)
            topic_ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    topic_headers = ["주제(한글)", "주제 코드(영문)", "어떤 문제를 넣나요?", "문제 소재 예시", "작성 팁"]
    topic_header_row = rule_header_row + len(rules) + 3
    for col_idx, header in enumerate(topic_headers, start=1):
        cell = topic_ws.cell(row=topic_header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    topic_ws.column_dimensions["A"].width = 24
    topic_ws.column_dimensions["B"].width = 24
    topic_ws.column_dimensions["C"].width = 40
    topic_ws.column_dimensions["D"].width = 42
    topic_ws.column_dimensions["E"].width = 42

    topic_data_start = topic_header_row + 1
    for offset, row in enumerate(_get_topic_template_rows()):
        current_row = topic_data_start + offset
        topic_ws.cell(row=current_row, column=1, value=row["label"])
        topic_ws.cell(row=current_row, column=2, value=row["key"])
        topic_ws.cell(row=current_row, column=3, value=row["focus"])
        topic_ws.cell(row=current_row, column=4, value=row["examples"])
        topic_ws.cell(row=current_row, column=5, value=row["tip"])
        for col_idx in range(1, 6):
            topic_ws.cell(row=current_row, column=col_idx).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
    topic_ws.freeze_panes = f"A{topic_header_row + 1}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="seed_quiz_template.xlsx"'
    return response


@login_required
def download_csv_sample_pack(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    if not classroom:
        return HttpResponseForbidden("권한이 없습니다.")

    headers = CSV_KOREAN_HEADERS

    buffer = BytesIO()
    with ZipFile(buffer, mode="w", compression=ZIP_DEFLATED) as zf:
        topic_rows = _get_topic_template_rows()
        topic_labels = ", ".join(row["label"] for row in topic_rows)
        topic_pairs = ", ".join(f"{row['label']}({row['key']})" for row in topic_rows)
        readme_lines = [
            "씨앗 퀴즈 CSV 샘플 팩",
            "",
            "- 모든 파일은 업로드 검증 규칙을 통과하도록 구성되어 있습니다.",
            "- 각 CSV 파일은 1세트로 저장됩니다.",
            "- 한 파일에 문항을 1~200개까지 원하는 개수로 넣을 수 있습니다.",
            "- 한 파일의 모든 행은 같은 주제/같은 학년이어야 합니다.",
            f"- 사용 가능한 주제(한글): {topic_labels}",
            f"- 사용 가능한 주제(코드): {topic_pairs}",
            "- 정답번호는 1~4 (보기1이 정답이면 1, 보기2이면 2).",
            "- 문항/보기/해설을 수정해 업로드하세요.",
        ]
        zf.writestr("README.txt", "\n".join(readme_lines))

        for idx, (topic_key, topic_label) in enumerate(SQQuizSet.PRESET_CHOICES, start=1):
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for guide_row in _build_template_guide_rows():
                writer.writerow(guide_row)
            for sample_row in _build_topic_sample_rows(topic_label=topic_label, grade=3):
                writer.writerow(sample_row)

            csv_text = output.getvalue()
            parsed_sets, errors = parse_csv_upload(csv_text.encode("utf-8"))
            if errors or not parsed_sets:
                return HttpResponse(
                    "샘플 CSV 팩 생성 중 검증 오류가 발생했습니다. 관리자에게 문의해 주세요.",
                    status=500,
                )
            zf.writestr(f"samples/{idx:02d}_{topic_key}.csv", "\ufeff" + csv_text)

    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="seed_quiz_sample_pack.zip"'
    response.write(buffer.getvalue())
    return response


@login_required
def download_csv_error_report(request, classroom_id, token):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    if not classroom:
        return HttpResponseForbidden("권한이 없습니다.")

    session_token = request.session.get(CSV_ERROR_REPORT_TOKEN_KEY)
    error_rows = request.session.get(CSV_ERROR_REPORT_ROWS_KEY) or []

    if not token or token != session_token or not error_rows:
        return HttpResponse("오류 리포트가 만료되었습니다. 다시 CSV를 업로드해 주세요.", status=404)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["no", "error_message"])
    for idx, message in enumerate(error_rows, start=1):
        writer.writerow([idx, message])

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="seed_quiz_error_report_{timezone.localdate().isoformat()}.csv"'
    )
    response.write("\ufeff")
    response.write(output.getvalue())
    return response


@login_required
def htmx_bank_browse(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    preset_type, grade = _parse_bank_filters(
        request.GET.get("preset_type", DEFAULT_TOPIC),
        request.GET.get("grade", "3"),
    )
    scope = (request.GET.get("scope", "official") or "official").strip().lower()
    if scope not in {"official", "public", "all", "mine"}:
        scope = "official"
    query = (request.GET.get("q") or "").strip()

    banks_qs = _build_bank_queryset(
        classroom=classroom,
        user=request.user,
        preset_type=preset_type,
        grade=grade,
        scope=scope,
    )
    if query:
        banks_qs = banks_qs.filter(
            Q(title__icontains=query) | Q(items__question_text__icontains=query)
        ).distinct()

    ordered_qs = banks_qs.order_by("-is_official", "-use_count", "-created_at")
    paginator = Paginator(ordered_qs, 20)
    page_raw = (request.GET.get("page") or "1").strip()
    page_obj = None
    banks = []
    prev_page = None
    next_page = None
    if paginator.count > 0:
        try:
            page_obj = paginator.page(page_raw)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        except (TypeError, ValueError):
            page_obj = paginator.page(1)
        banks = list(page_obj.object_list)
        prev_page = page_obj.previous_page_number() if page_obj.has_previous() else None
        next_page = page_obj.next_page_number() if page_obj.has_next() else None

    return render(
        request,
        "seed_quiz/partials/bank_browse.html",
        {
            "classroom": classroom,
            "banks": banks,
            "scope": scope,
            "preset_type": preset_type,
            "preset_label": TOPIC_LABELS.get(preset_type, "주제"),
            "grade": grade,
            "grade_str": "all" if grade is None else str(grade),
            "q": query,
            "total_count": paginator.count,
            "page_obj": page_obj,
            "prev_page": prev_page,
            "next_page": next_page,
            "allow_inline_ai": bool(getattr(settings, "SEED_QUIZ_ALLOW_INLINE_AI", False)),
        },
    )


@login_required
@require_POST
def htmx_bank_select(request, classroom_id, bank_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    try:
        quiz_set = copy_bank_to_draft(
            bank_id=bank_id,
            classroom=classroom,
            teacher=request.user,
        )
    except SQQuizBank.DoesNotExist:
        return HttpResponse(
            '<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">'
            "선택한 퀴즈 세트를 찾을 수 없습니다.</div>",
            status=404,
        )
    except ValueError as e:
        return HttpResponse(
            f'<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">{e}</div>',
            status=400,
        )

    items = quiz_set.items.order_by("order_no")
    return render(
        request,
        "seed_quiz/partials/teacher_preview.html",
        {
            "classroom": classroom,
            "quiz_set": quiz_set,
            "items": items,
        },
    )


@login_required
@require_POST
def htmx_csv_upload(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    _clear_csv_error_report(request)
    csv_limits = _get_csv_upload_limits()
    csv_file = request.FILES.get("csv_file")
    if not csv_file:
        errors = ["CSV 파일을 선택해 주세요."]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="CSV_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="CSV 미리보기 실패: 파일 미선택",
            payload={"error_count": 1, "errors": _trim_errors(errors)},
        )
        token = _store_csv_error_report(request, ["CSV 파일을 선택해 주세요."])
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "errors": ["CSV 파일을 선택해 주세요."],
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    if int(getattr(csv_file, "size", 0) or 0) > csv_limits["max_file_bytes"]:
        errors = [
            "CSV 파일 용량이 제한을 초과했습니다. "
            f"(최대 {_humanize_bytes(csv_limits['max_file_bytes'])})"
        ]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="CSV_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="CSV 미리보기 실패: 파일 용량 초과",
            payload={
                "filename": getattr(csv_file, "name", ""),
                "file_size": int(getattr(csv_file, "size", 0) or 0),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(
            request,
            errors,
        )
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": [
                    "CSV 파일 용량이 제한을 초과했습니다. "
                    f"(최대 {_humanize_bytes(csv_limits['max_file_bytes'])})"
                ],
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    parsed_sets, errors = parse_csv_upload(
        csv_file,
        max_rows=csv_limits["max_rows"],
        max_sets=csv_limits["max_sets"],
    )
    if errors:
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="CSV_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="CSV 미리보기 실패: 파싱/검증 오류",
            payload={
                "filename": getattr(csv_file, "name", ""),
                "file_size": int(getattr(csv_file, "size", 0) or 0),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    if not parsed_sets:
        errors = ["저장할 수 있는 문제 세트가 없습니다."]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="CSV_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="CSV 미리보기 실패: 유효 세트 없음",
            payload={
                "filename": getattr(csv_file, "name", ""),
                "file_size": int(getattr(csv_file, "size", 0) or 0),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, ["저장할 수 있는 문제 세트가 없습니다."])
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": ["저장할 수 있는 문제 세트가 없습니다."],
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    _log_csv_event(
        classroom=classroom,
        teacher=request.user,
        code="CSV_UPLOAD_PREVIEW_READY",
        level="info",
        message=f"CSV 미리보기 성공: {len(parsed_sets)}세트",
        payload={
            "filename": getattr(csv_file, "name", ""),
            "file_size": int(getattr(csv_file, "size", 0) or 0),
            "set_count": len(parsed_sets),
            "row_count": sum(len(s.get("items", [])) for s in parsed_sets),
        },
    )
    token = uuid4().hex
    request.session[CSV_PREVIEW_TOKEN_KEY] = token
    request.session[CSV_PREVIEW_PAYLOAD_KEY] = parsed_sets
    request.session.modified = True
    _clear_csv_error_report(request)

    return render(
        request,
        "seed_quiz/partials/csv_upload_preview.html",
        {
            "classroom": classroom,
            "preview_token": token,
            "parsed_sets": parsed_sets,
            "set_count": len(parsed_sets),
            "preview_title": "CSV 미리보기 완료",
        },
    )


@login_required
@require_POST
def htmx_text_upload(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    _clear_csv_error_report(request)
    text_limits = _get_text_upload_limits()
    submit_mode = (request.POST.get("submit_mode") or "preview").strip().lower()
    raw_text = (request.POST.get("pasted_text") or "").strip()
    if not raw_text:
        errors = ["붙여넣기 내용이 비어 있습니다."]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="붙여넣기 미리보기 실패: 빈 입력",
            payload={"error_count": 1, "errors": _trim_errors(errors)},
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    if len(raw_text) > text_limits["max_chars"]:
        errors = [f"붙여넣기 글자 수가 제한을 초과했습니다. (최대 {text_limits['max_chars']:,}자)"]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="붙여넣기 미리보기 실패: 입력 길이 초과",
            payload={
                "char_count": len(raw_text),
                "max_chars": text_limits["max_chars"],
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    converted_csv_bytes, source_format, convert_errors = parse_pasted_text_to_csv_bytes(raw_text)
    if convert_errors or not converted_csv_bytes:
        errors = convert_errors or ["붙여넣기 텍스트를 처리할 수 없습니다."]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="붙여넣기 미리보기 실패: 파싱/변환 오류",
            payload={
                "format": source_format,
                "char_count": len(raw_text),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    parsed_sets, errors = parse_csv_upload(
        converted_csv_bytes,
        max_rows=text_limits["max_rows"],
        max_sets=text_limits["max_sets"],
    )
    if errors:
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="붙여넣기 미리보기 실패: CSV 규칙 검증 오류",
            payload={
                "format": source_format,
                "char_count": len(raw_text),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    if not parsed_sets:
        errors = ["저장할 수 있는 문제 세트가 없습니다."]
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_PREVIEW_FAILED",
            level="warn",
            message="붙여넣기 미리보기 실패: 유효 세트 없음",
            payload={
                "format": source_format,
                "char_count": len(raw_text),
                "error_count": len(errors),
                "errors": _trim_errors(errors),
            },
        )
        token = _store_csv_error_report(request, errors)
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": errors,
                "error_report_url": reverse(
                    "seed_quiz:download_csv_error_report",
                    kwargs={"classroom_id": classroom.id, "token": token},
                ),
            },
            status=400,
        )

    if submit_mode == "save":
        created_count, updated_count, shared_count = save_parsed_sets_to_bank(
            parsed_sets=parsed_sets,
            created_by=request.user,
            share_opt_in=True,
        )
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="TEXT_UPLOAD_SAVE_SUCCESS",
            level="info",
            message=f"붙여넣기 즉시 저장 완료: 생성 {created_count}, 갱신 {updated_count}",
            payload={
                "format": source_format,
                "char_count": len(raw_text),
                "set_count": len(parsed_sets),
                "row_count": sum(len(s.get("items", [])) for s in parsed_sets),
                "created_count": created_count,
                "updated_count": updated_count,
                "shared_count": shared_count,
            },
        )
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": created_count,
                "updated_count": updated_count,
                "shared_count": shared_count,
                "success_title": "문제 만들기가 완료되었습니다.",
                "errors": [],
            },
            status=200,
        )

    _log_csv_event(
        classroom=classroom,
        teacher=request.user,
        code="TEXT_UPLOAD_PREVIEW_READY",
        level="info",
        message=f"붙여넣기 미리보기 성공: {len(parsed_sets)}세트",
        payload={
            "format": source_format,
            "char_count": len(raw_text),
            "set_count": len(parsed_sets),
            "row_count": sum(len(s.get("items", [])) for s in parsed_sets),
        },
    )
    token = uuid4().hex
    request.session[CSV_PREVIEW_TOKEN_KEY] = token
    request.session[CSV_PREVIEW_PAYLOAD_KEY] = parsed_sets
    request.session.modified = True
    _clear_csv_error_report(request)

    format_labels = {
        "tsv": "TSV",
        "csv_text": "CSV 텍스트",
        "markdown_table": "마크다운 표",
        "json": "JSON",
        "stacked_lines": "세로 나열",
    }
    return render(
        request,
        "seed_quiz/partials/csv_upload_preview.html",
        {
            "classroom": classroom,
            "preview_token": token,
            "parsed_sets": parsed_sets,
            "set_count": len(parsed_sets),
            "preview_title": "붙여넣기 미리보기 완료",
            "preview_source": format_labels.get(source_format, source_format),
        },
    )


@login_required
@require_POST
def htmx_csv_confirm(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    token = (request.POST.get("preview_token") or "").strip()
    session_token = request.session.get(CSV_PREVIEW_TOKEN_KEY)
    parsed_sets = request.session.get(CSV_PREVIEW_PAYLOAD_KEY) or []

    if not token or token != session_token or not parsed_sets:
        _log_csv_event(
            classroom=classroom,
            teacher=request.user,
            code="CSV_UPLOAD_CONFIRM_FAILED",
            level="warn",
            message="CSV 저장 실패: 미리보기 세션 만료",
            payload={"error_count": 1, "errors": ["CSV 미리보기 세션이 만료되었습니다."]},
        )
        return render(
            request,
            "seed_quiz/partials/csv_upload_result.html",
            {
                "classroom": classroom,
                "created_count": 0,
                "updated_count": 0,
                "shared_count": 0,
                "errors": ["CSV 미리보기 세션이 만료되었습니다. 다시 업로드해 주세요."],
            },
            status=400,
        )

    share_opt_in = True
    created_count, updated_count, shared_count = save_parsed_sets_to_bank(
        parsed_sets=parsed_sets,
        created_by=request.user,
        share_opt_in=share_opt_in,
    )
    _log_csv_event(
        classroom=classroom,
        teacher=request.user,
        code="CSV_UPLOAD_CONFIRM_SUCCESS",
        level="info",
        message=f"CSV 저장 완료: 생성 {created_count}, 갱신 {updated_count}",
        payload={
            "set_count": len(parsed_sets),
            "created_count": created_count,
            "updated_count": updated_count,
            "shared_count": shared_count,
            "share_opt_in": bool(share_opt_in),
        },
    )

    request.session.pop(CSV_PREVIEW_TOKEN_KEY, None)
    request.session.pop(CSV_PREVIEW_PAYLOAD_KEY, None)
    request.session.modified = True

    return render(
        request,
        "seed_quiz/partials/csv_upload_result.html",
        {
            "classroom": classroom,
            "created_count": created_count,
            "updated_count": updated_count,
            "shared_count": shared_count,
            "errors": [],
        },
        status=200,
    )


@login_required
def htmx_csv_history(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    logs = (
        SQGenerationLog.objects.filter(
            code__in=[
                "CSV_UPLOAD_PREVIEW_READY",
                "CSV_UPLOAD_PREVIEW_FAILED",
                "CSV_UPLOAD_CONFIRM_SUCCESS",
                "CSV_UPLOAD_CONFIRM_FAILED",
            ],
            payload__classroom_id=str(classroom.id),
            payload__teacher_id=int(request.user.id),
        )
        .order_by("-created_at")[:20]
    )
    return render(
        request,
        "seed_quiz/partials/csv_upload_history.html",
        {"classroom": classroom, "history_logs": logs},
    )


@login_required
@require_POST
def htmx_rag_generate(request, classroom_id):
    if not bool(getattr(settings, "SEED_QUIZ_ALLOW_RAG", False)):
        return HttpResponse(
            '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
            "현재 맞춤 생성 기능이 비활성화되어 있습니다.</div>",
            status=403,
        )

    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    preset_type, grade = _parse_bank_filters(
        request.POST.get("preset_type", DEFAULT_TOPIC),
        request.POST.get("grade", "3"),
    )
    source_text = (request.POST.get("source_text") or "").strip()
    if not source_text:
        return HttpResponse(
            '<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">'
            "지문 텍스트를 입력해 주세요.</div>",
            status=400,
        )

    daily_limit = max(0, int(getattr(settings, "SEED_QUIZ_RAG_DAILY_LIMIT", 1)))
    if daily_limit == 0:
        return HttpResponse(
            '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
            "현재 맞춤 생성 기능이 비활성화되어 있습니다.</div>",
            status=403,
        )

    allowed, remaining = consume_rag_quota(classroom=classroom, teacher=request.user, daily_limit=daily_limit)
    if not allowed:
        return HttpResponse(
            '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
            "오늘의 맞춤 생성 횟수를 모두 사용했습니다. 내일 다시 시도해 주세요.</div>",
            status=429,
        )

    try:
        result = generate_bank_from_context_ai(
            preset_type=preset_type,
            grade=grade,
            source_text=source_text,
            created_by=request.user,
        )
        if isinstance(result, tuple):
            bank, from_cache = result
        else:
            bank = result
            from_cache = False
        quiz_set = copy_bank_to_draft(bank_id=bank.id, classroom=classroom, teacher=request.user)
    except Exception as e:
        refund_rag_quota(classroom=classroom, teacher=request.user)
        return HttpResponse(
            f'<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">'
            f"맞춤 생성 오류: {e}</div>",
            status=400,
        )

    items = quiz_set.items.order_by("order_no")
    return render(
        request,
        "seed_quiz/partials/teacher_preview.html",
        {
            "classroom": classroom,
            "quiz_set": quiz_set,
            "items": items,
            "rag_notice": (
                f"지문 기반 맞춤 생성 완료 (오늘 남은 횟수: {remaining})"
                if not from_cache
                else f"지문 기반 캐시 세트 재사용 (오늘 남은 횟수: {remaining})"
            ),
        },
    )


@login_required
@require_POST
def htmx_generate(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    preset_type, grade = _parse_bank_filters(
        request.POST.get("preset_type", DEFAULT_TOPIC),
        request.POST.get("grade", "3"),
    )

    try:
        quiz_set = generate_and_save_draft(classroom, preset_type, grade, request.user)
    except RuntimeError as e:
        return HttpResponse(
            f'<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">'
            f"퀴즈 생성 오류: {e}</div>"
        )

    items = quiz_set.items.order_by("order_no")
    return render(
        request,
        "seed_quiz/partials/teacher_preview.html",
        {
            "classroom": classroom,
            "quiz_set": quiz_set,
            "items": items,
        },
    )


@login_required
@require_POST
def htmx_publish(request, classroom_id, set_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    quiz_set = get_object_or_404(
        SQQuizSet, id=set_id, classroom=classroom, status="draft"
    )
    force_replace = (request.POST.get("force_replace") or "").lower() in {"1", "true", "yes", "on"}
    existing_published = (
        SQQuizSet.objects.filter(
            classroom=classroom,
            target_date=quiz_set.target_date,
            preset_type=quiz_set.preset_type,
            status="published",
        )
        .exclude(id=quiz_set.id)
        .order_by("-published_at", "-updated_at")
        .first()
    )
    if existing_published and not force_replace:
        return render(
            request,
            "seed_quiz/partials/teacher_publish_confirm_replace.html",
            {
                "classroom": classroom,
                "quiz_set": quiz_set,
                "existing_published": existing_published,
            },
        )

    with transaction.atomic():
        replaced_set_id = None
        if existing_published:
            replaced_set_id = existing_published.id

        # 기존 published → closed
        SQQuizSet.objects.filter(
            classroom=classroom,
            target_date=quiz_set.target_date,
            preset_type=quiz_set.preset_type,
            status="published",
        ).exclude(id=quiz_set.id).update(status="closed")

        quiz_set.status = "published"
        quiz_set.published_at = timezone.now()
        quiz_set.published_by = request.user
        quiz_set.save(update_fields=["status", "published_at", "published_by"])
        if replaced_set_id:
            SQGenerationLog.objects.create(
                level="info",
                code="QUIZ_PUBLISH_REPLACED",
                message=f"퀴즈 재배포: {quiz_set.title}",
                payload={
                    "classroom_id": str(classroom.id),
                    "teacher_id": int(request.user.id),
                    "target_date": str(quiz_set.target_date),
                    "preset_type": quiz_set.preset_type,
                    "new_set_id": str(quiz_set.id),
                    "replaced_set_id": str(replaced_set_id),
                },
            )
        else:
            SQGenerationLog.objects.create(
                level="info",
                code="QUIZ_PUBLISH_NEW",
                message=f"퀴즈 배포: {quiz_set.title}",
                payload={
                    "classroom_id": str(classroom.id),
                    "teacher_id": int(request.user.id),
                    "target_date": str(quiz_set.target_date),
                    "preset_type": quiz_set.preset_type,
                    "new_set_id": str(quiz_set.id),
                },
            )

    student_url = request.build_absolute_uri(
        reverse(
            "seed_quiz:student_gate",
            kwargs={"class_slug": classroom.slug},
        )
    )
    student_qr_data_url = _build_qr_data_url(student_url)
    return render(
        request,
        "seed_quiz/partials/teacher_published.html",
        {
            "classroom": classroom,
            "quiz_set": quiz_set,
            "student_url": student_url,
            "student_qr_data_url": student_qr_data_url,
            "rollback_restore_set_id": str(replaced_set_id) if replaced_set_id else "",
            "rollback_from_set_id": str(quiz_set.id),
        },
    )


@login_required
@require_POST
def htmx_publish_rollback(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    restore_set_id = (request.POST.get("restore_set_id") or "").strip()
    current_set_id = (request.POST.get("current_set_id") or "").strip()
    if not restore_set_id:
        return HttpResponse(
            '<div class="p-4 bg-red-50 border border-red-200 rounded-xl text-red-600 text-sm">'
            "되돌릴 세트 정보가 없습니다.</div>",
            status=400,
        )

    restore_set = get_object_or_404(SQQuizSet, id=restore_set_id, classroom=classroom)

    with transaction.atomic():
        current_set = (
            SQQuizSet.objects.select_for_update()
            .filter(
                classroom=classroom,
                target_date=restore_set.target_date,
                preset_type=restore_set.preset_type,
                status="published",
            )
            .first()
        )
        if not current_set:
            return HttpResponse(
                '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
                "현재 배포 중인 세트를 찾을 수 없습니다. 이미 변경되었을 수 있습니다.</div>",
                status=409,
            )
        if current_set_id and str(current_set.id) != current_set_id:
            return HttpResponse(
                '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
                "배포 상태가 변경되어 되돌리기를 중단했습니다. 화면을 새로고침해 주세요.</div>",
                status=409,
            )
        if current_set.id == restore_set.id:
            return HttpResponse(
                '<div class="p-4 bg-amber-50 border border-amber-200 rounded-xl text-amber-700 text-sm">'
                "이미 해당 세트가 배포 중입니다.</div>",
                status=409,
            )

        SQQuizSet.objects.filter(id=current_set.id).update(status="closed")
        restore_set.status = "published"
        restore_set.published_at = timezone.now()
        restore_set.published_by = request.user
        restore_set.save(update_fields=["status", "published_at", "published_by"])

        SQGenerationLog.objects.create(
            level="info",
            code="QUIZ_PUBLISH_ROLLBACK",
            message=f"퀴즈 롤백: {restore_set.title}",
            payload={
                "classroom_id": str(classroom.id),
                "teacher_id": int(request.user.id),
                "target_date": str(restore_set.target_date),
                "preset_type": restore_set.preset_type,
                "restore_set_id": str(restore_set.id),
                "from_set_id": str(current_set.id),
            },
        )

    student_url = request.build_absolute_uri(
        reverse(
            "seed_quiz:student_gate",
            kwargs={"class_slug": classroom.slug},
        )
    )
    student_qr_data_url = _build_qr_data_url(student_url)
    return render(
        request,
        "seed_quiz/partials/teacher_published.html",
        {
            "classroom": classroom,
            "quiz_set": restore_set,
            "student_url": student_url,
            "student_qr_data_url": student_qr_data_url,
            "rollback_done": True,
        },
    )


@login_required
def htmx_progress(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    quiz_set = (
        SQQuizSet.objects.filter(
            classroom=classroom,
            target_date=timezone.localdate(),
            status="published",
        )
        .prefetch_related("attempts")
        .first()
    )

    stats = {}
    attempt_rows = []
    item_total = 0
    if quiz_set:
        total = classroom.students.filter(is_active=True).count()
        attempts = quiz_set.attempts.select_related("student").annotate(answer_count=Count("answers"))
        stats = {
            "total": total,
            "started": attempts.count(),
            "submitted": attempts.filter(status__in=["submitted", "rewarded"]).count(),
            "perfect": attempts.filter(score=F("max_score")).count(),
        }
        attempt_rows = list(attempts.order_by("-updated_at", "student__number", "student__name")[:30])
        item_total = quiz_set.items.count()
    return render(
        request,
        "seed_quiz/partials/teacher_progress.html",
        {
            "quiz_set": quiz_set,
            "stats": stats,
            "attempt_rows": attempt_rows,
            "item_total": item_total,
        },
    )


@login_required
def htmx_topic_summary(request, classroom_id):
    classroom = get_object_or_404(HSClassroom, id=classroom_id, teacher=request.user)
    counts_by_topic = {
        row["preset_type"]: row
        for row in SQQuizBank.objects.filter(is_active=True)
        .values("preset_type")
        .annotate(
            total_count=Count("id"),
            official_count=Count("id", filter=Q(is_official=True)),
            public_count=Count("id", filter=Q(is_public=True)),
            last_created=Max("created_at"),
        )
    }
    last_used_by_topic = {
        row["bank_source__preset_type"]: row["last_used"]
        for row in SQQuizSet.objects.filter(classroom=classroom, bank_source__isnull=False)
        .values("bank_source__preset_type")
        .annotate(last_used=Max("target_date"))
    }

    topics = []
    for topic_key, topic_label in SQQuizSet.PRESET_CHOICES:
        row = counts_by_topic.get(topic_key, {})
        topics.append(
            {
                "key": topic_key,
                "label": topic_label,
                "total_count": row.get("total_count", 0),
                "official_count": row.get("official_count", 0),
                "public_count": row.get("public_count", 0),
                "last_used": last_used_by_topic.get(topic_key),
            }
        )

    return render(
        request,
        "seed_quiz/partials/topic_summary.html",
        {
            "classroom": classroom,
            "topics": topics,
        },
    )


# ---------------------------------------------------------------------------
# 학생 뷰
# ---------------------------------------------------------------------------

def student_gate(request, class_slug):
    classroom = get_object_or_404(HSClassroom, slug=class_slug, is_active=True)
    quiz_set = get_today_published_set(classroom)
    error = request.session.pop("sq_gate_error", None)
    return render(
        request,
        "seed_quiz/student_gate.html",
        {
            "classroom": classroom,
            "quiz_set": quiz_set,
            "error": error,
        },
    )


@require_POST
def student_start(request, class_slug):
    classroom = get_object_or_404(HSClassroom, slug=class_slug, is_active=True)
    quiz_set = get_today_published_set(classroom)
    if not quiz_set:
        return redirect("seed_quiz:student_gate", class_slug=class_slug)

    number_raw = request.POST.get("number", "").strip()
    name = request.POST.get("name", "").strip()

    if not number_raw.isdigit() or not name:
        request.session["sq_gate_error"] = "번호와 이름을 모두 입력해 주세요."
        return redirect("seed_quiz:student_gate", class_slug=class_slug)

    student = HSStudent.objects.filter(
        classroom=classroom,
        number=int(number_raw),
        name=name,
        is_active=True,
    ).first()

    if not student:
        if int(number_raw) == 0 and name == "선생님":
            student, _ = HSStudent.objects.get_or_create(
                classroom=classroom,
                number=0,
                defaults={"name": "선생님", "is_active": True}
            )
            if student.name != "선생님" or not student.is_active:
                student.name = "선생님"
                student.is_active = True
                student.save()
        else:
            request.session["sq_gate_error"] = "번호와 이름을 다시 확인해 주세요."
            return redirect("seed_quiz:student_gate", class_slug=class_slug)

    attempt = find_or_create_attempt(quiz_set, student)
    set_session(request, classroom, student, attempt)

    if attempt.status in ("submitted", "rewarded"):
        return redirect("seed_quiz:htmx_play_result")

    return redirect("seed_quiz:student_play")


def student_play_shell(request):
    attempt_id = request.session.get("sq_attempt_id")
    if not attempt_id:
        return redirect("/")
    return render(request, "seed_quiz/student_play.html")


def htmx_play_current(request):
    attempt_id = request.session.get("sq_attempt_id")
    if not attempt_id:
        return HttpResponse(status=403)

    attempt = get_object_or_404(SQAttempt, id=attempt_id)

    if attempt.status in ("submitted", "rewarded"):
        return _render_result(request, attempt)

    item = get_current_item(attempt)
    if not item:
        # 모든 문항 답변됨 → 채점
        return _do_finish(request, attempt)

    answered_count = attempt.answers.count()
    total_count = attempt.quiz_set.items.count()
    return render(
        request,
        "seed_quiz/partials/play_item.html",
        {
            "item": item,
            "item_no": answered_count + 1,
            "total": total_count,
        },
    )


@require_POST
def htmx_play_answer(request):
    attempt_id = request.session.get("sq_attempt_id")
    if not attempt_id:
        return HttpResponse(status=403)

    attempt = get_object_or_404(SQAttempt, id=attempt_id)

    if attempt.status in ("submitted", "rewarded"):
        return _render_result(request, attempt)

    item_id = request.POST.get("item_id")
    selected_raw = request.POST.get("selected_index", "")

    if not selected_raw.lstrip("-").isdigit():
        return HttpResponse(status=400)
    selected_index = int(selected_raw)
    if selected_index not in [0, 1, 2, 3]:
        return HttpResponse(status=400)

    # 해당 item이 현재 퀴즈 세트 소속이고, 현재 순서여야 함
    item = get_object_or_404(SQQuizItem, id=item_id, quiz_set=attempt.quiz_set)

    # 순서 검증: 현재 풀어야 할 문항만 허용
    expected_item = get_current_item(attempt)
    if expected_item is None or str(expected_item.id) != str(item.id):
        # 이미 완료되었거나 순서가 다름 → 현재 상태 기준으로 다음 화면 반환
        next_item = get_current_item(attempt)
        if next_item is None:
            return _do_finish(request, attempt)
        answered_count = attempt.answers.count()
        total_count = attempt.quiz_set.items.count()
        return render(
            request,
            "seed_quiz/partials/play_item.html",
            {"item": next_item, "item_no": answered_count + 1, "total": total_count},
        )

    # 답변 저장 (이미 답변한 경우 get_or_create로 무시)
    SQAttempt.objects.select_for_update().filter(id=attempt.id)  # 잠금
    with transaction.atomic():
        attempt_locked = SQAttempt.objects.select_for_update().get(id=attempt.id)
        if attempt_locked.status in ("submitted", "rewarded"):
            return _render_result(request, attempt_locked)

        from seed_quiz.models import SQAttemptAnswer
        SQAttemptAnswer.objects.get_or_create(
            attempt=attempt_locked,
            item=item,
            defaults={
                "selected_index": selected_index,
                "is_correct": item.correct_index == selected_index,
            },
        )

    # 다음 문항 확인
    attempt.refresh_from_db()
    next_item = get_current_item(attempt)
    if next_item is None:
        return _do_finish(request, attempt)

    answered_count = attempt.answers.count()
    total_count = attempt.quiz_set.items.count()
    return render(
        request,
        "seed_quiz/partials/play_item.html",
        {"item": next_item, "item_no": answered_count + 1, "total": total_count},
    )


def htmx_play_result(request):
    attempt_id = request.session.get("sq_attempt_id")
    if not attempt_id:
        return HttpResponse(status=403)
    attempt = get_object_or_404(SQAttempt, id=attempt_id)
    return _render_result(request, attempt)


def _do_finish(request, attempt: SQAttempt) -> HttpResponse:
    """채점 + 보상 처리 후 결과 partial 반환."""
    answers = {
        a.item.order_no: a.selected_index
        for a in attempt.answers.select_related("item").all()
    }
    attempt = submit_and_reward(attempt_id=attempt.id, answers=answers)
    return _render_result(request, attempt)


def _render_result(request, attempt: SQAttempt) -> HttpResponse:
    attempt = SQAttempt.objects.select_related("quiz_set__classroom").get(id=attempt.id)
    answers_by_order = {
        a.item.order_no: a
        for a in attempt.answers.select_related("item")
    }
    items_with_answers = []
    for item in attempt.quiz_set.items.order_by("order_no"):
        ans = answers_by_order.get(item.order_no)
        choices = item.choices or []
        selected_choice = choices[ans.selected_index] if ans and 0 <= ans.selected_index < len(choices) else ""
        items_with_answers.append({
            "item": item,
            "answer": ans,
            "correct_choice": choices[item.correct_index] if 0 <= item.correct_index < len(choices) else "",
            "selected_choice": selected_choice,
        })
    return render(
        request,
        "seed_quiz/partials/play_result.html",
        {
            "attempt": attempt,
            "items_with_answers": items_with_answers,
        },
    )
