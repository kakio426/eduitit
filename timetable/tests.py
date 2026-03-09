from io import BytesIO
from datetime import datetime
from collections import defaultdict

import openpyxl
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.utils import timezone

from core.models import UserProfile
from reservations.models import RecurringSchedule, School, SpecialRoom
from timetable.models import TimetableSyncLog
from timetable.services import (
    REQUIRED_SHEETS,
    build_template_workbook,
    generate_timetable_schedule,
)


class TimetablePhaseOneTest(TestCase):
    def setUp(self):
        self.client = Client()

    def test_main_page_loads(self):
        response = self.client.get("/timetable/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "전담 시간표·특별실 배치 도우미")

    def test_main_page_uses_teacher_first_sections(self):
        response = self.client.get("/timetable/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "양식 받기")
        self.assertContains(response, "파일 올리기")
        self.assertContains(response, "배치 조건 메모")
        self.assertNotContains(response, "1단계: 입력 양식 내려받기")
        self.assertNotContains(response, "2단계: 작성 파일 점검하기")

    def test_template_download_has_required_sheets(self):
        response = self.client.get("/timetable/template/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        for sheet_name in REQUIRED_SHEETS:
            self.assertIn(sheet_name, wb.sheetnames)

    def test_upload_valid_template_passes_check(self):
        data = build_template_workbook()
        upload = SimpleUploadedFile(
            "timetable_template.xlsx",
            data,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post("/timetable/", {"excel_file": upload})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["check_result"]["is_valid"])
        self.assertIsNotNone(response.context["generated_result"])
        self.assertTrue(response.context["generated_result"]["is_success"])
        self.assertEqual(response.context["generated_result"]["summary"]["unplaced_count"], 0)

    def test_upload_missing_sheet_fails_check(self):
        wb = openpyxl.load_workbook(BytesIO(build_template_workbook()))
        del wb["전담배정표"]
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        upload = SimpleUploadedFile(
            "missing_sheet.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post("/timetable/", {"excel_file": upload})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["check_result"]["is_valid"])
        self.assertIn("전담배정표", " ".join(response.context["check_result"]["errors"]))
        self.assertIsNone(response.context["generated_result"])

    def test_upload_without_optional_sheet_still_passes_check(self):
        wb = openpyxl.load_workbook(BytesIO(build_template_workbook()))
        del wb["배치조건"]
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        upload = SimpleUploadedFile(
            "without_optional_sheet.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post("/timetable/", {"excel_file": upload})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["check_result"]["is_valid"])
        self.assertIsNotNone(response.context["generated_result"])

    def test_apply_to_reservation_when_school_selected(self):
        user = User.objects.create_user(
            username="teacher1",
            password="pw-123456",
            email="teacher1@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사"
        profile.save(update_fields=["nickname"])
        school = School.objects.create(name="테스트학교", slug="test-school", owner=user)

        self.client.login(username="teacher1", password="pw-123456")

        wb = openpyxl.load_workbook(BytesIO(build_template_workbook()))
        ws = wb["특별실설정"]
        ws["D2"] = "바로반영"
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        upload = SimpleUploadedFile(
            "apply_sync.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            "/timetable/",
            {"excel_file": upload, "reservation_school_slug": school.slug},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["generated_result"]["is_success"])
        self.assertIsNotNone(response.context["integration_result"])
        self.assertGreaterEqual(len(response.context["recent_sync_logs"]), 1)
        self.assertGreaterEqual(response.context["integration_result"]["applied_count"], 1)
        self.assertGreaterEqual(RecurringSchedule.objects.filter(room__school=school).count(), 1)
        self.assertEqual(TimetableSyncLog.objects.filter(user=user).count(), 1)
        latest = TimetableSyncLog.objects.filter(user=user).first()
        self.assertEqual(latest.sync_mode, "direct")
        self.assertEqual(latest.school_slug, school.slug)

    def test_preview_sync_manual_button_flow(self):
        user = User.objects.create_user(
            username="teacher2",
            password="pw-123456",
            email="teacher2@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사2"
        profile.save(update_fields=["nickname"])
        school = School.objects.create(name="미리보기학교", slug="preview-school", owner=user)
        self.client.login(username="teacher2", password="pw-123456")

        upload = SimpleUploadedFile(
            "preview_sync.xlsx",
            build_template_workbook(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        first = self.client.post(
            "/timetable/",
            {"excel_file": upload, "reservation_school_slug": school.slug, "action": "generate"},
        )
        self.assertEqual(first.status_code, 200)
        preview_info = first.context["preview_apply_info"]
        self.assertIsNotNone(preview_info)
        self.assertGreaterEqual(preview_info["preview_count"], 1)

        second = self.client.post(
            "/timetable/",
            {
                "action": "apply_preview",
                "sync_token": preview_info["token"],
                "school_slug": school.slug,
            },
        )
        self.assertEqual(second.status_code, 200)
        self.assertIsNotNone(second.context["integration_result"])
        self.assertGreaterEqual(second.context["integration_result"]["applied_count"], 1)
        self.assertGreaterEqual(RecurringSchedule.objects.filter(room__school=school).count(), 1)
        self.assertTrue(
            TimetableSyncLog.objects.filter(user=user, sync_mode="preview_manual").exists()
        )

    def test_overwrite_existing_is_applied_when_checked(self):
        user = User.objects.create_user(
            username="teacher3",
            password="pw-123456",
            email="teacher3@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사3"
        profile.save(update_fields=["nickname"])
        school = School.objects.create(name="덮어쓰기학교", slug="overwrite-school", owner=user)
        room = SpecialRoom.objects.create(school=school, name="과학실", icon="🔬")
        existing = RecurringSchedule.objects.create(
            room=room,
            day_of_week=1,  # 화
            period=2,
            name="기존 고정수업",
        )

        self.client.login(username="teacher3", password="pw-123456")

        wb = openpyxl.load_workbook(BytesIO(build_template_workbook()))
        ws_special = wb["특별실설정"]
        ws_special["D2"] = "바로반영"
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        upload_without_overwrite = SimpleUploadedFile(
            "without_overwrite.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        first = self.client.post(
            "/timetable/",
            {
                "excel_file": upload_without_overwrite,
                "reservation_school_slug": school.slug,
                "action": "generate",
            },
        )
        self.assertEqual(first.status_code, 200)
        existing.refresh_from_db()
        self.assertEqual(existing.name, "기존 고정수업")
        self.assertGreaterEqual(first.context["integration_result"]["conflict_count"], 1)

        upload_with_overwrite = SimpleUploadedFile(
            "with_overwrite.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        second = self.client.post(
            "/timetable/",
            {
                "excel_file": upload_with_overwrite,
                "reservation_school_slug": school.slug,
                "overwrite_existing": "on",
                "action": "generate",
            },
        )
        self.assertEqual(second.status_code, 200)
        existing.refresh_from_db()
        self.assertNotEqual(existing.name, "기존 고정수업")
        self.assertGreaterEqual(second.context["integration_result"]["updated_count"], 1)
        self.assertTrue(
            TimetableSyncLog.objects.filter(
                user=user,
                sync_mode="direct",
                overwrite_existing=True,
            ).exists()
        )

    def test_download_sync_logs_csv_requires_login(self):
        response = self.client.get("/timetable/sync-logs.csv")
        self.assertEqual(response.status_code, 302)

    def test_download_sync_logs_csv_contains_user_logs(self):
        user = User.objects.create_user(
            username="teacher4",
            password="pw-123456",
            email="teacher4@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사4"
        profile.save(update_fields=["nickname"])

        TimetableSyncLog.objects.create(
            user=user,
            school_slug="csv-school",
            school_name="CSV학교",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=3,
            updated_count=0,
            skipped_count=1,
            conflict_count=0,
            room_created_count=1,
            summary_text="새로 반영 3건",
            payload={},
        )

        self.client.login(username="teacher4", password="pw-123456")
        response = self.client.get("/timetable/sync-logs.csv")
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        content = response.content.decode("utf-8-sig")
        self.assertIn("실행시각", content)
        self.assertIn("CSV학교", content)
        self.assertIn("바로반영", content)

    def test_download_sync_logs_csv_filters_school_and_period(self):
        user = User.objects.create_user(
            username="teacher5",
            password="pw-123456",
            email="teacher5@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사5"
        profile.save(update_fields=["nickname"])

        in_range = TimetableSyncLog.objects.create(
            user=user,
            school_slug="school-a",
            school_name="가람초",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=2,
            updated_count=0,
            skipped_count=0,
            conflict_count=0,
            room_created_count=0,
            summary_text="포함건",
            payload={},
        )
        TimetableSyncLog.objects.filter(pk=in_range.pk).update(
            created_at=timezone.make_aware(datetime(2026, 1, 10, 9, 0, 0))
        )

        out_of_range = TimetableSyncLog.objects.create(
            user=user,
            school_slug="school-a",
            school_name="가람초",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=1,
            updated_count=0,
            skipped_count=0,
            conflict_count=0,
            room_created_count=0,
            summary_text="기간제외건",
            payload={},
        )
        TimetableSyncLog.objects.filter(pk=out_of_range.pk).update(
            created_at=timezone.make_aware(datetime(2026, 2, 10, 9, 0, 0))
        )

        other_school = TimetableSyncLog.objects.create(
            user=user,
            school_slug="school-b",
            school_name="나래초",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=1,
            updated_count=0,
            skipped_count=0,
            conflict_count=0,
            room_created_count=0,
            summary_text="학교제외건",
            payload={},
        )
        TimetableSyncLog.objects.filter(pk=other_school.pk).update(
            created_at=timezone.make_aware(datetime(2026, 1, 12, 9, 0, 0))
        )

        self.client.login(username="teacher5", password="pw-123456")
        response = self.client.get(
            "/timetable/sync-logs.csv",
            {
                "log_school_slug": "school-a",
                "log_date_from": "2026-01-01",
                "log_date_to": "2026-01-31",
            },
        )
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8-sig")
        self.assertIn("포함건", content)
        self.assertNotIn("기간제외건", content)
        self.assertNotIn("학교제외건", content)

    def test_main_recent_sync_logs_filters_and_csv_query(self):
        user = User.objects.create_user(
            username="teacher6",
            password="pw-123456",
            email="teacher6@example.com",
        )
        profile = UserProfile.objects.get(user=user)
        profile.nickname = "담당교사6"
        profile.save(update_fields=["nickname"])

        included = TimetableSyncLog.objects.create(
            user=user,
            school_slug="school-c",
            school_name="다온초",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=1,
            updated_count=0,
            skipped_count=0,
            conflict_count=0,
            room_created_count=0,
            summary_text="목록포함건",
            payload={},
        )
        TimetableSyncLog.objects.filter(pk=included.pk).update(
            created_at=timezone.make_aware(datetime(2026, 1, 5, 10, 0, 0))
        )

        excluded = TimetableSyncLog.objects.create(
            user=user,
            school_slug="school-c",
            school_name="다온초",
            sync_mode="direct",
            sync_options_text="바로반영",
            overwrite_existing=False,
            status="success",
            applied_count=1,
            updated_count=0,
            skipped_count=0,
            conflict_count=0,
            room_created_count=0,
            summary_text="목록제외건",
            payload={},
        )
        TimetableSyncLog.objects.filter(pk=excluded.pk).update(
            created_at=timezone.make_aware(datetime(2026, 2, 5, 10, 0, 0))
        )

        self.client.login(username="teacher6", password="pw-123456")
        response = self.client.get(
            "/timetable/",
            {
                "log_school_slug": "school-c",
                "log_date_from": "2026-01-01",
                "log_date_to": "2026-01-31",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["recent_sync_logs"]), 1)
        self.assertEqual(response.context["recent_sync_logs"][0].summary_text, "목록포함건")
        self.assertIn("log_school_slug=school-c", response.context["csv_download_query"])


class TimetableSchedulingConstraintsTest(TestCase):
    def _build_result(
        self,
        days,
        slot_labels,
        teachers,
        assignments,
        conditions=None,
        manual_rules=None,
    ):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_settings = wb.create_sheet("기본설정")
        ws_settings.append(["항목", "값"])
        ws_settings.append(["운영요일", ",".join(days)])
        ws_settings.append(["기본교시수", str(len(slot_labels))])
        ws_settings.append(["시간칸형식", "기본"])
        ws_settings.append(["시간칸목록", ",".join(slot_labels)])

        class_rows = sorted(
            {(row[3], row[4]) for row in assignments},
            key=lambda item: (item[0], item[1]),
        )
        ws_classes = wb.create_sheet("학급목록")
        ws_classes.append(["학년", "반", "사용여부"])
        for grade, class_no in class_rows:
            ws_classes.append([grade, class_no, "Y"])

        ws_teachers = wb.create_sheet("전담선생님목록")
        ws_teachers.append(["선생님코드", "이름", "담당교과", "하루최대수업칸"])
        for row in teachers:
            ws_teachers.append(list(row))

        ws_assignments = wb.create_sheet("전담배정표")
        ws_assignments.append(
            ["배정번호", "선생님코드", "교과", "학년", "반", "주당시수", "특별실처리", "특별실명"]
        )
        for row in assignments:
            ws_assignments.append(list(row))

        if conditions:
            ws_conditions = wb.create_sheet("배치조건")
            ws_conditions.append(["조건이름", "적용대상", "대상값", "배치방법", "세부값", "중요도"])
            for row in conditions:
                ws_conditions.append(list(row))

        if manual_rules:
            ws_manual = wb.create_sheet("수동고정")
            ws_manual.append(["배정번호", "요일", "시간칸", "처리", "메모"])
            for row in manual_rules:
                ws_manual.append(list(row))

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return generate_timetable_schedule(output)

    def _class_events(self, result, class_id):
        day_index = {day: idx for idx, day in enumerate(result["days"])}
        slot_index = {slot: idx for idx, slot in enumerate(result["slot_labels"])}
        events = []
        for table in result["class_tables"]:
            if table["class_id"] != class_id:
                continue
            for row in table["rows"]:
                slot_label = row["slot_label"]
                for day, text in zip(result["days"], row["cells"]):
                    if text:
                        events.append((day, slot_label, day_index[day], slot_index[slot_label]))
        events.sort(key=lambda item: (item[2], item[3]))
        return events

    def test_default_rule_blocks_same_subject_twice_in_one_day(self):
        result = self._build_result(
            days=["월", "화"],
            slot_labels=["1", "2"],
            teachers=[("T1", "영어전담", "영어", 4)],
            assignments=[("A1", "T1", "영어", 3, 1, 3, "해당없음", "")],
        )

        self.assertFalse(result["is_success"])
        self.assertGreaterEqual(result["summary"]["unplaced_count"], 1)

        events = self._class_events(result, "3-1")
        self.assertEqual(len(events), 2)
        by_day = defaultdict(int)
        for day, *_ in events:
            by_day[day] += 1
        self.assertTrue(all(count <= 1 for count in by_day.values()))

    def test_split_pattern_allows_only_requested_consecutive_double_block(self):
        result = self._build_result(
            days=["월", "화", "수"],
            slot_labels=["1", "2", "3"],
            teachers=[("T2", "과학전담", "과학", 6)],
            assignments=[("S1", "T2", "과학", 3, 1, 3, "해당없음", "")],
            conditions=[("과학 2+1", "교과", "과학", "나눠배치", "2+1", "반드시")],
        )

        self.assertTrue(result["is_success"])
        events = self._class_events(result, "3-1")
        self.assertEqual(len(events), 3)

        by_day_slots = defaultdict(list)
        for day, slot_label, _, slot_idx in events:
            by_day_slots[day].append((slot_label, slot_idx))

        self.assertTrue(all(len(items) <= 2 for items in by_day_slots.values()))

        has_consecutive_double = False
        for day in by_day_slots:
            ordered = sorted(by_day_slots[day], key=lambda item: item[1])
            if len(ordered) == 2 and ordered[1][1] - ordered[0][1] == 1:
                has_consecutive_double = True
        self.assertTrue(has_consecutive_double)

    def test_round_robin_prevents_second_turn_before_all_first_turns(self):
        result = self._build_result(
            days=["월", "화", "수", "목", "금"],
            slot_labels=["1", "2"],
            teachers=[("T1", "영어전담", "영어", 8)],
            assignments=[
                ("A1", "T1", "영어", 3, 1, 2, "해당없음", ""),
                ("A2", "T1", "영어", 3, 2, 2, "해당없음", ""),
                ("A3", "T1", "영어", 3, 3, 2, "해당없음", ""),
                ("A4", "T1", "영어", 3, 4, 2, "해당없음", ""),
            ],
        )

        self.assertTrue(result["is_success"])
        first_positions = []
        second_positions = []

        for class_id in ["3-1", "3-2", "3-3", "3-4"]:
            events = self._class_events(result, class_id)
            self.assertEqual(len(events), 2)
            first_positions.append((events[0][2], events[0][3]))
            second_positions.append((events[1][2], events[1][3]))

        self.assertGreater(min(second_positions), max(first_positions))

    def test_round_robin_default_on_blocks_early_repeat_when_later_classes_blocked(self):
        blocked = [
            ("A3", "월", "1", "배치금지", ""),
            ("A3", "월", "2", "배치금지", ""),
            ("A3", "화", "1", "배치금지", ""),
            ("A3", "화", "2", "배치금지", ""),
            ("A4", "월", "1", "배치금지", ""),
            ("A4", "월", "2", "배치금지", ""),
            ("A4", "화", "1", "배치금지", ""),
            ("A4", "화", "2", "배치금지", ""),
        ]

        result = self._build_result(
            days=["월", "화", "수", "목"],
            slot_labels=["1", "2"],
            teachers=[("T1", "영어전담", "영어", 8)],
            assignments=[
                ("A1", "T1", "영어", 3, 1, 2, "해당없음", ""),
                ("A2", "T1", "영어", 3, 2, 2, "해당없음", ""),
                ("A3", "T1", "영어", 3, 3, 2, "해당없음", ""),
                ("A4", "T1", "영어", 3, 4, 2, "해당없음", ""),
            ],
            manual_rules=blocked,
        )

        self.assertFalse(result["is_success"])
        self.assertGreater(result["summary"]["unplaced_count"], 0)

    def test_round_robin_can_be_disabled_with_condition_off(self):
        blocked = [
            ("A3", "월", "1", "배치금지", ""),
            ("A3", "월", "2", "배치금지", ""),
            ("A3", "화", "1", "배치금지", ""),
            ("A3", "화", "2", "배치금지", ""),
            ("A4", "월", "1", "배치금지", ""),
            ("A4", "월", "2", "배치금지", ""),
            ("A4", "화", "1", "배치금지", ""),
            ("A4", "화", "2", "배치금지", ""),
        ]

        result = self._build_result(
            days=["월", "화", "수", "목"],
            slot_labels=["1", "2"],
            teachers=[("T1", "영어전담", "영어", 8)],
            assignments=[
                ("A1", "T1", "영어", 3, 1, 2, "해당없음", ""),
                ("A2", "T1", "영어", 3, 2, 2, "해당없음", ""),
                ("A3", "T1", "영어", 3, 3, 2, "해당없음", ""),
                ("A4", "T1", "영어", 3, 4, 2, "해당없음", ""),
            ],
            conditions=[("영어 순환 해제", "교과", "영어", "순환배치", "OFF", "반드시")],
            manual_rules=blocked,
        )

        self.assertTrue(result["is_success"])
        self.assertEqual(result["summary"]["unplaced_count"], 0)
