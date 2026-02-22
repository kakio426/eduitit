from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


REQUIRED_SHEETS = {
    "기본설정": ["항목", "값"],
    "학급목록": ["학년", "반", "사용여부"],
    "전담선생님목록": ["선생님코드", "이름", "담당교과", "하루최대수업칸"],
    "전담배정표": ["배정번호", "선생님코드", "교과", "학년", "반", "주당시수", "특별실처리", "특별실명"],
}

OPTIONAL_SHEETS = {
    "배치불가시간": ["선생님코드", "요일", "시간칸", "사유"],
    "배치조건": ["조건이름", "적용대상", "대상값", "배치방법", "세부값", "중요도"],
    "수동고정": ["배정번호", "요일", "시간칸", "처리", "메모"],
    "특별실설정": ["특별실명", "대상교과", "운영방식", "연동선택"],
}

ALL_TEMPLATE_SHEETS = {**REQUIRED_SHEETS, **OPTIONAL_SHEETS}

REQUIRED_SETTING_KEYS = {"운영요일", "기본교시수", "시간칸형식"}

TEMPLATE_GUIDE_ROWS = {
    "기본설정": [
        ["※작성안내", "운영요일(예: 월,화,수,목,금) / 기본교시수(예: 6) / 시간칸목록(예: 1,2,3,4,5,6)"],
        ["※학년별사용시간칸", "선택 입력: 1=1,2,3,4;2=1,2,3,4,5 형태(세미콜론으로 학년 구분)"],
    ],
    "학급목록": [
        ["※작성안내: 학년 숫자(예: 3)", "반 숫자(예: 1)", "사용여부 Y/N (N이면 배치 제외)"],
    ],
    "전담선생님목록": [
        ["※작성안내: 선생님코드는 고유값(예: T001)", "이름", "담당교과는 참고용(복수 가능: 영어,체육)", "하루최대수업칸 숫자"],
        ["※운영팁: 같은 선생님코드로 전담배정표 여러 행 배정 가능", "", "", ""],
    ],
    "전담배정표": [
        ["※작성안내: 배정번호는 고유값(예: A001)", "선생님코드(T001)", "교과명", "학년 숫자", "반 숫자", "주당시수", "자동배치/예약연동/해당없음", "해당없음이면 빈칸, 아니면 특별실명 입력"],
        ["※예시: 전담 1명이 3학년/5학년 동시 담당이면 행을 2개로 작성", "같은 선생님코드 반복 입력", "", "", "", "", "", ""],
    ],
    "배치불가시간": [
        ["※작성안내: 선생님코드(T001)", "요일(월~금)", "시간칸(기본설정의 시간칸과 동일)", "사유 메모(선택)"],
    ],
    "배치조건": [
        ["※작성안내", "적용대상: 교과/선생님/학년", "대상값 예: 영어, T001, 3학년", "배치방법: 피하기/나눠배치", "피하기: 월,수 / 나눠배치: 2+1", "중요도: 반드시/권장"],
    ],
    "수동고정": [
        ["※작성안내: 배정번호(A001)", "요일(월~금)", "시간칸(예: 2)", "처리: 고정/배치금지", "메모(선택)"],
    ],
    "특별실설정": [
        ["※작성안내: 특별실명(예: 과학실)", "대상교과(예: 과학)", "운영방식: 자동배치/예약연동", "연동선택: 연동안함/미리보기/바로반영"],
    ],
}

ALLOWED_CHOICES = {
    ("전담배정표", "특별실처리"): {"자동배치", "예약연동", "해당없음"},
    ("배치조건", "중요도"): {"반드시", "권장"},
    ("특별실설정", "운영방식"): {"자동배치", "예약연동"},
    ("특별실설정", "연동선택"): {"연동안함", "미리보기", "바로반영"},
    ("수동고정", "처리"): {"고정", "배치금지"},
}


def build_template_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "기본설정", ALL_TEMPLATE_SHEETS["기본설정"], [
        ["운영요일", "월,화,수,목,금"],
        ["기본교시수", "6"],
        ["시간칸형식", "기본"],
        ["시간칸목록", "1,2,3,4,5,6"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("기본설정"))
    _add_sheet(wb, "학급목록", ALL_TEMPLATE_SHEETS["학급목록"], [
        [3, 1, "Y"],
        [3, 2, "Y"],
        [4, 1, "Y"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("학급목록"))
    _add_sheet(wb, "전담선생님목록", ALL_TEMPLATE_SHEETS["전담선생님목록"], [
        ["T001", "김과학", "과학", 5],
        ["T002", "이영어", "영어", 6],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("전담선생님목록"))
    _add_sheet(wb, "전담배정표", ALL_TEMPLATE_SHEETS["전담배정표"], [
        ["A001", "T001", "과학", 3, 1, 3, "자동배치", "과학실"],
        ["A002", "T002", "영어", 3, 1, 2, "해당없음", ""],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("전담배정표"))
    _add_sheet(wb, "배치불가시간", ALL_TEMPLATE_SHEETS["배치불가시간"], [
        ["T001", "월", "1", "회의"],
        ["T002", "수", "5", "학년협의"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("배치불가시간"))
    _add_sheet(wb, "배치조건", ALL_TEMPLATE_SHEETS["배치조건"], [
        ["과학 3시간 분리", "교과", "과학", "나눠배치", "2+1", "반드시"],
        ["영어 월요일 피하기", "교과", "영어", "피하기", "월", "권장"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("배치조건"))
    _add_sheet(wb, "수동고정", ALL_TEMPLATE_SHEETS["수동고정"], [
        ["A001", "화", "2", "고정", "교내 행사 연계"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("수동고정"))
    _add_sheet(wb, "특별실설정", ALL_TEMPLATE_SHEETS["특별실설정"], [
        ["과학실", "과학", "자동배치", "미리보기"],
        ["음악실", "음악", "예약연동", "미리보기"],
    ], guide_rows=TEMPLATE_GUIDE_ROWS.get("특별실설정"))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def validate_timetable_workbook(file_obj):
    result = {
        "is_valid": True,
        "errors": [],
        "warnings": [],
        "sheet_stats": [],
        "summary": {
            "필수시트수": len(REQUIRED_SHEETS),
            "확인된시트수": 0,
            "총입력행수": 0,
        },
    }

    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:
        result["is_valid"] = False
        result["errors"].append(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {exc}")
        return result

    existing_sheets = wb.sheetnames
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in existing_sheets]
    extra_sheets = [name for name in existing_sheets if name not in ALL_TEMPLATE_SHEETS]

    if missing_sheets:
        result["is_valid"] = False
        result["errors"].append(f"필수 시트가 빠져 있습니다: {', '.join(missing_sheets)}")
    if extra_sheets:
        result["warnings"].append(f"사용하지 않는 시트가 포함되어 있습니다: {', '.join(extra_sheets)}")

    for sheet_name, required_headers in ALL_TEMPLATE_SHEETS.items():
        if sheet_name not in existing_sheets:
            continue

        ws = wb[sheet_name]
        headers = _read_headers(ws)
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            result["is_valid"] = False
            result["errors"].append(
                f"[{sheet_name}] 필수 항목이 없습니다: {', '.join(missing_headers)}"
            )
            continue

        data_rows = _count_data_rows(ws)
        result["sheet_stats"].append({"sheet_name": sheet_name, "data_rows": data_rows})
        if sheet_name in REQUIRED_SHEETS:
            result["summary"]["확인된시트수"] += 1
        result["summary"]["총입력행수"] += data_rows

        if sheet_name in REQUIRED_SHEETS and sheet_name != "기본설정" and data_rows == 0:
            result["warnings"].append(f"[{sheet_name}] 입력된 데이터가 없습니다.")

        _validate_sheet_choices(ws, sheet_name, headers, result)

        if sheet_name == "기본설정":
            _validate_settings_key_rows(ws, headers, result)

    return result


def _add_sheet(wb, title, headers, rows, guide_rows=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    _style_header_row(ws, len(headers))
    for row in rows:
        ws.append(row)
    if guide_rows:
        for guide_row in guide_rows:
            padded = list(guide_row)[: len(headers)]
            if len(padded) < len(headers):
                padded.extend([""] * (len(headers) - len(padded)))
            ws.append(padded)
            _style_guide_row(ws, ws.max_row, len(headers))
    _set_column_widths(ws, headers)


def _style_header_row(ws, header_count):
    fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, header_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _style_guide_row(ws, row_idx, col_count):
    fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    font = Font(color="334155", bold=False)
    align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _set_column_widths(ws, headers):
    for idx, header in enumerate(headers, start=1):
        width = max(12, len(str(header)) * 2 + 4)
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width


def _read_headers(ws):
    values = []
    for cell in ws[1]:
        text = str(cell.value).strip() if cell.value is not None else ""
        values.append(text)
    return values


def _count_data_rows(ws):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_instruction_row(row):
            continue
        if any(value not in (None, "") for value in row):
            count += 1
    return count


def _header_index_map(headers):
    return {name: idx for idx, name in enumerate(headers)}


def _validate_settings_key_rows(ws, headers, result):
    idx_map = _header_index_map(headers)
    key_idx = idx_map.get("항목")
    if key_idx is None:
        return

    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_instruction_row(row):
            continue
        key_value = row[key_idx]
        if key_value in (None, ""):
            continue
        existing_keys.add(str(key_value).strip())

    missing_keys = sorted(REQUIRED_SETTING_KEYS - existing_keys)
    if missing_keys:
        result["warnings"].append(
            f"[기본설정] 권장 항목이 비어 있습니다: {', '.join(missing_keys)}"
        )


def _validate_sheet_choices(ws, sheet_name, headers, result):
    idx_map = _header_index_map(headers)

    for (target_sheet, column_name), allowed_values in ALLOWED_CHOICES.items():
        if target_sheet != sheet_name:
            continue
        if column_name not in idx_map:
            continue

        column_idx = idx_map[column_name]
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if _is_instruction_row(row):
                continue
            value = row[column_idx]
            if value in (None, ""):
                continue
            value_text = str(value).strip()
            if value_text not in allowed_values:
                result["is_valid"] = False
                result["errors"].append(
                    f"[{sheet_name}] {row_no}행 '{column_name}' 값이 올바르지 않습니다: {value_text}"
                )


def generate_timetable_schedule(file_obj):
    result = {
        "is_success": False,
        "errors": [],
        "warnings": [],
        "days": [],
        "slot_labels": [],
        "sync_candidates": [],
        "summary": {
            "total_needed": 0,
            "placed_count": 0,
            "unplaced_count": 0,
            "overlap_count": 0,
            "teacher_count": 0,
            "class_count": 0,
            "room_count": 0,
        },
        "teacher_tables": [],
        "class_tables": [],
        "room_tables": [],
        "reservation_rows": [],
        "assignment_status_rows": [],
    }

    try:
        file_obj.seek(0)
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:
        result["errors"].append(f"자동 배치용 파일을 읽는 중 오류가 발생했습니다: {exc}")
        return result

    parsed = _load_schedule_input(wb, result)
    if parsed is None:
        return result

    state = _empty_schedule_state()
    tasks = []

    for assignment in parsed["assignments"]:
        remaining = _place_fixed_slots(assignment, parsed, state, result)
        if remaining < 0:
            continue
        tasks.extend(_build_assignment_tasks(assignment, remaining, result))

    if tasks:
        solved, _ = _solve_tasks_backtracking(tasks, parsed, state)
        if not solved:
            result["warnings"].append(
                "모든 수업칸을 한 번에 배치하지 못해, 가능한 범위까지 먼저 배치했습니다."
            )
            unplaced_tasks = _solve_tasks_greedy(tasks, parsed, state)
            if unplaced_tasks:
                by_assignment = defaultdict(int)
                for task in unplaced_tasks:
                    by_assignment[task["assignment_id"]] += task["length"]
                for assignment_id, missing_count in by_assignment.items():
                    assignment = parsed["assignment_map"][assignment_id]
                    result["errors"].append(
                        f"{assignment['교과']} {assignment['학년']}-{assignment['반']} 배정에서 {missing_count}시간이 남았습니다."
                    )

    _check_split_conditions(parsed, state, result)
    _build_schedule_outputs(parsed, state, result)
    _build_summary(parsed, state, result)

    result["is_success"] = (
        not result["errors"] and result["summary"]["unplaced_count"] == 0
    )
    return result


def _empty_schedule_state():
    return {
        "teacher_busy": {},
        "class_busy": {},
        "room_busy": {},
        "room_slot_mode": {},
        "teacher_day_count": defaultdict(int),
        "assignment_slots": defaultdict(list),
    }


def _load_schedule_input(wb, result):
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            result["errors"].append(f"필수 시트 '{sheet_name}'가 없어 자동 배치를 시작할 수 없습니다.")
    if result["errors"]:
        return None

    setting_rows = _read_sheet_rows(wb["기본설정"], REQUIRED_SHEETS["기본설정"])
    settings = {
        row["항목"]: row["값"]
        for row in setting_rows
        if row.get("항목") and row.get("값")
    }

    days = _parse_days(settings.get("운영요일", "월,화,수,목,금"))
    base_periods = _parse_int(settings.get("기본교시수", "6"), default=6)
    slot_labels = _parse_slot_labels(settings.get("시간칸목록", ""), base_periods)
    grade_slot_map = _parse_grade_slot_map(settings.get("학년별사용시간칸", ""), slot_labels, result)

    teacher_rows = _read_sheet_rows(wb["전담선생님목록"], REQUIRED_SHEETS["전담선생님목록"])
    teachers = {}
    for row in teacher_rows:
        teacher_id = row.get("선생님코드", "")
        if not teacher_id:
            continue
        max_daily = _parse_int(row.get("하루최대수업칸", ""), default=len(slot_labels))
        if max_daily <= 0:
            max_daily = len(slot_labels)
        teachers[teacher_id] = {
            "teacher_id": teacher_id,
            "teacher_name": row.get("이름", teacher_id),
            "subject": row.get("담당교과", ""),
            "max_daily": max_daily,
        }

    if not teachers:
        result["errors"].append("전담선생님목록에 입력된 선생님이 없어 자동 배치를 진행할 수 없습니다.")
        return None

    special_room_rows = _read_optional_sheet_rows(
        wb,
        "특별실설정",
        OPTIONAL_SHEETS["특별실설정"],
    )
    special_room_settings = {}
    for row in special_room_rows:
        room_name = row.get("특별실명", "")
        if not room_name:
            continue
        special_room_settings[room_name] = {
            "mode": row.get("운영방식", "자동배치"),
            "sync_option": row.get("연동선택", "미리보기"),
            "subject": row.get("대상교과", ""),
        }

    condition_rows = _read_optional_sheet_rows(
        wb,
        "배치조건",
        OPTIONAL_SHEETS["배치조건"],
    )
    manual_rows = _read_optional_sheet_rows(
        wb,
        "수동고정",
        OPTIONAL_SHEETS["수동고정"],
    )

    manual_fixed = defaultdict(list)
    manual_blocked = defaultdict(set)
    for row in manual_rows:
        assignment_id = row.get("배정번호", "")
        if not assignment_id:
            continue
        day = row.get("요일", "")
        slot = row.get("시간칸", "")
        mode = row.get("처리", "")
        if day not in days or slot not in slot_labels:
            result["warnings"].append(
                f"[수동고정] {assignment_id}의 요일/시간칸 값({day} {slot})이 기본설정과 달라 무시됩니다."
            )
            continue
        if mode == "고정":
            manual_fixed[assignment_id].append((day, slot))
        elif mode == "배치금지":
            manual_blocked[assignment_id].add((day, slot))

    unavailable_rows = _read_optional_sheet_rows(
        wb,
        "배치불가시간",
        OPTIONAL_SHEETS["배치불가시간"],
    )
    teacher_unavailable = set()
    for row in unavailable_rows:
        teacher_id = row.get("선생님코드", "")
        day = row.get("요일", "")
        slot = row.get("시간칸", "")
        if not teacher_id or day not in days or slot not in slot_labels:
            continue
        teacher_unavailable.add((teacher_id, day, slot))

    assignment_rows = _read_sheet_rows(wb["전담배정표"], REQUIRED_SHEETS["전담배정표"])
    assignments = []
    assignment_ids = set()
    for row in assignment_rows:
        assignment_id = row.get("배정번호", "")
        if not assignment_id:
            continue
        if assignment_id in assignment_ids:
            result["errors"].append(f"배정번호가 중복됩니다: {assignment_id}")
            continue
        assignment_ids.add(assignment_id)

        teacher_id = row.get("선생님코드", "")
        if teacher_id not in teachers:
            result["errors"].append(f"[{assignment_id}] 선생님코드 '{teacher_id}'를 전담선생님목록에서 찾을 수 없습니다.")
            continue

        grade = _parse_int(row.get("학년", ""), default=0)
        class_no = _parse_int(row.get("반", ""), default=0)
        weekly_hours = _parse_int(row.get("주당시수", ""), default=0)
        if grade <= 0 or class_no <= 0:
            result["errors"].append(f"[{assignment_id}] 학년/반 값이 올바르지 않습니다.")
            continue
        if weekly_hours <= 0:
            result["errors"].append(f"[{assignment_id}] 주당시수는 1 이상이어야 합니다.")
            continue

        room_name = row.get("특별실명", "")
        room_mode = _resolve_room_mode(row.get("특별실처리", ""), room_name, special_room_settings)
        sync_option = special_room_settings.get(room_name, {}).get("sync_option", "미리보기")
        if room_mode != "해당없음" and not room_name:
            result["warnings"].append(f"[{assignment_id}] 특별실처리가 선택되었지만 특별실명이 비어 있어 해당없음으로 처리됩니다.")
            room_mode = "해당없음"

        assignment = {
            "배정번호": assignment_id,
            "선생님코드": teacher_id,
            "선생님이름": teachers[teacher_id]["teacher_name"],
            "교과": row.get("교과", ""),
            "학년": grade,
            "반": class_no,
            "학급ID": f"{grade}-{class_no}",
            "주당시수": weekly_hours,
            "특별실명": room_name,
            "특별실처리": room_mode,
            "연동선택": sync_option,
            "수동고정": manual_fixed.get(assignment_id, []),
            "배치금지": manual_blocked.get(assignment_id, set()),
            "hard_avoid_days": set(),
            "soft_avoid_days": set(),
            "split_pattern": None,
            "split_required": False,
            "split_applied": False,
        }
        _apply_conditions_to_assignment(assignment, condition_rows, days, result)
        assignments.append(assignment)

    if not assignments:
        result["errors"].append("전담배정표에 배치할 수업이 없습니다.")
        return None
    if result["errors"]:
        return None

    return {
        "settings": settings,
        "days": days,
        "day_index": {day: idx for idx, day in enumerate(days)},
        "slot_labels": slot_labels,
        "slot_index": {slot: idx for idx, slot in enumerate(slot_labels)},
        "grade_slot_map": grade_slot_map,
        "teachers": teachers,
        "assignments": assignments,
        "assignment_map": {a["배정번호"]: a for a in assignments},
        "teacher_unavailable": teacher_unavailable,
    }


def _resolve_room_mode(raw_mode, room_name, special_room_settings):
    normalized = (raw_mode or "").strip()
    if normalized in {"자동배치", "예약연동", "해당없음"}:
        return normalized
    if room_name and room_name in special_room_settings:
        mode = special_room_settings[room_name]["mode"]
        if mode in {"자동배치", "예약연동"}:
            return mode
    return "해당없음"


def _apply_conditions_to_assignment(assignment, condition_rows, days, result):
    matched_split = None
    matched_split_required = False

    for row in condition_rows:
        if not _condition_matches_assignment(row, assignment):
            continue

        method = row.get("배치방법", "")
        detail = row.get("세부값", "")
        importance = row.get("중요도", "권장")
        is_required = importance == "반드시"

        if method == "피하기":
            avoid_days = [item for item in _parse_csv_values(detail) if item in days]
            if not avoid_days:
                continue
            if is_required:
                assignment["hard_avoid_days"].update(avoid_days)
            else:
                assignment["soft_avoid_days"].update(avoid_days)
        elif method == "나눠배치":
            split = _parse_split_pattern(detail)
            if not split:
                result["warnings"].append(
                    f"[{assignment['배정번호']}] 나눠배치 세부값 '{detail}'을 읽지 못해 일반 배치로 진행합니다."
                )
                continue
            if matched_split is None or is_required:
                matched_split = split
                matched_split_required = is_required

    if matched_split:
        assignment["split_pattern"] = matched_split
        assignment["split_required"] = matched_split_required


def _condition_matches_assignment(condition_row, assignment):
    target_type = condition_row.get("적용대상", "")
    target_value = (condition_row.get("대상값", "") or "").strip()
    if not target_type or not target_value:
        return False

    if target_type == "교과":
        return target_value == assignment["교과"]
    if target_type == "선생님":
        return target_value in {assignment["선생님코드"], assignment["선생님이름"]}
    if target_type == "학년":
        return target_value in {str(assignment["학년"]), f"{assignment['학년']}학년"}
    return False


def _place_fixed_slots(assignment, parsed, state, result):
    fixed_slots = list(dict.fromkeys(assignment["수동고정"]))
    if len(fixed_slots) > assignment["주당시수"]:
        result["errors"].append(
            f"[{assignment['배정번호']}] 수동고정 칸이 주당시수보다 많습니다."
        )
        return -1

    for day, slot in fixed_slots:
        can_place = _can_place_block(assignment, day, [slot], parsed, state)
        if not can_place:
            result["errors"].append(
                f"[{assignment['배정번호']}] 수동고정 {day} {slot}칸을 배치할 수 없습니다. (겹침 또는 불가시간)"
            )
            continue
        _apply_block(assignment, day, [slot], state)

    return assignment["주당시수"] - len(state["assignment_slots"][assignment["배정번호"]])


def _build_assignment_tasks(assignment, remaining_count, result):
    if remaining_count <= 0:
        return []

    split_pattern = assignment["split_pattern"]
    block_lengths = [1] * remaining_count

    if split_pattern:
        if assignment["수동고정"]:
            result["warnings"].append(
                f"[{assignment['배정번호']}] 수동고정이 있어 나눠배치 조건은 참고만 하고 일반 배치로 진행합니다."
            )
        elif sum(split_pattern) == remaining_count:
            block_lengths = sorted(split_pattern, reverse=True)
            assignment["split_applied"] = True
        elif assignment["split_required"]:
            result["errors"].append(
                f"[{assignment['배정번호']}] 나눠배치 조건({'+'.join(str(v) for v in split_pattern)})과 남은 시수가 맞지 않습니다."
            )
        else:
            result["warnings"].append(
                f"[{assignment['배정번호']}] 나눠배치 조건과 남은 시수가 달라 일반 배치로 진행합니다."
            )

    tasks = []
    for idx, block_length in enumerate(block_lengths, start=1):
        tasks.append(
            {
                "task_id": f"{assignment['배정번호']}#{idx}",
                "assignment_id": assignment["배정번호"],
                "length": block_length,
            }
        )
    return tasks


def _solve_tasks_backtracking(tasks, parsed, state):
    max_steps = 50000
    step_count = {"value": 0}

    def _recurse(remaining_tasks):
        if not remaining_tasks:
            return True
        if step_count["value"] >= max_steps:
            return False

        best_idx = -1
        best_candidates = None
        for idx, task in enumerate(remaining_tasks):
            candidates = _collect_candidates(task, parsed, state)
            if not candidates:
                return False
            if best_candidates is None or len(candidates) < len(best_candidates):
                best_idx = idx
                best_candidates = candidates
            if len(best_candidates) == 1:
                break

        task = remaining_tasks[best_idx]
        for candidate in best_candidates:
            step_count["value"] += 1
            assignment = parsed["assignment_map"][task["assignment_id"]]
            _apply_block(assignment, candidate["day"], candidate["slots"], state)
            next_tasks = remaining_tasks[:best_idx] + remaining_tasks[best_idx + 1 :]
            if _recurse(next_tasks):
                return True
            _revert_block(assignment, candidate["day"], candidate["slots"], state)
        return False

    success = _recurse(tasks)
    return success, step_count["value"]


def _solve_tasks_greedy(tasks, parsed, state):
    sorted_tasks = sorted(tasks, key=lambda item: item["length"], reverse=True)
    unplaced = []
    for task in sorted_tasks:
        candidates = _collect_candidates(task, parsed, state)
        if not candidates:
            unplaced.append(task)
            continue
        assignment = parsed["assignment_map"][task["assignment_id"]]
        best = candidates[0]
        _apply_block(assignment, best["day"], best["slots"], state)
    return unplaced


def _collect_candidates(task, parsed, state):
    assignment = parsed["assignment_map"][task["assignment_id"]]
    days = parsed["days"]
    slot_labels = parsed["slot_labels"]
    block_length = task["length"]
    candidates = []

    for day in days:
        if day in assignment["hard_avoid_days"]:
            continue
        for start in range(0, len(slot_labels) - block_length + 1):
            block_slots = slot_labels[start : start + block_length]
            if not _can_place_block(assignment, day, block_slots, parsed, state):
                continue
            score = _candidate_score(assignment, day, start, parsed, state)
            candidates.append({"day": day, "slots": block_slots, "score": score})

    candidates.sort(key=lambda item: item["score"])
    return candidates


def _candidate_score(assignment, day, start, parsed, state):
    assignment_id = assignment["배정번호"]
    teacher_id = assignment["선생님코드"]
    used_days = {item[0] for item in state["assignment_slots"][assignment_id]}
    soft_penalty = 2 if day in assignment["soft_avoid_days"] else 0
    same_day_penalty = 1 if day in used_days else 0
    teacher_load = state["teacher_day_count"][(teacher_id, day)]
    return (
        soft_penalty,
        same_day_penalty,
        teacher_load,
        parsed["day_index"][day],
        start,
    )


def _can_place_block(assignment, day, block_slots, parsed, state):
    teacher_id = assignment["선생님코드"]
    class_id = assignment["학급ID"]
    grade = assignment["학년"]
    room_name = assignment["특별실명"]
    room_mode = assignment["특별실처리"]

    allowed_slots = parsed["grade_slot_map"].get(grade, set(parsed["slot_labels"]))
    max_daily = parsed["teachers"][teacher_id]["max_daily"]
    if state["teacher_day_count"][(teacher_id, day)] + len(block_slots) > max_daily:
        return False

    for slot in block_slots:
        if slot not in allowed_slots:
            return False
        if (day, slot) in assignment["배치금지"]:
            return False
        if (teacher_id, day, slot) in parsed["teacher_unavailable"]:
            return False
        if (teacher_id, day, slot) in state["teacher_busy"]:
            return False
        if (class_id, day, slot) in state["class_busy"]:
            return False
        if room_mode != "해당없음" and room_name and (room_name, day, slot) in state["room_busy"]:
            return False
    return True


def _apply_block(assignment, day, block_slots, state):
    assignment_id = assignment["배정번호"]
    teacher_id = assignment["선생님코드"]
    class_id = assignment["학급ID"]
    room_name = assignment["특별실명"]
    room_mode = assignment["특별실처리"]

    for slot in block_slots:
        state["teacher_busy"][(teacher_id, day, slot)] = assignment_id
        state["class_busy"][(class_id, day, slot)] = assignment_id
        if room_mode != "해당없음" and room_name:
            state["room_busy"][(room_name, day, slot)] = assignment_id
            state["room_slot_mode"][(room_name, day, slot)] = room_mode
        state["assignment_slots"][assignment_id].append((day, slot))
        state["teacher_day_count"][(teacher_id, day)] += 1


def _revert_block(assignment, day, block_slots, state):
    assignment_id = assignment["배정번호"]
    teacher_id = assignment["선생님코드"]
    class_id = assignment["학급ID"]
    room_name = assignment["특별실명"]
    room_mode = assignment["특별실처리"]

    for slot in block_slots:
        state["teacher_busy"].pop((teacher_id, day, slot), None)
        state["class_busy"].pop((class_id, day, slot), None)
        if room_mode != "해당없음" and room_name:
            state["room_busy"].pop((room_name, day, slot), None)
            state["room_slot_mode"].pop((room_name, day, slot), None)
        if (day, slot) in state["assignment_slots"][assignment_id]:
            state["assignment_slots"][assignment_id].remove((day, slot))
        state["teacher_day_count"][(teacher_id, day)] -= 1


def _check_split_conditions(parsed, state, result):
    slot_index = parsed["slot_index"]
    for assignment in parsed["assignments"]:
        if not assignment["split_applied"]:
            continue
        pattern = assignment["split_pattern"] or []
        if not pattern:
            continue
        placements = state["assignment_slots"][assignment["배정번호"]]
        if _is_split_pattern_satisfied(placements, pattern, slot_index):
            continue
        message = (
            f"[{assignment['배정번호']}] 나눠배치 조건({'+'.join(str(v) for v in pattern)})을 "
            "완전히 만족하지 못했습니다."
        )
        if assignment["split_required"]:
            result["errors"].append(message)
        else:
            result["warnings"].append(message)


def _is_split_pattern_satisfied(placements, pattern, slot_index):
    if not placements:
        return False
    by_day = defaultdict(list)
    for day, slot in placements:
        by_day[day].append(slot_index[slot])
    for day in by_day:
        by_day[day].sort()

    required_runs = sorted([size for size in pattern if size > 1], reverse=True)
    for size in required_runs:
        found = False
        for day, indexes in by_day.items():
            for pos in range(0, len(indexes) - size + 1):
                window = indexes[pos : pos + size]
                if _is_consecutive(window):
                    found = True
                    by_day[day] = indexes[:pos] + indexes[pos + size :]
                    break
            if found:
                break
        if not found:
            return False
    return True


def _is_consecutive(values):
    if not values:
        return False
    return all(values[idx + 1] - values[idx] == 1 for idx in range(len(values) - 1))


def _build_schedule_outputs(parsed, state, result):
    days = parsed["days"]
    slots = parsed["slot_labels"]
    assignment_map = parsed["assignment_map"]

    result["days"] = days
    result["slot_labels"] = slots

    teacher_tables = []
    for teacher_id, teacher in sorted(parsed["teachers"].items(), key=lambda item: item[1]["teacher_name"]):
        rows = []
        for slot in slots:
            cells = []
            for day in days:
                assignment_id = state["teacher_busy"].get((teacher_id, day, slot))
                cells.append(_lesson_text(assignment_map.get(assignment_id)))
            rows.append({"slot_label": slot, "cells": cells})
        teacher_tables.append(
            {
                "teacher_id": teacher_id,
                "teacher_name": teacher["teacher_name"],
                "rows": rows,
            }
        )
    result["teacher_tables"] = teacher_tables

    class_ids = sorted({assignment["학급ID"] for assignment in parsed["assignments"]}, key=_class_sort_key)
    class_tables = []
    for class_id in class_ids:
        rows = []
        for slot in slots:
            cells = []
            for day in days:
                assignment_id = state["class_busy"].get((class_id, day, slot))
                cells.append(_lesson_text(assignment_map.get(assignment_id), include_teacher=True))
            rows.append({"slot_label": slot, "cells": cells})
        class_tables.append({"class_id": class_id, "rows": rows})
    result["class_tables"] = class_tables

    room_names = sorted({assignment["특별실명"] for assignment in parsed["assignments"] if assignment["특별실명"]})
    room_tables = []
    for room_name in room_names:
        rows = []
        has_auto_item = False
        for slot in slots:
            cells = []
            for day in days:
                assignment_id = state["room_busy"].get((room_name, day, slot))
                mode = state["room_slot_mode"].get((room_name, day, slot))
                if assignment_id and mode == "자동배치":
                    has_auto_item = True
                    cells.append(_lesson_text(assignment_map.get(assignment_id), include_teacher=False))
                else:
                    cells.append("")
            rows.append({"slot_label": slot, "cells": cells})
        if has_auto_item:
            room_tables.append({"room_name": room_name, "rows": rows})
    result["room_tables"] = room_tables

    reservation_rows = []
    for assignment in parsed["assignments"]:
        if assignment["특별실처리"] != "예약연동":
            continue
        for day, slot in sorted(state["assignment_slots"][assignment["배정번호"]], key=lambda item: (parsed["day_index"][item[0]], parsed["slot_index"][item[1]])):
            reservation_rows.append(
                {
                    "room_name": assignment["특별실명"],
                    "day": day,
                    "slot_label": slot,
                    "class_id": assignment["학급ID"],
                    "subject": assignment["교과"],
                    "teacher_name": assignment["선생님이름"],
                    "sync_option": assignment["연동선택"],
                }
            )
    result["reservation_rows"] = reservation_rows

    assignment_status_rows = []
    sync_candidates = []
    for assignment in parsed["assignments"]:
        placed_count = len(state["assignment_slots"][assignment["배정번호"]])
        assignment_status_rows.append(
            {
                "assignment_id": assignment["배정번호"],
                "teacher_name": assignment["선생님이름"],
                "subject": assignment["교과"],
                "class_id": assignment["학급ID"],
                "needed_hours": assignment["주당시수"],
                "placed_hours": placed_count,
                "remaining_hours": max(0, assignment["주당시수"] - placed_count),
            }
        )
        if assignment["특별실처리"] == "해당없음" or not assignment["특별실명"]:
            continue
        for day, slot in sorted(
            state["assignment_slots"][assignment["배정번호"]],
            key=lambda item: (parsed["day_index"][item[0]], parsed["slot_index"][item[1]]),
        ):
            sync_candidates.append(
                {
                    "assignment_id": assignment["배정번호"],
                    "room_name": assignment["특별실명"],
                    "day": day,
                    "slot_label": slot,
                    "slot_number": parsed["slot_index"][slot] + 1,
                    "class_id": assignment["학급ID"],
                    "subject": assignment["교과"],
                    "teacher_name": assignment["선생님이름"],
                    "mode": assignment["특별실처리"],
                    "sync_option": assignment["연동선택"],
                }
            )
    result["assignment_status_rows"] = assignment_status_rows
    result["sync_candidates"] = sync_candidates


def _build_summary(parsed, state, result):
    total_needed = sum(assignment["주당시수"] for assignment in parsed["assignments"])
    placed_count = sum(len(state["assignment_slots"][assignment["배정번호"]]) for assignment in parsed["assignments"])
    room_names = {assignment["특별실명"] for assignment in parsed["assignments"] if assignment["특별실처리"] != "해당없음" and assignment["특별실명"]}

    result["summary"] = {
        "total_needed": total_needed,
        "placed_count": placed_count,
        "unplaced_count": max(0, total_needed - placed_count),
        "overlap_count": 0,
        "teacher_count": len(parsed["teachers"]),
        "class_count": len({assignment["학급ID"] for assignment in parsed["assignments"]}),
        "room_count": len(room_names),
    }


def _lesson_text(assignment, include_teacher=False):
    if not assignment:
        return ""
    text = f"{assignment['교과']} {assignment['학년']}-{assignment['반']}"
    if include_teacher:
        text = f"{text} ({assignment['선생님이름']})"
    return text


def _class_sort_key(class_id):
    if "-" not in class_id:
        return (999, 999)
    left, right = class_id.split("-", 1)
    return (_parse_int(left, 999), _parse_int(right, 999))


def _read_sheet_rows(ws, headers):
    rows = []
    idx_map = _header_index_map(_read_headers(ws))
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if _is_instruction_row(excel_row):
            continue
        if not any(value not in (None, "") for value in excel_row):
            continue
        row_dict = {}
        for header in headers:
            idx = idx_map.get(header)
            cell_value = excel_row[idx] if idx is not None and idx < len(excel_row) else None
            row_dict[header] = _normalize_cell(cell_value)
        rows.append(row_dict)
    return rows


def _read_optional_sheet_rows(wb, sheet_name, headers):
    if sheet_name not in wb.sheetnames:
        return []
    return _read_sheet_rows(wb[sheet_name], headers)


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _is_instruction_row(values):
    if not values:
        return False
    first = values[0]
    first_text = str(first).strip() if first is not None else ""
    return first_text.startswith("※")


def _parse_int(value, default=0):
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _parse_csv_values(text):
    raw = (text or "").replace("|", ",").replace("/", ",").replace(";", ",")
    return [item.strip() for item in raw.split(",") if item and item.strip()]


def _parse_days(text):
    values = _parse_csv_values(text)
    if values:
        return values
    compact = (text or "").strip()
    if not compact:
        return ["월", "화", "수", "목", "금"]
    return [char for char in compact if char.strip()]


def _parse_slot_labels(raw_slots, base_periods):
    values = _parse_csv_values(raw_slots)
    if values:
        return values
    return [str(i) for i in range(1, max(1, base_periods) + 1)]


def _parse_grade_slot_map(raw_text, slot_labels, result):
    default_slots = set(slot_labels)
    grade_slot_map = defaultdict(lambda: set(default_slots))
    if not raw_text:
        return grade_slot_map

    entries = [item.strip() for item in raw_text.split(";") if item.strip()]
    for entry in entries:
        if "=" not in entry:
            result["warnings"].append(
                f"[기본설정] 학년별사용시간칸 형식이 올바르지 않습니다: {entry}"
            )
            continue
        key, value = entry.split("=", 1)
        grade_text = key.strip().replace("학년", "")
        grade = _parse_int(grade_text, default=0)
        slots = [item for item in _parse_csv_values(value) if item in default_slots]
        if grade <= 0 or not slots:
            result["warnings"].append(
                f"[기본설정] 학년별사용시간칸 값이 올바르지 않아 무시됩니다: {entry}"
            )
            continue
        grade_slot_map[grade] = set(slots)
    return grade_slot_map


def _parse_split_pattern(text):
    values = []
    for item in _parse_csv_values((text or "").replace("+", ",")):
        value = _parse_int(item, default=0)
        if value <= 0:
            return []
        values.append(value)
    return values


def apply_schedule_to_reservations(
    generated_result,
    school,
    overwrite_existing=False,
    sync_options=None,
):
    from reservations.models import RecurringSchedule, SchoolConfig, SpecialRoom

    response = {
        "status": "skipped",
        "applied_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
        "room_created_count": 0,
        "conflict_count": 0,
        "messages": [],
        "school_name": school.name,
        "school_slug": school.slug,
    }

    if sync_options is None:
        sync_options = {"바로반영"}

    candidates = [
        item
        for item in generated_result.get("sync_candidates", [])
        if item.get("sync_option") in sync_options
    ]
    if not candidates:
        response["messages"].append("바로반영으로 지정된 항목이 없어 예약 시스템 반영을 건너뛰었습니다.")
        return response

    slot_labels = generated_result.get("slot_labels", [])
    config, _ = SchoolConfig.objects.get_or_create(school=school)
    if slot_labels:
        config.period_labels = ",".join(slot_labels)
        config.max_periods = len(slot_labels)
        config.save(update_fields=["period_labels", "max_periods"])

    room_cache = {}
    day_order = generated_result.get("days", [])

    for item in candidates:
        room_name = item.get("room_name", "").strip()
        if not room_name:
            response["skipped_count"] += 1
            continue

        room = room_cache.get(room_name)
        if room is None:
            room, created = SpecialRoom.objects.get_or_create(
                school=school,
                name=room_name,
                defaults={"icon": "🏫"},
            )
            room_cache[room_name] = room
            if created:
                response["room_created_count"] += 1

        day_idx = _to_weekday_index(item.get("day", ""), day_order)
        if day_idx is None:
            response["skipped_count"] += 1
            response["messages"].append(
                f"{room_name} {item.get('day')} {item.get('slot_label')}칸은 요일 값을 읽지 못해 건너뜀"
            )
            continue

        period = int(item.get("slot_number", 0))
        if period <= 0:
            response["skipped_count"] += 1
            continue

        schedule_name = _build_recurring_name(item)
        existing = RecurringSchedule.objects.filter(
            room=room,
            day_of_week=day_idx,
            period=period,
        ).first()
        if existing is None:
            RecurringSchedule.objects.create(
                room=room,
                day_of_week=day_idx,
                period=period,
                name=schedule_name,
            )
            response["applied_count"] += 1
            continue

        if existing.name == schedule_name:
            response["skipped_count"] += 1
            continue

        if overwrite_existing:
            existing.name = schedule_name
            existing.save(update_fields=["name"])
            response["updated_count"] += 1
        else:
            response["conflict_count"] += 1
            response["messages"].append(
                f"{room_name} {item.get('day')} {item.get('slot_label')}칸은 기존 고정시간표가 있어 반영하지 않았습니다."
            )

    if response["conflict_count"] > 0:
        response["status"] = "partial"
    elif response["applied_count"] > 0 or response["updated_count"] > 0:
        response["status"] = "success"
    else:
        response["status"] = "skipped"
    return response


def _to_weekday_index(day_text, day_order):
    day_text = (day_text or "").strip()
    day_map = {
        "월": 0,
        "화": 1,
        "수": 2,
        "목": 3,
        "금": 4,
        "토": 5,
        "일": 6,
        "mon": 0,
        "tue": 1,
        "wed": 2,
        "thu": 3,
        "fri": 4,
        "sat": 5,
        "sun": 6,
    }
    lower = day_text.lower()
    if day_text in day_map:
        return day_map[day_text]
    if lower in day_map:
        return day_map[lower]
    if day_text in day_order:
        idx = day_order.index(day_text)
        if 0 <= idx <= 6:
            return idx
    return None


def _build_recurring_name(item):
    class_id = item.get("class_id", "")
    subject = item.get("subject", "")
    teacher_name = item.get("teacher_name", "")
    text = f"{class_id} {subject} ({teacher_name})".strip()
    return text[:50]
