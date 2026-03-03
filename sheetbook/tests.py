from argparse import Namespace
from datetime import date, timedelta
import csv
from io import BytesIO, StringIO
import json
from pathlib import Path
import subprocess
import tempfile
from urllib.parse import parse_qs, urlparse
from unittest.mock import call, patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError
from django.test import SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from classcalendar.models import CalendarEvent
from collect.models import CollectionRequest
from core.models import UserProfile
from handoff.models import HandoffRosterGroup, HandoffSession
from scripts.run_sheetbook_signoff_decision import (
    _build_next_actions,
    _compute_decision,
    _default_manual_payload,
)
from scripts.run_sheetbook_archive_bulk_snapshot import (
    _build_markdown as _build_archive_bulk_snapshot_markdown,
    _collect_snapshot as _collect_archive_bulk_snapshot,
)
from scripts.run_sheetbook_consent_freeze_snapshot import (
    _build_markdown as _build_consent_freeze_snapshot_markdown,
    _build_report as _build_consent_freeze_snapshot_report,
)
from scripts.run_sheetbook_daily_start_bundle import (
    _build_bundle_markdown as _build_daily_start_bundle_markdown,
    _build_bundle_next_actions as _build_daily_start_bundle_next_actions,
    _build_bundle_summary as _build_daily_start_bundle_summary,
)
from scripts.run_sheetbook_ops_index_report import (
    _build_markdown as _build_ops_index_markdown,
    _build_summary as _build_ops_index_summary,
)
from scripts.run_sheetbook_sample_gap_summary import (
    _build_sample_gap_markdown as _build_sample_gap_markdown_payload,
    _build_sample_gap_summary as _build_sample_gap_summary_payload,
)
from scripts.run_sheetbook_release_signoff_log import (
    _build_markdown as _build_release_signoff_log_markdown,
)
from scripts.run_sheetbook_guarded_commit import (
    run as _run_sheetbook_guarded_commit,
)
from scripts.run_sheetbook_pilot_log_snapshot import (
    _build_markdown as _build_pilot_log_markdown,
    _collect_snapshot as _collect_pilot_log_snapshot,
)
from scripts.run_sheetbook_seed_metric_samples import (
    SEED_TAG as SHEETBOOK_METRIC_SEED_TAG,
    _clear_seed_events,
    _seed_metric_events,
)
from sheetbook.models import (
    ActionInvocation,
    SavedView,
    SheetCell,
    SheetColumn,
    SheetbookMetricEvent,
    SheetRow,
    SheetTab,
    Sheetbook,
)


User = get_user_model()


class SheetbookFlagTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_t1",
            password="pw123456",
            email="sheetbook_t1@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_t1", "role": "school"},
        )
        self.other_user = User.objects.create_user(
            username="sheetbook_t2",
            password="pw123456",
            email="sheetbook_t2@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.other_user,
            defaults={"nickname": "sheetbook_t2", "role": "school"},
        )

    @override_settings(SHEETBOOK_ENABLED=False)
    def test_sheetbook_returns_404_when_flag_disabled(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sheetbook_returns_200_when_flag_enabled(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "교무수첩")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sheetbook_requires_login(self):
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 302)

    @override_settings(SHEETBOOK_ENABLED=False, SHEETBOOK_BETA_USERNAMES=["sheetbook_t1"])
    def test_sheetbook_allows_beta_user_when_flag_disabled(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "교무수첩")

    @override_settings(SHEETBOOK_ENABLED=False, SHEETBOOK_BETA_EMAILS="sheetbook_t1@example.com")
    def test_sheetbook_allows_beta_user_by_email_when_flag_disabled(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 200)

    @override_settings(SHEETBOOK_ENABLED=False, SHEETBOOK_BETA_USER_IDS="9999")
    def test_sheetbook_keeps_non_allowlisted_user_blocked_when_flag_disabled(self):
        self.client.force_login(self.other_user)
        response = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=False, SHEETBOOK_BETA_USERNAMES=["sheetbook_t1"])
    def test_sheetbook_beta_user_can_create_when_flag_disabled(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:create"),
            data={"title": "베타 수첩", "academic_year": 2026},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Sheetbook.objects.filter(owner=self.user, title="베타 수첩").exists())


class SheetbookModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_model_t1",
            password="pw123456",
            email="sheetbook_model_t1@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_model_t1", "role": "school"},
        )

    def test_create_sheetbook_tab_column_row_cell(self):
        sheetbook = Sheetbook.objects.create(
            owner=self.user,
            title="2026 2-3반 교무수첩",
            academic_year=2026,
        )
        tab = SheetTab.objects.create(
            sheetbook=sheetbook,
            name="일정",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=1,
        )
        col_title = SheetColumn.objects.create(
            tab=tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=1,
        )
        col_date = SheetColumn.objects.create(
            tab=tab,
            key="date",
            label="날짜",
            column_type=SheetColumn.TYPE_DATE,
            sort_order=2,
        )
        row = SheetRow.objects.create(
            tab=tab,
            sort_order=1,
            created_by=self.user,
            updated_by=self.user,
        )

        SheetCell.objects.create(row=row, column=col_title, value_text="생태체험 학부모 도우미 안내")
        SheetCell.objects.create(row=row, column=col_date, value_date=date(2026, 3, 12))

        self.assertEqual(sheetbook.tabs.count(), 1)
        self.assertEqual(tab.columns.count(), 2)
        self.assertEqual(tab.rows.count(), 1)
        self.assertEqual(row.cells.count(), 2)

    def test_cell_unique_constraint_enforced(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="중복 제약 테스트")
        tab = SheetTab.objects.create(sheetbook=sheetbook, name="메모")
        column = SheetColumn.objects.create(tab=tab, key="note", label="메모")
        row = SheetRow.objects.create(tab=tab)

        SheetCell.objects.create(row=row, column=column, value_text="첫 값")
        with self.assertRaises(IntegrityError):
            SheetCell.objects.create(row=row, column=column, value_text="중복 값")


class SheetbookOwnershipTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_owner",
            password="pw123456",
            email="sheetbook_owner@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_owner", "role": "school"},
        )
        self.other_user = User.objects.create_user(
            username="sheetbook_other",
            password="pw123456",
            email="sheetbook_other@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.other_user,
            defaults={"nickname": "sheetbook_other", "role": "school"},
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_shows_only_my_sheetbooks(self):
        Sheetbook.objects.create(owner=self.user, title="내 수첩")
        Sheetbook.objects.create(owner=self.other_user, title="남의 수첩")

        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "내 수첩")
        self.assertNotContains(response, "남의 수첩")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_can_filter_by_title(self):
        Sheetbook.objects.create(owner=self.user, title="2026 2-3반 운영")
        Sheetbook.objects.create(owner=self.user, title="체험학습 안내")

        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"), data={"q": "체험"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "체험학습 안내")
        self.assertNotContains(response, "2026 2-3반 운영")
        self.assertEqual(response.context["search_query"], "체험")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_paginates_sheetbooks(self):
        for idx in range(23):
            Sheetbook.objects.create(owner=self.user, title=f"수첩 {idx + 1:02d}")

        self.client.force_login(self.user)
        page1 = self.client.get(reverse("sheetbook:index"))
        self.assertEqual(page1.status_code, 200)
        self.assertEqual(len(page1.context["sheetbooks"]), 20)
        self.assertTrue(page1.context["sheetbook_page"].has_next())

        page2 = self.client.get(reverse("sheetbook:index"), data={"page": 2})
        self.assertEqual(page2.status_code, 200)
        self.assertEqual(len(page2.context["sheetbooks"]), 3)
        self.assertTrue(page2.context["sheetbook_page"].has_previous())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_can_filter_archived_sheetbooks(self):
        Sheetbook.objects.create(owner=self.user, title="활성 수첩", is_archived=False)
        Sheetbook.objects.create(owner=self.user, title="보관 수첩", is_archived=True)

        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"), data={"status": "archived"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "보관 수첩")
        self.assertNotContains(response, "활성 수첩")
        self.assertEqual(response.context["status_filter"], "archived")
        self.assertEqual(response.context["status_counts"]["active"], 1)
        self.assertEqual(response.context["status_counts"]["archived"], 1)
        self.assertEqual(response.context["status_counts"]["all"], 2)
        self.assertContains(response, "index_status=archived")
        self.assertContains(response, "아카이브는 최근 보관한 순서로 보여요.")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_archived_orders_by_recent_archived_at(self):
        now = timezone.now()
        older = Sheetbook.objects.create(owner=self.user, title="오래된 보관", is_archived=True, archived_at=now - timedelta(days=3))
        recent = Sheetbook.objects.create(owner=self.user, title="최근 보관", is_archived=True, archived_at=now - timedelta(hours=2))
        # Intentionally invert updated_at to ensure archived_at ordering is used for archived filter.
        Sheetbook.objects.filter(id=older.id).update(updated_at=now)
        Sheetbook.objects.filter(id=recent.id).update(updated_at=now - timedelta(days=5))

        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"), data={"status": "archived"})

        self.assertEqual(response.status_code, 200)
        ordered_titles = [item.title for item in response.context["sheetbooks"]]
        self.assertEqual(ordered_titles[:2], ["최근 보관", "오래된 보관"])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_all_lists_active_before_archived(self):
        archived = Sheetbook.objects.create(owner=self.user, title="보관 수첩", is_archived=True, archived_at=timezone.now())
        active = Sheetbook.objects.create(owner=self.user, title="활성 수첩", is_archived=False)
        now = timezone.now()
        Sheetbook.objects.filter(id=archived.id).update(updated_at=now)
        Sheetbook.objects.filter(id=active.id).update(updated_at=now - timedelta(days=10))

        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"), data={"status": "all"})

        self.assertEqual(response.status_code, 200)
        items = list(response.context["sheetbooks"])
        self.assertTrue(items)
        self.assertFalse(items[0].is_archived)
        self.assertContains(response, "전체 보기에서는 활성 수첩이 먼저 보여요.")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_archive_and_unarchive_sheetbook(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="아카이브 대상", academic_year=2026)
        self.client.force_login(self.user)

        archive_response = self.client.post(
            reverse("sheetbook:archive", kwargs={"pk": sheetbook.pk}),
            data={"status": "active"},
        )
        self.assertEqual(archive_response.status_code, 302)
        sheetbook.refresh_from_db()
        self.assertTrue(sheetbook.is_archived)
        self.assertIsNotNone(sheetbook.archived_at)

        unarchive_response = self.client.post(
            reverse("sheetbook:unarchive", kwargs={"pk": sheetbook.pk}),
            data={"status": "archived"},
        )
        self.assertEqual(unarchive_response.status_code, 302)
        sheetbook.refresh_from_db()
        self.assertFalse(sheetbook.is_archived)
        self.assertIsNone(sheetbook.archived_at)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_bulk_archive_update_archives_selected_sheetbooks(self):
        first = Sheetbook.objects.create(owner=self.user, title="일괄 대상 1", is_archived=False)
        second = Sheetbook.objects.create(owner=self.user, title="일괄 대상 2", is_archived=False)
        already_archived = Sheetbook.objects.create(
            owner=self.user,
            title="이미 보관",
            is_archived=True,
            archived_at=timezone.now() - timedelta(days=2),
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:bulk_archive_update"),
            data={
                "sheetbook_ids": [str(first.id), str(second.id), str(already_archived.id)],
                "archive_action": "archive",
                "q": "운영",
                "status": "active",
                "source": "workspace_home",
                "page": "2",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("q=%EC%9A%B4%EC%98%81", response.url)
        self.assertIn("source=workspace_home", response.url)
        self.assertIn("page=2", response.url)

        first.refresh_from_db()
        second.refresh_from_db()
        already_archived.refresh_from_db()
        self.assertTrue(first.is_archived)
        self.assertIsNotNone(first.archived_at)
        self.assertTrue(second.is_archived)
        self.assertIsNotNone(second.archived_at)
        self.assertTrue(already_archived.is_archived)

        event = (
            SheetbookMetricEvent.objects.filter(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(event)
        self.assertEqual(event.metadata.get("selected_count"), 3)
        self.assertEqual(event.metadata.get("matched_count"), 3)
        self.assertEqual(event.metadata.get("changed_count"), 2)
        self.assertEqual(event.metadata.get("unchanged_count"), 1)
        self.assertEqual(event.metadata.get("ignored_count"), 0)
        self.assertEqual(event.metadata.get("archive_action"), "archive")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_bulk_archive_update_unarchives_selected_sheetbooks_and_excludes_inaccessible_ids(self):
        target = Sheetbook.objects.create(
            owner=self.user,
            title="복구 대상",
            is_archived=True,
            archived_at=timezone.now() - timedelta(days=1),
        )
        already_active = Sheetbook.objects.create(owner=self.user, title="이미 활성", is_archived=False)
        other_user_sheetbook = Sheetbook.objects.create(
            owner=self.other_user,
            title="타 유저 보관",
            is_archived=True,
            archived_at=timezone.now() - timedelta(days=1),
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:bulk_archive_update"),
            data={
                "sheetbook_ids": [
                    str(target.id),
                    str(already_active.id),
                    str(other_user_sheetbook.id),
                    "not-a-number",
                    str(target.id),
                ],
                "archive_action": "unarchive",
                "status": "archived",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "1개 수첩을 활성 상태로 되돌렸어요.")
        self.assertContains(response, "접근 불가·삭제됨 1개 제외")

        target.refresh_from_db()
        already_active.refresh_from_db()
        other_user_sheetbook.refresh_from_db()
        self.assertFalse(target.is_archived)
        self.assertIsNone(target.archived_at)
        self.assertFalse(already_active.is_archived)
        self.assertTrue(other_user_sheetbook.is_archived)

        event = (
            SheetbookMetricEvent.objects.filter(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(event)
        self.assertEqual(event.metadata.get("selected_count"), 3)
        self.assertEqual(event.metadata.get("matched_count"), 2)
        self.assertEqual(event.metadata.get("changed_count"), 1)
        self.assertEqual(event.metadata.get("unchanged_count"), 1)
        self.assertEqual(event.metadata.get("ignored_count"), 1)
        self.assertEqual(event.metadata.get("archive_action"), "unarchive")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_bulk_archive_update_requires_selected_sheetbook_ids(self):
        Sheetbook.objects.create(owner=self.user, title="선택 필요", is_archived=False)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:bulk_archive_update"),
            data={"archive_action": "archive", "status": "active"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "선택된 수첩이 없어요. 목록에서 수첩을 먼저 선택해 주세요.")
        self.assertFalse(
            SheetbookMetricEvent.objects.filter(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
            ).exists()
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_bulk_archive_update_all_inaccessible_ids_records_zero_change(self):
        other_user_sheetbook = Sheetbook.objects.create(
            owner=self.other_user,
            title="타 유저 수첩",
            is_archived=False,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:bulk_archive_update"),
            data={
                "sheetbook_ids": [str(other_user_sheetbook.id)],
                "archive_action": "archive",
                "status": "active",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "선택한 수첩을 찾을 수 없어요. 목록을 새로고침한 뒤 다시 선택해 주세요.")

        event = (
            SheetbookMetricEvent.objects.filter(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(event)
        self.assertEqual(event.metadata.get("selected_count"), 1)
        self.assertEqual(event.metadata.get("matched_count"), 0)
        self.assertEqual(event.metadata.get("changed_count"), 0)
        self.assertEqual(event.metadata.get("unchanged_count"), 0)
        self.assertEqual(event.metadata.get("ignored_count"), 1)
        self.assertEqual(event.metadata.get("archive_action"), "archive")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_renders_bulk_archive_controls(self):
        Sheetbook.objects.create(owner=self.user, title="목록 수첩")
        self.client.force_login(self.user)

        response = self.client.get(reverse("sheetbook:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="sheetbook-bulk-select-all"')
        self.assertContains(response, 'id="sheetbook-selected-count"')
        self.assertContains(response, 'id="sheetbook-bulk-apply-button"')
        self.assertContains(response, 'name="sheetbook_ids"')
        self.assertContains(response, "현재 페이지에서 선택한 수첩만 일괄 처리됩니다.")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_archived_sheetbook_blocks_tab_create(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="보관 수첩", is_archived=True)
        tab = SheetTab.objects.create(sheetbook=sheetbook, name="기존 탭", sort_order=1)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:create_tab", kwargs={"pk": sheetbook.pk}),
            data={"name": "새 탭", "tab_type": SheetTab.TYPE_GRID},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(sheetbook.tabs.count(), 1)
        self.assertTrue(sheetbook.tabs.filter(id=tab.id).exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_shows_archive_read_only_banner_and_copy_cta(self):
        sheetbook = Sheetbook.objects.create(
            owner=self.user,
            title="보관 수첩",
            academic_year=2025,
            is_archived=True,
        )
        SheetTab.objects.create(sheetbook=sheetbook, name="일정", tab_type=SheetTab.TYPE_GRID, sort_order=1)
        self.client.force_login(self.user)

        response = self.client.get(reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "아카이브 읽기 전용 수첩")
        self.assertContains(response, "행 포함 이어쓰기")
        self.assertContains(response, 'name="source" value="sheetbook_detail_archive_banner"')
        self.assertContains(response, 'name="include_rows" value="1"')
        self.assertTrue(
            SheetbookMetricEvent.objects.filter(
                event_name="sheetbook_archive_read_mode_opened",
                user=self.user,
                sheetbook=sheetbook,
            ).exists()
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_back_link_preserves_index_filter_context(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="컨텍스트 수첩")
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk}),
            data={
                "source": "workspace_home",
                "index_status": "archived",
                "index_q": "운영",
                "index_page": "2",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'href="{reverse("sheetbook:index")}?q=%EC%9A%B4%EC%98%81&amp;source=workspace_home&amp;status=archived&amp;page=2"',
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_empty_state_shows_sample_onboarding_cta(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "샘플 수첩으로 60초 시작")
        self.assertContains(response, reverse("sheetbook:quick_sample"))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_create_sheetbook_sets_owner_and_default_tabs(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:create"),
            data={
                "title": "신규 수첩",
                "academic_year": 2026,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        sheetbook = Sheetbook.objects.get(title="신규 수첩")
        self.assertEqual(sheetbook.owner_id, self.user.id)
        self.assertEqual(
            list(sheetbook.tabs.values_list("name", flat=True).order_by("sort_order", "id")),
            ["달력", "일정", "학생명부", "메모"],
        )
        self.assertContains(response, "탭 4개")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_quick_create_sheetbook_creates_default_tabs(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_create"),
            data={"source": "workspace_home_create"},
        )
        self.assertEqual(response.status_code, 302)

        created = Sheetbook.objects.filter(owner=self.user).order_by("-id").first()
        self.assertIsNotNone(created)
        self.assertEqual(created.academic_year, timezone.localdate().year)
        self.assertEqual(
            list(created.tabs.values_list("name", flat=True).order_by("sort_order", "id")),
            ["달력", "일정", "학생명부", "메모"],
        )
        created_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=created,
        ).order_by("-id").first()
        self.assertIsNotNone(created_event)
        self.assertEqual(created_event.metadata.get("entry_source"), "workspace_home_create")
        self.assertEqual(created_event.metadata.get("quick_flow"), "workspace_quick_create")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_quick_sample_sheetbook_creates_seeded_rows_and_onboarding_redirect(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_sample"),
            data={"source": "sheetbook_index_sample"},
        )
        self.assertEqual(response.status_code, 302)
        location = response.headers.get("Location", "")
        self.assertIn("onboarding=sample", location)

        created = Sheetbook.objects.filter(owner=self.user).order_by("-id").first()
        self.assertIsNotNone(created)
        schedule_tab = created.tabs.get(name="일정")
        title_column = schedule_tab.columns.get(key="title")
        first_row = schedule_tab.rows.order_by("sort_order", "id").first()
        self.assertIsNotNone(first_row)
        first_title = SheetCell.objects.get(row=first_row, column=title_column).value_text
        self.assertEqual(first_title, "학기 시작 안내")

        roster_tab = created.tabs.get(name="학생명부")
        name_column = roster_tab.columns.get(key="name")
        roster_names = list(
            SheetCell.objects.filter(row__tab=roster_tab, column=name_column)
            .values_list("value_text", flat=True)
        )
        self.assertIn("김하늘", roster_names)
        self.assertIn("박나래", roster_names)

        created_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=created,
        ).order_by("-id").first()
        self.assertIsNotNone(created_event)
        self.assertEqual(created_event.metadata.get("quick_flow"), "workspace_quick_sample")
        self.assertTrue(created_event.metadata.get("sample_seeded"))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_renders_sample_onboarding_guide_when_requested(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_sample"),
            data={"source": "sheetbook_index_sample"},
        )
        self.assertEqual(response.status_code, 302)
        detail_response = self.client.get(response.headers.get("Location", ""))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, "샘플 수첩 60초 시작 가이드")
        self.assertContains(detail_response, "가이드 닫기")
        self.assertContains(detail_response, "서명 요청")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_quick_copy_sheetbook_clones_tab_structure(self):
        source = Sheetbook.objects.create(
            owner=self.user,
            title="2025 2-3반 교무수첩",
            academic_year=2025,
        )
        source_calendar_tab = SheetTab.objects.create(
            sheetbook=source,
            name="달력",
            tab_type=SheetTab.TYPE_CALENDAR,
            sort_order=1,
        )
        source_grid_tab = SheetTab.objects.create(
            sheetbook=source,
            name="운영 기록",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=2,
        )
        SheetColumn.objects.create(
            tab=source_grid_tab,
            key="date",
            label="날짜",
            column_type=SheetColumn.TYPE_DATE,
            sort_order=1,
        )
        SheetColumn.objects.create(
            tab=source_grid_tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=2,
        )
        SheetRow.objects.create(tab=source_grid_tab, sort_order=1, created_by=self.user, updated_by=self.user)
        SheetRow.objects.create(tab=source_grid_tab, sort_order=2, created_by=self.user, updated_by=self.user)

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_copy"),
            data={"source": "workspace_home_copy"},
        )
        self.assertEqual(response.status_code, 302)

        copied = Sheetbook.objects.filter(owner=self.user).exclude(id=source.id).order_by("-id").first()
        self.assertIsNotNone(copied)
        self.assertEqual(copied.academic_year, 2026)
        self.assertEqual(
            list(copied.tabs.values_list("name", "tab_type").order_by("sort_order", "id")),
            [(source_calendar_tab.name, source_calendar_tab.tab_type), (source_grid_tab.name, source_grid_tab.tab_type)],
        )
        copied_grid_tab = copied.tabs.get(name="운영 기록")
        self.assertEqual(copied_grid_tab.columns.count(), 2)
        self.assertEqual(copied_grid_tab.rows.count(), 1)
        self.assertEqual(
            list(copied_grid_tab.columns.values_list("key", flat=True).order_by("sort_order", "id")),
            ["date", "title"],
        )
        created_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=copied,
        ).order_by("-id").first()
        self.assertIsNotNone(created_event)
        self.assertEqual(created_event.metadata.get("entry_source"), "workspace_home_copy")
        self.assertEqual(created_event.metadata.get("quick_flow"), "workspace_quick_copy")
        self.assertEqual(created_event.metadata.get("copied_from_sheetbook_id"), source.id)
        self.assertFalse(created_event.metadata.get("copied_with_rows"))
        self.assertEqual(created_event.metadata.get("cloned_row_count"), 1)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_quick_copy_sheetbook_can_include_rows_and_cells(self):
        source = Sheetbook.objects.create(
            owner=self.user,
            title="2025 2-3반 교무수첩",
            academic_year=2025,
        )
        SheetTab.objects.create(
            sheetbook=source,
            name="달력",
            tab_type=SheetTab.TYPE_CALENDAR,
            sort_order=1,
        )
        source_grid_tab = SheetTab.objects.create(
            sheetbook=source,
            name="운영 기록",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=2,
        )
        source_col_date = SheetColumn.objects.create(
            tab=source_grid_tab,
            key="date",
            label="날짜",
            column_type=SheetColumn.TYPE_DATE,
            sort_order=1,
        )
        source_col_title = SheetColumn.objects.create(
            tab=source_grid_tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=2,
        )
        row1 = SheetRow.objects.create(tab=source_grid_tab, sort_order=1, created_by=self.user, updated_by=self.user)
        row2 = SheetRow.objects.create(tab=source_grid_tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=row1, column=source_col_date, value_date=date(2025, 3, 4))
        SheetCell.objects.create(row=row1, column=source_col_title, value_text="학기 시작 안내")
        SheetCell.objects.create(row=row2, column=source_col_date, value_date=date(2025, 3, 8))
        SheetCell.objects.create(row=row2, column=source_col_title, value_text="학부모 상담")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_copy"),
            data={"source": "workspace_home_copy", "include_rows": "1"},
        )
        self.assertEqual(response.status_code, 302)

        copied = Sheetbook.objects.filter(owner=self.user).exclude(id=source.id).order_by("-id").first()
        self.assertIsNotNone(copied)
        copied_grid_tab = copied.tabs.get(name="운영 기록")
        self.assertEqual(copied_grid_tab.rows.count(), 2)
        copied_col_date = copied_grid_tab.columns.get(key="date")
        copied_col_title = copied_grid_tab.columns.get(key="title")
        copied_rows = list(copied_grid_tab.rows.order_by("sort_order", "id"))
        self.assertEqual(
            SheetCell.objects.get(row=copied_rows[0], column=copied_col_title).value_text,
            "학기 시작 안내",
        )
        self.assertEqual(
            SheetCell.objects.get(row=copied_rows[1], column=copied_col_title).value_text,
            "학부모 상담",
        )
        self.assertEqual(
            SheetCell.objects.get(row=copied_rows[0], column=copied_col_date).value_date,
            date(2025, 3, 4),
        )

        created_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=copied,
        ).order_by("-id").first()
        self.assertIsNotNone(created_event)
        self.assertTrue(created_event.metadata.get("copied_with_rows"))
        self.assertEqual(created_event.metadata.get("cloned_row_count"), 2)
        self.assertEqual(created_event.metadata.get("cloned_cell_count"), 4)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_quick_copy_without_source_sheetbook_redirects_to_index(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:quick_copy"),
            data={"source": "workspace_home_copy"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("sheetbook:index"), response.headers.get("Location", ""))
        self.assertEqual(Sheetbook.objects.filter(owner=self.user).count(), 0)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_rejects_other_users_sheetbook(self):
        sheetbook = Sheetbook.objects.create(owner=self.other_user, title="외부 수첩")
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk}))
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=False)
    def test_create_is_blocked_by_feature_flag(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:create"),
            data={"title": "플래그OFF", "academic_year": 2026},
        )
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=False)
    def test_detail_is_blocked_by_feature_flag(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="플래그 테스트")
        self.client.force_login(self.user)
        response = self.client.get(reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk}))
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_create_tab_rejects_other_users_sheetbook(self):
        foreign_sheetbook = Sheetbook.objects.create(owner=self.other_user, title="외부 수첩")
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:create_tab", kwargs={"pk": foreign_sheetbook.pk}),
            data={"name": "침입탭", "tab_type": SheetTab.TYPE_GRID},
        )
        self.assertEqual(response.status_code, 404)
        self.assertFalse(foreign_sheetbook.tabs.filter(name="침입탭").exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_rename_tab_updates_name_for_owner(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="이름 변경 수첩")
        tab = SheetTab.objects.create(sheetbook=sheetbook, name="기존 탭", sort_order=1)
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:rename_tab", kwargs={"pk": sheetbook.pk, "tab_pk": tab.pk}),
            data={"name": "변경된 탭"},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        tab.refresh_from_db()
        self.assertEqual(tab.name, "변경된 탭")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_delete_tab_removes_tab_and_reorders(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="삭제 수첩")
        tab1 = SheetTab.objects.create(sheetbook=sheetbook, name="탭1", sort_order=1)
        tab2 = SheetTab.objects.create(sheetbook=sheetbook, name="탭2", sort_order=2)
        tab3 = SheetTab.objects.create(sheetbook=sheetbook, name="탭3", sort_order=3)

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:delete_tab", kwargs={"pk": sheetbook.pk, "tab_pk": tab2.pk}),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(SheetTab.objects.filter(pk=tab2.pk).exists())
        orders = list(sheetbook.tabs.order_by("sort_order", "id").values_list("name", "sort_order"))
        self.assertEqual(orders, [("탭1", 1), ("탭3", 2)])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_move_tab_up_and_down_changes_order(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="정렬 수첩")
        tab1 = SheetTab.objects.create(sheetbook=sheetbook, name="탭1", sort_order=1)
        tab2 = SheetTab.objects.create(sheetbook=sheetbook, name="탭2", sort_order=2)
        tab3 = SheetTab.objects.create(sheetbook=sheetbook, name="탭3", sort_order=3)
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:move_tab_up", kwargs={"pk": sheetbook.pk, "tab_pk": tab3.pk}),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        tab2.refresh_from_db()
        tab3.refresh_from_db()
        self.assertEqual(tab3.sort_order, 2)
        self.assertEqual(tab2.sort_order, 3)

        response = self.client.post(
            reverse("sheetbook:move_tab_down", kwargs={"pk": sheetbook.pk, "tab_pk": tab1.pk}),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        tab1.refresh_from_db()
        tab3.refresh_from_db()
        self.assertEqual(tab1.sort_order, 2)
        self.assertEqual(tab3.sort_order, 1)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_htmx_create_tab_returns_partial_without_redirect(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="HTMX 수첩")
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:create_tab", kwargs={"pk": sheetbook.pk}),
            data={"name": "즉시추가", "tab_type": SheetTab.TYPE_GRID},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("sheetbook-tab-list", response.content.decode("utf-8"))
        self.assertTrue(sheetbook.tabs.filter(name="즉시추가").exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_htmx_rename_tab_returns_partial_without_redirect(self):
        sheetbook = Sheetbook.objects.create(owner=self.user, title="HTMX 이름변경 수첩")
        tab = SheetTab.objects.create(sheetbook=sheetbook, name="원래이름", sort_order=1)
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:rename_tab", kwargs={"pk": sheetbook.pk, "tab_pk": tab.pk}),
            data={"name": "바뀐이름"},
            HTTP_HX_REQUEST="true",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("sheetbook-tab-list", response.content.decode("utf-8"))
        tab.refresh_from_db()
        self.assertEqual(tab.name, "바뀐이름")


class SheetbookGridApiTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_api_owner",
            password="pw123456",
            email="sheetbook_api_owner@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_api_owner", "role": "school"},
        )
        self.other_user = User.objects.create_user(
            username="sheetbook_api_other",
            password="pw123456",
            email="sheetbook_api_other@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.other_user,
            defaults={"nickname": "sheetbook_api_other", "role": "school"},
        )

        self.sheetbook = Sheetbook.objects.create(owner=self.user, title="API 테스트 수첩")
        self.tab = SheetTab.objects.create(sheetbook=self.sheetbook, name="일정", tab_type=SheetTab.TYPE_GRID, sort_order=1)
        self.col_title = SheetColumn.objects.create(
            tab=self.tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=1,
        )
        self.col_cost = SheetColumn.objects.create(
            tab=self.tab,
            key="cost",
            label="비용",
            column_type=SheetColumn.TYPE_NUMBER,
            sort_order=2,
        )
        self.col_day = SheetColumn.objects.create(
            tab=self.tab,
            key="day",
            label="날짜",
            column_type=SheetColumn.TYPE_DATE,
            sort_order=3,
        )
        self.row = SheetRow.objects.create(tab=self.tab, sort_order=1, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=self.row, column=self.col_title, value_text="생태체험")
        SheetCell.objects.create(row=self.row, column=self.col_cost, value_number="12000")
        self.calendar_tab = SheetTab.objects.create(
            sheetbook=self.sheetbook,
            name="달력",
            tab_type=SheetTab.TYPE_CALENDAR,
            sort_order=2,
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_grid_data_returns_rows_columns_values(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"offset": 0, "limit": 20},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["tab"]["id"], self.tab.id)
        self.assertEqual(len(payload["columns"]), 3)
        self.assertEqual(len(payload["rows"]), 1)
        self.assertEqual(payload["rows"][0]["values"][str(self.col_title.id)], "생태체험")
        self.assertEqual(payload["rows"][0]["values"][str(self.col_cost.id)], 12000.0)
        self.assertIsNone(payload["rows"][0]["values"][str(self.col_day.id)])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_grid_data_limit_is_capped_to_1000(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"offset": 0, "limit": 5000},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["limit"], 1000)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_grid_data_rejects_other_users_access(self):
        self.client.force_login(self.other_user)
        response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        )
        self.assertEqual(response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_supports_text_number_date(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_title.id,
                "value": "학부모 도우미 모집",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], "학부모 도우미 모집")

        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_cost.id,
                "value": "34000",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], 34000.0)

        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_day.id,
                "value": "2026-03-14",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], "2026-03-14")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_accepts_when_client_original_matches(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_title.id,
                "value": "학부모 도우미 모집",
                "client_original": "생태체험",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], "학부모 도우미 모집")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_detects_conflict_with_stale_client_original(self):
        SheetCell.objects.filter(row=self.row, column=self.col_title).update(value_text="선수정")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_title.id,
                "value": "내 수정",
                "client_original": "생태체험",
            },
        )
        self.assertEqual(response.status_code, 409)
        payload = response.json()
        self.assertTrue(payload["conflict"])
        self.assertEqual(payload["current_value"], "선수정")

        title_cell = SheetCell.objects.get(row=self.row, column=self.col_title)
        self.assertEqual(title_cell.value_text, "선수정")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_accepts_flexible_date_formats(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_day.id,
                "value": "2026/03/14",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["value"], "2026-03-14")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_returns_400_for_invalid_date(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_day.id,
                "value": "2026/13/40",
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("날짜를 읽지 못했어요", response.json()["error"])

    @override_settings(SHEETBOOK_ENABLED=False)
    def test_grid_and_update_blocked_when_feature_disabled(self):
        self.client.force_login(self.user)
        grid_response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        )
        self.assertEqual(grid_response.status_code, 404)
        update_response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"row_id": self.row.id, "column_id": self.col_title.id, "value": "테스트"},
        )
        self.assertEqual(update_response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_create_grid_row_endpoint_adds_row(self):
        self.client.force_login(self.user)
        before_count = self.tab.rows.count()
        response = self.client.post(
            reverse("sheetbook:create_grid_row", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.tab.rows.count(), before_count + 1)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_create_grid_column_endpoint_adds_column(self):
        self.client.force_login(self.user)
        before_count = self.tab.columns.count()
        response = self.client.post(
            reverse("sheetbook:create_grid_column", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "label": "비고",
                "column_type": "text",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.tab.columns.count(), before_count + 1)
        self.assertTrue(self.tab.columns.filter(label="비고").exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_paste_cells_bulk_updates_and_creates_rows(self):
        self.client.force_login(self.user)
        # Ensure we start with one row.
        self.assertEqual(self.tab.rows.count(), 1)
        response = self.client.post(
            reverse("sheetbook:paste_cells", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "raw_text": "체험학습\t15000\t2026-03-20\n보건수업\t20000\t2026-03-21",
                    "start_row_index": 0,
                    "start_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertGreaterEqual(payload["updated"], 6)
        self.assertEqual(payload["invalid_rows"], [])
        self.assertEqual(self.tab.rows.count(), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_paste_cells_reports_invalid_rows_on_column_overflow(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:paste_cells", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "raw_text": "a\tb\tc\td",
                    "start_row_index": 0,
                    "start_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertGreater(payload["skipped"], 0)
        self.assertIn(1, payload["invalid_rows"])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_paste_cells_accepts_csv_fallback(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:paste_cells", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "raw_text": "체험학습,15000,2026-03-20\n보건수업,20000,2026-03-21",
                    "start_row_index": 0,
                    "start_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["invalid_rows"], [])
        self.assertEqual(self.tab.rows.count(), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_paste_cells_skips_invalid_number_cell(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:paste_cells", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "raw_text": "수업A\t숫자아님\t2026-03-31",
                    "start_row_index": 0,
                    "start_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["skipped"], 1)
        self.assertEqual(payload["invalid_rows"], [1])

        title_cell = SheetCell.objects.get(row=self.row, column=self.col_title)
        cost_cell = SheetCell.objects.get(row=self.row, column=self.col_cost)
        day_cell = SheetCell.objects.get(row=self.row, column=self.col_day)
        self.assertEqual(title_cell.value_text, "수업A")
        self.assertEqual(float(cost_cell.value_number), 12000.0)
        self.assertEqual(day_cell.value_date.isoformat(), "2026-03-31")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_replaces_rows_from_csv_with_header(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile(
            "sheetbook_import.csv",
            "제목,비용,날짜\n체험학습,15000,2026-03-20\n보건수업,20000,2026-03-21".encode("utf-8"),
            content_type="text/csv",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "has_header": "on",
                "replace_existing": "on",
                "auto_add_columns": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        rows = list(self.tab.rows.order_by("sort_order", "id"))
        self.assertEqual(len(rows), 2)
        first_row = rows[0]
        first_title = SheetCell.objects.get(row=first_row, column=self.col_title)
        first_cost = SheetCell.objects.get(row=first_row, column=self.col_cost)
        first_day = SheetCell.objects.get(row=first_row, column=self.col_day)
        self.assertEqual(first_title.value_text, "체험학습")
        self.assertEqual(float(first_cost.value_number), 15000.0)
        self.assertEqual(first_day.value_date.isoformat(), "2026-03-20")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_auto_adds_missing_columns_from_header(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile(
            "sheetbook_import_extra_cols.csv",
            "제목,비용,날짜,반,비고\n체험학습,15000,2026-03-20,2-3,준비물 확인".encode("utf-8"),
            content_type="text/csv",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "has_header": "on",
                "replace_existing": "on",
                "auto_add_columns": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        columns = list(self.tab.columns.order_by("sort_order", "id"))
        self.assertEqual(len(columns), 5)
        self.assertEqual(columns[3].label, "반")
        self.assertEqual(columns[4].label, "비고")

        row = self.tab.rows.order_by("sort_order", "id").first()
        self.assertIsNotNone(row)
        group_cell = SheetCell.objects.get(row=row, column=columns[3])
        note_cell = SheetCell.objects.get(row=row, column=columns[4])
        self.assertEqual(group_cell.value_text, "2-3")
        self.assertEqual(note_cell.value_text, "준비물 확인")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_accepts_xlsx(self):
        try:
            from openpyxl import Workbook
        except Exception:
            self.skipTest("openpyxl not installed")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["제목", "비용", "날짜"])
        worksheet.append(["체험학습", 18000, "2026-04-05"])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()

        self.client.force_login(self.user)
        upload = SimpleUploadedFile(
            "sheetbook_import.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "has_header": "on",
                "replace_existing": "on",
                "auto_add_columns": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        row = self.tab.rows.order_by("sort_order", "id").first()
        self.assertIsNotNone(row)
        title_cell = SheetCell.objects.get(row=row, column=self.col_title)
        cost_cell = SheetCell.objects.get(row=row, column=self.col_cost)
        day_cell = SheetCell.objects.get(row=row, column=self.col_day)
        self.assertEqual(title_cell.value_text, "체험학습")
        self.assertEqual(float(cost_cell.value_number), 18000.0)
        self.assertEqual(day_cell.value_date.isoformat(), "2026-04-05")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_accepts_cp949_csv(self):
        self.client.force_login(self.user)
        cp949_bytes = "제목,비용,날짜\n현장학습,22000,2026-04-12".encode("cp949")
        upload = SimpleUploadedFile(
            "sheetbook_import_cp949.csv",
            cp949_bytes,
            content_type="text/csv",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "has_header": "on",
                "replace_existing": "on",
                "auto_add_columns": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        row = self.tab.rows.order_by("sort_order", "id").first()
        self.assertIsNotNone(row)
        title_cell = SheetCell.objects.get(row=row, column=self.col_title)
        cost_cell = SheetCell.objects.get(row=row, column=self.col_cost)
        day_cell = SheetCell.objects.get(row=row, column=self.col_day)
        self.assertEqual(title_cell.value_text, "현장학습")
        self.assertEqual(float(cost_cell.value_number), 22000.0)
        self.assertEqual(day_cell.value_date.isoformat(), "2026-04-12")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_without_header_keeps_first_row_as_data(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile(
            "sheetbook_import_no_header.csv",
            "체험학습,15000,2026-03-20\n보건수업,20000,2026-03-21".encode("utf-8"),
            content_type="text/csv",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "replace_existing": "on",
                "auto_add_columns": "on",
            },
        )
        self.assertEqual(response.status_code, 302)

        rows = list(self.tab.rows.order_by("sort_order", "id"))
        self.assertEqual(len(rows), 2)
        first_title = SheetCell.objects.get(row=rows[0], column=self.col_title)
        second_title = SheetCell.objects.get(row=rows[1], column=self.col_title)
        self.assertEqual(first_title.value_text, "체험학습")
        self.assertEqual(second_title.value_text, "보건수업")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_grid_tab_file_shows_warning_when_some_cells_skipped(self):
        self.client.force_login(self.user)
        upload = SimpleUploadedFile(
            "sheetbook_import_with_errors.csv",
            "제목,비용,날짜\n체험학습,숫자아님,2026-03-20\n보건수업,20000,2026-03-21,초과".encode("utf-8"),
            content_type="text/csv",
        )
        response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "file": upload,
                "has_header": "on",
                "replace_existing": "on",
                "auto_add_columns": "off",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        messages = [str(message) for message in response.context["messages"]]
        self.assertTrue(any("제외" in message for message in messages))
        self.assertTrue(any("확인 줄" in message for message in messages))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_export_grid_tab_csv_downloads_header_and_rows(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:export_grid_tab_csv", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])

        decoded = response.content.decode("utf-8-sig")
        rows = list(csv.reader(StringIO(decoded)))
        self.assertEqual(rows[0], ["제목", "비용", "날짜"])
        self.assertEqual(rows[1][0], "생태체험")
        self.assertEqual(rows[1][2], "2026-03-20")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_export_grid_tab_xlsx_downloads_binary_file(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:export_grid_tab_xlsx", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertTrue(response.content.startswith(b"PK"))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_import_export_rejects_other_users_access(self):
        self.client.force_login(self.other_user)
        upload = SimpleUploadedFile(
            "sheetbook_import.csv",
            "제목,비용,날짜\n체험학습,15000,2026-03-20".encode("utf-8"),
            content_type="text/csv",
        )
        import_response = self.client.post(
            reverse("sheetbook:import_grid_tab_file", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"file": upload, "has_header": "on"},
        )
        export_response = self.client.get(
            reverse("sheetbook:export_grid_tab_csv", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        )
        self.assertEqual(import_response.status_code, 404)
        self.assertEqual(export_response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_search_returns_tab_cell_action_results(self):
        self.client.force_login(self.user)
        search_tab = SheetTab.objects.create(
            sheetbook=self.sheetbook,
            name="체험 기록",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=3,
        )
        ActionInvocation.objects.create(
            sheetbook=self.sheetbook,
            tab=self.tab,
            actor=self.user,
            action_type=ActionInvocation.ACTION_NOTICE,
            status=ActionInvocation.STATUS_SUCCESS,
            selection_start_row=0,
            selection_start_col=0,
            selection_end_row=0,
            selection_end_col=0,
            selected_cell_count=1,
            summary="체험 준비물 안내",
            result_label="안내문 열기",
        )

        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"q": "체험"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "통합 검색")
        results = response.context["search_results"]
        self.assertGreaterEqual(len(results["tabs"]), 1)
        self.assertGreaterEqual(len(results["cells"]), 1)
        self.assertGreaterEqual(len(results["actions"]), 1)
        self.assertTrue(any(item["title"] == search_tab.name for item in results["tabs"]))
        cell_urls = [item["url"] for item in results["cells"]]
        self.assertTrue(any("focus_row_id=" in url and "focus_col_id=" in url for url in cell_urls))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_search_focus_params_are_embedded_in_grid_editor(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={
                "tab": self.tab.id,
                "q": "생태",
                "focus_row_id": self.row.id,
                "focus_col_id": self.col_title.id,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'data-focus-row-id="{self.row.id}"')
        self.assertContains(response, f'data-focus-col-id="{self.col_title.id}"')
        self.assertContains(response, 'id="sheetbook-search-input"')
        self.assertContains(response, "bindSheetbookSearchShortcut")
        self.assertEqual(response.context["grid_limit"], 1000)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_grid_data_applies_view_filter_and_sort_params(self):
        row2 = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=row2, column=self.col_title, value_text="보건수업")
        SheetCell.objects.create(row=row2, column=self.col_cost, value_number="5000")

        self.client.force_login(self.user)
        filtered_response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"view_filter": "보건"},
        )
        self.assertEqual(filtered_response.status_code, 200)
        filtered_payload = filtered_response.json()
        self.assertEqual(filtered_payload["total_rows"], 1)
        self.assertEqual(filtered_payload["rows"][0]["id"], row2.id)

        sorted_desc_response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"sort_col": self.col_cost.id, "sort_dir": "desc"},
        )
        self.assertEqual(sorted_desc_response.status_code, 200)
        sorted_desc_payload = sorted_desc_response.json()
        self.assertEqual(sorted_desc_payload["rows"][0]["id"], self.row.id)

        sorted_asc_response = self.client.get(
            reverse("sheetbook:grid_data", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"sort_col": self.col_cost.id, "sort_dir": "asc"},
        )
        self.assertEqual(sorted_asc_response.status_code, 200)
        sorted_asc_payload = sorted_asc_response.json()
        self.assertEqual(sorted_asc_payload["rows"][0]["id"], row2.id)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_applies_default_saved_view_context(self):
        default_view = SavedView.objects.create(
            tab=self.tab,
            name="체험만 보기",
            filter_text="생태",
            sort_column=self.col_title,
            sort_direction=SavedView.SORT_ASC,
            is_default=True,
            created_by=self.user,
        )
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["active_saved_view"].id, default_view.id)
        self.assertEqual(response.context["grid_view_filter"], "생태")
        self.assertEqual(response.context["grid_view_sort_column_id"], self.col_title.id)
        self.assertEqual(response.context["grid_view_sort_direction"], "asc")
        self.assertContains(response, 'data-view-filter="생태"')
        self.assertContains(response, f'data-view-sort-column-id="{self.col_title.id}"')

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_includes_mobile_row_editor_controls(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="grid-mobile-row-open-btn"')
        self.assertContains(response, 'id="grid-mobile-row-panel"')
        self.assertContains(response, "모바일 행 편집")
        self.assertContains(response, "touchstart")
        self.assertContains(response, "touchmove")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_shows_mobile_read_only_banner_for_phone_user_agent(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.id},
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
            ),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["sheetbook_mobile_read_only"])
        self.assertContains(response, "휴대폰 읽기 모드")
        self.assertContains(response, 'data-mobile-read-only="1"')

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_keeps_edit_mode_for_tablet_user_agent(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.id},
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (iPad; CPU OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
            ),
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["sheetbook_mobile_read_only"])
        self.assertContains(response, 'data-mobile-read-only="0"')

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_includes_selection_recommendation_parser_script(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.id},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "analyzeSelectionForRecommendations")
        self.assertContains(response, "recommendation_primary")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_update_cell_blocks_phone_user_agent_with_403(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "row_id": self.row.id,
                "column_id": self.col_title.id,
                "value": "휴대폰 수정 시도",
            },
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
            ),
        )
        self.assertEqual(response.status_code, 403)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertTrue(payload["mobile_read_only"])
        self.assertIn("휴대폰", payload["error"])
        title_cell = SheetCell.objects.get(row=self.row, column=self.col_title)
        self.assertEqual(title_cell.value_text, "생태체험")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_create_grid_row_blocks_phone_and_allows_tablet(self):
        self.client.force_login(self.user)
        phone_response = self.client.post(
            reverse("sheetbook:create_grid_row", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
            ),
        )
        self.assertEqual(phone_response.status_code, 403)
        self.assertEqual(self.tab.rows.count(), 1)

        tablet_response = self.client.post(
            reverse("sheetbook:create_grid_row", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            HTTP_USER_AGENT=(
                "Mozilla/5.0 (iPad; CPU OS 17_0 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
            ),
        )
        self.assertEqual(tablet_response.status_code, 200)
        self.assertTrue(tablet_response.json()["ok"])
        self.assertEqual(self.tab.rows.count(), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_saved_view_endpoints_create_and_manage(self):
        self.client.force_login(self.user)
        create_url = reverse("sheetbook:create_saved_view", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        first_create = self.client.post(
            create_url,
            data={
                "name": "상담 일정",
                "view_filter": "생태",
                "sort_col": self.col_title.id,
                "sort_dir": "asc",
                "is_default": "on",
                "is_favorite": "on",
            },
        )
        self.assertEqual(first_create.status_code, 302)
        first_view = SavedView.objects.get(tab=self.tab, name="상담 일정")
        self.assertTrue(first_view.is_default)
        self.assertTrue(first_view.is_favorite)
        self.assertEqual(first_view.sort_column_id, self.col_title.id)

        second_create = self.client.post(
            create_url,
            data={
                "name": "비용 높은 순",
                "sort_col": self.col_cost.id,
                "sort_dir": "desc",
                "is_default": "on",
            },
        )
        self.assertEqual(second_create.status_code, 302)
        second_view = SavedView.objects.get(tab=self.tab, name="비용 높은 순")
        first_view.refresh_from_db()
        self.assertTrue(second_view.is_default)
        self.assertFalse(first_view.is_default)

        favorite_toggle = self.client.post(
            reverse(
                "sheetbook:toggle_saved_view_favorite",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": second_view.pk},
            ),
            data={},
        )
        self.assertEqual(favorite_toggle.status_code, 302)
        second_view.refresh_from_db()
        self.assertTrue(second_view.is_favorite)

        default_toggle = self.client.post(
            reverse(
                "sheetbook:set_saved_view_default",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": second_view.pk},
            ),
            data={},
        )
        self.assertEqual(default_toggle.status_code, 302)
        second_view.refresh_from_db()
        self.assertFalse(second_view.is_default)

        delete_response = self.client.post(
            reverse(
                "sheetbook:delete_saved_view",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": second_view.pk},
            ),
            data={},
        )
        self.assertEqual(delete_response.status_code, 302)
        self.assertFalse(SavedView.objects.filter(pk=second_view.pk).exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_saved_view_endpoints_reject_other_users_access(self):
        saved_view = SavedView.objects.create(
            tab=self.tab,
            name="내부 보기",
            created_by=self.user,
        )
        self.client.force_login(self.other_user)
        create_response = self.client.post(
            reverse("sheetbook:create_saved_view", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={"name": "침입 보기"},
        )
        delete_response = self.client.post(
            reverse(
                "sheetbook:delete_saved_view",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": saved_view.pk},
            ),
            data={},
        )
        favorite_response = self.client.post(
            reverse(
                "sheetbook:toggle_saved_view_favorite",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": saved_view.pk},
            ),
            data={},
        )
        default_response = self.client.post(
            reverse(
                "sheetbook:set_saved_view_default",
                kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk, "view_pk": saved_view.pk},
            ),
            data={},
        )
        self.assertEqual(create_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)
        self.assertEqual(favorite_response.status_code, 404)
        self.assertEqual(default_response.status_code, 404)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_search_suggest_returns_grouped_results(self):
        self.client.force_login(self.user)
        SheetTab.objects.create(
            sheetbook=self.sheetbook,
            name="체험 기록",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=3,
        )
        ActionInvocation.objects.create(
            sheetbook=self.sheetbook,
            tab=self.tab,
            actor=self.user,
            action_type=ActionInvocation.ACTION_NOTICE,
            status=ActionInvocation.STATUS_SUCCESS,
            selection_start_row=0,
            selection_start_col=0,
            selection_end_row=0,
            selection_end_col=0,
            selected_cell_count=1,
            summary="체험 준비물 안내",
            result_label="안내문 열기",
        )
        response = self.client.get(
            reverse("sheetbook:search_suggest"),
            data={"q": "체험", "limit": 5},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertGreaterEqual(len(payload["tabs"]), 1)
        self.assertGreaterEqual(len(payload["cells"]), 1)
        self.assertGreaterEqual(len(payload["actions"]), 1)
        self.assertTrue(any("focus_row_id=" in item["url"] for item in payload["cells"]))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_search_suggest_is_scoped_to_current_user(self):
        other_sheetbook = Sheetbook.objects.create(owner=self.other_user, title="다른 사용자 수첩")
        other_tab = SheetTab.objects.create(
            sheetbook=other_sheetbook,
            name="비밀체험",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=1,
        )
        other_col = SheetColumn.objects.create(
            tab=other_tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=1,
        )
        other_row = SheetRow.objects.create(tab=other_tab, sort_order=1, created_by=self.other_user, updated_by=self.other_user)
        SheetCell.objects.create(row=other_row, column=other_col, value_text="다른사용자전용키워드")

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:search_suggest"),
            data={"q": "다른사용자전용키워드"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["tabs"], [])
        self.assertEqual(payload["cells"], [])
        self.assertEqual(payload["actions"], [])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_search_suggest_requires_login(self):
        response = self.client.get(
            reverse("sheetbook:search_suggest"),
            data={"q": "체험"},
        )
        self.assertEqual(response.status_code, 302)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_creates_updates_and_deletes_events(self):
        SheetCell.objects.create(row=self.row, column=self.col_day, value_date=date(2026, 3, 20))

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["created"], 1)

        event = CalendarEvent.objects.filter(
            author=self.user,
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        ).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.title, "생태체험")
        self.assertTrue(event.is_all_day)

        note_col = SheetColumn.objects.create(
            tab=self.tab,
            key="note",
            label="메모",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "보건수업"},
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=note_col,
            defaults={"value_text": "교실 환기 필수"},
        )

        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertGreaterEqual(payload["updated"], 1)
        event.refresh_from_db()
        self.assertEqual(event.title, "보건수업")
        self.assertEqual(event.blocks.filter(block_type="text").first().content.get("text"), "교실 환기 필수")

        date_cell = SheetCell.objects.get(row=self.row, column=self.col_day)
        date_cell.value_date = None
        date_cell.save(update_fields=["value_date", "updated_at"])
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertGreaterEqual(payload["deleted"], 1)
        self.assertFalse(CalendarEvent.objects.filter(id=event.id).exists())

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_uses_time_columns_when_present(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="start_time",
            label="시작 시간",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        end_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="end_time",
            label="종료 시간",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=5,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="09:10")
        SheetCell.objects.create(row=self.row, column=end_time_col, value_text="10:00")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)
        self.assertEqual(payload["all_day_synced"], 0)

        event = CalendarEvent.objects.filter(
            author=self.user,
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        ).first()
        self.assertIsNotNone(event)
        self.assertFalse(event.is_all_day)
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (9, 10))
        self.assertEqual((local_end.hour, local_end.minute), (10, 0))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_handles_mixed_all_day_and_timed_rows(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="start_time",
            label="시간",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="13:20")

        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="운동회")
        SheetCell.objects.create(row=second_row, column=self.col_day, value_date=date(2026, 3, 21))

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)
        self.assertEqual(payload["all_day_synced"], 1)

        event_row1 = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        event_row2 = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{second_row.id}",
        )
        self.assertFalse(event_row1.is_all_day)
        self.assertTrue(event_row2.is_all_day)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_parses_korean_meridiem_time_notation(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="start_time",
            label="시간",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="오후 1시 20분")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)
        self.assertEqual(payload["all_day_synced"], 0)

        event = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (13, 20))
        self.assertEqual((local_end.hour, local_end.minute), (14, 10))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_parses_school_period_notation(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="period",
            label="교시",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="3교시")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)

        event = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (11, 0))
        self.assertEqual((local_end.hour, local_end.minute), (11, 50))

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=8,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=45,
    )
    def test_sync_calendar_from_schedule_uses_configured_period_start_hour(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="period",
            label="교시",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="3교시")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)

        event = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (10, 0))
        self.assertEqual((local_end.hour, local_end.minute), (10, 45))

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=8,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=10,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=45,
    )
    def test_sync_calendar_from_schedule_uses_configured_period_start_minute(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="period",
            label="교시",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="3교시")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)

        event = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (10, 10))
        self.assertEqual((local_end.hour, local_end.minute), (10, 55))

    @override_settings(SHEETBOOK_ENABLED=True, SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=45)
    def test_sync_calendar_from_schedule_uses_configured_default_duration(self):
        start_time_col = SheetColumn.objects.create(
            tab=self.tab,
            key="start_time",
            label="시간",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 3, 20)},
        )
        SheetCell.objects.create(row=self.row, column=start_time_col, value_text="09:00")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["timed_synced"], 1)

        event = CalendarEvent.objects.get(
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{self.sheetbook.id}:{self.tab.id}:{self.row.id}",
        )
        local_start = timezone.localtime(event.start_time)
        local_end = timezone.localtime(event.end_time)
        self.assertEqual((local_start.hour, local_start.minute), (9, 0))
        self.assertEqual((local_end.hour, local_end.minute), (9, 45))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_sync_calendar_from_schedule_rejects_non_calendar_tab(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("달력 탭", response.json().get("error", ""))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_calendar_creates_events_and_log(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_day,
            defaults={"value_date": date(2026, 4, 1)},
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "calendar",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 2,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_CALENDAR)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("달력", invocation.result_label)
        self.assertEqual(invocation.selected_cell_count, 3)

        event = CalendarEvent.objects.filter(
            author=self.user,
            integration_source="sheetbook_action_calendar",
            integration_key__contains=f":{invocation.id}:",
        ).first()
        self.assertIsNotNone(event)
        self.assertEqual(event.title, "생태체험")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_collect_creates_request_and_log(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하나"},
        )
        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="이둘")
        SheetCell.objects.create(row=second_row, column=self.col_cost, value_number="23000")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "collect",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 1,
                    "end_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_COLLECT)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("/collect/", invocation.result_url)

        created_request = CollectionRequest.objects.get(creator=self.user)
        self.assertIn("김하나", created_request.expected_submitters)
        self.assertIn("이둘", created_request.expected_submitters)
        self.assertIn(str(created_request.id), invocation.result_url)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_collect_non_json_redirects_with_message(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하나"},
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "action": "collect",
                "start_row_index": 0,
                "start_col_index": 0,
                "end_row_index": 0,
                "end_col_index": 1,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "간편 수합을 만들었어요")
        self.assertTrue(
            ActionInvocation.objects.filter(
                sheetbook=self.sheetbook,
                tab=self.tab,
                action_type=ActionInvocation.ACTION_COLLECT,
                status=ActionInvocation.STATUS_SUCCESS,
            ).exists()
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_handoff_creates_group_and_session(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김교사"},
        )
        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="박교사")
        SheetCell.objects.create(row=second_row, column=self.col_cost, value_number="1")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "handoff",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 1,
                    "end_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_HANDOFF)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("/handoff/sessions/", invocation.result_url)

        group = HandoffRosterGroup.objects.get(owner=self.user)
        self.assertEqual(group.members.count(), 2)
        self.assertTrue(group.members.filter(display_name="김교사").exists())
        self.assertTrue(group.members.filter(display_name="박교사").exists())

        session = HandoffSession.objects.get(owner=self.user)
        self.assertEqual(session.roster_group_id, group.id)
        self.assertEqual(session.receipts.count(), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_consent_returns_guide_link(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_CONSENT)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("/actions/consent/review/", invocation.result_url)
        parsed = urlparse(invocation.result_url)
        query = parse_qs(parsed.query)
        self.assertTrue(query.get("sb_seed"))
        self.assertEqual(invocation.payload.get("prefilled_recipients"), 1)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_signature_returns_prefilled_link(self):
        self.col_title.label = "학생명"
        self.col_title.save(update_fields=["label"])
        affiliation_col = SheetColumn.objects.create(
            tab=self.tab,
            key="student_class",
            label="학년반",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=affiliation_col,
            defaults={"value_text": "3-1"},
        )
        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="박나래")
        SheetCell.objects.create(row=second_row, column=affiliation_col, value_text="3-2")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "signature",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 1,
                    "end_col_index": 4,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_SIGNATURE)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("/signatures/create/", invocation.result_url)
        parsed = urlparse(invocation.result_url)
        query = parse_qs(parsed.query)
        token = (query.get("sb_seed") or [""])[0]
        self.assertTrue(token)

        seeds = self.client.session.get("sheetbook_action_seeds", {})
        seed_data = ((seeds.get(token) or {}).get("data") or {})
        participants_text = seed_data.get("participants_text", "")
        self.assertIn("김하늘,3-1", participants_text)
        self.assertIn("박나래,3-2", participants_text)
        self.assertEqual(invocation.payload.get("prefilled_participants"), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_notice_returns_prefilled_link(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "준비물: 실내화"},
        )
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "notice",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.action_type, ActionInvocation.ACTION_NOTICE)
        self.assertEqual(invocation.status, ActionInvocation.STATUS_SUCCESS)
        self.assertIn("/noticegen/", invocation.result_url)
        parsed = urlparse(invocation.result_url)
        query = parse_qs(parsed.query)
        self.assertTrue(query.get("sb_seed"))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_consent_prefill_detects_parent_and_phone(self):
        self.col_title.label = "학생명"
        self.col_title.save(update_fields=["label"])
        parent_col = SheetColumn.objects.create(
            tab=self.tab,
            key="parent_name",
            label="학부모명",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=4,
        )
        phone_col = SheetColumn.objects.create(
            tab=self.tab,
            key="contact_phone",
            label="연락처",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=5,
        )

        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=parent_col,
            defaults={"value_text": "김하늘 어머니"},
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=phone_col,
            defaults={"value_text": "010-1234-5678"},
        )

        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="박나래")
        SheetCell.objects.create(row=second_row, column=phone_col, value_text="010 9988 7766")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 1,
                    "end_col_index": 4,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])

        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        parsed = urlparse(invocation.result_url)
        query = parse_qs(parsed.query)
        token = (query.get("sb_seed") or [""])[0]
        self.assertTrue(token)

        seeds = self.client.session.get("sheetbook_action_seeds", {})
        seed_data = ((seeds.get(token) or {}).get("data") or {})
        recipients_text = seed_data.get("recipients_text", "")
        self.assertIn("김하늘,김하늘 어머니,01012345678", recipients_text)
        self.assertIn("박나래,박나래 보호자,01099887766", recipients_text)
        self.assertEqual(invocation.payload.get("prefilled_recipients"), 2)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_consent_prefill_ignores_number_column_for_student_name(self):
        number_col = SheetColumn.objects.create(
            tab=self.tab,
            key="student_no",
            label="번호",
            column_type=SheetColumn.TYPE_NUMBER,
            sort_order=0,
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=number_col,
            defaults={"value_number": "1"},
        )
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "한지민"},
        )

        second_row = SheetRow.objects.create(tab=self.tab, sort_order=2, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=second_row, column=number_col, value_number="2")
        SheetCell.objects.create(row=second_row, column=self.col_title, value_text="윤서준")

        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 1,
                    "end_col_index": 2,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])

        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        parsed = urlparse(invocation.result_url)
        query = parse_qs(parsed.query)
        token = (query.get("sb_seed") or [""])[0]
        self.assertTrue(token)

        seeds = self.client.session.get("sheetbook_action_seeds", {})
        seed_data = ((seeds.get(token) or {}).get("data") or {})
        recipients_text = seed_data.get("recipients_text", "")
        self.assertIn("한지민,한지민 보호자,", recipients_text)
        self.assertIn("윤서준,윤서준 보호자,", recipients_text)
        self.assertNotIn("1,1 보호자,", recipients_text)
        self.assertNotIn("2,2 보호자,", recipients_text)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_consent_seed_review_get_renders_prefilled_context(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        self.client.force_login(self.user)
        execute_response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(execute_response.status_code, 200)
        invocation = ActionInvocation.objects.get(id=execute_response.json()["invocation"]["id"])

        review_response = self.client.get(invocation.result_url)
        self.assertEqual(review_response.status_code, 200)
        self.assertContains(review_response, "동의서 보내기 전에 확인")
        self.assertContains(review_response, "김하늘,김하늘 보호자,")
        self.assertEqual(review_response.context["recipients_count"], 1)
        self.assertEqual(review_response.context["initial_title"], f"{self.tab.name} 동의서")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_consent_seed_review_shows_recipient_parse_summary(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        self.client.force_login(self.user)
        execute_response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        invocation = ActionInvocation.objects.get(id=execute_response.json()["invocation"]["id"])
        parsed_review = urlparse(invocation.result_url)
        token = (parse_qs(parsed_review.query).get("sb_seed") or [""])[0]

        session = self.client.session
        seeds = session.get("sheetbook_action_seeds", {})
        seeds[token]["data"]["recipients_text"] = (
            "김하늘,김하늘 보호자,01012345678\n"
            "김하늘,김하늘 보호자,01012345678\n"
            "형식 오류 줄"
        )
        session["sheetbook_action_seeds"] = seeds
        session.save()

        review_response = self.client.get(invocation.result_url)
        self.assertEqual(review_response.status_code, 200)
        self.assertEqual(review_response.context["recipients_input_line_count"], 3)
        self.assertEqual(review_response.context["recipients_count"], 1)
        self.assertEqual(review_response.context["recipients_duplicate_count"], 1)
        self.assertEqual(review_response.context["recipients_skipped_count"], 1)
        self.assertEqual(
            review_response.context["recipients_duplicate_samples"],
            ["김하늘,김하늘 보호자,01012345678"],
        )
        self.assertEqual(
            review_response.context["recipients_skipped_samples"],
            ["형식 오류 줄"],
        )
        self.assertEqual(review_response.context["recipients_issue_line_numbers"], [2, 3])
        preview_items = review_response.context["recipients_issue_preview_items"]
        self.assertEqual(len(preview_items), 2)
        self.assertEqual(preview_items[0]["line_no"], 2)
        self.assertEqual(preview_items[0]["issue_type"], "duplicate")
        self.assertEqual(preview_items[1]["line_no"], 3)
        self.assertEqual(preview_items[1]["issue_type"], "format")
        self.assertContains(review_response, "입력 3줄 중 1명 반영")
        self.assertContains(review_response, "중복으로 제외된 줄 예시")
        self.assertContains(review_response, "형식 확인이 필요한 줄 예시")
        self.assertContains(review_response, 'data-recipients-line="2"')
        self.assertContains(review_response, 'data-recipients-line="3"')
        self.assertContains(review_response, 'id="recipients-issue-minimap"')
        self.assertContains(review_response, 'id="recipients-active-line"')
        self.assertContains(review_response, 'data-testid="recipients-textarea"')
        self.assertContains(review_response, 'id="recipients-cleanup-btn"')
        self.assertContains(review_response, 'data-testid="recipients-cleanup-btn"')
        self.assertContains(review_response, 'id="recipients-cleanup-undo-btn"')
        self.assertContains(review_response, 'data-testid="recipients-cleanup-undo-btn"')
        self.assertContains(review_response, 'id="recipients-cleanup-undo-used"')
        self.assertContains(review_response, 'id="recipients-copy-issues-btn"')
        self.assertContains(review_response, 'data-testid="recipients-copy-issues-btn"')
        self.assertContains(review_response, 'id="recipients-prev-issue-btn"')
        self.assertContains(review_response, 'data-testid="recipients-prev-issue-btn"')
        self.assertContains(review_response, 'id="recipients-next-issue-btn"')
        self.assertContains(review_response, 'data-testid="recipients-next-issue-btn"')
        self.assertContains(review_response, 'data-testid="recipients-jump-top-btn"')
        self.assertContains(review_response, 'data-testid="recipients-jump-bottom-btn"')
        self.assertContains(review_response, 'id="recipients-action-guide"')
        self.assertContains(review_response, 'id="recipients-issue-tip"')
        self.assertContains(review_response, 'id="recipients-cleanup-status"')
        self.assertContains(review_response, 'id="recipients-issue-copy-used"')
        self.assertContains(review_response, 'id="recipients-issue-jump-count"')
        self.assertContains(review_response, 'id="recipients-submit-btn"')
        self.assertContains(review_response, 'data-testid="recipients-submit-btn"')
        self.assertContains(review_response, "문제 줄 제외하고 정리")
        self.assertContains(review_response, "문제 줄만 복사")
        self.assertContains(review_response, "이전 문제 줄")
        self.assertContains(review_response, "다음 문제 줄")
        review_html = review_response.content.decode("utf-8")
        cleanup_idx = review_html.find('id="recipients-cleanup-btn"')
        undo_idx = review_html.find('id="recipients-cleanup-undo-btn"')
        copy_idx = review_html.find('id="recipients-copy-issues-btn"')
        prev_idx = review_html.find('id="recipients-prev-issue-btn"')
        next_idx = review_html.find('id="recipients-next-issue-btn"')
        jump_top_idx = review_html.find('data-testid="recipients-jump-top-btn"')
        jump_bottom_idx = review_html.find('data-testid="recipients-jump-bottom-btn"')
        self.assertTrue(0 <= cleanup_idx < undo_idx < copy_idx)
        self.assertTrue(0 <= prev_idx < next_idx < jump_top_idx < jump_bottom_idx)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_consent_seed_review_post_updates_seed_and_redirects_step1(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        self.client.force_login(self.user)
        execute_response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        invocation = ActionInvocation.objects.get(id=execute_response.json()["invocation"]["id"])
        parsed_review = urlparse(invocation.result_url)
        token = (parse_qs(parsed_review.query).get("sb_seed") or [""])[0]
        self.assertTrue(token)

        review_url = reverse("sheetbook:consent_seed_review", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        response = self.client.post(
            review_url,
            data={
                "sb_seed": token,
                "title": "학부모 회신 동의서",
                "message": "가정에서 확인 후 회신 부탁드립니다.",
                "document_title": "가정통신문",
                "recipients_text": "김하늘,김하늘 어머니,010-1234-5678\n김하늘,김하늘 어머니,01012345678",
                "recipients_cleanup_applied": "1",
                "recipients_cleanup_removed_count": "1",
                "recipients_cleanup_undo_used": "1",
                "recipients_issue_copy_used": "1",
                "recipients_issue_jump_count": "3",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/consent/create/step1/", response["Location"])
        redirected_query = parse_qs(urlparse(response["Location"]).query)
        self.assertEqual((redirected_query.get("sb_seed") or [""])[0], token)

        seeds = self.client.session.get("sheetbook_action_seeds", {})
        seed_data = ((seeds.get(token) or {}).get("data") or {})
        self.assertEqual(seed_data.get("title"), "학부모 회신 동의서")
        self.assertEqual(seed_data.get("message"), "가정에서 확인 후 회신 부탁드립니다.")
        self.assertEqual(seed_data.get("document_title"), "가정통신문")
        self.assertEqual(seed_data.get("recipients_text"), "김하늘,김하늘 어머니,01012345678")
        submitted_event = SheetbookMetricEvent.objects.filter(
            event_name="consent_review_submitted",
            user=self.user,
            sheetbook=self.sheetbook,
            tab=self.tab,
        ).order_by("-id").first()
        self.assertIsNotNone(submitted_event)
        self.assertTrue(submitted_event.metadata.get("recipients_cleanup_applied"))
        self.assertEqual(submitted_event.metadata.get("recipients_cleanup_removed_count"), 1)
        self.assertTrue(submitted_event.metadata.get("recipients_cleanup_undo_used"))
        self.assertTrue(submitted_event.metadata.get("recipients_issue_copy_used"))
        self.assertEqual(submitted_event.metadata.get("recipients_issue_jump_count"), 3)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_consent_seed_review_post_clamps_cleanup_removed_count_metric(self):
        SheetCell.objects.update_or_create(
            row=self.row,
            column=self.col_title,
            defaults={"value_text": "김하늘"},
        )
        self.client.force_login(self.user)
        execute_response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "consent",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        invocation = ActionInvocation.objects.get(id=execute_response.json()["invocation"]["id"])
        token = (parse_qs(urlparse(invocation.result_url).query).get("sb_seed") or [""])[0]
        self.assertTrue(token)

        review_url = reverse("sheetbook:consent_seed_review", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        response = self.client.post(
            review_url,
            data={
                "sb_seed": token,
                "title": "학부모 회신 동의서",
                "message": "가정에서 확인 후 회신 부탁드립니다.",
                "document_title": "가정통신문",
                "recipients_text": "김하늘,김하늘 어머니,01012345678\n김하늘,김하늘 어머니,01012345678\n형식 오류 줄",
                "recipients_cleanup_applied": "1",
                "recipients_cleanup_removed_count": "999",
            },
        )
        self.assertEqual(response.status_code, 302)
        submitted_event = SheetbookMetricEvent.objects.filter(
            event_name="consent_review_submitted",
            user=self.user,
            sheetbook=self.sheetbook,
            tab=self.tab,
        ).order_by("-id").first()
        self.assertIsNotNone(submitted_event)
        self.assertTrue(submitted_event.metadata.get("recipients_cleanup_applied"))
        self.assertEqual(submitted_event.metadata.get("recipients_cleanup_removed_count"), 2)
        self.assertFalse(submitted_event.metadata.get("recipients_cleanup_undo_used"))
        self.assertFalse(submitted_event.metadata.get("recipients_issue_copy_used"))
        self.assertEqual(submitted_event.metadata.get("recipients_issue_jump_count"), 0)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_consent_seed_review_missing_seed_redirects_back(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:consent_seed_review", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "동의서 준비 정보를 찾지 못했어요")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_rejects_unknown_action(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "unsupported_action",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 0,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("알 수 없는", response.json().get("error", ""))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_rejects_unknown_action_non_json_with_redirect(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data={
                "action": "unsupported_action",
                "start_row_index": 0,
                "start_col_index": 0,
                "end_row_index": 0,
                "end_col_index": 0,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "알 수 없는 실행 버튼입니다")
        self.assertEqual(
            ActionInvocation.objects.filter(sheetbook=self.sheetbook, tab=self.tab).count(),
            0,
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    @patch("sheetbook.views._execute_collect_creation", side_effect=RuntimeError("boom"))
    def test_execute_grid_action_collect_unexpected_error_returns_action_specific_message(self, mocked_collect):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "collect",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 500)
        payload = response.json()
        self.assertFalse(payload["ok"])
        self.assertIn("간편 수합 만들기 중 문제가 생겼어요", payload.get("error", ""))

        invocation = ActionInvocation.objects.get(id=payload["invocation"]["id"])
        self.assertEqual(invocation.status, ActionInvocation.STATUS_FAILED)
        self.assertIn("간편 수합 만들기 중 문제가 생겼어요", invocation.summary)
        mocked_collect.assert_called_once()

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_grid_renders_action_layer_ui(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.pk},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="grid-action-layer"')
        self.assertContains(response, 'data-grid-action="calendar"')
        self.assertContains(response, 'data-grid-action="collect"')
        self.assertContains(response, 'data-grid-action="consent"')
        self.assertContains(response, 'data-grid-action="signature"')
        self.assertContains(response, 'data-grid-action="handoff"')
        self.assertContains(response, 'data-grid-action="notice"')
        self.assertContains(response, 'id="grid-action-preview-modal"')
        self.assertContains(response, 'id="grid-action-history"')
        self.assertContains(response, "바로 만들기 전에 확인")
        self.assertContains(response, "선택한 칸")
        self.assertContains(response, "최근 만든 기록")
        self.assertContains(response, 'data-history-filter="success"')
        self.assertContains(response, 'id="grid-action-history-more"')
        self.assertContains(response, "data-action-history-url=")
        self.assertContains(response, "data-sheetbook-id=")
        self.assertContains(response, "data-tab-id=")
        self.assertContains(response, 'id="grid-retry-failed-btn"')
        self.assertContains(response, 'id="grid-save-guide"')
        self.assertContains(response, "저장 안 된 칸 다시 저장")
        self.assertContains(response, "<noscript>")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_grid_uses_sanitized_grid_limit_for_smoke(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.pk, "grid_limit": "5000"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-grid-limit="1000"')

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_action_history_endpoint_returns_paginated_items(self):
        for idx in range(11):
            ActionInvocation.objects.create(
                sheetbook=self.sheetbook,
                tab=self.tab,
                actor=self.user,
                action_type=ActionInvocation.ACTION_COLLECT,
                status=ActionInvocation.STATUS_SUCCESS,
                selection_start_row=0,
                selection_start_col=0,
                selection_end_row=0,
                selection_end_col=0,
                selected_cell_count=1,
                summary=f"기록 {idx}",
            )

        self.client.force_login(self.user)
        url = reverse("sheetbook:action_history", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        response = self.client.get(url, data={"limit": 4})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(len(payload["items"]), 4)
        self.assertTrue(payload["has_more"])
        self.assertTrue(payload["next_cursor"])
        self.assertEqual(payload["items"][0]["summary"], "기록 10")

        response2 = self.client.get(
            url,
            data={
                "limit": 4,
                "cursor_id": payload["next_cursor"],
            },
        )
        self.assertEqual(response2.status_code, 200)
        payload2 = response2.json()
        self.assertTrue(payload2["ok"])
        self.assertEqual(len(payload2["items"]), 4)
        self.assertTrue(payload2["has_more"])
        self.assertEqual(payload2["items"][0]["summary"], "기록 6")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_action_history_endpoint_returns_last_page_without_more(self):
        for idx in range(3):
            ActionInvocation.objects.create(
                sheetbook=self.sheetbook,
                tab=self.tab,
                actor=self.user,
                action_type=ActionInvocation.ACTION_NOTICE,
                status=ActionInvocation.STATUS_FAILED,
                selection_start_row=0,
                selection_start_col=0,
                selection_end_row=0,
                selection_end_col=0,
                selected_cell_count=1,
                summary=f"마지막 {idx}",
            )

        self.client.force_login(self.user)
        url = reverse("sheetbook:action_history", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk})
        response = self.client.get(url, data={"limit": 8})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(len(payload["items"]), 3)
        self.assertFalse(payload["has_more"])
        self.assertIsNone(payload["next_cursor"])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_action_history_endpoint_rejects_calendar_tab(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:action_history", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.calendar_tab.pk})
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("그리드", response.json().get("error", ""))

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_shows_history_more_button_when_items_exceed_initial_limit(self):
        for idx in range(10):
            ActionInvocation.objects.create(
                sheetbook=self.sheetbook,
                tab=self.tab,
                actor=self.user,
                action_type=ActionInvocation.ACTION_CALENDAR,
                status=ActionInvocation.STATUS_SUCCESS,
                selection_start_row=0,
                selection_start_col=0,
                selection_end_row=0,
                selection_end_col=0,
                selected_cell_count=1,
                summary=f"이력 {idx}",
            )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"tab": self.tab.pk},
        )
        self.assertEqual(response.status_code, 200)
        html = response.content.decode("utf-8")
        self.assertIn('id="grid-action-history-more"', html)
        self.assertNotRegex(html, r'id="grid-action-history-more"[^>]*\shidden\b')


class SheetbookP0IntegrationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_p0_owner",
            password="pw123456",
            email="sheetbook_p0_owner@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_p0_owner", "role": "school"},
        )

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_teacher_core_flow_create_edit_action_and_sync(self):
        self.client.force_login(self.user)

        index_response = self.client.get(
            reverse("sheetbook:index"),
            data={"source": "workspace_home_create"},
        )
        self.assertEqual(index_response.status_code, 200)

        create_response = self.client.post(
            reverse("sheetbook:create"),
            data={
                "title": "2-3반 운영 수첩",
                "academic_year": 2026,
                "source": "workspace_home_create",
            },
        )
        self.assertEqual(create_response.status_code, 302)

        sheetbook = Sheetbook.objects.filter(owner=self.user, title="2-3반 운영 수첩").first()
        self.assertIsNotNone(sheetbook)
        schedule_tab = sheetbook.tabs.filter(name="일정").first()
        calendar_tab = sheetbook.tabs.filter(name="달력").first()
        self.assertIsNotNone(schedule_tab)
        self.assertIsNotNone(calendar_tab)

        columns = {column.key: column for column in schedule_tab.columns.all()}
        row = schedule_tab.rows.order_by("sort_order", "id").first()
        self.assertIsNotNone(row)

        update_title = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": sheetbook.pk, "tab_pk": schedule_tab.pk}),
            data={
                "row_id": row.id,
                "column_id": columns["title"].id,
                "value": "학부모 상담",
            },
        )
        self.assertEqual(update_title.status_code, 200)

        update_date = self.client.post(
            reverse("sheetbook:update_cell", kwargs={"pk": sheetbook.pk, "tab_pk": schedule_tab.pk}),
            data={
                "row_id": row.id,
                "column_id": columns["date"].id,
                "value": "2026-03-10",
            },
        )
        self.assertEqual(update_date.status_code, 200)

        action_response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": sheetbook.pk, "tab_pk": schedule_tab.pk}),
            data=json.dumps(
                {
                    "action": "calendar",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 2,
                    "entry_source": "workspace_home_recent",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(action_response.status_code, 200)
        self.assertTrue(action_response.json()["ok"])

        sync_response = self.client.post(
            reverse("sheetbook:sync_calendar_from_schedule", kwargs={"pk": sheetbook.pk, "tab_pk": calendar_tab.pk}),
            data={},
        )
        self.assertEqual(sync_response.status_code, 200)
        self.assertTrue(sync_response.json()["ok"])
        self.assertGreaterEqual(sync_response.json()["synced"], 1)

        history_response = self.client.get(
            reverse("sheetbook:action_history", kwargs={"pk": sheetbook.pk, "tab_pk": schedule_tab.pk}),
            data={"limit": 8},
        )
        self.assertEqual(history_response.status_code, 200)
        self.assertTrue(history_response.json()["ok"])
        self.assertGreaterEqual(len(history_response.json()["items"]), 1)

        sync_event = CalendarEvent.objects.filter(
            author=self.user,
            integration_source="sheetbook_schedule_sync",
            integration_key=f"{sheetbook.id}:{schedule_tab.id}:{row.id}",
        ).first()
        self.assertIsNotNone(sync_event)

        created_metric = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=sheetbook,
        ).order_by("-id").first()
        self.assertIsNotNone(created_metric)
        self.assertEqual(created_metric.metadata.get("entry_source"), "workspace_home_create")

        action_metric = SheetbookMetricEvent.objects.filter(
            event_name="action_execute_succeeded",
            user=self.user,
            action_type="calendar",
        ).order_by("-id").first()
        self.assertIsNotNone(action_metric)
        self.assertEqual(action_metric.metadata.get("entry_source"), "workspace_home_recent")


class SheetbookMetricTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_metric_owner",
            password="pw123456",
            email="sheetbook_metric_owner@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_metric_owner", "role": "school"},
        )
        self.admin_user = User.objects.create_superuser(
            username="sheetbook_metric_admin",
            password="pw123456",
            email="sheetbook_metric_admin@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.admin_user,
            defaults={"nickname": "sheetbook_metric_admin", "role": "school"},
        )
        self.sheetbook = Sheetbook.objects.create(owner=self.user, title="지표 테스트 수첩")
        self.tab = SheetTab.objects.create(sheetbook=self.sheetbook, name="일정", tab_type=SheetTab.TYPE_GRID, sort_order=1)
        self.col_date = SheetColumn.objects.create(
            tab=self.tab,
            key="date",
            label="날짜",
            column_type=SheetColumn.TYPE_DATE,
            sort_order=1,
        )
        self.col_title = SheetColumn.objects.create(
            tab=self.tab,
            key="title",
            label="제목",
            column_type=SheetColumn.TYPE_TEXT,
            sort_order=2,
        )
        self.row = SheetRow.objects.create(tab=self.tab, sort_order=1, created_by=self.user, updated_by=self.user)
        SheetCell.objects.create(row=self.row, column=self.col_date, value_date=date(2026, 3, 15))
        SheetCell.objects.create(row=self.row, column=self.col_title, value_text="상담 주간")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_index_and_create_store_metric_events(self):
        self.client.force_login(self.user)
        self.client.get(reverse("sheetbook:index"), data={"source": "workspace_home_create"})
        self.client.post(
            reverse("sheetbook:create"),
            data={
                "title": "새 지표 수첩",
                "academic_year": 2026,
                "source": "workspace_home_create",
            },
        )

        index_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_index_opened",
            user=self.user,
        ).order_by("-id").first()
        self.assertIsNotNone(index_event)
        self.assertEqual(index_event.metadata.get("entry_source"), "workspace_home_create")
        created_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_created",
            user=self.user,
        ).order_by("-id").first()
        self.assertIsNotNone(created_event)
        self.assertIsNotNone(created_event.sheetbook_id)
        self.assertEqual(created_event.metadata.get("entry_source"), "workspace_home_create")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_detail_stores_entry_source_in_metric_event(self):
        self.client.force_login(self.user)
        self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"source": "workspace_home_recent"},
        )
        detail_event = SheetbookMetricEvent.objects.filter(
            event_name="sheetbook_detail_opened",
            user=self.user,
        ).order_by("-id").first()
        self.assertIsNotNone(detail_event)
        self.assertEqual(detail_event.sheetbook_id, self.sheetbook.id)
        self.assertEqual(detail_event.metadata.get("entry_source"), "workspace_home_recent")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_persists_metric_events(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "calendar",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 1,
                    "entry_source": "workspace_home_recent",
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        requested_event = (
            SheetbookMetricEvent.objects.filter(
                event_name="action_execute_requested",
                user=self.user,
                action_type="calendar",
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(requested_event)
        self.assertEqual(requested_event.metadata.get("entry_source"), "workspace_home_recent")
        succeeded_event = (
            SheetbookMetricEvent.objects.filter(
                event_name="action_execute_succeeded",
                user=self.user,
                action_type="calendar",
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(succeeded_event)
        self.assertEqual(succeeded_event.metadata.get("entry_source"), "workspace_home_recent")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_persists_recommendation_metric_metadata(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "calendar",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 1,
                    "recommendation_primary": "calendar",
                    "recommendation_signals": {
                        "token_count": 4,
                        "date_ratio": 0.5,
                        "phone_ratio": 0.0,
                        "name_ratio": 0.25,
                    },
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        requested_event = (
            SheetbookMetricEvent.objects.filter(
                event_name="action_execute_requested",
                user=self.user,
                action_type="calendar",
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(requested_event)
        self.assertEqual(requested_event.metadata.get("recommended_action"), "calendar")
        self.assertEqual(requested_event.metadata.get("recommendation_token_count"), 4)
        self.assertEqual(requested_event.metadata.get("recommendation_date_ratio"), 0.5)
        self.assertEqual(requested_event.metadata.get("recommendation_phone_ratio"), 0.0)
        self.assertEqual(requested_event.metadata.get("recommendation_name_ratio"), 0.25)

        invocation = ActionInvocation.objects.filter(
            sheetbook=self.sheetbook,
            tab=self.tab,
            action_type=ActionInvocation.ACTION_CALENDAR,
        ).order_by("-id").first()
        self.assertIsNotNone(invocation)
        recommendation_meta = (invocation.payload or {}).get("recommendation", {})
        self.assertEqual(recommendation_meta.get("primary_action"), "calendar")
        self.assertEqual(recommendation_meta.get("token_count"), 4)
        self.assertEqual(recommendation_meta.get("date_ratio"), 0.5)
        self.assertEqual(recommendation_meta.get("name_ratio"), 0.25)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_execute_grid_action_uses_recent_sheetbook_entry_source_from_session(self):
        self.client.force_login(self.user)
        self.client.get(
            reverse("sheetbook:detail", kwargs={"pk": self.sheetbook.pk}),
            data={"source": "workspace_home_recent"},
        )
        response = self.client.post(
            reverse("sheetbook:execute_grid_action", kwargs={"pk": self.sheetbook.pk, "tab_pk": self.tab.pk}),
            data=json.dumps(
                {
                    "action": "calendar",
                    "start_row_index": 0,
                    "start_col_index": 0,
                    "end_row_index": 0,
                    "end_col_index": 1,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        requested_event = (
            SheetbookMetricEvent.objects.filter(
                event_name="action_execute_requested",
                user=self.user,
                action_type="calendar",
            )
            .order_by("-id")
            .first()
        )
        self.assertIsNotNone(requested_event)
        self.assertEqual(requested_event.metadata.get("entry_source"), "workspace_home_recent")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_is_superuser_only(self):
        SheetbookMetricEvent.objects.create(
            event_name="sheetbook_created",
            user=self.user,
            sheetbook=self.sheetbook,
            metadata={"source": "test"},
        )

        self.client.force_login(self.user)
        denied = self.client.get(reverse("sheetbook:metrics_dashboard"))
        self.assertEqual(denied.status_code, 404)

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "교무수첩 사용 지표")
        self.assertGreaterEqual(response.context["summary"]["event_total"], 1)

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_calculates_revisit_and_quick_create_rate(self):
        other_user = User.objects.create_user(
            username="sheetbook_metric_other",
            password="pw123456",
            email="sheetbook_metric_other@example.com",
        )
        UserProfile.objects.update_or_create(
            user=other_user,
            defaults={"nickname": "sheetbook_metric_other", "role": "school"},
        )
        base_ts = timezone.now() - timedelta(days=2)

        def _create_event(user, event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        _create_event(self.user, "workspace_home_opened", -5)
        _create_event(self.user, "sheetbook_index_opened", 0, metadata={"entry_source": "workspace_home_create"})
        _create_event(self.user, "sheetbook_created", 5, metadata={"entry_source": "workspace_home_create"})
        _create_event(self.user, "sheetbook_archived", 6)
        _create_event(self.user, "sheetbook_unarchived", 7)
        _create_event(self.user, "sheetbook_archive_read_mode_opened", 8)
        _create_event(
            self.user,
            "sheetbook_archive_bulk_updated",
            9,
            metadata={
                "archive_action": "archive",
                "selected_count": 4,
                "matched_count": 4,
                "changed_count": 3,
                "unchanged_count": 1,
                "ignored_count": 0,
            },
        )
        _create_event(self.user, "sheetbook_detail_opened", 10)
        _create_event(self.user, "action_execute_requested", 12, metadata={"entry_source": "workspace_home_create"})
        _create_event(self.user, "action_execute_succeeded", 13, metadata={"entry_source": "workspace_home_create"})
        _create_event(self.user, "sheetbook_detail_opened", 24 * 60 + 10)

        _create_event(other_user, "workspace_home_opened", 20)
        _create_event(other_user, "sheetbook_detail_opened", 25, metadata={"entry_source": "workspace_home_recent"})

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        summary = response.context["summary"]
        self.assertEqual(summary["quick_create_base_count"], 2)
        self.assertEqual(summary["quick_create_user_count"], 1)
        self.assertEqual(summary["quick_create_rate"], 50.0)
        self.assertEqual(summary["revisitor_user_count"], 1)
        self.assertEqual(summary["revisit_rate"], 50.0)
        self.assertEqual(summary["workspace_home_opened_count"], 2)
        self.assertEqual(summary["workspace_source_index_count"], 1)
        self.assertEqual(summary["workspace_source_detail_count"], 1)
        self.assertEqual(summary["workspace_source_create_count"], 1)
        self.assertEqual(summary["workspace_source_action_requested_count"], 1)
        self.assertEqual(summary["workspace_source_action_success_count"], 1)
        self.assertEqual(summary["sheetbook_archived_count"], 1)
        self.assertEqual(summary["sheetbook_unarchived_count"], 1)
        self.assertEqual(summary["sheetbook_archive_bulk_updated_count"], 1)
        self.assertEqual(summary["sheetbook_bulk_archive_changed_count"], 3)
        self.assertEqual(summary["sheetbook_bulk_unarchive_changed_count"], 0)
        self.assertEqual(summary["sheetbook_bulk_ignored_count"], 0)
        self.assertEqual(summary["sheetbook_bulk_unchanged_count"], 1)
        self.assertEqual(summary["sheetbook_archive_read_mode_opened_count"], 1)
        self.assertEqual(summary["workspace_to_index_rate"], 50.0)
        self.assertEqual(summary["workspace_to_detail_rate"], 50.0)
        self.assertEqual(summary["workspace_to_create_rate"], 50.0)
        self.assertEqual(summary["workspace_to_action_requested_rate"], 50.0)
        self.assertEqual(summary["workspace_to_action_success_rate"], 50.0)
        self.assertEqual(summary["workspace_create_to_action_requested_rate"], 100.0)
        self.assertEqual(summary["workspace_to_create_target_rate"], 60.0)
        self.assertEqual(summary["workspace_create_to_action_target_rate"], 50.0)
        self.assertEqual(summary["workspace_to_create_min_sample"], 5)
        self.assertEqual(summary["workspace_create_to_action_min_sample"], 5)
        self.assertFalse(summary["workspace_to_create_needs_attention"])
        self.assertFalse(summary["workspace_create_to_action_needs_attention"])
        self.assertContains(response, "다건 처리 1회 (보관 3건 / 복구 0건)")
        self.assertContains(response, "다건 제외 0건 · 동일 상태 1건")
        self.assertContains(response, "교무수첩 다건 아카이브/복구")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_flags_attention_when_funnel_below_threshold(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        for idx in range(10):
            _create_event("workspace_home_opened", idx)
        for idx in range(5):
            _create_event("sheetbook_created", 20 + idx, metadata={"entry_source": "workspace_home_create"})
        _create_event("action_execute_requested", 40, metadata={"entry_source": "workspace_home_create"})

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        summary = response.context["summary"]
        self.assertEqual(summary["workspace_source_create_count"], 5)
        self.assertEqual(summary["workspace_source_action_requested_count"], 1)
        self.assertEqual(summary["workspace_to_create_rate"], 50.0)
        self.assertEqual(summary["workspace_create_to_action_requested_rate"], 20.0)
        self.assertTrue(summary["workspace_to_create_needs_attention"])
        self.assertTrue(summary["workspace_create_to_action_needs_attention"])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_archive_bulk_snapshot_collects_counts_rates_and_attention(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(minutes, metadata):
            event = SheetbookMetricEvent.objects.create(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata,
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        _create_event(
            0,
            {
                "selected_count": 4,
                "matched_count": 4,
                "changed_count": 3,
                "unchanged_count": 1,
                "ignored_count": 0,
                "archive_action": "archive",
            },
        )
        _create_event(
            1,
            {
                "selected_count": 2,
                "matched_count": 1,
                "changed_count": 1,
                "unchanged_count": 0,
                "ignored_count": 1,
                "archive_action": "unarchive",
            },
        )
        _create_event(
            2,
            {
                "selected_count": 2,
                "matched_count": 2,
                "changed_count": 0,
                "unchanged_count": 2,
                "ignored_count": 0,
                "archive_action": "archive",
            },
        )
        _create_event(
            3,
            {
                "selected_count": 2,
                "matched_count": 2,
                "changed_count": 0,
                "unchanged_count": 2,
                "ignored_count": 0,
                "archive_action": "archive",
            },
        )
        _create_event(
            4,
            {
                "selected_count": 2,
                "matched_count": 2,
                "changed_count": 0,
                "unchanged_count": 1,
                "ignored_count": 0,
                "archive_action": "archive",
            },
        )

        snapshot = _collect_archive_bulk_snapshot(days=14)

        self.assertEqual(snapshot["event_count"], 5)
        self.assertEqual(snapshot["counts"]["selected_count_total"], 12)
        self.assertEqual(snapshot["counts"]["matched_count_total"], 11)
        self.assertEqual(snapshot["counts"]["changed_total"], 4)
        self.assertEqual(snapshot["counts"]["unchanged_total"], 6)
        self.assertEqual(snapshot["counts"]["ignored_total"], 1)
        self.assertEqual(snapshot["counts"]["archive_changed_total"], 3)
        self.assertEqual(snapshot["counts"]["unarchive_changed_total"], 1)
        self.assertEqual(snapshot["rates"]["changed_rate_pct"], 36.4)
        self.assertEqual(snapshot["rates"]["unchanged_rate_pct"], 54.5)
        self.assertEqual(snapshot["rates"]["ignored_rate_pct"], 9.1)
        self.assertTrue(snapshot["quality"]["has_enough_samples"])
        self.assertEqual(snapshot["quality"]["sample_gap_count"], 0)
        self.assertTrue(snapshot["quality"]["needs_attention"])
        self.assertIn("unchanged_rate_over_50pct", snapshot["quality"]["attention_reasons"])
        self.assertEqual(snapshot["quality"]["thresholds"]["min_events"], 5)
        self.assertEqual(snapshot["quality"]["thresholds"]["ignored_rate_threshold_pct"], 10.0)
        self.assertEqual(snapshot["quality"]["thresholds"]["unchanged_rate_threshold_pct"], 50.0)
        self.assertEqual(snapshot["quality"]["next_step"], "investigate_bulk_flow")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_archive_bulk_snapshot_reports_sample_gap_and_supports_custom_thresholds(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(minutes, metadata):
            event = SheetbookMetricEvent.objects.create(
                event_name="sheetbook_archive_bulk_updated",
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata,
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        _create_event(
            0,
            {
                "selected_count": 2,
                "matched_count": 1,
                "changed_count": 0,
                "unchanged_count": 1,
                "ignored_count": 1,
                "archive_action": "archive",
            },
        )
        _create_event(
            1,
            {
                "selected_count": 2,
                "matched_count": 1,
                "changed_count": 0,
                "unchanged_count": 1,
                "ignored_count": 1,
                "archive_action": "unarchive",
            },
        )

        default_snapshot = _collect_archive_bulk_snapshot(days=14)
        self.assertEqual(default_snapshot["event_count"], 2)
        self.assertFalse(default_snapshot["quality"]["has_enough_samples"])
        self.assertEqual(default_snapshot["quality"]["sample_gap_count"], 3)
        self.assertFalse(default_snapshot["quality"]["needs_attention"])
        self.assertEqual(default_snapshot["quality"]["attention_reasons"], [])
        self.assertEqual(default_snapshot["quality"]["next_step"], "collect_more_samples")

        custom_snapshot = _collect_archive_bulk_snapshot(
            days=14,
            min_events=2,
            ignored_rate_threshold=40.0,
            unchanged_rate_threshold=40.0,
        )
        self.assertTrue(custom_snapshot["quality"]["has_enough_samples"])
        self.assertEqual(custom_snapshot["quality"]["sample_gap_count"], 0)
        self.assertTrue(custom_snapshot["quality"]["needs_attention"])
        self.assertIn("ignored_rate_over_40pct", custom_snapshot["quality"]["attention_reasons"])
        self.assertIn("unchanged_rate_over_40pct", custom_snapshot["quality"]["attention_reasons"])
        self.assertEqual(custom_snapshot["quality"]["next_step"], "investigate_bulk_flow")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_summarizes_consent_cleanup_usage(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        _create_event(
            "consent_review_submitted",
            0,
            metadata={
                "recipients_cleanup_applied": True,
                "recipients_cleanup_removed_count": 2,
                "recipients_cleanup_undo_used": True,
                "recipients_issue_copy_used": True,
                "recipients_issue_jump_count": 5,
            },
        )
        _create_event(
            "consent_review_submitted",
            1,
            metadata={
                "recipients_cleanup_applied": "1",
                "recipients_cleanup_removed_count": "4",
                "recipients_cleanup_undo_used": "0",
                "recipients_issue_copy_used": "1",
                "recipients_issue_jump_count": "1",
            },
        )
        _create_event(
            "consent_review_submitted",
            2,
            metadata={
                "recipients_cleanup_applied": False,
                "recipients_cleanup_removed_count": 99,
                "recipients_cleanup_undo_used": False,
                "recipients_issue_copy_used": False,
            },
        )

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "아카이브/복구")
        self.assertContains(response, "동의서 자동 정리 사용률")
        self.assertContains(response, "자동 정리 취소")
        self.assertContains(response, "문제 줄 복사")
        self.assertContains(response, "문제 줄 이동")
        summary = response.context["summary"]
        self.assertEqual(summary["consent_review_submitted_count"], 3)
        self.assertEqual(summary["consent_cleanup_applied_count"], 2)
        self.assertEqual(summary["consent_cleanup_apply_rate"], 66.7)
        self.assertEqual(summary["consent_cleanup_removed_avg"], 3.0)
        self.assertEqual(summary["consent_cleanup_undo_used_count"], 1)
        self.assertEqual(summary["consent_cleanup_undo_use_rate"], 33.3)
        self.assertEqual(summary["consent_cleanup_undo_alert_rate"], 20.0)
        self.assertEqual(summary["consent_cleanup_undo_min_sample"], 5)
        self.assertEqual(summary["consent_cleanup_undo_gap"], 13.3)
        self.assertFalse(summary["consent_cleanup_undo_needs_attention"])
        self.assertEqual(summary["consent_issue_copy_used_count"], 2)
        self.assertEqual(summary["consent_issue_copy_use_rate"], 66.7)
        self.assertEqual(summary["consent_issue_jump_total"], 6)
        self.assertEqual(summary["consent_issue_jump_used_count"], 2)
        self.assertEqual(summary["consent_issue_jump_use_rate"], 66.7)
        self.assertEqual(summary["consent_issue_jump_avg"], 2.0)
        self.assertFalse(summary["consent_cleanup_needs_attention"])

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_flags_consent_cleanup_attention_when_low(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        for idx in range(5):
            _create_event(
                "consent_review_submitted",
                idx,
                metadata={
                    "recipients_cleanup_applied": idx == 0,
                    "recipients_cleanup_removed_count": 2 if idx == 0 else 0,
                },
            )

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        summary = response.context["summary"]
        self.assertEqual(summary["consent_review_submitted_count"], 5)
        self.assertEqual(summary["consent_cleanup_applied_count"], 1)
        self.assertEqual(summary["consent_cleanup_apply_rate"], 20.0)
        self.assertTrue(summary["consent_cleanup_needs_attention"])
        self.assertContains(response, "보완 필요")

    @override_settings(SHEETBOOK_ENABLED=True)
    def test_metrics_dashboard_flags_cleanup_undo_attention_when_high(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        for idx in range(5):
            _create_event(
                "consent_review_submitted",
                idx,
                metadata={
                    "recipients_cleanup_applied": idx < 3,
                    "recipients_cleanup_removed_count": 1 if idx < 3 else 0,
                    "recipients_cleanup_undo_used": idx < 4,
                },
            )

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        summary = response.context["summary"]
        self.assertEqual(summary["consent_review_submitted_count"], 5)
        self.assertEqual(summary["consent_cleanup_undo_used_count"], 4)
        self.assertEqual(summary["consent_cleanup_undo_use_rate"], 80.0)
        self.assertTrue(summary["consent_cleanup_undo_needs_attention"])
        self.assertContains(response, "자동 정리 취소")
        self.assertContains(response, "보완 필요")

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=40,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=30,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=1,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=1,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=25,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=3,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_ALERT_RATE=10,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_MIN_SAMPLE=2,
    )
    def test_metrics_dashboard_uses_configured_thresholds(self):
        base_ts = timezone.now() - timedelta(days=1)

        def _create_event(event_name, minutes, metadata=None):
            event = SheetbookMetricEvent.objects.create(
                event_name=event_name,
                user=self.user,
                sheetbook=self.sheetbook,
                tab=self.tab,
                metadata=metadata or {"from_test": True},
            )
            event.created_at = base_ts + timedelta(minutes=minutes)
            event.save(update_fields=["created_at"])

        _create_event("workspace_home_opened", 0)
        _create_event("sheetbook_created", 1, metadata={"entry_source": "workspace_home_create"})
        _create_event("action_execute_requested", 2, metadata={"entry_source": "workspace_home_create"})

        self.client.force_login(self.admin_user)
        response = self.client.get(reverse("sheetbook:metrics_dashboard"), data={"days": 7})
        self.assertEqual(response.status_code, 200)
        summary = response.context["summary"]
        self.assertEqual(summary["workspace_to_create_target_rate"], 40.0)
        self.assertEqual(summary["workspace_create_to_action_target_rate"], 30.0)
        self.assertEqual(summary["workspace_to_create_min_sample"], 1)
        self.assertEqual(summary["workspace_create_to_action_min_sample"], 1)
        self.assertEqual(summary["consent_cleanup_target_rate"], 25.0)
        self.assertEqual(summary["consent_cleanup_min_sample"], 3)
        self.assertEqual(summary["consent_cleanup_undo_alert_rate"], 10.0)
        self.assertEqual(summary["consent_cleanup_undo_min_sample"], 2)
        self.assertFalse(summary["workspace_to_create_needs_attention"])
        self.assertFalse(summary["workspace_create_to_action_needs_attention"])


class SheetbookRolloutCommandTests(SimpleTestCase):
    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_ALERT_RATE=20,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_passes_when_ready(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        call_command("check_sheetbook_rollout", stdout=out)

        value = out.getvalue()
        self.assertIn("rollout 점검 통과", value)
        self.assertIn("consent cleanup target", value)
        self.assertIn("consent cleanup undo alert", value)

    @override_settings(
        SHEETBOOK_ENABLED=False,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_warns_when_no_access_path(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        call_command("check_sheetbook_rollout", stdout=out)

        self.assertIn("경고", out.getvalue())
        self.assertIn("allowlist가 비어", out.getvalue())

    @override_settings(
        SHEETBOOK_ENABLED=False,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_strict_fails_on_warning(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command("check_sheetbook_rollout", "--strict", stdout=out, stderr=out)

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_fails_when_collect_schema_not_ready(self, mocked_schema):
        mocked_schema.return_value = (
            False,
            [],
            {"collect_collectionrequest": ["bti_integration_source"]},
            "",
        )
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command("check_sheetbook_rollout", stdout=out, stderr=out)

        self.assertIn("collect 스키마", out.getvalue())

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=77,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_fails_when_period_start_minute_invalid(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command("check_sheetbook_rollout", stdout=out, stderr=out)

        self.assertIn("SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE", out.getvalue())

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=30,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=5,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
        SHEETBOOK_GRID_BULK_BATCH_SIZE=10,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_fails_when_grid_batch_size_invalid(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command("check_sheetbook_rollout", stdout=out, stderr=out)

        self.assertIn("SHEETBOOK_GRID_BULK_BATCH_SIZE", out.getvalue())

    @override_settings(
        SHEETBOOK_ENABLED=True,
        SHEETBOOK_BETA_USERNAMES=[],
        SHEETBOOK_BETA_EMAILS=[],
        SHEETBOOK_BETA_USER_IDS=[],
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
        SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE=120,
        SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE=0,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_ALERT_RATE=130,
        SHEETBOOK_CONSENT_CLEANUP_UNDO_MIN_SAMPLE=0,
        SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES=50,
        SHEETBOOK_PERIOD_FIRST_CLASS_HOUR=9,
        SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE=0,
    )
    @patch("sheetbook.management.commands.check_sheetbook_rollout.get_collect_schema_status")
    def test_check_sheetbook_rollout_fails_when_consent_cleanup_threshold_invalid(self, mocked_schema):
        mocked_schema.return_value = (True, [], {}, "")
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command("check_sheetbook_rollout", stdout=out, stderr=out)

        value = out.getvalue()
        self.assertIn("SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE", value)
        self.assertIn("SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE", value)
        self.assertIn("SHEETBOOK_CONSENT_CLEANUP_UNDO_ALERT_RATE", value)
        self.assertIn("SHEETBOOK_CONSENT_CLEANUP_UNDO_MIN_SAMPLE", value)


class SheetbookSignoffDecisionScriptTests(SimpleTestCase):
    def _build_readiness(self, status="HOLD", automated_gate_pass=True):
        return {
            "overall": {
                "status": status,
                "automated_gate_pass": automated_gate_pass,
                "manual_pending": [
                    "staging_real_account_signoff",
                    "production_real_account_signoff",
                ],
            }
        }

    def _build_manual(self, *, staging_status="HOLD", production_status="HOLD", real_device_status="PASS"):
        manual = _default_manual_payload()
        for key in ["staging_allowlisted", "staging_non_allowlisted"]:
            manual["checks"][key]["status"] = staging_status
        for key in ["production_allowlisted", "production_non_allowlisted"]:
            manual["checks"][key]["status"] = production_status
        manual["checks"]["real_device_grid_1000"]["status"] = real_device_status
        return manual

    def test_next_actions_suggest_manual_signoff_commands_when_pending(self):
        readiness = self._build_readiness()
        manual = self._build_manual(
            staging_status="HOLD",
            production_status="HOLD",
            real_device_status="PASS",
        )
        decision = _compute_decision(
            readiness,
            manual,
            real_device_waived=True,
            allow_pilot_hold_for_beta=False,
        )
        actions = _build_next_actions(
            readiness,
            decision,
            allow_pilot_hold_for_beta=False,
        )
        commands = [str(item.get("command") or "") for item in actions]
        self.assertTrue(any(str(item.get("type")) == "manual_signoff_batch" for item in actions))
        self.assertTrue(
            any(
                "--set staging_real_account_signoff=PASS:staging-ok" in cmd
                and "--set production_real_account_signoff=PASS:prod-ok" in cmd
                for cmd in commands
            )
        )
        self.assertTrue(
            any("staging_real_account_signoff=PASS:staging-ok" in cmd for cmd in commands)
        )
        self.assertTrue(
            any("production_real_account_signoff=PASS:prod-ok" in cmd for cmd in commands)
        )
        self.assertTrue(any("run_sheetbook_release_readiness.py --days 14" in cmd for cmd in commands))
        self.assertFalse(any(str(item.get("type")) == "optional_beta_go" for item in actions))

    def test_next_actions_suggest_optional_beta_go_only_when_pilot_hold_and_manual_pass(self):
        readiness = self._build_readiness(status="HOLD", automated_gate_pass=True)
        manual = self._build_manual(
            staging_status="PASS",
            production_status="PASS",
            real_device_status="PASS",
        )
        decision_hold = _compute_decision(
            readiness,
            manual,
            real_device_waived=True,
            allow_pilot_hold_for_beta=False,
        )
        actions_hold = _build_next_actions(
            readiness,
            decision_hold,
            allow_pilot_hold_for_beta=False,
        )
        self.assertTrue(any(str(item.get("type")) == "optional_beta_go" for item in actions_hold))
        self.assertTrue(any(str(item.get("type")) == "optional_beta_restore_hold" for item in actions_hold))
        self.assertFalse(any(str(item.get("type")) == "manual_signoff" for item in actions_hold))
        self.assertFalse(any(str(item.get("type")) == "manual_signoff_batch" for item in actions_hold))

        decision_go = _compute_decision(
            readiness,
            manual,
            real_device_waived=True,
            allow_pilot_hold_for_beta=True,
        )
        actions_go = _build_next_actions(
            readiness,
            decision_go,
            allow_pilot_hold_for_beta=True,
        )
        self.assertEqual(decision_go["decision"], "GO")
        self.assertFalse(any(str(item.get("type")) == "optional_beta_go" for item in actions_go))
        self.assertFalse(any(str(item.get("type")) == "optional_beta_restore_hold" for item in actions_go))

    def test_compute_decision_includes_manual_alias_statuses(self):
        readiness = self._build_readiness(status="HOLD", automated_gate_pass=True)
        manual = self._build_manual(
            staging_status="PASS",
            production_status="HOLD",
            real_device_status="FAIL",
        )

        decision = _compute_decision(
            readiness,
            manual,
            real_device_waived=False,
            allow_pilot_hold_for_beta=False,
        )

        alias_statuses = decision.get("manual_alias_statuses") or {}
        self.assertEqual(alias_statuses.get("staging_real_account_signoff"), "PASS")
        self.assertEqual(alias_statuses.get("production_real_account_signoff"), "HOLD")
        self.assertEqual(alias_statuses.get("real_device_grid_1000_smoke"), "FAIL")


class SheetbookReleaseSignoffLogScriptTests(SimpleTestCase):
    def test_build_release_signoff_markdown_includes_gate_summary_and_manual_rows(self):
        readiness = {
            "overall": {
                "status": "HOLD",
                "blocking_reasons": [],
                "manual_pending": [
                    "staging_real_account_signoff",
                    "production_real_account_signoff",
                ],
                "waived_manual_checks": ["real_device_grid_1000_smoke"],
            }
        }
        manual = {
            "checks": {
                "staging_allowlisted": {"status": "PASS", "notes": "staging-ok"},
                "staging_non_allowlisted": {"status": "HOLD", "notes": "pending"},
                "production_allowlisted": {"status": "HOLD", "notes": "pending"},
                "production_non_allowlisted": {"status": "HOLD", "notes": "pending"},
                "real_device_grid_1000": {
                    "status": "PASS",
                    "notes": "waived_by_policy(device-unavailable)",
                },
            }
        }
        decision = {
            "decision": "HOLD",
            "manual_checks": manual["checks"],
            "decision_context": {
                "manual_alias_statuses": {
                    "staging_real_account_signoff": "HOLD",
                    "production_real_account_signoff": "HOLD",
                    "real_device_grid_1000_smoke": "PASS",
                },
                "waivers": {"pilot_hold_for_beta": False},
            },
            "next_actions": [
                {
                    "description": "스테이징 점검 후 PASS 반영",
                    "command": "python scripts/run_sheetbook_signoff_decision.py --set staging_real_account_signoff=PASS:staging-ok",
                }
            ],
        }

        markdown = _build_release_signoff_log_markdown(
            record_date=date(2026, 3, 2),
            author="qa-owner",
            readiness=readiness,
            manual=manual,
            decision=decision,
            owner="ops-team",
            next_action="staging/prod 실계정 점검",
            due_date="2026-03-03",
        )

        self.assertIn("작성일: 2026-03-02", markdown)
        self.assertIn("- `overall.status`: HOLD", markdown)
        self.assertIn("- `manual_pending`: staging_real_account_signoff, production_real_account_signoff", markdown)
        self.assertIn(
            "- `manual_pending_raw(readiness)`: staging_real_account_signoff, production_real_account_signoff",
            markdown,
        )
        self.assertIn("- `waived_manual_checks`: real_device_grid_1000_smoke", markdown)
        self.assertIn("- `pilot_hold_for_beta`: False", markdown)
        self.assertIn(
            "| staging_real_account_signoff | staging | allowlisted | PASS | staging-ok |",
            markdown,
        )
        self.assertIn(
            "`python scripts/run_sheetbook_signoff_decision.py --set staging_real_account_signoff=PASS:staging-ok`",
            markdown,
        )
        self.assertIn("- decision: `HOLD`", markdown)
        self.assertIn("- owner: ops-team", markdown)
        self.assertIn("- due_date: 2026-03-03", markdown)

    def test_build_release_signoff_markdown_uses_defaults_for_missing_values(self):
        markdown = _build_release_signoff_log_markdown(
            record_date=date(2026, 3, 2),
            author="",
            readiness={},
            manual={},
            decision={},
            owner="",
            next_action="",
            due_date="",
        )

        self.assertIn("- `blocking_reasons`: (없음)", markdown)
        self.assertIn("- `manual_pending`: (없음)", markdown)
        self.assertIn("- `manual_pending_raw(readiness)`: (없음)", markdown)
        self.assertIn("- `waived_manual_checks`: (없음)", markdown)
        self.assertIn("- `pilot_hold_for_beta`: False", markdown)
        self.assertIn("- `next_actions` (decision json 자동 추천 명령):", markdown)
        self.assertIn("- decision: `HOLD`", markdown)
        self.assertIn("- owner: -", markdown)
        self.assertIn("- next_action: -", markdown)
        self.assertIn("- due_date: -", markdown)

    def test_build_release_signoff_markdown_uses_effective_pending_from_alias_statuses(self):
        readiness = {
            "overall": {
                "status": "HOLD",
                "manual_pending": [
                    "staging_real_account_signoff",
                    "production_real_account_signoff",
                ],
                "waived_manual_checks": ["real_device_grid_1000_smoke"],
            }
        }
        manual = {
            "checks": {
                "staging_allowlisted": {"status": "PASS", "notes": "staging-ok"},
                "staging_non_allowlisted": {"status": "PASS", "notes": "staging-ok"},
                "production_allowlisted": {"status": "PASS", "notes": "prod-ok"},
                "production_non_allowlisted": {"status": "PASS", "notes": "prod-ok"},
                "real_device_grid_1000": {
                    "status": "PASS",
                    "notes": "waived_by_policy(device-unavailable)",
                },
            }
        }
        decision = {
            "decision": "GO",
            "manual_checks": manual["checks"],
            "decision_context": {
                "manual_alias_statuses": {
                    "staging_real_account_signoff": "PASS",
                    "production_real_account_signoff": "PASS",
                    "real_device_grid_1000_smoke": "PASS",
                },
                "waivers": {"pilot_hold_for_beta": True},
            },
            "next_actions": [],
        }

        markdown = _build_release_signoff_log_markdown(
            record_date=date(2026, 3, 3),
            author="qa-owner",
            readiness=readiness,
            manual=manual,
            decision=decision,
            owner="ops-team",
            next_action="pilot 표본 보강",
            due_date="2026-03-04",
        )

        self.assertIn("- `manual_pending`: (없음)", markdown)
        self.assertIn(
            "- `manual_pending_raw(readiness)`: staging_real_account_signoff, production_real_account_signoff",
            markdown,
        )
        self.assertIn("- `pilot_hold_for_beta`: True", markdown)


class SheetbookGuardedCommitScriptTests(SimpleTestCase):
    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    def test_run_blocks_on_branch_override_mismatch(self, mock_current_branch, mock_repo_root):
        mock_repo_root.return_value = Path("C:/repo")
        mock_current_branch.return_value = "feature/sheetbook"

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="hotfix/main-ops",
                expected_branch="feature/sheetbook",
                message="",
                allow_empty=False,
                guard_only=True,
                push=False,
                remote="origin",
                push_retries=2,
                push_retry_delay=1.0,
            )
        )

        self.assertEqual(code, 2)

    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    def test_run_blocks_on_branch_mismatch(self, mock_current_branch, mock_repo_root):
        mock_repo_root.return_value = Path("C:/repo")
        mock_current_branch.return_value = "hotfix/main-ops"

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="",
                allow_empty=False,
                guard_only=True,
                push=False,
                remote="origin",
                push_retries=2,
                push_retry_delay=1.0,
            )
        )

        self.assertEqual(code, 2)

    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    @patch("scripts.run_sheetbook_guarded_commit._staged_files")
    @patch("scripts.run_sheetbook_guarded_commit._run_guard")
    def test_run_guard_only_passes(self, mock_run_guard, mock_staged_files, mock_current_branch, mock_repo_root):
        mock_repo_root.return_value = Path("C:/repo")
        mock_current_branch.return_value = "feature/sheetbook"
        mock_staged_files.return_value = ["scripts/run_sheetbook_guarded_commit.py"]
        mock_run_guard.return_value = 0

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="",
                allow_empty=False,
                guard_only=True,
                push=False,
                remote="origin",
                push_retries=2,
                push_retry_delay=1.0,
            )
        )

        self.assertEqual(code, 0)
        mock_run_guard.assert_called_once_with(Path("C:/repo"), "feature/sheetbook")

    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    @patch("scripts.run_sheetbook_guarded_commit._staged_files")
    @patch("scripts.run_sheetbook_guarded_commit._run_guard")
    @patch("scripts.run_sheetbook_guarded_commit._run")
    def test_run_commit_and_push_when_guard_passes(
        self,
        mock_run,
        mock_run_guard,
        mock_staged_files,
        mock_current_branch,
        mock_repo_root,
    ):
        root = Path("C:/repo")
        mock_repo_root.return_value = root
        mock_current_branch.return_value = "feature/sheetbook"
        mock_staged_files.return_value = ["scripts/run_sheetbook_guarded_commit.py"]
        mock_run_guard.return_value = 0
        mock_run.side_effect = [
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="abc123\n", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
        ]

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="feat(sheetbook): add helper",
                allow_empty=False,
                guard_only=False,
                push=True,
                remote="origin",
                push_retries=2,
                push_retry_delay=1.0,
            )
        )

        self.assertEqual(code, 0)
        self.assertEqual(
            mock_run.call_args_list[0].args,
            (root, ["git", "commit", "-m", "feat(sheetbook): add helper"]),
        )
        self.assertEqual(
            mock_run.call_args_list[1].args,
            (root, ["git", "rev-parse", "--short", "HEAD"]),
        )
        self.assertEqual(
            mock_run.call_args_list[2].args,
            (root, ["git", "push", "origin", "feature/sheetbook"]),
        )

    @patch("scripts.run_sheetbook_guarded_commit.time.sleep")
    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    @patch("scripts.run_sheetbook_guarded_commit._staged_files")
    @patch("scripts.run_sheetbook_guarded_commit._run_guard")
    @patch("scripts.run_sheetbook_guarded_commit._run")
    def test_run_retries_push_on_transient_failure(
        self,
        mock_run,
        mock_run_guard,
        mock_staged_files,
        mock_current_branch,
        mock_repo_root,
        mock_sleep,
    ):
        root = Path("C:/repo")
        mock_repo_root.return_value = root
        mock_current_branch.return_value = "feature/sheetbook"
        mock_staged_files.return_value = ["scripts/run_sheetbook_guarded_commit.py"]
        mock_run_guard.return_value = 0
        mock_run.side_effect = [
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="def456\n", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=1, stdout="", stderr="push failed"),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
        ]

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="feat(sheetbook): retry push",
                allow_empty=False,
                guard_only=False,
                push=True,
                remote="origin",
                push_retries=2,
                push_retry_delay=0.0,
            )
        )

        self.assertEqual(code, 0)
        self.assertEqual(mock_run.call_count, 4)
        mock_sleep.assert_called_once_with(0.0)

    @patch("builtins.print")
    @patch("scripts.run_sheetbook_guarded_commit.time.sleep")
    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    @patch("scripts.run_sheetbook_guarded_commit._staged_files")
    @patch("scripts.run_sheetbook_guarded_commit._run_guard")
    @patch("scripts.run_sheetbook_guarded_commit._run")
    def test_run_reports_manual_push_command_when_all_retries_fail(
        self,
        mock_run,
        mock_run_guard,
        mock_staged_files,
        mock_current_branch,
        mock_repo_root,
        mock_sleep,
        mock_print,
    ):
        root = Path("C:/repo")
        mock_repo_root.return_value = root
        mock_current_branch.return_value = "feature/sheetbook"
        mock_staged_files.return_value = ["scripts/run_sheetbook_guarded_commit.py"]
        mock_run_guard.return_value = 0
        mock_run.side_effect = [
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="fedcba\n", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=1, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=1, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=1, stdout="", stderr=""),
        ]

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="feat(sheetbook): retry push",
                allow_empty=False,
                guard_only=False,
                push=True,
                remote="origin",
                push_retries=2,
                push_retry_delay=0.0,
            )
        )

        self.assertEqual(code, 1)
        self.assertEqual(mock_run.call_count, 5)
        self.assertEqual(mock_sleep.call_count, 2)
        printed_texts = [" ".join(str(arg) for arg in args) for args, _ in mock_print.call_args_list]
        self.assertTrue(
            any("retry manually: `git push origin feature/sheetbook`" in text for text in printed_texts)
        )
        self.assertTrue(any("local commit: fedcba" in text for text in printed_texts))

    @patch("builtins.print")
    @patch("scripts.run_sheetbook_guarded_commit.time.sleep")
    @patch("scripts.run_sheetbook_guarded_commit._repo_root")
    @patch("scripts.run_sheetbook_guarded_commit._current_branch")
    @patch("scripts.run_sheetbook_guarded_commit._staged_files")
    @patch("scripts.run_sheetbook_guarded_commit._run_guard")
    @patch("scripts.run_sheetbook_guarded_commit._run")
    def test_run_stops_retry_on_non_retryable_push_failure(
        self,
        mock_run,
        mock_run_guard,
        mock_staged_files,
        mock_current_branch,
        mock_repo_root,
        mock_sleep,
        mock_print,
    ):
        root = Path("C:/repo")
        mock_repo_root.return_value = root
        mock_current_branch.return_value = "feature/sheetbook"
        mock_staged_files.return_value = ["scripts/run_sheetbook_guarded_commit.py"]
        mock_run_guard.return_value = 0
        mock_run.side_effect = [
            subprocess.CompletedProcess(args=[], returncode=0, stdout="", stderr=""),
            subprocess.CompletedProcess(args=[], returncode=0, stdout="aa11bb\n", stderr=""),
            subprocess.CompletedProcess(
                args=[],
                returncode=1,
                stdout="",
                stderr="fatal: Authentication failed for 'https://github.com/x/y.git/'",
            ),
        ]

        code = _run_sheetbook_guarded_commit(
            Namespace(
                branch="",
                expected_branch="feature/sheetbook",
                message="feat(sheetbook): retry push",
                allow_empty=False,
                guard_only=False,
                push=True,
                remote="origin",
                push_retries=2,
                push_retry_delay=0.0,
            )
        )

        self.assertEqual(code, 1)
        self.assertEqual(mock_run.call_count, 3)
        mock_sleep.assert_not_called()
        printed_texts = [" ".join(str(arg) for arg in args) for args, _ in mock_print.call_args_list]
        self.assertTrue(any("non-retryable push failure detected" in text for text in printed_texts))
        self.assertTrue(any("local commit: aa11bb" in text for text in printed_texts))


class SheetbookDailyStartBundleScriptTests(SimpleTestCase):
    def test_build_daily_start_bundle_summary_uses_decision_when_commands_ok(self):
        summary = _build_daily_start_bundle_summary(
            generated_at="2026-03-03 09:00:00",
            days=14,
            ops_index_report="docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
            command_results=[
                {"command": "cmd1", "ok": True, "returncode": 0, "tail": []},
                {"command": "cmd2", "ok": True, "returncode": 0, "tail": []},
            ],
            readiness={
                "overall": {
                    "status": "HOLD",
                    "manual_pending": [
                        "staging_real_account_signoff",
                        "production_real_account_signoff",
                    ],
                },
                "pilot": {
                    "counts": {
                        "workspace_home_opened": 3,
                        "home_source_sheetbook_created": 1,
                        "home_source_action_execute_requested": 0,
                    }
                },
            },
            decision={
                "decision": "HOLD",
                "decision_context": {
                    "manual_alias_statuses": {
                        "staging_real_account_signoff": "HOLD",
                        "production_real_account_signoff": "HOLD",
                    },
                    "waivers": {"pilot_hold_for_beta": False},
                },
            },
            archive_snapshot={
                "event_count": 2,
                "quality": {"next_step": "collect_more_samples", "needs_attention": False},
                "md_output": "docs/runbooks/logs/SHEETBOOK_ARCHIVE_BULK_2026-03-03.md",
            },
            consent_freeze_snapshot={
                "status": "PASS",
                "reasons": [],
                "md_output": "docs/runbooks/logs/SHEETBOOK_CONSENT_FREEZE_2026-03-03.md",
            },
            sample_gap_summary={
                "overall": {
                    "ready": False,
                    "blockers": ["pilot_home_opened_gap:2"],
                    "next_actions": [
                        {
                            "type": "collect_pilot_samples",
                            "description": "파일럿 이벤트 추가 확보: workspace_home_opened 2건",
                            "command": "python scripts/run_sheetbook_release_readiness.py --days 14",
                        }
                    ],
                }
            },
            allow_pilot_hold_for_beta=True,
            due_date="2026-03-04",
        )

        self.assertEqual(summary["overall"], "HOLD")
        self.assertFalse(summary["has_command_failures"])
        self.assertEqual(summary["decision"], "HOLD")
        self.assertEqual(summary["decision_waivers"], {"pilot_hold_for_beta": False})
        self.assertEqual(summary["pilot_counts"]["workspace_home_opened"], 3)
        self.assertEqual(summary["archive"]["event_count"], 2)
        self.assertEqual(
            summary["ops_index_report"],
            "docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
        )
        self.assertEqual(
            summary["archive"]["md_output"],
            "docs/runbooks/logs/SHEETBOOK_ARCHIVE_BULK_2026-03-03.md",
        )
        self.assertEqual(summary["consent_freeze"]["status"], "PASS")
        self.assertEqual(
            summary["consent_freeze"]["md_output"],
            "docs/runbooks/logs/SHEETBOOK_CONSENT_FREEZE_2026-03-03.md",
        )
        self.assertFalse(summary["sample_gap"]["ready"])
        self.assertIn("pilot_home_opened_gap:2", summary["sample_gap"]["blockers"])
        self.assertEqual(summary["sample_gap"]["next_actions"][0]["type"], "collect_pilot_samples")
        next_actions = summary.get("next_actions") or []
        action_types = {str(item.get("type")) for item in next_actions if isinstance(item, dict)}
        self.assertIn("manual_signoff_pending", action_types)
        self.assertIn("collect_samples", action_types)
        self.assertTrue(
            any(
                "--days 14" in str(item.get("command") or "")
                for item in next_actions
                if isinstance(item, dict) and str(item.get("type")) == "collect_samples"
            )
        )
        self.assertTrue(
            any(
                "--allow-pilot-hold-for-beta" in str(item.get("command") or "")
                for item in next_actions
                if isinstance(item, dict) and str(item.get("type")) == "collect_samples"
            )
        )
        self.assertTrue(
            any(
                "--due-date 2026-03-04" in str(item.get("command") or "")
                for item in next_actions
                if isinstance(item, dict) and str(item.get("type")) == "collect_samples"
            )
        )

    def test_build_daily_start_bundle_summary_clears_manual_pending_when_aliases_pass(self):
        summary = _build_daily_start_bundle_summary(
            generated_at="2026-03-03 11:45:00",
            days=14,
            ops_index_report="docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
            command_results=[
                {"command": "cmd1", "ok": True, "returncode": 0, "tail": []},
            ],
            readiness={
                "overall": {
                    "status": "HOLD",
                    "manual_pending": [
                        "staging_real_account_signoff",
                        "production_real_account_signoff",
                    ],
                },
                "pilot": {"counts": {}},
            },
            decision={
                "decision": "GO",
                "decision_context": {
                    "manual_alias_statuses": {
                        "staging_real_account_signoff": "PASS",
                        "production_real_account_signoff": "PASS",
                        "real_device_grid_1000_smoke": "PASS",
                    },
                    "waivers": {"pilot_hold_for_beta": True},
                },
            },
            archive_snapshot={"event_count": 0, "quality": {"next_step": "collect_more_samples"}},
            consent_freeze_snapshot={"status": "PASS", "reasons": []},
            sample_gap_summary={"overall": {"ready": True, "blockers": []}},
        )

        self.assertEqual(
            summary["manual_pending_raw"],
            ["staging_real_account_signoff", "production_real_account_signoff"],
        )
        self.assertEqual(summary["decision_waivers"], {"pilot_hold_for_beta": True})
        self.assertEqual(summary["manual_pending"], [])
        next_actions = summary.get("next_actions") or []
        action_types = {str(item.get("type")) for item in next_actions if isinstance(item, dict)}
        self.assertNotIn("manual_signoff_pending", action_types)

    def test_build_daily_start_bundle_summary_forces_hold_on_command_failure(self):
        summary = _build_daily_start_bundle_summary(
            generated_at="2026-03-03 09:00:00",
            days=14,
            ops_index_report="docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
            command_results=[
                {"command": "cmd1", "ok": True, "returncode": 0, "tail": []},
                {"command": "cmd2", "ok": False, "returncode": 1, "tail": ["boom"]},
            ],
            readiness={"overall": {"status": "PASS", "manual_pending": []}, "pilot": {"counts": {}}},
            decision={"decision": "GO", "decision_context": {"manual_alias_statuses": {}}},
            archive_snapshot={
                "event_count": 10,
                "quality": {"next_step": "continue_monitoring"},
                "md_output": "archive.md",
            },
            consent_freeze_snapshot={"status": "PASS", "reasons": [], "md_output": "freeze.md"},
            sample_gap_summary={"overall": {"ready": True, "blockers": []}},
        )

        self.assertTrue(summary["has_command_failures"])
        self.assertEqual(summary["overall"], "HOLD")
        self.assertEqual(summary["decision"], "GO")
        next_actions = summary.get("next_actions") or []
        self.assertTrue(any(str(item.get("type")) == "rerun_failed_commands" for item in next_actions))

    def test_build_daily_start_bundle_next_actions_defaults_to_monitoring(self):
        actions = _build_daily_start_bundle_next_actions(
            {
                "has_command_failures": False,
                "manual_pending": [],
                "sample_gap": {"blockers": []},
                "decision": "GO",
            }
        )
        self.assertTrue(any(str(item.get("type")) == "monitoring" for item in actions))

    def test_build_daily_start_bundle_next_actions_carries_allow_and_due_date(self):
        actions = _build_daily_start_bundle_next_actions(
            {
                "days": 14,
                "has_command_failures": False,
                "manual_pending": [],
                "sample_gap": {"blockers": ["pilot_home_opened_gap:1"]},
                "decision": "GO",
            },
            allow_pilot_hold_for_beta=True,
            due_date="2026-03-04",
        )
        collect_samples = next(
            item
            for item in actions
            if isinstance(item, dict) and str(item.get("type")) == "collect_samples"
        )
        self.assertIn("--allow-pilot-hold-for-beta", str(collect_samples.get("command") or ""))
        self.assertIn("--due-date 2026-03-04", str(collect_samples.get("command") or ""))

    def test_build_daily_start_bundle_markdown_includes_commands_and_actions(self):
        summary = {
            "generated_at": "2026-03-03 09:00:00",
            "days": 14,
            "ops_index_report": "docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
            "overall": "HOLD",
            "decision": "HOLD",
            "decision_waivers": {"pilot_hold_for_beta": True},
            "readiness_status": "HOLD",
            "manual_pending": ["staging_real_account_signoff"],
            "sample_gap": {
                "ready": False,
                "blockers": ["pilot_home_opened_gap:2"],
                "next_actions": [
                    {
                        "description": "파일럿 이벤트 추가 확보: workspace_home_opened 2건",
                        "command": "python scripts/run_sheetbook_release_readiness.py --days 14",
                    }
                ],
            },
            "archive": {
                "next_step": "collect_more_samples",
                "md_output": "docs/runbooks/logs/SHEETBOOK_ARCHIVE_BULK_2026-03-03.md",
            },
            "consent_freeze": {
                "status": "PASS",
                "reasons": ["unexpected_extra_tokens"],
                "md_output": "docs/runbooks/logs/SHEETBOOK_CONSENT_FREEZE_2026-03-03.md",
            },
            "commands": [
                {"command": "python cmd1", "ok": True},
                {"command": "python cmd2", "ok": False},
            ],
            "next_actions": [
                {
                    "description": "수동 signoff 완료 후 PASS 반영",
                    "command": "python scripts/run_sheetbook_signoff_decision.py --set ...",
                }
            ],
        }
        markdown = _build_daily_start_bundle_markdown(
            summary=summary,
            json_output_path=Path("docs/handoff/sheetbook_daily_start_bundle_latest.json"),
        )
        self.assertIn("Sheetbook Daily Start Bundle", markdown)
        self.assertIn("## Commands", markdown)
        self.assertIn("`python cmd1`", markdown)
        self.assertIn("`python cmd2`", markdown)
        self.assertIn("## Next Actions", markdown)
        self.assertIn("수동 signoff 완료 후 PASS 반영", markdown)
        self.assertIn("## Sample Gap Next Actions", markdown)
        self.assertIn("파일럿 이벤트 추가 확보: workspace_home_opened 2건", markdown)
        self.assertIn("SHEETBOOK_OPS_INDEX_2026-03-03.md", markdown)
        self.assertIn("SHEETBOOK_ARCHIVE_BULK_2026-03-03.md", markdown)
        self.assertIn("SHEETBOOK_CONSENT_FREEZE_2026-03-03.md", markdown)
        self.assertIn("unexpected_extra_tokens", markdown)
        self.assertIn("- pilot_hold_for_beta: `True`", markdown)


class SheetbookOpsIndexReportScriptTests(SimpleTestCase):
    def test_build_ops_index_summary_dedupes_next_actions_by_command(self):
        summary = _build_ops_index_summary(
            readiness={"overall": {"status": "HOLD", "manual_pending": ["staging_real_account_signoff"]}},
            decision={
                "decision": "HOLD",
                "decision_context": {"waivers": {"pilot_hold_for_beta": True}},
                "next_actions": [
                    {
                        "type": "review_hold_reasons",
                        "description": "자동/수동 게이트 상태 재검토 후 GO/HOLD 재판정",
                        "command": "python scripts/run_sheetbook_signoff_decision.py",
                    },
                    {
                        "type": "collect_samples",
                        "description": "표본 수집",
                        "command": "python scripts/run_sheetbook_release_readiness.py --days 14",
                    },
                ],
            },
            daily_start={
                "overall": "HOLD",
                "next_actions": [
                    {
                        "type": "collect_samples",
                        "description": "표본 수집 후 bundle 재실행",
                        "command": "python scripts/run_sheetbook_release_readiness.py --days 14",
                    }
                ],
            },
            archive_snapshot={"quality": {"next_step": "collect_more_samples"}},
            sample_gap_summary={
                "overall": {
                    "blockers": ["pilot_home_opened_gap:2"],
                    "next_actions": [
                        {
                            "type": "refresh_gap_summary",
                            "description": "gap summary 재생성",
                            "command": "python scripts/run_sheetbook_sample_gap_summary.py --days 14",
                        }
                    ],
                }
            },
            consent_freeze_snapshot={"status": "PASS", "reasons": []},
        )

        commands = [str(item.get("command") or "") for item in summary["next_actions"]]
        self.assertEqual(
            commands,
            [
                "python scripts/run_sheetbook_release_readiness.py --days 14",
                "python scripts/run_sheetbook_sample_gap_summary.py --days 14",
                "python scripts/run_sheetbook_signoff_decision.py",
            ],
        )
        self.assertEqual(summary["next_actions"][0]["source"], "daily_start")
        self.assertEqual(summary["next_actions"][1]["source"], "sample_gap")
        self.assertEqual(summary["next_actions"][2]["source"], "decision")
        self.assertEqual(summary["archive_next_step"], "collect_more_samples")
        self.assertEqual(summary["decision"], "HOLD")
        self.assertEqual(summary["overall"], "HOLD")
        self.assertTrue(summary["pilot_hold_for_beta"])

    def test_build_ops_index_summary_prefers_daily_start_effective_manual_pending(self):
        summary = _build_ops_index_summary(
            readiness={
                "overall": {
                    "status": "HOLD",
                    "manual_pending": [
                        "staging_real_account_signoff",
                        "production_real_account_signoff",
                    ],
                }
            },
            decision={"decision": "GO", "next_actions": []},
            daily_start={"overall": "HOLD", "manual_pending": [], "next_actions": []},
            archive_snapshot={"quality": {"next_step": "collect_more_samples"}},
            sample_gap_summary={"overall": {"blockers": []}},
            consent_freeze_snapshot={"status": "PASS", "reasons": []},
        )

        self.assertEqual(summary["manual_pending"], [])
        self.assertEqual(
            summary["manual_pending_raw"],
            ["staging_real_account_signoff", "production_real_account_signoff"],
        )

    def test_build_ops_index_markdown_includes_reports_and_actions(self):
        markdown = _build_ops_index_markdown(
            record_date="2026-03-03",
            summary={
                "overall": "HOLD",
                "decision": "HOLD",
                "pilot_hold_for_beta": True,
                "readiness_status": "HOLD",
                "manual_pending": ["staging_real_account_signoff"],
                "sample_gap_blockers": ["pilot_home_opened_gap:2"],
                "archive_next_step": "collect_more_samples",
                "consent_freeze_status": "PASS",
                "consent_freeze_reasons": [],
                "next_actions": [
                    {
                        "source": "daily_start",
                        "description": "표본 수집 후 bundle 재실행",
                        "command": "python scripts/run_sheetbook_release_readiness.py --days 14",
                    }
                ],
            },
            report_paths={
                "daily_start": "docs/runbooks/logs/SHEETBOOK_DAILY_START_2026-03-03.md",
                "ops_index": "docs/runbooks/logs/SHEETBOOK_OPS_INDEX_2026-03-03.md",
            },
        )

        self.assertIn("Sheetbook Ops Index (2026-03-03)", markdown)
        self.assertIn("## Reports", markdown)
        self.assertIn("SHEETBOOK_DAILY_START_2026-03-03.md", markdown)
        self.assertIn("SHEETBOOK_OPS_INDEX_2026-03-03.md", markdown)
        self.assertIn("## Next Actions", markdown)
        self.assertIn("[daily_start] 표본 수집 후 bundle 재실행", markdown)
        self.assertIn("run_sheetbook_release_readiness.py --days 14", markdown)
        self.assertIn("- pilot_hold_for_beta: `True`", markdown)


class SheetbookArchiveBulkSnapshotScriptTests(SimpleTestCase):
    def test_build_archive_bulk_snapshot_markdown_includes_quality_fields(self):
        snapshot = {
            "days": 14,
            "event_count": 3,
            "counts": {"archive_changed_total": 2, "unarchive_changed_total": 1},
            "rates": {"changed_rate_pct": 50.0, "unchanged_rate_pct": 40.0, "ignored_rate_pct": 10.0},
            "quality": {
                "has_enough_samples": False,
                "sample_gap_count": 2,
                "needs_attention": False,
                "attention_reasons": [],
                "next_step": "collect_more_samples",
            },
            "md_output": "docs/runbooks/logs/SHEETBOOK_ARCHIVE_BULK_2026-03-03.md",
        }
        markdown = _build_archive_bulk_snapshot_markdown(
            snapshot=snapshot,
            json_output_path=Path("docs/handoff/sheetbook_archive_bulk_snapshot_latest.json"),
        )

        self.assertIn("Sheetbook Archive Bulk Snapshot", markdown)
        self.assertIn("sample_gap_count", markdown)
        self.assertIn("collect_more_samples", markdown)
        self.assertIn("SHEETBOOK_ARCHIVE_BULK_2026-03-03.md", markdown)


class SheetbookSeedMetricSamplesScriptTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_seed_script_t1",
            password="pw123456",
            email="sheetbook_seed_script_t1@example.com",
        )

    def test_seed_metric_events_creates_expected_event_counts(self):
        created = _seed_metric_events(
            user=self.user,
            home_count=2,
            create_count=2,
            action_count=1,
            archive_event_count=3,
        )
        self.assertEqual(
            created,
            {
                "workspace_home_opened": 2,
                "sheetbook_created": 2,
                "action_execute_requested": 1,
                "sheetbook_archive_bulk_updated": 3,
            },
        )

        self.assertEqual(
            SheetbookMetricEvent.objects.filter(event_name="workspace_home_opened").count(),
            2,
        )
        self.assertEqual(
            SheetbookMetricEvent.objects.filter(event_name="sheetbook_created").count(),
            2,
        )
        self.assertEqual(
            SheetbookMetricEvent.objects.filter(event_name="action_execute_requested").count(),
            1,
        )
        self.assertEqual(
            SheetbookMetricEvent.objects.filter(event_name="sheetbook_archive_bulk_updated").count(),
            3,
        )

        create_metadata = list(
            SheetbookMetricEvent.objects.filter(event_name="sheetbook_created").values_list(
                "metadata", flat=True
            )
        )
        self.assertTrue(
            all(
                str((row or {}).get("entry_source") or "").startswith("workspace_home")
                for row in create_metadata
            )
        )
        self.assertTrue(
            all(
                str((row or {}).get("seeded_by") or "") == SHEETBOOK_METRIC_SEED_TAG
                for row in create_metadata
            )
        )

    def test_clear_seed_events_removes_only_seed_tagged_events(self):
        SheetbookMetricEvent.objects.create(
            event_name="workspace_home_opened",
            user=self.user,
            metadata={"seeded_by": SHEETBOOK_METRIC_SEED_TAG},
        )
        SheetbookMetricEvent.objects.create(
            event_name="sheetbook_created",
            user=self.user,
            metadata={"seeded_by": SHEETBOOK_METRIC_SEED_TAG},
        )
        SheetbookMetricEvent.objects.create(
            event_name="workspace_home_opened",
            user=self.user,
            metadata={"seeded_by": "another_seed_tag"},
        )

        removed = _clear_seed_events()
        self.assertEqual(removed, 2)
        self.assertEqual(SheetbookMetricEvent.objects.count(), 1)
        self.assertEqual(
            SheetbookMetricEvent.objects.first().metadata.get("seeded_by"),
            "another_seed_tag",
        )


class SheetbookSampleGapSummaryScriptTests(SimpleTestCase):
    def test_build_sample_gap_summary_reports_gaps(self):
        summary = _build_sample_gap_summary_payload(
            days=21,
            generated_at="2026-03-03 09:00:00",
            readiness={
                "pilot": {
                    "counts": {
                        "workspace_home_opened": 2,
                        "home_source_sheetbook_created": 1,
                        "home_source_action_execute_requested": 0,
                    },
                    "minimum_samples": {
                        "workspace_home_opened": 5,
                        "home_source_sheetbook_created": 4,
                    },
                }
            },
            archive_snapshot={
                "event_count": 1,
                "quality": {
                    "sample_gap_count": 4,
                    "next_step": "collect_more_samples",
                },
            },
        )

        self.assertFalse(summary["overall"]["ready"])
        self.assertEqual(summary["pilot"]["gaps"]["workspace_home_opened_gap"], 3)
        self.assertEqual(summary["pilot"]["gaps"]["home_source_sheetbook_created_gap"], 3)
        self.assertEqual(summary["archive"]["event_gap"], 4)
        self.assertIn("pilot_home_opened_gap:3", summary["overall"]["blockers"])
        self.assertIn("archive_event_gap:4", summary["overall"]["blockers"])
        next_actions = summary["overall"]["next_actions"]
        next_action_types = {str(item.get("type")) for item in next_actions if isinstance(item, dict)}
        self.assertIn("collect_pilot_samples", next_action_types)
        self.assertIn("collect_archive_events", next_action_types)
        self.assertIn("refresh_gap_summary", next_action_types)
        self.assertTrue(
            any("--days 21" in str(item.get("command") or "") for item in next_actions if isinstance(item, dict))
        )

    def test_build_sample_gap_summary_ready_when_gaps_zero(self):
        summary = _build_sample_gap_summary_payload(
            days=7,
            generated_at="2026-03-03 09:00:00",
            readiness={
                "pilot": {
                    "counts": {
                        "workspace_home_opened": 5,
                        "home_source_sheetbook_created": 5,
                        "home_source_action_execute_requested": 2,
                    },
                    "minimum_samples": {
                        "workspace_home_opened": 5,
                        "home_source_sheetbook_created": 5,
                    },
                }
            },
            archive_snapshot={
                "event_count": 9,
                "quality": {
                    "sample_gap_count": 0,
                    "next_step": "continue_monitoring",
                },
            },
        )

        self.assertTrue(summary["pilot"]["ready"])
        self.assertTrue(summary["archive"]["ready"])
        self.assertTrue(summary["overall"]["ready"])
        self.assertEqual(summary["overall"]["blockers"], [])
        self.assertEqual(summary["overall"]["next_actions"][0]["type"], "monitoring")
        self.assertIn("--days 7", summary["overall"]["next_actions"][0]["command"])

    def test_build_sample_gap_markdown_includes_blockers_and_actions(self):
        summary = {
            "generated_at": "2026-03-03 09:00:00",
            "pilot": {
                "counts": {
                    "workspace_home_opened": 2,
                    "home_source_sheetbook_created": 1,
                    "home_source_action_execute_requested": 0,
                },
                "gaps": {
                    "workspace_home_opened_gap": 3,
                    "home_source_sheetbook_created_gap": 4,
                },
            },
            "archive": {
                "event_count": 1,
                "event_gap": 4,
                "next_step": "collect_more_samples",
            },
            "overall": {
                "ready": False,
                "blockers": ["pilot_home_opened_gap:3", "archive_event_gap:4"],
                "next_actions": [
                    {
                        "description": "표본 수집 후 gap summary 재생성",
                        "command": "python scripts/run_sheetbook_sample_gap_summary.py",
                    }
                ],
            },
        }
        markdown = _build_sample_gap_markdown_payload(
            summary=summary,
            json_output_path=Path("docs/handoff/sheetbook_sample_gap_summary_latest.json"),
        )
        self.assertIn("Sheetbook Sample Gap Summary", markdown)
        self.assertIn("pilot_home_opened_gap:3", markdown)
        self.assertIn("archive_event_gap:4", markdown)
        self.assertIn("## Next Actions", markdown)
        self.assertIn("표본 수집 후 gap summary 재생성", markdown)


class SheetbookPreflightCommandTests(SimpleTestCase):
    @patch("sheetbook.management.commands.check_sheetbook_preflight.call_command")
    def test_check_sheetbook_preflight_runs_full_flow_with_strict(self, mocked_call_command):
        out = StringIO()

        call_command(
            "check_sheetbook_preflight",
            "--strict",
            "--recommend-days",
            "21",
            stdout=out,
        )

        mocked_call_command.assert_has_calls(
            [
                call("check_collect_schema"),
                call("check_sheetbook_rollout", "--strict"),
                call("check_sheetbook_consent_freeze"),
                call("recommend_sheetbook_thresholds", "--days", "21"),
            ]
        )
        self.assertIn("preflight done", out.getvalue())

    @patch("sheetbook.management.commands.check_sheetbook_preflight.call_command")
    def test_check_sheetbook_preflight_can_skip_recommendation(self, mocked_call_command):
        call_command(
            "check_sheetbook_preflight",
            "--skip-recommend",
        )

        mocked_call_command.assert_has_calls(
            [
                call("check_collect_schema"),
                call("check_sheetbook_rollout"),
                call("check_sheetbook_consent_freeze"),
            ]
        )
        self.assertEqual(mocked_call_command.call_count, 3)

    @patch("sheetbook.management.commands.check_sheetbook_preflight.call_command")
    def test_check_sheetbook_preflight_can_skip_consent_freeze(self, mocked_call_command):
        call_command(
            "check_sheetbook_preflight",
            "--skip-consent-freeze",
        )

        mocked_call_command.assert_has_calls(
            [
                call("check_collect_schema"),
                call("check_sheetbook_rollout"),
                call("recommend_sheetbook_thresholds", "--days", "14"),
            ]
        )
        self.assertEqual(mocked_call_command.call_count, 3)

    def test_check_sheetbook_preflight_rejects_invalid_days(self):
        with self.assertRaises(CommandError):
            call_command("check_sheetbook_preflight", "--recommend-days", "0")


class SheetbookConsentFreezeCommandTests(SimpleTestCase):
    def test_check_sheetbook_consent_freeze_passes_for_current_template(self):
        out = StringIO()

        call_command("check_sheetbook_consent_freeze", stdout=out)

        value = out.getvalue()
        self.assertIn("consent freeze 점검 통과", value)
        self.assertIn("consent_review.html", value)

    def test_check_sheetbook_consent_freeze_fails_when_required_token_missing(self):
        invalid_template = """
        <form>
          <textarea id="recipients-textarea" data-testid="recipients-textarea"></textarea>
        </form>
        """.strip()
        with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as tmp:
            tmp.write(invalid_template)
            tmp_path = tmp.name
        try:
            with self.assertRaises(CommandError) as cm:
                call_command(
                    "check_sheetbook_consent_freeze",
                    "--template-path",
                    tmp_path,
                )
            message = str(cm.exception)
            self.assertIn("consent freeze 점검 실패", message)
            self.assertIn('누락: id="recipients-cleanup-btn"', message)
        finally:
            Path(tmp_path).unlink(missing_ok=True)


class SheetbookConsentFreezeSnapshotScriptTests(SimpleTestCase):
    def test_consent_freeze_snapshot_report_passes_for_current_template(self):
        template_path = (
            Path(__file__).resolve().parents[1]
            / "sheetbook"
            / "templates"
            / "sheetbook"
            / "consent_review.html"
        )
        content = template_path.read_text(encoding="utf-8")

        report = _build_consent_freeze_snapshot_report(
            content=content,
            template_path=str(template_path),
            strict_extras=False,
        )

        self.assertEqual(report.get("status"), "PASS")
        self.assertFalse((report.get("missing") or {}).get("ids"))
        self.assertFalse((report.get("missing") or {}).get("testids"))
        self.assertFalse((report.get("missing") or {}).get("jump_values"))
        self.assertFalse((report.get("missing") or {}).get("hidden_names"))
        self.assertTrue(all(bool(item.get("ok")) for item in report.get("order_checks") or []))

    def test_consent_freeze_snapshot_report_detects_missing_required_token(self):
        invalid_template = """
        <form>
          <textarea id="recipients-textarea" data-testid="recipients-textarea"></textarea>
        </form>
        """.strip()
        report = _build_consent_freeze_snapshot_report(
            content=invalid_template,
            template_path="inline-invalid",
            strict_extras=False,
        )

        self.assertEqual(report.get("status"), "HOLD")
        self.assertIn("missing_required_tokens", report.get("reasons") or [])
        self.assertIn("recipients-cleanup-btn", (report.get("missing") or {}).get("ids") or [])

    def test_consent_freeze_snapshot_report_can_hold_on_extras_when_strict(self):
        extra_template = """
        <div>
          <textarea id="recipients-textarea" data-testid="recipients-textarea"></textarea>
          <button id="recipients-cleanup-btn" data-testid="recipients-cleanup-btn"></button>
          <button id="recipients-cleanup-undo-btn" data-testid="recipients-cleanup-undo-btn"></button>
          <button id="recipients-copy-issues-btn" data-testid="recipients-copy-issues-btn"></button>
          <button id="recipients-prev-issue-btn" data-testid="recipients-prev-issue-btn"></button>
          <button id="recipients-next-issue-btn" data-testid="recipients-next-issue-btn"></button>
          <button data-testid="recipients-jump-top-btn" data-recipients-jump="top"></button>
          <button data-testid="recipients-jump-bottom-btn" data-recipients-jump="bottom"></button>
          <button id="recipients-submit-btn" data-testid="recipients-submit-btn"></button>
          <button id="recipients-extra-btn" data-testid="recipients-extra-btn"></button>
          <input type="hidden" name="recipients_cleanup_applied" />
          <input type="hidden" name="recipients_cleanup_removed_count" />
          <input type="hidden" name="recipients_cleanup_undo_used" />
          <input type="hidden" name="recipients_issue_copy_used" />
          <input type="hidden" name="recipients_issue_jump_count" />
        </div>
        """.strip()
        report = _build_consent_freeze_snapshot_report(
            content=extra_template,
            template_path="inline-extra",
            strict_extras=True,
        )

        self.assertEqual(report.get("status"), "HOLD")
        self.assertIn("unexpected_extra_tokens", report.get("reasons") or [])
        self.assertIn("recipients-extra-btn", (report.get("extra") or {}).get("ids") or [])

    def test_consent_freeze_snapshot_markdown_includes_status_and_checks(self):
        report = {
            "generated_at": "2026-03-03 09:00:00",
            "status": "HOLD",
            "strict_extras": True,
            "reasons": ["missing_required_tokens"],
            "template_path": "sheetbook/templates/sheetbook/consent_review.html",
            "missing": {"ids": ["recipients-submit-btn"], "testids": [], "jump_values": [], "hidden_names": []},
            "extra": {"ids": [], "testids": ["recipients-extra-btn"], "jump_values": [], "hidden_names": []},
            "order_checks": [
                {"name": "cleanup button order", "ok": True, "error": ""},
                {"name": "issue navigation order", "ok": False, "error": "order mismatch"},
            ],
        }
        markdown = _build_consent_freeze_snapshot_markdown(
            report=report,
            json_output_path=Path("docs/handoff/sheetbook_consent_freeze_snapshot_latest.json"),
        )

        self.assertIn("Sheetbook Consent Freeze Snapshot", markdown)
        self.assertIn("`HOLD`", markdown)
        self.assertIn("missing_required_tokens", markdown)
        self.assertIn("recipients-submit-btn", markdown)
        self.assertIn("recipients-extra-btn", markdown)
        self.assertIn("[PASS] cleanup button order", markdown)
        self.assertIn("[FAIL] issue navigation order", markdown)


class SheetbookThresholdRecommendationCommandTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="sheetbook_threshold_user",
            password="pw123456",
            email="sheetbook_threshold_user@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.user,
            defaults={"nickname": "sheetbook_threshold_user", "role": "school"},
        )
        self.sheetbook = Sheetbook.objects.create(owner=self.user, title="임계치 테스트 수첩")
        self.tab = SheetTab.objects.create(
            sheetbook=self.sheetbook,
            name="일정",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=1,
        )

    def _create_event(self, event_name, minutes, metadata=None, user=None):
        event = SheetbookMetricEvent.objects.create(
            event_name=event_name,
            user=user or self.user,
            sheetbook=self.sheetbook,
            tab=self.tab,
            metadata=metadata or {},
        )
        event.created_at = timezone.now() - timedelta(days=1) + timedelta(minutes=minutes)
        event.save(update_fields=["created_at"])

    @override_settings(
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60.0,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50.0,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
    )
    def test_recommend_sheetbook_thresholds_outputs_adjusted_values(self):
        for idx in range(10):
            self._create_event("workspace_home_opened", idx)
        for idx in range(6):
            self._create_event(
                "sheetbook_created",
                30 + idx,
                metadata={"entry_source": "workspace_home_create"},
            )
        for idx in range(3):
            self._create_event(
                "action_execute_requested",
                60 + idx,
                metadata={"entry_source": "workspace_home_create"},
            )

        out = StringIO()
        call_command("recommend_sheetbook_thresholds", "--days", "14", stdout=out)
        value = out.getvalue()

        self.assertIn("관측 전환율: 홈->수첩 생성=60.0%", value)
        self.assertIn("SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=50.0", value)
        self.assertIn("SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=35.0", value)
        self.assertIn("SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5", value)
        self.assertIn("SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5", value)

    @override_settings(
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=62.0,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=41.0,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=5,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=5,
    )
    def test_recommend_sheetbook_thresholds_keeps_current_values_when_sample_low(self):
        for idx in range(2):
            self._create_event("workspace_home_opened", idx)
        self._create_event(
            "sheetbook_created",
            10,
            metadata={"entry_source": "workspace_home_create"},
        )

        out = StringIO()
        call_command("recommend_sheetbook_thresholds", "--days", "14", stdout=out)
        value = out.getvalue()

        self.assertIn("샘플 부족(2 < 5)", value)
        self.assertIn("SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=62.0", value)
        self.assertIn("SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=41.0", value)

    @override_settings(
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60.0,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50.0,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=2,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=2,
    )
    def test_recommend_sheetbook_thresholds_outputs_role_breakdown(self):
        instructor = User.objects.create_user(
            username="sheetbook_threshold_instructor",
            password="pw123456",
            email="sheetbook_threshold_instructor@example.com",
        )
        UserProfile.objects.update_or_create(
            user=instructor,
            defaults={"nickname": "sheetbook_threshold_instructor", "role": "instructor"},
        )

        for idx in range(4):
            self._create_event("workspace_home_opened", idx, user=self.user)
        for idx in range(2):
            self._create_event(
                "sheetbook_created",
                20 + idx,
                metadata={"entry_source": "workspace_home_create"},
                user=self.user,
            )
        self._create_event(
            "action_execute_requested",
            40,
            metadata={"entry_source": "workspace_home_create"},
            user=self.user,
        )

        for idx in range(2):
            self._create_event("workspace_home_opened", 100 + idx, user=instructor)
        self._create_event(
            "sheetbook_created",
            130,
            metadata={"entry_source": "workspace_home_create"},
            user=instructor,
        )

        out = StringIO()
        call_command("recommend_sheetbook_thresholds", "--days", "14", "--group-by-role", stdout=out)
        value = out.getvalue()

        self.assertIn("[sheetbook] role별 재보정 참고", value)
        self.assertIn("role=school: home=4, create=2, action=1, rate=50.0%/50.0%", value)
        self.assertIn("role=instructor: home=2, create=1, action=0, rate=50.0%/0.0%", value)


class SheetbookPilotLogSnapshotScriptTests(TestCase):
    def setUp(self):
        self.school_user = User.objects.create_user(
            username="sheetbook_pilot_school",
            password="pw123456",
            email="sheetbook_pilot_school@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.school_user,
            defaults={"nickname": "sheetbook_pilot_school", "role": "school"},
        )
        self.instructor_user = User.objects.create_user(
            username="sheetbook_pilot_instructor",
            password="pw123456",
            email="sheetbook_pilot_instructor@example.com",
        )
        UserProfile.objects.update_or_create(
            user=self.instructor_user,
            defaults={"nickname": "sheetbook_pilot_instructor", "role": "instructor"},
        )
        self.sheetbook = Sheetbook.objects.create(owner=self.school_user, title="파일럿 로그 테스트 수첩")
        self.tab = SheetTab.objects.create(
            sheetbook=self.sheetbook,
            name="일정",
            tab_type=SheetTab.TYPE_GRID,
            sort_order=1,
        )

    def _create_event(self, user, event_name, minutes, metadata=None):
        event = SheetbookMetricEvent.objects.create(
            event_name=event_name,
            user=user,
            sheetbook=self.sheetbook,
            tab=self.tab,
            metadata=metadata or {},
        )
        event.created_at = timezone.now() - timedelta(days=1) + timedelta(minutes=minutes)
        event.save(update_fields=["created_at"])

    @override_settings(
        SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE=60.0,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE=50.0,
        SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE=2,
        SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE=2,
    )
    def test_pilot_snapshot_includes_role_breakdown(self):
        for idx in range(4):
            self._create_event(self.school_user, "workspace_home_opened", idx)
        for idx in range(2):
            self._create_event(
                self.school_user,
                "sheetbook_created",
                20 + idx,
                metadata={"entry_source": "workspace_home_create"},
            )
        self._create_event(
            self.school_user,
            "action_execute_requested",
            40,
            metadata={"entry_source": "workspace_home_create"},
        )

        for idx in range(2):
            self._create_event(self.instructor_user, "workspace_home_opened", 100 + idx)
        self._create_event(
            self.instructor_user,
            "sheetbook_created",
            130,
            metadata={"entry_source": "workspace_home_create"},
        )

        snapshot = _collect_pilot_log_snapshot(days=14)
        role_breakdown = snapshot.get("role_breakdown") or {}

        self.assertIn("school", role_breakdown)
        self.assertIn("instructor", role_breakdown)
        self.assertEqual(
            role_breakdown["school"]["counts"]["workspace_home_opened_count"],
            4,
        )
        self.assertEqual(
            role_breakdown["school"]["counts"]["workspace_source_create_count"],
            2,
        )
        self.assertEqual(
            role_breakdown["school"]["counts"]["workspace_source_action_requested_count"],
            1,
        )
        self.assertEqual(
            role_breakdown["instructor"]["counts"]["workspace_home_opened_count"],
            2,
        )
        self.assertEqual(
            role_breakdown["instructor"]["counts"]["workspace_source_create_count"],
            1,
        )
        self.assertEqual(
            role_breakdown["instructor"]["counts"]["workspace_source_action_requested_count"],
            0,
        )
        self.assertEqual(role_breakdown["school"]["rates"]["home_to_create"], 50.0)
        self.assertEqual(role_breakdown["school"]["rates"]["create_to_action"], 50.0)
        self.assertEqual(role_breakdown["instructor"]["rates"]["home_to_create"], 50.0)
        self.assertEqual(role_breakdown["instructor"]["rates"]["create_to_action"], 0.0)

    def test_pilot_markdown_role_section_uses_actual_newlines(self):
        snapshot = {
            "days": 14,
            "counts": {
                "workspace_home_opened_count": 6,
                "workspace_source_create_count": 3,
                "workspace_source_action_requested_count": 1,
            },
            "rates": {"home_to_create": 50.0, "create_to_action": 33.3},
            "current": {
                "to_create_target": 60.0,
                "create_to_action_target": 50.0,
                "to_create_min_sample": 5,
                "create_to_action_min_sample": 5,
            },
            "recommended": {
                "to_create_target": 42.0,
                "create_to_action_target": 30.0,
                "to_create_min_sample": 5,
                "create_to_action_min_sample": 5,
                "to_create_reason": "관측치 50.0% - 안정 마진 8.0%",
                "create_to_action_reason": "샘플 부족(3 < 5)",
            },
            "role_breakdown": {
                "school": {
                    "counts": {
                        "workspace_home_opened_count": 4,
                        "workspace_source_create_count": 2,
                        "workspace_source_action_requested_count": 1,
                    },
                    "rates": {"home_to_create": 50.0, "create_to_action": 50.0},
                    "recommended": {
                        "to_create_target": 40.0,
                        "create_to_action_target": 35.0,
                        "to_create_reason": "관측치 50.0% - 안정 마진 10.0%",
                        "create_to_action_reason": "샘플 부족(2 < 5)",
                    },
                },
                "instructor": {
                    "counts": {
                        "workspace_home_opened_count": 2,
                        "workspace_source_create_count": 1,
                        "workspace_source_action_requested_count": 0,
                    },
                    "rates": {"home_to_create": 50.0, "create_to_action": 0.0},
                    "recommended": {
                        "to_create_target": 45.0,
                        "create_to_action_target": 35.0,
                        "to_create_reason": "현재 설정 유지",
                        "create_to_action_reason": "샘플 부족(1 < 5)",
                    },
                },
            },
        }
        row = {
            "date": "2026-03-02",
            "school_or_group": "pilot-school",
            "class_scope": "5학년",
            "active_teachers": "3",
            "workspace_home_opened": "6",
            "home_source_sheetbook_created": "3",
            "home_source_action_execute_requested": "1",
            "home_to_create_rate_pct": "50.0",
            "create_to_action_rate_pct": "33.3",
            "blockers": "없음",
            "next_action": "파일럿 계속",
        }

        markdown = _build_pilot_log_markdown(
            run_datetime=timezone.now(),
            record_date=date(2026, 3, 2),
            row=row,
            snapshot=snapshot,
            reflected_env=False,
            reflected_reason="기본값 유지",
        )

        self.assertIn("## 3) 역할별 스냅샷 참고", markdown)
        self.assertIn("\n- role=instructor:", markdown)
        self.assertIn("\n- role=school:", markdown)
        self.assertNotIn("\\n- role=school", markdown)


class SheetbookBenchmarkCommandTests(TestCase):
    def test_benchmark_sheetbook_grid_runs_small_case(self):
        out = StringIO()

        call_command(
            "benchmark_sheetbook_grid",
            "--cells",
            "12",
            "--cols",
            "4",
            "--runs",
            "1",
            "--batch-sizes",
            "50",
            "--skip-read",
            stdout=out,
        )

        value = out.getvalue()
        self.assertIn("benchmark summary", value)
        self.assertIn("best batch=50", value)

    def test_benchmark_sheetbook_grid_rejects_invalid_cells_option(self):
        out = StringIO()

        with self.assertRaises(CommandError):
            call_command(
                "benchmark_sheetbook_grid",
                "--cells",
                "abc",
                "--cols",
                "4",
                "--runs",
                "1",
                "--batch-sizes",
                "50",
                "--skip-read",
                stdout=out,
                stderr=out,
            )
