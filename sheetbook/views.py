import json
import logging
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
import csv
import io
import re
import secrets
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Min, Q
from django.db.models.functions import TruncDate
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from classcalendar.models import CalendarEvent, EventPageBlock
from collect.models import CollectionRequest
from handoff.models import HandoffReceipt, HandoffRosterGroup, HandoffRosterMember, HandoffSession

from .forms import SheetTabCreateForm, SheetTabRenameForm, SheetbookCreateForm
from .models import ActionInvocation, SavedView, SheetCell, SheetColumn, SheetRow, SheetTab, Sheetbook, SheetbookMetricEvent

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - optional dependency guard
    Workbook = None
    load_workbook = None

SHEETBOOK_CALENDAR_SYNC_SOURCE = "sheetbook_schedule_sync"
SHEETBOOK_CALENDAR_ACTION_SOURCE = "sheetbook_action_calendar"
SHEETBOOK_ACTION_SEED_SESSION_KEY = "sheetbook_action_seeds"
SHEETBOOK_ENTRY_SOURCE_SESSION_KEY = "sheetbook_entry_source_map"
CONSENT_REVIEW_PREVIEW_LIMIT = 10
CONSENT_REVIEW_ISSUE_SAMPLE_LIMIT = 5
CONSENT_REVIEW_ISSUE_LINE_LIMIT = 20
SHEETBOOK_IMPORT_MAX_ROWS = 5000
logger = logging.getLogger(__name__)


def _log_sheetbook_metric(event_name, **fields):
    payload = {"event": event_name}
    payload.update(fields or {})
    logger.info("[sheetbook_metric] %s", json.dumps(payload, ensure_ascii=False, default=str))
    try:
        user_id = fields.get("user_id")
        sheetbook_id = fields.get("sheetbook_id")
        tab_id = fields.get("tab_id")
        action_type = str(fields.get("action") or fields.get("action_type") or "").strip()[:20]

        metadata = dict(fields or {})
        metadata.pop("user_id", None)
        metadata.pop("sheetbook_id", None)
        metadata.pop("tab_id", None)
        metadata.pop("action_type", None)
        metadata.pop("action", None)

        SheetbookMetricEvent.objects.create(
            event_name=str(event_name or "").strip()[:80] or "unknown_event",
            user_id=user_id if user_id else None,
            sheetbook_id=sheetbook_id if sheetbook_id else None,
            tab_id=tab_id if tab_id else None,
            action_type=action_type,
            metadata=metadata,
        )
    except Exception:
        logger.exception("[sheetbook_metric] failed to persist metric event: %s", event_name)


def _normalize_allowlist_values(raw):
    if raw is None:
        return []
    if isinstance(raw, (list, tuple, set)):
        source = raw
    else:
        source = str(raw).split(",")
    values = []
    for item in source:
        value = str(item or "").strip()
        if value:
            values.append(value)
    return values


def _is_sheetbook_beta_user(user):
    if not user or not getattr(user, "is_authenticated", False):
        return False

    username_allowlist = {
        item.lower()
        for item in _normalize_allowlist_values(getattr(settings, "SHEETBOOK_BETA_USERNAMES", ()))
    }
    email_allowlist = {
        item.lower()
        for item in _normalize_allowlist_values(getattr(settings, "SHEETBOOK_BETA_EMAILS", ()))
    }
    id_allowlist = set()
    for item in _normalize_allowlist_values(getattr(settings, "SHEETBOOK_BETA_USER_IDS", ())):
        try:
            id_allowlist.add(int(item))
        except (TypeError, ValueError):
            continue

    username = str(getattr(user, "username", "") or "").strip().lower()
    email = str(getattr(user, "email", "") or "").strip().lower()
    if username and username in username_allowlist:
        return True
    if email and email in email_allowlist:
        return True
    return bool(user.id in id_allowlist if user.id else False)


def _ensure_sheetbook_enabled(user=None):
    if settings.SHEETBOOK_ENABLED:
        return
    if _is_sheetbook_beta_user(user):
        return
    raise Http404("Sheetbook is disabled.")


def _is_htmx_request(request):
    return request.headers.get("HX-Request") == "true"


def _get_owner_sheetbook_or_404(user, pk):
    return get_object_or_404(
        Sheetbook.objects.prefetch_related("tabs"),
        pk=pk,
        owner=user,
    )


def _get_owner_tab_or_404(user, sheetbook_pk, tab_pk):
    return get_object_or_404(
        SheetTab.objects.select_related("sheetbook"),
        pk=tab_pk,
        sheetbook_id=sheetbook_pk,
        sheetbook__owner=user,
    )


def _get_owner_saved_view_or_404(user, sheetbook_pk, tab_pk, view_pk):
    return get_object_or_404(
        SavedView.objects.select_related("tab", "sort_column", "tab__sheetbook"),
        pk=view_pk,
        tab_id=tab_pk,
        tab__sheetbook_id=sheetbook_pk,
        tab__sheetbook__owner=user,
    )


def _create_default_tabs(sheetbook):
    defaults = [
        ("달력", SheetTab.TYPE_CALENDAR),
        ("일정", SheetTab.TYPE_GRID),
        ("학생명부", SheetTab.TYPE_GRID),
        ("메모", SheetTab.TYPE_GRID),
    ]
    for sort_order, (name, tab_type) in enumerate(defaults, start=1):
        tab = SheetTab.objects.create(
            sheetbook=sheetbook,
            name=name,
            tab_type=tab_type,
            sort_order=sort_order,
        )
        if tab_type == SheetTab.TYPE_GRID:
            _seed_default_columns(tab)


def _upsert_grid_cell_value(row, column, raw_value):
    cell, _ = SheetCell.objects.get_or_create(row=row, column=column)
    ok, _ = _apply_cell_value(cell, column, raw_value)
    if not ok:
        cell.value_text = "" if raw_value in (None, "") else str(raw_value)
        cell.value_number = None
        cell.value_bool = None
        cell.value_date = None
        cell.value_json = None
    cell.save()
    return cell


def _seed_sample_rows_for_tab(tab, actor, rows_by_key):
    if tab.tab_type != SheetTab.TYPE_GRID:
        return {"row_count": 0, "cell_count": 0}

    column_map = {column.key: column for column in tab.columns.order_by("sort_order", "id")}
    existing_rows = list(tab.rows.order_by("sort_order", "id"))
    seeded_row_count = 0
    seeded_cell_count = 0
    for index, row_payload in enumerate(rows_by_key or [], start=1):
        if index <= len(existing_rows):
            row = existing_rows[index - 1]
            changed_fields = []
            if row.sort_order != index:
                row.sort_order = index
                changed_fields.append("sort_order")
            if actor and row.updated_by_id != actor.id:
                row.updated_by = actor
                changed_fields.append("updated_by")
            if changed_fields:
                changed_fields.append("updated_at")
                row.save(update_fields=changed_fields)
        else:
            row = SheetRow.objects.create(
                tab=tab,
                sort_order=index,
                created_by=actor,
                updated_by=actor,
            )

        for key, raw_value in (row_payload or {}).items():
            column = column_map.get(str(key or "").strip())
            if not column:
                continue
            _upsert_grid_cell_value(row, column, raw_value)
            seeded_cell_count += 1
        seeded_row_count += 1
    return {"row_count": seeded_row_count, "cell_count": seeded_cell_count}


def _seed_onboarding_sample_sheetbook(sheetbook, actor):
    base_year = sheetbook.academic_year or timezone.localdate().year
    schedule_rows = [
        {"date": f"{base_year}-03-04", "title": "학기 시작 안내", "note": "안내문 액션으로 학부모 안내 발송"},
        {"date": f"{base_year}-03-06", "title": "기초 학력 진단", "note": "체크 결과를 메모 탭에 기록"},
        {"date": f"{base_year}-03-12", "title": "학부모 상담 주간", "note": "학생명부 선택 후 동의서 발송"},
        {"date": f"{base_year}-03-21", "title": "체험학습 사전 점검", "note": "준비물/도우미 확정 후 달력 등록"},
        {"date": f"{base_year}-03-28", "title": "월말 학급 회의", "note": "배부 체크로 회수물 확인"},
    ]
    student_rows = [
        {"number": 1, "name": "김하늘", "contact": "010-1234-1001"},
        {"number": 2, "name": "박나래", "contact": "010-1234-1002"},
        {"number": 3, "name": "이서준", "contact": "010-1234-1003"},
        {"number": 4, "name": "최지우", "contact": "010-1234-1004"},
        {"number": 5, "name": "정민호", "contact": "010-1234-1005"},
    ]
    memo_rows = [
        {"memo": "[60초 시작 가이드]"},
        {"memo": "1) 학생명부에서 2~3명 선택 후 동의서 버튼 클릭"},
        {"memo": "2) 일정 탭에서 날짜+제목 선택 후 달력 등록 클릭"},
        {"memo": "3) 필요한 줄 선택 후 안내문/서명 요청으로 바로 연결"},
    ]

    tabs = list(sheetbook.tabs.all().order_by("sort_order", "id"))
    schedule_tab = next((tab for tab in tabs if tab.name == "일정"), None)
    student_tab = next((tab for tab in tabs if tab.name == "학생명부"), None)
    memo_tab = next((tab for tab in tabs if tab.name == "메모"), None)

    summary = {"seeded_tabs": 0, "seeded_rows": 0, "seeded_cells": 0}
    for tab, rows in (
        (schedule_tab, schedule_rows),
        (student_tab, student_rows),
        (memo_tab, memo_rows),
    ):
        if not tab:
            continue
        seeded = _seed_sample_rows_for_tab(tab, actor, rows)
        if seeded["row_count"] > 0:
            summary["seeded_tabs"] += 1
        summary["seeded_rows"] += seeded["row_count"]
        summary["seeded_cells"] += seeded["cell_count"]
    return summary


def _build_unique_sheetbook_title(owner, base_title):
    normalized = str(base_title or "").strip()[:200]
    if not normalized:
        normalized = f"{timezone.localdate().year} 교무수첩"
    if not Sheetbook.objects.filter(owner=owner, title=normalized).exists():
        return normalized
    for index in range(2, 1000):
        suffix = f" ({index})"
        candidate = f"{normalized[: 200 - len(suffix)]}{suffix}".strip()
        if not Sheetbook.objects.filter(owner=owner, title=candidate).exists():
            return candidate
    stamp = timezone.localtime().strftime("%Y%m%d%H%M%S")
    return f"{normalized[:180]} {stamp}"[:200]


def _clone_sheetbook_structure(source_sheetbook, target_sheetbook, actor, *, include_rows=False):
    tabs_qs = source_sheetbook.tabs.prefetch_related("columns")
    if include_rows:
        tabs_qs = tabs_qs.prefetch_related("rows__cells")
    source_tabs = list(tabs_qs.order_by("sort_order", "id"))
    if not source_tabs:
        _create_default_tabs(target_sheetbook)
        return 4, 0, 0, 0

    cloned_tab_count = 0
    cloned_column_count = 0
    cloned_row_count = 0
    cloned_cell_count = 0
    for source_tab in source_tabs:
        target_tab = SheetTab.objects.create(
            sheetbook=target_sheetbook,
            name=source_tab.name,
            tab_type=source_tab.tab_type,
            sort_order=source_tab.sort_order,
        )
        cloned_tab_count += 1
        if source_tab.tab_type != SheetTab.TYPE_GRID:
            continue

        source_columns = list(source_tab.columns.order_by("sort_order", "id"))
        if not source_columns:
            _seed_default_columns(target_tab)
            cloned_column_count += target_tab.columns.count()
            cloned_row_count += target_tab.rows.count()
            continue

        SheetColumn.objects.bulk_create(
            [
                SheetColumn(
                    tab=target_tab,
                    key=column.key,
                    label=column.label,
                    column_type=column.column_type,
                    sort_order=column.sort_order,
                    width=column.width,
                    is_required=column.is_required,
                    options=column.options or {},
                )
                for column in source_columns
            ],
            batch_size=200,
        )
        cloned_column_count += len(source_columns)
        target_columns = {
            column.key: column
            for column in target_tab.columns.order_by("sort_order", "id")
        }
        if not include_rows:
            SheetRow.objects.create(
                tab=target_tab,
                sort_order=1,
                created_by=actor,
                updated_by=actor,
            )
            cloned_row_count += 1
            continue

        source_rows = list(source_tab.rows.order_by("sort_order", "id"))
        if not source_rows:
            SheetRow.objects.create(
                tab=target_tab,
                sort_order=1,
                created_by=actor,
                updated_by=actor,
            )
            cloned_row_count += 1
            continue

        created_rows = SheetRow.objects.bulk_create(
            [
                SheetRow(
                    tab=target_tab,
                    sort_order=index,
                    created_by=actor,
                    updated_by=actor,
                )
                for index, _ in enumerate(source_rows, start=1)
            ],
            batch_size=200,
        )
        cloned_row_count += len(created_rows)

        cells_to_clone = []
        for source_row, target_row in zip(source_rows, created_rows):
            source_cells = {
                cell.column_id: cell
                for cell in source_row.cells.all()
            }
            for source_column in source_columns:
                source_cell = source_cells.get(source_column.id)
                if not source_cell:
                    continue
                target_column = target_columns.get(source_column.key)
                if not target_column:
                    continue
                cells_to_clone.append(
                    SheetCell(
                        row=target_row,
                        column=target_column,
                        value_text=source_cell.value_text,
                        value_number=source_cell.value_number,
                        value_bool=source_cell.value_bool,
                        value_date=source_cell.value_date,
                        value_json=source_cell.value_json,
                    )
                )
        if cells_to_clone:
            SheetCell.objects.bulk_create(cells_to_clone, batch_size=400)
            cloned_cell_count += len(cells_to_clone)
    return cloned_tab_count, cloned_column_count, cloned_row_count, cloned_cell_count


def _next_tab_order(sheetbook):
    last_tab = sheetbook.tabs.order_by("-sort_order", "-id").first()
    return (last_tab.sort_order + 1) if last_tab else 1


def _next_row_order(tab):
    last_row = tab.rows.order_by("-sort_order", "-id").first()
    return (last_row.sort_order + 1) if last_row else 1


def _next_column_order(tab):
    last_column = tab.columns.order_by("-sort_order", "-id").first()
    return (last_column.sort_order + 1) if last_column else 1


def _get_sheetbook_grid_bulk_batch_size():
    raw = getattr(settings, "SHEETBOOK_GRID_BULK_BATCH_SIZE", 400)
    try:
        parsed = int(raw)
    except (TypeError, ValueError):
        return 400
    return min(max(parsed, 50), 2000)


def _build_unique_column_key(tab, label, fallback_prefix="col"):
    base = slugify(label or "", allow_unicode=True).replace("-", "_")
    if not base:
        base = fallback_prefix
    key = base[:64]
    existing = set(tab.columns.values_list("key", flat=True))
    if key not in existing:
        return key
    for i in range(2, 2000):
        candidate = f"{base[:58]}_{i}"[:64]
        if candidate not in existing:
            return candidate
    return f"{fallback_prefix}_{tab.columns.count()+1}"


def _seed_default_columns(tab):
    presets = {
        "일정": [
            ("date", "날짜", SheetColumn.TYPE_DATE),
            ("title", "제목", SheetColumn.TYPE_TEXT),
            ("note", "메모", SheetColumn.TYPE_TEXT),
        ],
        "학생명부": [
            ("number", "번호", SheetColumn.TYPE_NUMBER),
            ("name", "이름", SheetColumn.TYPE_TEXT),
            ("contact", "연락처", SheetColumn.TYPE_TEXT),
        ],
        "메모": [
            ("memo", "메모", SheetColumn.TYPE_TEXT),
        ],
    }
    default_columns = presets.get(tab.name, [("item", "항목", SheetColumn.TYPE_TEXT), ("value", "내용", SheetColumn.TYPE_TEXT)])
    for sort_order, (key, label, col_type) in enumerate(default_columns, start=1):
        SheetColumn.objects.get_or_create(
            tab=tab,
            key=key,
            defaults={
                "label": label,
                "column_type": col_type,
                "sort_order": sort_order,
            },
        )
    if not tab.rows.exists():
        SheetRow.objects.create(
            tab=tab,
            sort_order=1,
        )


def _normalize_tab_sort_orders(sheetbook):
    for index, tab in enumerate(sheetbook.tabs.order_by("sort_order", "id"), start=1):
        if tab.sort_order != index:
            tab.sort_order = index
            tab.save(update_fields=["sort_order", "updated_at"])


def _render_tab_list_partial(request, sheetbook, *, tab_error="", status=200):
    return render(
        request,
        "sheetbook/_tab_list.html",
        {
            "sheetbook": sheetbook,
            "tabs": sheetbook.tabs.all(),
            "tab_error": tab_error,
        },
        status=status,
    )


def _render_grid_editor_partial(request, sheetbook, selected_tab):
    saved_views = []
    if selected_tab and selected_tab.tab_type == SheetTab.TYPE_GRID:
        saved_views = _list_saved_views_for_tab(selected_tab)
    sheetbook_mobile_read_only = _is_sheetbook_mobile_read_only_request(request)
    return render(
        request,
        "sheetbook/_grid_editor.html",
        {
            "sheetbook": sheetbook,
            "selected_tab": selected_tab,
            "saved_views": saved_views,
            "active_saved_view": None,
            "grid_view_filter": "",
            "grid_view_sort_column_id": 0,
            "grid_view_sort_direction": SavedView.SORT_ASC,
            "sheetbook_mobile_read_only": sheetbook_mobile_read_only,
            "sheetbook_mobile_read_only_message": _sheetbook_mobile_read_only_message(),
        },
    )


def _normalize_saved_view_filter_text(value):
    return str(value or "").strip()[:120]


def _normalize_saved_view_sort_direction(value):
    normalized = str(value or "").strip().lower()
    if normalized == SavedView.SORT_DESC:
        return SavedView.SORT_DESC
    return SavedView.SORT_ASC


def _coerce_saved_view_sort_column_id(tab, raw_value, default=0):
    parsed = _parse_positive_int(raw_value, default=default)
    if not parsed:
        return 0
    return parsed if tab.columns.filter(id=parsed).exists() else 0


def _list_saved_views_for_tab(tab):
    return list(
        SavedView.objects.filter(tab=tab)
        .select_related("sort_column")
        .order_by("-is_favorite", "name", "id")
    )


def _serialize_saved_view(saved_view):
    return {
        "id": saved_view.id,
        "name": saved_view.name,
        "filter_text": saved_view.filter_text or "",
        "sort_column_id": saved_view.sort_column_id or 0,
        "sort_direction": saved_view.sort_direction,
        "is_favorite": bool(saved_view.is_favorite),
        "is_default": bool(saved_view.is_default),
        "sort_column_label": saved_view.sort_column.label if saved_view.sort_column else "",
    }


def _resolve_schedule_source_tab(sheetbook):
    grid_tabs = list(sheetbook.tabs.filter(tab_type=SheetTab.TYPE_GRID).order_by("sort_order", "id"))
    if not grid_tabs:
        return None
    for tab in grid_tabs:
        if (tab.name or "").strip() == "일정":
            return tab
    for tab in grid_tabs:
        if tab.columns.filter(column_type=SheetColumn.TYPE_DATE).exists():
            return tab
    return grid_tabs[0]


def _parse_request_payload(request):
    if request.content_type and "application/json" in request.content_type:
        try:
            return json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return {}
    return request.POST


def _request_prefers_json(request):
    content_type = (request.content_type or "").lower()
    if "application/json" in content_type:
        return True
    if request.headers.get("HX-Request") == "true":
        return True
    accept = (request.headers.get("Accept") or "").lower()
    return "application/json" in accept and "text/html" not in accept


def _is_sheetbook_phone_user_agent(user_agent):
    ua = str(user_agent or "").strip().lower()
    if not ua:
        return False
    if "iphone" in ua or "ipod" in ua:
        return True
    if "android" in ua and "mobile" in ua:
        return True
    if "mobile" in ua and "ipad" not in ua and "tablet" not in ua:
        return True
    return False


def _is_sheetbook_mobile_read_only_request(request):
    return _is_sheetbook_phone_user_agent(request.META.get("HTTP_USER_AGENT", ""))


def _sheetbook_mobile_read_only_message():
    return "휴대폰에서는 읽기 모드로 제공돼요. 수정/생성은 태블릿이나 PC에서 진행해 주세요."


def _sheetbook_archive_read_only_message():
    return "아카이브된 수첩은 읽기 전용입니다. 이어쓰기(복제) 후 수정해 주세요."


def _normalize_sheetbook_status_filter(raw):
    value = str(raw or "").strip().lower()
    if value in {"active", "archived", "all"}:
        return value
    return "active"


def _apply_sheetbook_status_filter_and_order(sheetbook_qs, status_filter):
    normalized = _normalize_sheetbook_status_filter(status_filter)
    if normalized == "active":
        return sheetbook_qs.filter(is_archived=False).order_by("-updated_at", "-id")
    if normalized == "archived":
        return sheetbook_qs.filter(is_archived=True).order_by("-archived_at", "-updated_at", "-id")
    return sheetbook_qs.order_by("is_archived", "-updated_at", "-id")


def _maybe_block_mobile_read_only_edit(
    request,
    *,
    wants_json=False,
    sheetbook_id=0,
    tab_id=0,
    entry_source="",
    blocked_action="",
):
    is_mobile_read_only = _is_sheetbook_mobile_read_only_request(request)
    is_archived_read_only = False
    if sheetbook_id:
        is_archived_read_only = Sheetbook.objects.filter(
            id=sheetbook_id,
            owner=request.user,
            is_archived=True,
        ).exists()
    if not is_mobile_read_only and not is_archived_read_only:
        return None

    if is_archived_read_only:
        _log_sheetbook_metric(
            "sheetbook_archive_read_mode_blocked",
            user_id=getattr(request.user, "id", None),
            sheetbook_id=sheetbook_id or None,
            tab_id=tab_id or None,
            blocked_action=str(blocked_action or "")[:80],
        )
        message = _sheetbook_archive_read_only_message()
    else:
        _log_sheetbook_metric(
            "sheetbook_mobile_read_mode_blocked",
            user_id=getattr(request.user, "id", None),
            sheetbook_id=sheetbook_id or None,
            tab_id=tab_id or None,
            blocked_action=str(blocked_action or "")[:80],
        )
        message = _sheetbook_mobile_read_only_message()
    if wants_json:
        return JsonResponse(
            {
                "ok": False,
                "error": message,
                "mobile_read_only": bool(is_mobile_read_only),
                "archived_read_only": bool(is_archived_read_only),
            },
            status=403,
        )

    messages.info(request, message)
    if sheetbook_id and tab_id:
        return _redirect_sheetbook_tab_detail(sheetbook_id, tab_id, entry_source=entry_source)
    if sheetbook_id:
        detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook_id})
        params = _build_sheetbook_detail_query_params(tab_id=tab_id, entry_source=entry_source)
        if params:
            return redirect(f"{detail_url}?{urlencode(params)}")
        return redirect(detail_url)
    return redirect("sheetbook:index")


def _parse_decimal_or_none(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_date_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    first_chunk = text.split("T", 1)[0]
    first_chunk = first_chunk.split(" ", 1)[0].strip()
    if not first_chunk:
        return None

    def _safe_date(year, month, day):
        try:
            return date(int(year), int(month), int(day))
        except (ValueError, TypeError):
            return None

    compact_digits = re.sub(r"\D", "", first_chunk)
    if len(compact_digits) == 8 and first_chunk.isdigit():
        parsed = _safe_date(compact_digits[:4], compact_digits[4:6], compact_digits[6:8])
        if parsed:
            return parsed

    normalized = first_chunk.replace(".", "-").replace("/", "-")
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", normalized)
    if m:
        return _safe_date(*m.groups())

    m = re.fullmatch(r"(\d{2})-(\d{1,2})-(\d{1,2})", normalized)
    if m:
        yy, month, day = m.groups()
        return _safe_date(2000 + int(yy), month, day)

    m = re.fullmatch(r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일?", first_chunk)
    if m:
        return _safe_date(*m.groups())

    m = re.fullmatch(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일?", first_chunk)
    if m:
        month, day = m.groups()
        return _safe_date(timezone.localdate().year, month, day)

    m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", normalized)
    if m:
        month, day = m.groups()
        return _safe_date(timezone.localdate().year, month, day)

    return None


def _parse_time_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)

    raw = str(value).strip().lower()
    if not raw:
        return None

    normalized = raw.replace(" ", "")
    meridiem = ""
    for prefix, marker in (("오전", "am"), ("오후", "pm"), ("am", "am"), ("pm", "pm")):
        if normalized.startswith(prefix):
            meridiem = marker
            normalized = normalized[len(prefix):]
            break

    hour = None
    minute = 0

    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})", normalized)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2))
    else:
        m = re.fullmatch(r"(\d{1,2})시(?:(\d{1,2})분?)?", normalized)
        if m:
            hour = int(m.group(1))
            minute = int(m.group(2) or 0)
        else:
            period_match = re.search(r"(\d{1,2})교시", normalized)
            if period_match:
                period = int(period_match.group(1))
                if period <= 0:
                    return None
                first_class_hour = _sheetbook_period_first_class_hour()
                first_class_minute = _sheetbook_period_first_class_minute()
                hour = min(22, first_class_hour + (period - 1))
                minute = first_class_minute
            else:
                digits = re.sub(r"\D", "", normalized)
                if re.fullmatch(r"\d{3,4}", digits):
                    if len(digits) == 3:
                        hour = int(digits[:1])
                        minute = int(digits[1:])
                    else:
                        hour = int(digits[:2])
                        minute = int(digits[2:])
                elif re.fullmatch(r"\d{1,2}", digits):
                    hour = int(digits)

    if hour is None:
        return None
    if not (0 <= minute <= 59):
        return None

    if meridiem == "am":
        if hour == 12:
            hour = 0
    elif meridiem == "pm":
        if 1 <= hour <= 11:
            hour += 12

    if not (0 <= hour <= 23):
        return None
    return time(hour=hour, minute=minute)


def _sheetbook_schedule_default_duration_minutes():
    raw = getattr(settings, "SHEETBOOK_SCHEDULE_DEFAULT_DURATION_MINUTES", 50)
    try:
        minutes = int(raw)
    except (TypeError, ValueError):
        return 50
    return max(10, min(240, minutes))


def _sheetbook_period_first_class_hour():
    raw = getattr(settings, "SHEETBOOK_PERIOD_FIRST_CLASS_HOUR", 9)
    try:
        hour = int(raw)
    except (TypeError, ValueError):
        return 9
    return max(6, min(18, hour))


def _sheetbook_period_first_class_minute():
    raw = getattr(settings, "SHEETBOOK_PERIOD_FIRST_CLASS_MINUTE", 0)
    try:
        minute = int(raw)
    except (TypeError, ValueError):
        return 0
    return max(0, min(59, minute))


def _parse_bool_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in ("true", "1", "yes", "y", "on"):
        return True
    if normalized in ("false", "0", "no", "n", "off"):
        return False
    return None


def _serialize_cell_value(cell, column):
    if not cell:
        return None
    if column.column_type == SheetColumn.TYPE_NUMBER:
        return float(cell.value_number) if cell.value_number is not None else None
    if column.column_type == SheetColumn.TYPE_DATE:
        return cell.value_date.isoformat() if cell.value_date else None
    if column.column_type == SheetColumn.TYPE_CHECKBOX:
        return cell.value_bool
    if column.column_type == SheetColumn.TYPE_MULTI_SELECT:
        return cell.value_json or []
    return cell.value_text


def _serialize_cell_value_from_data(cell_data, column_type):
    if not cell_data:
        return None
    if column_type == SheetColumn.TYPE_NUMBER:
        value_number = cell_data.get("value_number")
        return float(value_number) if value_number is not None else None
    if column_type == SheetColumn.TYPE_DATE:
        value_date = cell_data.get("value_date")
        return value_date.isoformat() if value_date else None
    if column_type == SheetColumn.TYPE_CHECKBOX:
        return cell_data.get("value_bool")
    if column_type == SheetColumn.TYPE_MULTI_SELECT:
        return cell_data.get("value_json") or []
    return cell_data.get("value_text", "")


def _serialize_cell_value_for_compare(cell, column):
    value = _serialize_cell_value(cell, column)
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    return str(value)


def _apply_cell_value(cell, column, raw_value):
    # Reset typed fields to avoid stale values between type changes.
    cell.value_text = ""
    cell.value_number = None
    cell.value_bool = None
    cell.value_date = None
    cell.value_json = None

    if column.column_type == SheetColumn.TYPE_NUMBER:
        parsed = _parse_decimal_or_none(raw_value)
        if raw_value not in (None, "") and parsed is None:
            return False, "숫자 칸이라 숫자만 입력할 수 있어요."
        cell.value_number = parsed
        return True, ""
    if column.column_type == SheetColumn.TYPE_DATE:
        parsed = _parse_date_or_none(raw_value)
        if raw_value not in (None, "") and parsed is None:
            return False, "날짜를 읽지 못했어요. 예: 2026-03-14, 2026/3/14, 3/14"
        cell.value_date = parsed
        return True, ""
    if column.column_type == SheetColumn.TYPE_CHECKBOX:
        parsed = _parse_bool_or_none(raw_value)
        if raw_value not in (None, "") and parsed is None:
            return False, "체크 칸은 체크/해제 값만 사용할 수 있어요."
        cell.value_bool = parsed
        return True, ""
    if column.column_type == SheetColumn.TYPE_MULTI_SELECT:
        if raw_value in (None, ""):
            cell.value_json = []
            return True, ""
        if isinstance(raw_value, list):
            cell.value_json = raw_value
            return True, ""
        if isinstance(raw_value, str):
            tokens = [item.strip() for item in re.split(r"[,\n;/|]+", raw_value) if item.strip()]
            cell.value_json = tokens
            return True, ""
        return False, "여러 선택 칸은 목록 형태로 입력해 주세요."

    # text/select/file/grid default
    cell.value_text = "" if raw_value in (None, "") else str(raw_value)
    return True, ""


def _parse_clipboard_matrix(raw_text):
    text = (raw_text or "").replace("\r\n", "\n").replace("\r", "\n")
    text = text.strip("\n")
    if not text:
        return []
    # Prefer tab-separated rows from spreadsheet paste.
    if "\t" in text:
        return [row.split("\t") for row in text.split("\n")]
    # Fallback CSV parsing for comma-based content.
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader if row]


def _parse_csv_upload_matrix(uploaded_file):
    raw_bytes = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("CSV 파일 인코딩을 읽지 못했어요. UTF-8 또는 CP949로 저장해 주세요.")

    matrix = _parse_clipboard_matrix(text)
    if len(matrix) > SHEETBOOK_IMPORT_MAX_ROWS:
        raise ValueError(f"한 번에 가져올 수 있는 행 수는 최대 {SHEETBOOK_IMPORT_MAX_ROWS}줄입니다.")
    return matrix


def _normalize_upload_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date()
    return value


def _parse_xlsx_upload_matrix(uploaded_file):
    if load_workbook is None:
        raise ValueError("엑셀 가져오기에 필요한 openpyxl이 설치되어 있지 않습니다.")

    try:
        workbook = load_workbook(uploaded_file, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("엑셀 파일을 읽지 못했어요. 다시 저장한 뒤 시도해 주세요.") from exc

    try:
        worksheet = workbook.active
        matrix = []
        for row in worksheet.iter_rows(values_only=True):
            values = [_normalize_upload_cell(value) for value in row]
            while values and values[-1] in ("", None):
                values.pop()
            if not values:
                continue
            matrix.append(values)
            if len(matrix) > SHEETBOOK_IMPORT_MAX_ROWS:
                raise ValueError(f"한 번에 가져올 수 있는 행 수는 최대 {SHEETBOOK_IMPORT_MAX_ROWS}줄입니다.")
        return matrix
    finally:
        workbook.close()


def _parse_uploaded_grid_file(uploaded_file):
    filename = str(getattr(uploaded_file, "name", "") or "").strip().lower()
    if filename.endswith(".csv"):
        return _parse_csv_upload_matrix(uploaded_file), "csv"
    if filename.endswith(".xlsx"):
        return _parse_xlsx_upload_matrix(uploaded_file), "xlsx"
    if filename.endswith(".xls"):
        raise ValueError("옛날 엑셀(.xls)은 지원하지 않아요. .xlsx로 저장해서 올려 주세요.")
    raise ValueError("CSV 또는 XLSX 파일만 가져올 수 있어요.")


def _normalize_import_header_labels(header_row):
    labels = []
    for index, value in enumerate(header_row, start=1):
        label = str(value or "").strip()[:120]
        if not label:
            label = f"열{index}"
        labels.append(label)
    return labels


def _ensure_grid_columns_for_import(tab, header_labels):
    existing_columns = list(tab.columns.order_by("sort_order", "id"))
    missing_count = max(0, len(header_labels) - len(existing_columns))
    if missing_count <= 0:
        return 0

    base_order = _next_column_order(tab)
    existing_keys = set(tab.columns.values_list("key", flat=True))
    created_columns = []

    for offset in range(missing_count):
        header_index = len(existing_columns) + offset
        label = header_labels[header_index]
        base_key = slugify(label or "", allow_unicode=True).replace("-", "_") or "col"
        key = base_key[:64]
        if key in existing_keys:
            for suffix in range(2, 2000):
                candidate = f"{base_key[:58]}_{suffix}"[:64]
                if candidate not in existing_keys:
                    key = candidate
                    break
        existing_keys.add(key)
        created_columns.append(
            SheetColumn(
                tab=tab,
                key=key,
                label=label,
                column_type=SheetColumn.TYPE_TEXT,
                sort_order=base_order + offset,
            )
        )

    if created_columns:
        SheetColumn.objects.bulk_create(created_columns, batch_size=200)
    return len(created_columns)


def _serialize_cell_value_for_export(cell_data, column_type):
    if not cell_data:
        return ""

    if column_type == SheetColumn.TYPE_NUMBER:
        value_number = cell_data.get("value_number")
        return "" if value_number is None else str(value_number)
    if column_type == SheetColumn.TYPE_DATE:
        value_date = cell_data.get("value_date")
        return value_date.isoformat() if value_date else ""
    if column_type == SheetColumn.TYPE_CHECKBOX:
        value_bool = cell_data.get("value_bool")
        if value_bool is None:
            return ""
        return "true" if value_bool else "false"
    if column_type == SheetColumn.TYPE_MULTI_SELECT:
        value_json = cell_data.get("value_json")
        if isinstance(value_json, list):
            return ", ".join(str(item).strip() for item in value_json if str(item).strip())
        return ""

    return str(cell_data.get("value_text") or "")


def _build_grid_export_matrix(tab):
    columns = list(tab.columns.order_by("sort_order", "id"))
    if not columns:
        return []

    rows = list(tab.rows.order_by("sort_order", "id").values("id"))
    row_ids = [row["id"] for row in rows]
    column_ids = [column.id for column in columns]
    cell_map = {}
    if row_ids:
        cells = SheetCell.objects.filter(row_id__in=row_ids, column_id__in=column_ids).values(
            "row_id",
            "column_id",
            "value_text",
            "value_number",
            "value_bool",
            "value_date",
            "value_json",
        )
        cell_map = {(cell["row_id"], cell["column_id"]): cell for cell in cells}

    matrix = [[column.label for column in columns]]
    for row in rows:
        row_id = row["id"]
        matrix.append(
            [
                _serialize_cell_value_for_export(cell_map.get((row_id, column.id)), column.column_type)
                for column in columns
            ]
        )
    return matrix


def _build_grid_export_filename(tab, extension):
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    return f"sheetbook_tab_{tab.id}_{timestamp}.{extension}"


def _normalize_column_hint(value):
    return re.sub(r"[\s_]+", "", str(value or "").strip().lower())


def _is_schedule_end_time_column(column):
    key_hint = _normalize_column_hint(column.key)
    label_hint = _normalize_column_hint(column.label)
    if key_hint in {"endtime", "timeend", "finishtime"}:
        return True
    if label_hint in {"종료시간", "끝시간", "마감시간", "종료시각", "끝시각"}:
        return True
    hint = f"{key_hint} {label_hint}"
    return any(token in hint for token in ("종료", "끝", "마감"))


def _is_schedule_start_time_column(column):
    key_hint = _normalize_column_hint(column.key)
    label_hint = _normalize_column_hint(column.label)
    if key_hint in {"starttime", "timestart", "time", "classtime", "period"}:
        return True
    if label_hint in {"시간", "시작시간", "수업시간", "시각", "교시", "시작시각"}:
        return True
    hint = f"{key_hint} {label_hint}"
    if any(token in hint for token in ("종료", "끝", "마감")):
        return False
    return any(token in hint for token in ("시작", "시간", "교시", "시각", "time", "period"))


def _extract_schedule_cells(tab, rows):
    columns = list(tab.columns.order_by("sort_order", "id"))
    row_ids = [row.id for row in rows]
    if not row_ids:
        return [], {}, {}

    cells = list(
        SheetCell.objects.filter(row_id__in=row_ids, column_id__in=[column.id for column in columns])
    )
    cell_map = {(cell.row_id, cell.column_id): cell for cell in cells}
    column_by_key = {column.key: column for column in columns}

    date_column = column_by_key.get("date")
    if not date_column:
        date_column = next((column for column in columns if column.column_type == SheetColumn.TYPE_DATE), None)

    title_column = column_by_key.get("title")
    note_column = column_by_key.get("note")

    if not title_column:
        title_column = next(
            (
                column
                for column in columns
                if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT}
            ),
            None,
        )
    if not note_column:
        text_candidates = [
            column
            for column in columns
            if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT}
            and (not title_column or column.id != title_column.id)
        ]
        note_column = text_candidates[0] if text_candidates else None

    time_like_columns = [
        column
        for column in columns
        if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT, SheetColumn.TYPE_NUMBER}
    ]
    start_time_column = column_by_key.get("start_time") or column_by_key.get("time")
    end_time_column = column_by_key.get("end_time")
    if not end_time_column:
        end_time_column = next(
            (column for column in time_like_columns if _is_schedule_end_time_column(column)),
            None,
        )
    if not start_time_column:
        start_time_column = next(
            (
                column
                for column in time_like_columns
                if (not end_time_column or column.id != end_time_column.id)
                and _is_schedule_start_time_column(column)
            ),
            None,
        )
    if start_time_column and end_time_column and start_time_column.id == end_time_column.id:
        end_time_column = None

    return columns, cell_map, {
        "date": date_column,
        "title": title_column,
        "note": note_column,
        "start_time": start_time_column,
        "end_time": end_time_column,
    }


def _persist_calendar_note(event, note_value):
    note_text = (note_value or "").strip()
    text_blocks = event.blocks.filter(block_type="text").order_by("order", "id")
    primary_block = text_blocks.first()

    if not note_text:
        text_blocks.delete()
        return

    if primary_block:
        primary_block.content = {"text": note_text}
        primary_block.order = 0
        primary_block.save(update_fields=["content", "order"])
        text_blocks.exclude(id=primary_block.id).delete()
        return

    EventPageBlock.objects.create(
        event=event,
        block_type="text",
        content={"text": note_text},
        order=0,
    )


def _column_index_to_label(col_index):
    value = col_index + 1
    label = ""
    while value > 0:
        remainder = (value - 1) % 26
        label = chr(65 + remainder) + label
        value = (value - 1) // 26
    return label or "A"


def _build_sheet_range_label(bounds):
    start_ref = f"{_column_index_to_label(bounds['min_col'])}{bounds['min_row'] + 1}"
    end_ref = f"{_column_index_to_label(bounds['max_col'])}{bounds['max_row'] + 1}"
    return f"{start_ref}:{end_ref}"


def _cell_value_as_text(cell, column):
    value = _serialize_cell_value(cell, column)
    if value in (None, ""):
        return ""
    if isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip())
    return str(value).strip()


def _normalize_selection_bounds(payload, row_count, col_count):
    if row_count <= 0 or col_count <= 0:
        return None, "선택할 칸이 없어요."

    try:
        start_row = int(payload.get("start_row_index", 0))
        start_col = int(payload.get("start_col_index", 0))
        end_row = int(payload.get("end_row_index", start_row))
        end_col = int(payload.get("end_col_index", start_col))
    except (TypeError, ValueError):
        return None, "선택한 칸 정보를 다시 확인해 주세요."

    if start_row < 0 or start_col < 0 or end_row < 0 or end_col < 0:
        return None, "선택한 칸 정보를 다시 확인해 주세요."

    if start_row >= row_count or start_col >= col_count:
        return None, "선택한 칸이 표 범위를 벗어났어요."

    end_row = min(end_row, row_count - 1)
    end_col = min(end_col, col_count - 1)
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    min_col = min(start_col, end_col)
    max_col = max(start_col, end_col)

    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "count": (max_row - min_row + 1) * (max_col - min_col + 1),
    }, ""


def _parse_positive_int(value, default=0):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _parse_percentage_or_default(value, default=0.0):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return float(default)
    if parsed < 0:
        return 0.0
    if parsed > 100:
        return 100.0
    return round(parsed, 1)


def _parse_ratio_or_default(value, default=0.0):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return float(default)
    if parsed < 0:
        return 0.0
    if parsed > 1:
        return 1.0
    return round(parsed, 3)


def _sanitize_entry_source(value):
    raw = str(value or "").strip().lower()
    if not raw:
        return "direct"
    cleaned = re.sub(r"[^a-z0-9_:-]", "", raw)
    return cleaned[:40] or "direct"


def _parse_grid_limit(value, default=50, min_value=20, max_value=1000):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = int(default)
    if parsed < min_value:
        return min_value
    if parsed > max_value:
        return max_value
    return parsed


def _remember_sheetbook_entry_source(request, sheetbook_id, entry_source):
    if not sheetbook_id:
        return
    normalized = _sanitize_entry_source(entry_source)
    if normalized == "direct":
        return
    source_map = request.session.get(SHEETBOOK_ENTRY_SOURCE_SESSION_KEY, {})
    if not isinstance(source_map, dict):
        source_map = {}
    key = str(sheetbook_id)
    source_map.pop(key, None)
    source_map[key] = normalized
    while len(source_map) > 80:
        oldest_key = next(iter(source_map))
        source_map.pop(oldest_key, None)
    request.session[SHEETBOOK_ENTRY_SOURCE_SESSION_KEY] = source_map
    request.session.modified = True


def _resolve_sheetbook_entry_source(request, sheetbook_id, entry_source):
    normalized = _sanitize_entry_source(entry_source)
    if normalized != "direct":
        _remember_sheetbook_entry_source(request, sheetbook_id, normalized)
        return normalized
    source_map = request.session.get(SHEETBOOK_ENTRY_SOURCE_SESSION_KEY, {})
    if isinstance(source_map, dict):
        stored = _sanitize_entry_source(source_map.get(str(sheetbook_id)))
        if stored != "direct":
            return stored
    return normalized


def _build_sheetbook_search_results(sheetbook, query, limit_per_group=8):
    normalized = str(query or "").strip()
    if not normalized:
        return {"tabs": [], "cells": [], "actions": []}

    detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook.id})
    results = {"tabs": [], "cells": [], "actions": []}

    tab_hits = list(
        sheetbook.tabs.filter(name__icontains=normalized).order_by("sort_order", "id")[:limit_per_group]
    )
    for tab in tab_hits:
        params = {"tab": tab.id, "q": normalized}
        results["tabs"].append(
            {
                "title": tab.name,
                "meta": "탭 이름",
                "url": f"{detail_url}?{urlencode(params)}",
            }
        )

    cell_filters = Q(value_text__icontains=normalized) | Q(value_json__icontains=normalized)
    parsed_decimal = _parse_decimal_or_none(normalized)
    parsed_date = _parse_date_or_none(normalized)
    if parsed_decimal is not None:
        cell_filters |= Q(value_number=parsed_decimal)
    if parsed_date is not None:
        cell_filters |= Q(value_date=parsed_date)

    cell_hits = list(
        SheetCell.objects.filter(row__tab__sheetbook=sheetbook)
        .select_related("row__tab", "column")
        .filter(cell_filters)
        .order_by("row__tab__sort_order", "row__sort_order", "column__sort_order", "id")[:limit_per_group]
    )
    for cell in cell_hits:
        params = {
            "tab": cell.row.tab_id,
            "q": normalized,
            "focus_row_id": cell.row_id,
            "focus_col_id": cell.column_id,
        }
        row_label = (cell.row.sort_order or 0) + 1
        results["cells"].append(
            {
                "title": f"{cell.row.tab.name} · {row_label}행 · {cell.column.label}",
                "meta": _cell_value_as_text(cell, cell.column)[:100] or "(값 없음)",
                "url": f"{detail_url}?{urlencode(params)}",
            }
        )

    action_keyword_map = {
        "달력": ActionInvocation.ACTION_CALENDAR,
        "수합": ActionInvocation.ACTION_COLLECT,
        "동의": ActionInvocation.ACTION_CONSENT,
        "서명": ActionInvocation.ACTION_SIGNATURE,
        "배부": ActionInvocation.ACTION_HANDOFF,
        "안내": ActionInvocation.ACTION_NOTICE,
    }
    action_filters = Q(summary__icontains=normalized) | Q(result_label__icontains=normalized)
    for token, action_type in action_keyword_map.items():
        if token in normalized:
            action_filters |= Q(action_type=action_type)

    action_hits = list(
        ActionInvocation.objects.filter(sheetbook=sheetbook)
        .select_related("tab")
        .filter(action_filters)
        .order_by("-created_at", "-id")[:limit_per_group]
    )
    for invocation in action_hits:
        params = {"tab": invocation.tab_id, "q": normalized}
        results["actions"].append(
            {
                "title": (
                    f"{invocation.get_action_type_display()} "
                    f"({timezone.localtime(invocation.created_at).strftime('%m-%d %H:%M')})"
                ),
                "meta": (invocation.summary or invocation.result_label or "최근 실행 기록")[:100],
                "url": invocation.result_url or f"{detail_url}?{urlencode(params)}",
            }
        )
    return results


def _build_global_sheetbook_search_results(user, query, limit_per_group=6):
    normalized = str(query or "").strip()
    if not normalized:
        return {"tabs": [], "cells": [], "actions": []}

    results = {"tabs": [], "cells": [], "actions": []}

    tab_hits = list(
        SheetTab.objects.filter(
            sheetbook__owner=user,
            name__icontains=normalized,
        )
        .select_related("sheetbook")
        .order_by("-sheetbook__updated_at", "sort_order", "id")[:limit_per_group]
    )
    for tab in tab_hits:
        detail_url = reverse("sheetbook:detail", kwargs={"pk": tab.sheetbook_id})
        params = {"tab": tab.id, "q": normalized}
        results["tabs"].append(
            {
                "title": f"{tab.sheetbook.title} · {tab.name}",
                "meta": "탭 이름",
                "url": f"{detail_url}?{urlencode(params)}",
            }
        )

    cell_filters = Q(value_text__icontains=normalized) | Q(value_json__icontains=normalized)
    parsed_decimal = _parse_decimal_or_none(normalized)
    parsed_date = _parse_date_or_none(normalized)
    if parsed_decimal is not None:
        cell_filters |= Q(value_number=parsed_decimal)
    if parsed_date is not None:
        cell_filters |= Q(value_date=parsed_date)

    cell_hits = list(
        SheetCell.objects.filter(row__tab__sheetbook__owner=user)
        .select_related("row__tab__sheetbook", "column")
        .filter(cell_filters)
        .order_by(
            "-row__tab__sheetbook__updated_at",
            "row__tab__sort_order",
            "row__sort_order",
            "column__sort_order",
            "id",
        )[:limit_per_group]
    )
    for cell in cell_hits:
        detail_url = reverse("sheetbook:detail", kwargs={"pk": cell.row.tab.sheetbook_id})
        params = {
            "tab": cell.row.tab_id,
            "q": normalized,
            "focus_row_id": cell.row_id,
            "focus_col_id": cell.column_id,
        }
        row_label = (cell.row.sort_order or 0) + 1
        results["cells"].append(
            {
                "title": f"{cell.row.tab.sheetbook.title} · {cell.row.tab.name} · {row_label}행 {cell.column.label}",
                "meta": _cell_value_as_text(cell, cell.column)[:100] or "(값 없음)",
                "url": f"{detail_url}?{urlencode(params)}",
            }
        )

    action_keyword_map = {
        "달력": ActionInvocation.ACTION_CALENDAR,
        "수합": ActionInvocation.ACTION_COLLECT,
        "동의": ActionInvocation.ACTION_CONSENT,
        "서명": ActionInvocation.ACTION_SIGNATURE,
        "배부": ActionInvocation.ACTION_HANDOFF,
        "안내": ActionInvocation.ACTION_NOTICE,
    }
    action_filters = Q(summary__icontains=normalized) | Q(result_label__icontains=normalized)
    for token, action_type in action_keyword_map.items():
        if token in normalized:
            action_filters |= Q(action_type=action_type)

    action_hits = list(
        ActionInvocation.objects.filter(sheetbook__owner=user)
        .select_related("sheetbook", "tab")
        .filter(action_filters)
        .order_by("-created_at", "-id")[:limit_per_group]
    )
    for invocation in action_hits:
        detail_url = reverse("sheetbook:detail", kwargs={"pk": invocation.sheetbook_id})
        params = {"tab": invocation.tab_id, "q": normalized}
        results["actions"].append(
            {
                "title": (
                    f"{invocation.sheetbook.title} · {invocation.get_action_type_display()} "
                    f"({timezone.localtime(invocation.created_at).strftime('%m-%d %H:%M')})"
                ),
                "meta": (invocation.summary or invocation.result_label or "최근 실행 기록")[:100],
                "url": invocation.result_url or f"{detail_url}?{urlencode(params)}",
            }
        )
    return results


def _collect_selection_snapshot(tab, bounds):
    rows = list(
        tab.rows.order_by("sort_order", "id")[bounds["min_row"] : bounds["max_row"] + 1]
    )
    columns = list(
        tab.columns.order_by("sort_order", "id")[bounds["min_col"] : bounds["max_col"] + 1]
    )
    if not rows or not columns:
        return [], [], {}

    row_ids = [row.id for row in rows]
    column_ids = [column.id for column in columns]
    cells = list(
        SheetCell.objects.filter(row_id__in=row_ids, column_id__in=column_ids)
    )
    cell_map = {(cell.row_id, cell.column_id): cell for cell in cells}
    return rows, columns, cell_map


def _build_selection_sample(rows, columns, cell_map, max_rows=5):
    preview_lines = []
    for row in rows[:max_rows]:
        values = [
            _cell_value_as_text(cell_map.get((row.id, column.id)), column)
            for column in columns
        ]
        cleaned = [value for value in values if value]
        if cleaned:
            preview_lines.append(" | ".join(cleaned))
    return preview_lines


def _extract_selection_names(rows, columns, cell_map, max_count=200):
    if not columns:
        return []
    first_column = columns[0]
    names = []
    seen = set()
    for row in rows:
        name = _cell_value_as_text(cell_map.get((row.id, first_column.id)), first_column)
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name[:100])
        if len(names) >= max_count:
            break
    return names


def _normalize_header_token(value):
    token = str(value or "").strip().lower()
    for ch in (" ", "_", "-", "/", ".", "(", ")", "[", "]"):
        token = token.replace(ch, "")
    return token


def _is_parent_header(label):
    token = _normalize_header_token(label)
    if not token:
        return False
    keywords = (
        "보호자",
        "학부모",
        "부모",
        "parent",
        "guardian",
        "어머니",
        "아버지",
    )
    return any(key in token for key in keywords)


def _is_phone_header(label):
    token = _normalize_header_token(label)
    if not token:
        return False
    keywords = (
        "연락처",
        "전화",
        "휴대폰",
        "핸드폰",
        "phone",
        "mobile",
        "tel",
    )
    return any(key in token for key in keywords)


def _is_student_header(label):
    token = _normalize_header_token(label)
    if not token:
        return False
    if _is_parent_header(token) or _is_phone_header(token):
        return False
    keywords = (
        "학생명",
        "학생",
        "아동",
        "이름",
        "성명",
        "student",
        "name",
    )
    return any(key in token for key in keywords)


def _is_affiliation_header(label):
    token = _normalize_header_token(label)
    if not token:
        return False
    keywords = (
        "학년반",
        "학반",
        "반",
        "소속",
        "부서",
        "직위",
        "담당",
        "class",
        "affiliation",
        "group",
    )
    return any(key in token for key in keywords)


def _normalize_phone_value(value):
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""
    if digits.startswith("82") and len(digits) in (11, 12):
        digits = f"0{digits[2:]}"
    if len(digits) < 8 or len(digits) > 11:
        return ""
    if not digits.startswith("0"):
        return ""
    return digits


def _column_nonempty_values(rows, column, cell_map, max_samples=20):
    values = []
    for row in rows:
        raw = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
        if not raw:
            continue
        values.append(raw)
        if len(values) >= max_samples:
            break
    return values


def _contains_parent_marker(value):
    text = str(value or "").strip().lower()
    if not text:
        return False
    markers = (
        "보호자",
        "학부모",
        "어머니",
        "아버지",
        "parent",
        "guardian",
        "mother",
        "father",
    )
    return any(marker in text for marker in markers)


def _is_name_like_value(value):
    text = str(value or "").strip()
    if not text:
        return False
    if _normalize_phone_value(text):
        return False

    compact = re.sub(r"\s+", "", text)
    if not compact:
        return False
    if len(compact) > 24:
        return False
    if _contains_parent_marker(compact):
        return False
    if re.fullmatch(r"[0-9\-\(\)\.,_/]+", compact):
        return False

    digit_count = sum(1 for ch in compact if ch.isdigit())
    if digit_count and (digit_count / len(compact)) > 0.4:
        return False

    return any("가" <= ch <= "힣" for ch in compact) or any(ch.isalpha() for ch in compact)


def _score_student_column(rows, column, cell_map):
    values = _column_nonempty_values(rows, column, cell_map)
    if not values:
        return 0.0
    name_like = sum(1 for value in values if _is_name_like_value(value))
    return name_like / len(values)


def _extract_signature_participants(rows, columns, cell_map, max_count=300):
    if not rows or not columns:
        return []

    text_like_columns = [
        column
        for column in columns
        if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT, SheetColumn.TYPE_NUMBER}
    ]
    if not text_like_columns:
        text_like_columns = list(columns)

    name_column = next((column for column in text_like_columns if _is_student_header(column.label)), None)
    if not name_column:
        best_score = 0.0
        best_column = None
        for column in text_like_columns:
            score = _score_student_column(rows, column, cell_map)
            if score > best_score:
                best_score = score
                best_column = column
        if best_column and best_score >= 0.35:
            name_column = best_column
        elif text_like_columns:
            name_column = text_like_columns[0]

    affiliation_column = next(
        (
            column
            for column in text_like_columns
            if _is_affiliation_header(column.label)
            and (not name_column or column.id != name_column.id)
        ),
        None,
    )
    if not affiliation_column:
        for column in text_like_columns:
            if name_column and column.id == name_column.id:
                continue
            sample_values = _column_nonempty_values(rows, column, cell_map)
            if not sample_values:
                continue
            non_phone = [value for value in sample_values if not _normalize_phone_value(value)]
            if non_phone:
                affiliation_column = column
                break

    participants = []
    seen = set()
    for row in rows:
        name = ""
        if name_column:
            name = _cell_value_as_text(cell_map.get((row.id, name_column.id)), name_column)
        if not _is_name_like_value(name):
            name = ""
            for column in text_like_columns:
                candidate = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if _is_name_like_value(candidate):
                    name = candidate
                    break
        if not name:
            continue

        affiliation = ""
        if affiliation_column:
            affiliation = _cell_value_as_text(cell_map.get((row.id, affiliation_column.id)), affiliation_column)
            if _normalize_phone_value(affiliation):
                affiliation = ""

        dedupe_key = (name, affiliation)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        participants.append({"name": name[:100], "affiliation": affiliation[:100]})
        if len(participants) >= max_count:
            break
    return participants


def _extract_signature_datetime_seed(rows, columns, cell_map):
    selected_date = None
    selected_time = None
    for row in rows:
        row_date = None
        row_time = None
        for column in columns:
            value = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
            if not value:
                continue
            if row_date is None:
                row_date = _parse_date_or_none(value)
            if row_time is None:
                row_time = _parse_time_or_none(value)
            if row_date and row_time:
                break
        if row_date and selected_date is None:
            selected_date = row_date
        if row_time and selected_time is None:
            selected_time = row_time
        if selected_date and selected_time:
            break

    if selected_date is None:
        return ""
    if selected_time is None:
        selected_time = time(hour=15, minute=30)
    return datetime.combine(selected_date, selected_time).strftime("%Y-%m-%dT%H:%M")


def _extract_consent_recipients(rows, columns, cell_map, max_count=200):
    if not rows or not columns:
        return []

    text_like_columns = [
        column
        for column in columns
        if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT, SheetColumn.TYPE_NUMBER}
    ]
    if not text_like_columns:
        text_like_columns = list(columns)

    phone_column = next((column for column in text_like_columns if _is_phone_header(column.label)), None)
    parent_column = next((column for column in text_like_columns if _is_parent_header(column.label)), None)
    student_column = next(
        (
            column
            for column in text_like_columns
            if _is_student_header(column.label)
            and (not parent_column or column.id != parent_column.id)
            and (not phone_column or column.id != phone_column.id)
        ),
        None,
    )

    excluded_ids = {
        column.id
        for column in (phone_column, parent_column, student_column)
        if column is not None
    }

    if not phone_column:
        best_phone_ratio = 0.0
        best_phone_column = None
        for column in text_like_columns:
            if column.id in excluded_ids:
                continue
            values = _column_nonempty_values(rows, column, cell_map)
            if not values:
                continue
            valid = [value for value in values if _normalize_phone_value(value)]
            ratio = len(valid) / len(values)
            if ratio > best_phone_ratio and ratio >= 0.6:
                best_phone_ratio = ratio
                best_phone_column = column
        phone_column = best_phone_column
        if phone_column:
            excluded_ids.add(phone_column.id)

    if not parent_column:
        best_parent_ratio = 0.0
        best_parent_column = None
        for column in text_like_columns:
            if column.id in excluded_ids:
                continue
            values = _column_nonempty_values(rows, column, cell_map)
            if not values:
                continue
            ratio = sum(1 for value in values if _contains_parent_marker(value)) / len(values)
            if ratio > best_parent_ratio and ratio >= 0.4:
                best_parent_ratio = ratio
                best_parent_column = column
        parent_column = best_parent_column
        if parent_column:
            excluded_ids.add(parent_column.id)

    if not student_column:
        best_student_score = 0.0
        best_student_column = None
        for column in text_like_columns:
            if phone_column and column.id == phone_column.id:
                continue
            if parent_column and column.id == parent_column.id:
                continue
            score = _score_student_column(rows, column, cell_map)
            if score > best_student_score:
                best_student_score = score
                best_student_column = column
        if best_student_column and best_student_score >= 0.4:
            student_column = best_student_column
        else:
            for column in text_like_columns:
                if phone_column and column.id == phone_column.id:
                    continue
                if parent_column and column.id == parent_column.id:
                    continue
                student_column = column
                break

    recipients = []
    seen = set()
    for row in rows:
        student_name = ""
        if student_column:
            student_name = _cell_value_as_text(cell_map.get((row.id, student_column.id)), student_column)
        if not student_name:
            for column in text_like_columns:
                if parent_column and column.id == parent_column.id:
                    continue
                if phone_column and column.id == phone_column.id:
                    continue
                candidate = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if candidate and _is_name_like_value(candidate):
                    student_name = candidate
                    break
        if not student_name:
            for column in text_like_columns:
                if parent_column and column.id == parent_column.id:
                    continue
                if phone_column and column.id == phone_column.id:
                    continue
                candidate = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if candidate:
                    student_name = candidate
                    break
        student_name = student_name[:100].strip()
        if not student_name:
            continue

        parent_name = ""
        if parent_column:
            parent_name = _cell_value_as_text(cell_map.get((row.id, parent_column.id)), parent_column)
        parent_name = parent_name[:100].strip() if parent_name else ""
        if not parent_name:
            parent_name = f"{student_name} 보호자"

        phone_value = ""
        if phone_column:
            phone_raw = _cell_value_as_text(cell_map.get((row.id, phone_column.id)), phone_column)
            phone_value = _normalize_phone_value(phone_raw)

        dedupe_key = (student_name, parent_name, phone_value)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        recipients.append(
            {
                "student_name": student_name,
                "parent_name": parent_name,
                "phone_number": phone_value,
            }
        )
        if len(recipients) >= max_count:
            break

    return recipients


def _stash_sheetbook_action_seed(request, *, action_type, data):
    seeds = request.session.get(SHEETBOOK_ACTION_SEED_SESSION_KEY, {})
    if not isinstance(seeds, dict):
        seeds = {}

    token = secrets.token_urlsafe(12)
    seeds[token] = {
        "action": action_type,
        "data": data or {},
        "created_at": timezone.now().isoformat(),
    }
    if len(seeds) > 40:
        for old_key in list(seeds.keys())[: len(seeds) - 40]:
            seeds.pop(old_key, None)
    request.session[SHEETBOOK_ACTION_SEED_SESSION_KEY] = seeds
    request.session.modified = True
    return token


def _peek_sheetbook_action_seed(request, token, *, expected_action=""):
    token = (token or "").strip()
    if not token:
        return None
    seeds = request.session.get(SHEETBOOK_ACTION_SEED_SESSION_KEY, {})
    if not isinstance(seeds, dict):
        return None
    seed = seeds.get(token)
    if not isinstance(seed, dict):
        return None
    if expected_action and seed.get("action") != expected_action:
        return None
    return seed


def _update_sheetbook_action_seed_data(request, token, *, expected_action="", data=None):
    token = (token or "").strip()
    if not token:
        return False
    seeds = request.session.get(SHEETBOOK_ACTION_SEED_SESSION_KEY, {})
    if not isinstance(seeds, dict):
        return False
    seed = seeds.get(token)
    if not isinstance(seed, dict):
        return False
    if expected_action and seed.get("action") != expected_action:
        return False
    seed["data"] = data or {}
    seeds[token] = seed
    request.session[SHEETBOOK_ACTION_SEED_SESSION_KEY] = seeds
    request.session.modified = True
    return True


def _parse_recipients_text_block_with_meta(text):
    recipients = []
    seen = set()
    input_line_count = 0
    duplicate_count = 0
    skipped_line_count = 0
    duplicate_samples = []
    skipped_samples = []
    duplicate_line_numbers = []
    skipped_line_numbers = []
    duplicate_issue_items = []
    skipped_issue_items = []
    for line_no, raw in enumerate((text or "").splitlines(), start=1):
        line = raw.strip()
        if not line:
            continue
        input_line_count += 1
        parts = [item.strip() for item in line.split(",")]
        if len(parts) < 2:
            skipped_line_count += 1
            if len(skipped_samples) < CONSENT_REVIEW_ISSUE_SAMPLE_LIMIT:
                skipped_samples.append(line[:120])
            if len(skipped_line_numbers) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                skipped_line_numbers.append(line_no)
            if len(skipped_issue_items) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                skipped_issue_items.append({"line_no": line_no, "text": line[:120]})
            continue
        student_name = parts[0][:100]
        parent_name = parts[1][:100]
        if not student_name or not parent_name:
            skipped_line_count += 1
            if len(skipped_samples) < CONSENT_REVIEW_ISSUE_SAMPLE_LIMIT:
                skipped_samples.append(line[:120])
            if len(skipped_line_numbers) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                skipped_line_numbers.append(line_no)
            if len(skipped_issue_items) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                skipped_issue_items.append({"line_no": line_no, "text": line[:120]})
            continue
        phone_number = (parts[2] if len(parts) >= 3 else "").strip()
        normalized_phone = _normalize_phone_value(phone_number) if phone_number else ""
        dedupe_key = (student_name, parent_name, normalized_phone)
        if dedupe_key in seen:
            duplicate_count += 1
            if len(duplicate_samples) < CONSENT_REVIEW_ISSUE_SAMPLE_LIMIT:
                duplicate_samples.append(line[:120])
            if len(duplicate_line_numbers) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                duplicate_line_numbers.append(line_no)
            if len(duplicate_issue_items) < CONSENT_REVIEW_ISSUE_LINE_LIMIT:
                duplicate_issue_items.append({"line_no": line_no, "text": line[:120]})
            continue
        seen.add(dedupe_key)
        recipients.append(
            {
                "student_name": student_name,
                "parent_name": parent_name,
                "phone_number": normalized_phone,
            }
        )
    return recipients, {
        "input_line_count": input_line_count,
        "accepted_count": len(recipients),
        "duplicate_count": duplicate_count,
        "skipped_line_count": skipped_line_count,
        "duplicate_samples": duplicate_samples,
        "skipped_samples": skipped_samples,
        "duplicate_line_numbers": duplicate_line_numbers,
        "skipped_line_numbers": skipped_line_numbers,
        "issue_line_numbers": sorted(set(duplicate_line_numbers + skipped_line_numbers)),
        "duplicate_issue_items": duplicate_issue_items,
        "skipped_issue_items": skipped_issue_items,
    }


def _parse_recipients_text_block(text):
    recipients, _ = _parse_recipients_text_block_with_meta(text)
    return recipients


def _build_recipients_text_block(recipients):
    lines = []
    seen = set()
    for rec in recipients or []:
        student_name = str(rec.get("student_name") or "").strip()[:100]
        parent_name = str(rec.get("parent_name") or "").strip()[:100]
        phone_number = _normalize_phone_value(rec.get("phone_number") or "")
        if not student_name or not parent_name:
            continue
        dedupe_key = (student_name, parent_name, phone_number)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        lines.append(f"{student_name},{parent_name},{phone_number}")
    return "\n".join(lines)


def _serialize_action_invocation(invocation):
    return {
        "id": invocation.id,
        "action": invocation.action_type,
        "action_label": invocation.get_action_type_display(),
        "status": invocation.status,
        "status_label": invocation.get_status_display(),
        "summary": invocation.summary,
        "result_label": invocation.result_label,
        "result_url": invocation.result_url,
        "created_at": timezone.localtime(invocation.created_at).strftime("%m-%d %H:%M"),
    }


def _action_failure_fallback_message(action_type):
    if action_type == ActionInvocation.ACTION_CALENDAR:
        return "달력 등록 중 문제가 생겼어요. 날짜 칸을 확인한 뒤 다시 시도해 주세요."
    if action_type == ActionInvocation.ACTION_COLLECT:
        return "간편 수합 만들기 중 문제가 생겼어요. 선택한 칸을 다시 확인해 주세요."
    if action_type == ActionInvocation.ACTION_CONSENT:
        return "동의서 연결 중 문제가 생겼어요. 잠시 후 다시 시도해 주세요."
    if action_type == ActionInvocation.ACTION_SIGNATURE:
        return "서명 요청 연결 중 문제가 생겼어요. 잠시 후 다시 시도해 주세요."
    if action_type == ActionInvocation.ACTION_HANDOFF:
        return "배부 체크 만들기 중 문제가 생겼어요. 이름 칸을 확인해 주세요."
    if action_type == ActionInvocation.ACTION_NOTICE:
        return "안내문 연결 중 문제가 생겼어요. 잠시 후 다시 시도해 주세요."
    return "처리 중 문제가 생겼어요. 잠시 후 다시 시도해 주세요."


def _build_sheetbook_detail_query_params(
    *,
    tab_id=0,
    entry_source="",
    view_id=0,
    view_filter="",
    sort_col=0,
    sort_dir=SavedView.SORT_ASC,
):
    params = {}
    if tab_id:
        params["tab"] = int(tab_id)

    normalized_source = _sanitize_entry_source(entry_source)
    if normalized_source != "direct":
        params["source"] = normalized_source

    normalized_view_id = _parse_positive_int(view_id, default=0)
    if normalized_view_id:
        params["view"] = normalized_view_id

    normalized_filter = _normalize_saved_view_filter_text(view_filter)
    if normalized_filter:
        params["view_filter"] = normalized_filter

    normalized_sort_col = _parse_positive_int(sort_col, default=0)
    if normalized_sort_col:
        params["sort_col"] = normalized_sort_col

    normalized_sort_dir = _normalize_saved_view_sort_direction(sort_dir)
    if normalized_sort_col and normalized_sort_dir == SavedView.SORT_DESC:
        params["sort_dir"] = normalized_sort_dir

    return params


def _build_sheetbook_index_query_params(
    *,
    search_query="",
    entry_source="",
    status_filter="active",
    page=0,
):
    params = {}
    query = str(search_query or "").strip()[:80]
    if query:
        params["q"] = query

    normalized_source = _sanitize_entry_source(entry_source)
    if normalized_source != "direct":
        params["source"] = normalized_source

    normalized_status = _normalize_sheetbook_status_filter(status_filter)
    if normalized_status != "active":
        params["status"] = normalized_status

    page_no = _parse_positive_int(page, default=0)
    if page_no > 1:
        params["page"] = page_no

    return params


def _redirect_sheetbook_tab_detail(
    sheetbook_id,
    tab_id,
    *,
    entry_source="",
    view_id=0,
    view_filter="",
    sort_col=0,
    sort_dir=SavedView.SORT_ASC,
):
    detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook_id})
    params = _build_sheetbook_detail_query_params(
        tab_id=tab_id,
        entry_source=entry_source,
        view_id=view_id,
        view_filter=view_filter,
        sort_col=sort_col,
        sort_dir=sort_dir,
    )
    if params:
        return redirect(f"{detail_url}?{urlencode(params)}")
    return redirect(detail_url)


def _execute_calendar_registration(*, user, sheetbook, tab, rows, columns, cell_map, bounds, invocation):
    date_column = next(
        (column for column in columns if column.column_type == SheetColumn.TYPE_DATE),
        None,
    )
    text_columns = [
        column
        for column in columns
        if column.column_type in {SheetColumn.TYPE_TEXT, SheetColumn.TYPE_SELECT}
    ]
    title_column = text_columns[0] if text_columns else None
    note_column = text_columns[1] if len(text_columns) > 1 else None

    tz = timezone.get_current_timezone()
    created_events = []
    skipped_rows = []

    for row in rows:
        row_date = None
        if date_column:
            date_cell = cell_map.get((row.id, date_column.id))
            if date_cell and date_cell.value_date:
                row_date = date_cell.value_date
            elif date_cell:
                row_date = _parse_date_or_none(date_cell.value_text)

        if not row_date:
            for column in columns:
                candidate_text = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if not candidate_text:
                    continue
                parsed = _parse_date_or_none(candidate_text)
                if parsed:
                    row_date = parsed
                    break

        if not row_date:
            skipped_rows.append(row.sort_order)
            continue

        title = ""
        if title_column:
            title = _cell_value_as_text(cell_map.get((row.id, title_column.id)), title_column)

        if not title:
            for column in columns:
                if date_column and column.id == date_column.id:
                    continue
                candidate_title = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if candidate_title:
                    title = candidate_title
                    break

        if not title:
            title = "일정"

        note = ""
        if note_column:
            note = _cell_value_as_text(cell_map.get((row.id, note_column.id)), note_column)
        if not note:
            note_parts = []
            for column in columns:
                if date_column and column.id == date_column.id:
                    continue
                if title_column and column.id == title_column.id:
                    continue
                token = _cell_value_as_text(cell_map.get((row.id, column.id)), column)
                if token:
                    note_parts.append(token)
            note = " / ".join(note_parts[:3])

        start_dt = timezone.make_aware(datetime.combine(row_date, time.min), tz)
        end_dt = start_dt + timedelta(days=1)
        event = CalendarEvent.objects.create(
            title=title,
            author=user,
            start_time=start_dt,
            end_time=end_dt,
            is_all_day=True,
            visibility=CalendarEvent.VISIBILITY_TEACHER,
            source=CalendarEvent.SOURCE_LOCAL,
            color="indigo",
            integration_source=SHEETBOOK_CALENDAR_ACTION_SOURCE,
            integration_key=f"{sheetbook.id}:{tab.id}:{invocation.id}:{row.id}",
            is_locked=False,
        )
        _persist_calendar_note(event, note)
        created_events.append(event)

    if not created_events:
        raise ValueError("선택한 칸에서 날짜를 찾지 못했어요. 날짜 칸을 포함해 다시 선택해 주세요.")

    created_count = len(created_events)
    skipped_count = len(skipped_rows)
    range_label = _build_sheet_range_label(bounds)
    summary = f"{range_label} 선택 칸에서 일정 {created_count}건을 달력에 넣었어요."
    if skipped_count:
        summary = f"{summary} 날짜가 없는 줄 {skipped_count}개는 제외했어요."

    return {
        "summary": summary,
        "result_label": f"달력 일정 {created_count}건",
        "result_url": reverse("classcalendar:legacy_main"),
        "payload": {
            "created": created_count,
            "skipped_rows": skipped_rows,
            "event_ids": [str(event.id) for event in created_events[:20]],
        },
    }


def _execute_collect_creation(*, user, sheetbook, tab, rows, columns, cell_map, bounds):
    first_column = columns[0] if columns else None
    submitter_names = []
    seen_names = set()
    if first_column:
        for row in rows:
            name = _cell_value_as_text(cell_map.get((row.id, first_column.id)), first_column)
            if not name or name in seen_names:
                continue
            seen_names.add(name)
            submitter_names.append(name)

    preview_lines = _build_selection_sample(rows, columns, cell_map, max_rows=5)
    range_label = _build_sheet_range_label(bounds)
    summary_line = f"교무수첩 범위 {range_label}에서 생성됨"
    description_lines = [
        summary_line,
        f"- 수첩: {sheetbook.title}",
        f"- 탭: {tab.name}",
    ]
    if preview_lines:
        description_lines.append("")
        description_lines.append("선택 데이터 미리보기:")
        description_lines.extend(f"- {line}" for line in preview_lines)

    title = f"{sheetbook.title} · {tab.name} 수합"
    collection_request = CollectionRequest.objects.create(
        creator=user,
        title=title[:200],
        description="\n".join(description_lines).strip(),
        expected_submitters="\n".join(submitter_names[:200]),
    )

    return {
        "summary": f"{range_label} 선택 칸으로 간편 수합을 만들었어요.",
        "result_label": collection_request.title,
        "result_url": reverse(
            "collect:request_detail",
            kwargs={"request_id": collection_request.id},
        ),
        "payload": {
            "request_id": str(collection_request.id),
            "expected_submitter_count": len(submitter_names),
            "preview_lines": preview_lines,
        },
    }


def _build_unique_handoff_group_name(user, base_name):
    normalized = (base_name or "").strip()[:120] or "교무수첩 명단"
    if not HandoffRosterGroup.objects.filter(owner=user, name=normalized).exists():
        return normalized
    for idx in range(2, 200):
        suffix = f" ({idx})"
        candidate = f"{normalized[: max(1, 120 - len(suffix))]}{suffix}"
        if not HandoffRosterGroup.objects.filter(owner=user, name=candidate).exists():
            return candidate
    return f"{normalized[:116]} (새)"


def _execute_handoff_creation(*, user, sheetbook, tab, rows, columns, cell_map, bounds):
    if not columns:
        raise ValueError("배부 체크를 만들 칸이 없습니다.")

    name_column = columns[0]
    note_column = columns[1] if len(columns) > 1 else None
    members_payload = []
    seen = set()

    for order, row in enumerate(rows, start=1):
        name = _cell_value_as_text(cell_map.get((row.id, name_column.id)), name_column)
        if not name or name in seen:
            continue
        seen.add(name)
        note = ""
        if note_column:
            note = _cell_value_as_text(cell_map.get((row.id, note_column.id)), note_column)[:120]
        members_payload.append(
            {
                "display_name": name[:100],
                "sort_order": order,
                "note": note,
            }
        )

    if not members_payload:
        raise ValueError("첫 번째 열에 이름이 없어 배부 체크를 만들 수 없어요.")

    range_label = _build_sheet_range_label(bounds)
    group_name = _build_unique_handoff_group_name(
        user,
        f"{sheetbook.title} {tab.name} 명단",
    )
    group = HandoffRosterGroup.objects.create(
        owner=user,
        name=group_name,
        description=f"교무수첩 {sheetbook.title} / {tab.name} ({range_label})",
    )

    members = [
        HandoffRosterMember(
            group=group,
            display_name=item["display_name"],
            sort_order=item["sort_order"],
            note=item["note"],
            is_active=True,
        )
        for item in members_payload
    ]
    HandoffRosterMember.objects.bulk_create(members)

    session = HandoffSession.objects.create(
        owner=user,
        roster_group=group,
        roster_group_name=group.name,
        title=f"{tab.name} 배부 체크",
        note=f"교무수첩 {sheetbook.title}의 {range_label} 칸에서 자동 생성",
    )

    created_members = list(group.members.filter(is_active=True).order_by("sort_order", "id"))
    HandoffReceipt.objects.bulk_create(
        [
            HandoffReceipt(
                session=session,
                member=member,
                member_name_snapshot=member.display_name,
                member_order_snapshot=member.sort_order,
            )
            for member in created_members
        ]
    )

    return {
        "summary": f"{range_label} 선택 칸으로 배부 체크를 시작했어요. ({len(created_members)}명)",
        "result_label": session.title,
        "result_url": reverse("handoff:session_detail", kwargs={"session_id": session.id}),
        "payload": {
            "group_id": str(group.id),
            "session_id": str(session.id),
            "member_count": len(created_members),
        },
    }


def _execute_guide_action(
    *,
    request,
    action_type,
    sheetbook,
    tab,
    rows,
    columns,
    cell_map,
    bounds,
):
    range_label = _build_sheet_range_label(bounds)
    selection_lines = _build_selection_sample(rows, columns, cell_map, max_rows=5)

    if action_type == ActionInvocation.ACTION_CONSENT:
        consent_recipients = _extract_consent_recipients(rows, columns, cell_map, max_count=200)
        consent_title = f"{tab.name} 동의서"
        consent_message_lines = [
            f"{sheetbook.title}의 {tab.name} 내용을 바탕으로 동의서를 보냅니다.",
            "내용 확인 후 동의 부탁드립니다.",
        ]
        if selection_lines:
            consent_message_lines.append("")
            consent_message_lines.append("참고 내용:")
            consent_message_lines.extend(f"- {line}" for line in selection_lines[:3])
        recipients_text = "\n".join(
            f"{rec['student_name']},{rec['parent_name']},{rec['phone_number']}"
            for rec in consent_recipients
        )
        seed_token = _stash_sheetbook_action_seed(
            request,
            action_type=ActionInvocation.ACTION_CONSENT,
            data={
                "title": consent_title,
                "message": "\n".join(consent_message_lines).strip(),
                "document_title": f"{tab.name} 안내문",
                "recipients_text": recipients_text,
                "range_label": range_label,
            },
        )
        consent_url = f"{reverse('sheetbook:consent_seed_review', kwargs={'pk': sheetbook.id, 'tab_pk': tab.id})}?{urlencode({'sb_seed': seed_token})}"
        return {
            "summary": f"{range_label} 선택 칸 기준으로 수신자 확인 화면을 열어요.",
            "result_label": "동의서 확인 화면 열기",
            "result_url": consent_url,
            "payload": {
                "guide": "consent_create",
                "seed_token": seed_token,
                "prefilled_recipients": len(consent_recipients),
            },
        }
    if action_type == ActionInvocation.ACTION_SIGNATURE:
        signature_participants = _extract_signature_participants(
            rows,
            columns,
            cell_map,
            max_count=300,
        )
        participants_text = "\n".join(
            f"{item['name']},{item['affiliation']}" if item["affiliation"] else item["name"]
            for item in signature_participants
        )
        description_lines = [
            f"{sheetbook.title}의 {tab.name} 선택 칸을 기준으로 서명 요청을 준비합니다.",
            "제목/일시를 확인한 뒤 연수 서명 링크를 배포해 주세요.",
        ]
        if selection_lines:
            description_lines.append("")
            description_lines.append("참고 내용:")
            description_lines.extend(f"- {line}" for line in selection_lines[:3])
        instructor_seed = request.user.get_full_name().strip() or request.user.username
        datetime_seed = _extract_signature_datetime_seed(rows, columns, cell_map)
        seed_token = _stash_sheetbook_action_seed(
            request,
            action_type=ActionInvocation.ACTION_SIGNATURE,
            data={
                "title": f"{tab.name} 서명 요청",
                "print_title": f"{tab.name} 참석 서명",
                "instructor": instructor_seed[:100],
                "location": "교실",
                "datetime": datetime_seed,
                "description": "\n".join(description_lines).strip()[:1200],
                "participants_text": participants_text,
                "expected_count": len(signature_participants),
                "range_label": range_label,
            },
        )
        signature_url = f"{reverse('signatures:create')}?{urlencode({'sb_seed': seed_token, 'from': 'sheetbook'})}"
        return {
            "summary": f"{range_label} 선택 칸을 바탕으로 서명 요청 화면을 열어요.",
            "result_label": "서명 요청 만들기 열기",
            "result_url": signature_url,
            "payload": {
                "guide": "signature_create",
                "seed_token": seed_token,
                "prefilled_participants": len(signature_participants),
            },
        }
    if action_type == ActionInvocation.ACTION_NOTICE:
        keywords = "\n".join(selection_lines).strip()
        if not keywords:
            keywords = f"{sheetbook.title} {tab.name} 관련 안내사항"
        seed_token = _stash_sheetbook_action_seed(
            request,
            action_type=ActionInvocation.ACTION_NOTICE,
            data={
                "target": "parent",
                "topic": "notice",
                "length_style": "medium",
                "keywords": keywords[:1200],
                "range_label": range_label,
            },
        )
        notice_url = f"{reverse('noticegen:main')}?{urlencode({'sb_seed': seed_token, 'from': 'sheetbook'})}"
        return {
            "summary": f"{range_label} 선택 칸 내용을 넣어둔 안내문 멘트 화면을 열어요.",
            "result_label": "안내문 만들기 열기",
            "result_url": notice_url,
            "payload": {
                "guide": "notice_create",
                "seed_token": seed_token,
            },
        }
    raise ValueError("지원하지 않는 액션입니다.")


@login_required
def index(request):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.GET.get("source"))
    search_query = str(request.GET.get("q") or "").strip()[:80]
    status_filter = _normalize_sheetbook_status_filter(request.GET.get("status"))
    owner_sheetbook_qs = Sheetbook.objects.filter(owner=request.user)
    status_counts = {
        "active": owner_sheetbook_qs.filter(is_archived=False).count(),
        "archived": owner_sheetbook_qs.filter(is_archived=True).count(),
        "all": owner_sheetbook_qs.count(),
    }
    sheetbook_qs = _apply_sheetbook_status_filter_and_order(
        owner_sheetbook_qs.prefetch_related("tabs"),
        status_filter,
    )
    if search_query:
        sheetbook_qs = sheetbook_qs.filter(title__icontains=search_query)
    paginator = Paginator(sheetbook_qs, 20)
    sheetbook_page = paginator.get_page(request.GET.get("page") or 1)
    sheetbooks = list(sheetbook_page.object_list)
    form = SheetbookCreateForm()
    _log_sheetbook_metric(
        "sheetbook_index_opened",
        user_id=request.user.id,
        sheetbook_count=sheetbook_page.paginator.count,
        has_search=bool(search_query),
        entry_source=entry_source,
        status_filter=status_filter,
    )
    return render(
        request,
        "sheetbook/index.html",
        {
            "form": form,
            "sheetbooks": sheetbooks,
            "sheetbook_page": sheetbook_page,
            "search_query": search_query,
            "entry_source": entry_source,
            "status_filter": status_filter,
            "status_counts": status_counts,
            "show_sample_onboarding_cta": (
                status_filter == "active"
                and not search_query
                and sheetbook_page.paginator.count == 0
            ),
        },
    )


@login_required
@require_POST
def quick_create_sheetbook(request):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.POST.get("source") or "workspace_home_create")
    current_year = timezone.localdate().year
    title = _build_unique_sheetbook_title(request.user, f"{current_year} 교무수첩")

    with transaction.atomic():
        sheetbook = Sheetbook.objects.create(
            owner=request.user,
            title=title,
            academic_year=current_year,
        )
        _create_default_tabs(sheetbook)

    _remember_sheetbook_entry_source(request, sheetbook.id, entry_source)
    _log_sheetbook_metric(
        "sheetbook_created",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        has_academic_year=True,
        entry_source=entry_source,
        quick_flow="workspace_quick_create",
    )
    messages.success(request, "새 교무수첩을 만들었어요. 바로 입력을 시작해 보세요.")
    detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk})
    return redirect(f"{detail_url}?source={entry_source}")


@login_required
@require_POST
def quick_copy_sheetbook(request):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.POST.get("source") or "workspace_home_copy")
    include_rows = bool(_parse_bool_or_none(request.POST.get("include_rows")))
    source_sheetbook = (
        Sheetbook.objects.filter(owner=request.user)
        .prefetch_related("tabs__columns")
        .order_by("-updated_at", "-id")
        .first()
    )
    if not source_sheetbook:
        messages.info(request, "먼저 교무수첩을 만든 뒤 이어쓰기를 사용할 수 있어요.")
        index_url = reverse("sheetbook:index")
        return redirect(f"{index_url}?source={entry_source}")

    next_year = source_sheetbook.academic_year + 1 if source_sheetbook.academic_year else timezone.localdate().year
    next_year = max(2000, min(2100, next_year))
    base_title = source_sheetbook.title
    if source_sheetbook.academic_year:
        replaced = re.sub(
            rf"(?<!\d){source_sheetbook.academic_year}(?!\d)",
            str(next_year),
            base_title,
            count=1,
        )
        base_title = replaced if replaced != source_sheetbook.title else f"{source_sheetbook.title} ({next_year})"
    else:
        base_title = f"{source_sheetbook.title} (복제)"
    title = _build_unique_sheetbook_title(request.user, base_title)

    with transaction.atomic():
        sheetbook = Sheetbook.objects.create(
            owner=request.user,
            title=title,
            academic_year=next_year,
        )
        cloned_tab_count, cloned_column_count, cloned_row_count, cloned_cell_count = _clone_sheetbook_structure(
            source_sheetbook,
            sheetbook,
            request.user,
            include_rows=include_rows,
        )

    _remember_sheetbook_entry_source(request, sheetbook.id, entry_source)
    _log_sheetbook_metric(
        "sheetbook_created",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        has_academic_year=bool(sheetbook.academic_year),
        entry_source=entry_source,
        quick_flow="workspace_quick_copy",
        copied_from_sheetbook_id=source_sheetbook.id,
        copied_with_rows=include_rows,
        copied_from_tab_count=source_sheetbook.tabs.count(),
        cloned_tab_count=cloned_tab_count,
        cloned_column_count=cloned_column_count,
        cloned_row_count=cloned_row_count,
        cloned_cell_count=cloned_cell_count,
    )
    if include_rows:
        messages.success(request, f"'{source_sheetbook.title}' 내용을 포함해 새 교무수첩을 만들었어요.")
    else:
        messages.success(request, f"'{source_sheetbook.title}' 구성을 복제해 새 교무수첩을 만들었어요.")
    detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk})
    return redirect(f"{detail_url}?source={entry_source}")


def _redirect_sheetbook_index_with_state(request):
    entry_source = _sanitize_entry_source(request.POST.get("source") or request.GET.get("source"))
    search_query = str(request.POST.get("q") or request.GET.get("q") or "").strip()[:80]
    status_filter = _normalize_sheetbook_status_filter(request.POST.get("status") or request.GET.get("status"))
    page = _parse_positive_int(request.POST.get("page") or request.GET.get("page"), default=0)
    index_url = reverse("sheetbook:index")
    params = _build_sheetbook_index_query_params(
        search_query=search_query,
        entry_source=entry_source,
        status_filter=status_filter,
        page=page,
    )
    if not params:
        return redirect(index_url)
    return redirect(f"{index_url}?{urlencode(params)}")


@login_required
@require_POST
def archive_sheetbook(request, pk):
    _ensure_sheetbook_enabled(request.user)
    sheetbook = _get_owner_sheetbook_or_404(request.user, pk)
    if sheetbook.is_archived:
        messages.info(request, "이미 아카이브된 수첩입니다.")
        return _redirect_sheetbook_index_with_state(request)

    sheetbook.is_archived = True
    sheetbook.archived_at = timezone.now()
    sheetbook.save(update_fields=["is_archived", "archived_at", "updated_at"])
    _log_sheetbook_metric(
        "sheetbook_archived",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        archived=True,
    )
    messages.success(request, f"'{sheetbook.title}' 수첩을 아카이브했어요.")
    return _redirect_sheetbook_index_with_state(request)


@login_required
@require_POST
def unarchive_sheetbook(request, pk):
    _ensure_sheetbook_enabled(request.user)
    sheetbook = _get_owner_sheetbook_or_404(request.user, pk)
    if not sheetbook.is_archived:
        messages.info(request, "이미 활성 상태인 수첩입니다.")
        return _redirect_sheetbook_index_with_state(request)

    sheetbook.is_archived = False
    sheetbook.archived_at = None
    sheetbook.save(update_fields=["is_archived", "archived_at", "updated_at"])
    _log_sheetbook_metric(
        "sheetbook_unarchived",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        archived=False,
    )
    messages.success(request, f"'{sheetbook.title}' 수첩을 다시 활성화했어요.")
    return _redirect_sheetbook_index_with_state(request)


@login_required
@require_POST
def bulk_archive_update(request):
    _ensure_sheetbook_enabled(request.user)
    raw_ids = request.POST.getlist("sheetbook_ids")
    selected_ids = []
    for raw in raw_ids:
        parsed = _parse_positive_int(raw, default=0)
        if parsed > 0:
            selected_ids.append(parsed)
    selected_ids = list(dict.fromkeys(selected_ids))
    archive_action = str(request.POST.get("archive_action") or "").strip().lower()

    if not selected_ids:
        messages.info(request, "선택된 수첩이 없어요. 목록에서 수첩을 먼저 선택해 주세요.")
        return _redirect_sheetbook_index_with_state(request)

    if archive_action not in {"archive", "unarchive"}:
        messages.error(request, "일괄 처리 동작을 다시 선택해 주세요.")
        return _redirect_sheetbook_index_with_state(request)

    qs = Sheetbook.objects.filter(owner=request.user, id__in=selected_ids)
    matched_count = qs.count()
    ignored_count = max(0, len(selected_ids) - matched_count)
    if matched_count <= 0:
        messages.info(request, "선택한 수첩을 찾을 수 없어요. 목록을 새로고침한 뒤 다시 선택해 주세요.")
        _log_sheetbook_metric(
            "sheetbook_archive_bulk_updated",
            user_id=request.user.id,
            selected_count=len(selected_ids),
            matched_count=0,
            changed_count=0,
            unchanged_count=0,
            ignored_count=ignored_count,
            archive_action=archive_action,
        )
        return _redirect_sheetbook_index_with_state(request)

    now = timezone.now()
    if archive_action == "archive":
        changed = qs.filter(is_archived=False).update(
            is_archived=True,
            archived_at=now,
            updated_at=now,
        )
        unchanged = max(0, matched_count - changed)
        message = f"{changed}개 수첩을 아카이브했어요. (이미 아카이브 {unchanged}개)"
    else:
        changed = qs.filter(is_archived=True).update(
            is_archived=False,
            archived_at=None,
            updated_at=now,
        )
        unchanged = max(0, matched_count - changed)
        message = f"{changed}개 수첩을 활성 상태로 되돌렸어요. (이미 활성 {unchanged}개)"

    if ignored_count > 0:
        message = f"{message} / 접근 불가·삭제됨 {ignored_count}개 제외"
    messages.success(request, message)

    _log_sheetbook_metric(
        "sheetbook_archive_bulk_updated",
        user_id=request.user.id,
        selected_count=len(selected_ids),
        matched_count=matched_count,
        changed_count=changed,
        unchanged_count=unchanged,
        ignored_count=ignored_count,
        archive_action=archive_action,
    )
    return _redirect_sheetbook_index_with_state(request)


@login_required
@require_POST
def quick_sample_sheetbook(request):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.POST.get("source") or "workspace_home_sample")
    current_year = timezone.localdate().year
    title = _build_unique_sheetbook_title(request.user, f"{current_year} 새학기 샘플 수첩")

    with transaction.atomic():
        sheetbook = Sheetbook.objects.create(
            owner=request.user,
            title=title,
            academic_year=current_year,
        )
        _create_default_tabs(sheetbook)
        seeded_summary = _seed_onboarding_sample_sheetbook(sheetbook, request.user)

    schedule_tab = (
        sheetbook.tabs.filter(tab_type=SheetTab.TYPE_GRID, name="일정")
        .order_by("sort_order", "id")
        .first()
    )
    selected_tab_id = schedule_tab.id if schedule_tab else 0
    _remember_sheetbook_entry_source(request, sheetbook.id, entry_source)
    _log_sheetbook_metric(
        "sheetbook_created",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        has_academic_year=True,
        entry_source=entry_source,
        quick_flow="workspace_quick_sample",
        sample_seeded=True,
        sample_seeded_tabs=seeded_summary["seeded_tabs"],
        sample_seeded_rows=seeded_summary["seeded_rows"],
        sample_seeded_cells=seeded_summary["seeded_cells"],
    )
    _log_sheetbook_metric(
        "sheetbook_sample_onboarding_started",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        tab_id=selected_tab_id or None,
        entry_source=entry_source,
    )
    messages.success(request, "샘플 수첩을 만들었어요. 60초 가이드대로 바로 시작해 보세요.")
    detail_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk})
    params = _build_sheetbook_detail_query_params(
        tab_id=selected_tab_id,
        entry_source=entry_source,
    )
    params["onboarding"] = "sample"
    return redirect(f"{detail_url}?{urlencode(params)}")


@login_required
@require_POST
def create_sheetbook(request):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.POST.get("source"))
    status_filter = _normalize_sheetbook_status_filter(request.POST.get("status"))
    form = SheetbookCreateForm(request.POST)
    if not form.is_valid():
        owner_sheetbook_qs = Sheetbook.objects.filter(owner=request.user)
        status_counts = {
            "active": owner_sheetbook_qs.filter(is_archived=False).count(),
            "archived": owner_sheetbook_qs.filter(is_archived=True).count(),
            "all": owner_sheetbook_qs.count(),
        }
        sheetbook_qs = _apply_sheetbook_status_filter_and_order(
            owner_sheetbook_qs.prefetch_related("tabs"),
            status_filter,
        )
        sheetbook_page = Paginator(sheetbook_qs, 20).get_page(request.POST.get("page") or 1)
        sheetbooks = list(sheetbook_page.object_list)
        _log_sheetbook_metric(
            "sheetbook_create_validation_failed",
            user_id=request.user.id,
            error_keys=sorted(list(form.errors.keys())),
            entry_source=entry_source,
            status_filter=status_filter,
        )
        return render(
            request,
            "sheetbook/index.html",
            {
                "form": form,
                "sheetbooks": sheetbooks,
                "sheetbook_page": sheetbook_page,
                "search_query": "",
                "entry_source": entry_source,
                "status_filter": status_filter,
                "status_counts": status_counts,
                "show_sample_onboarding_cta": status_filter == "active" and sheetbook_page.paginator.count == 0,
            },
            status=400,
        )

    sheetbook = form.save(commit=False)
    sheetbook.owner = request.user
    sheetbook.save()
    _create_default_tabs(sheetbook)
    _remember_sheetbook_entry_source(request, sheetbook.id, entry_source)
    _log_sheetbook_metric(
        "sheetbook_created",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        has_academic_year=bool(sheetbook.academic_year),
        entry_source=entry_source,
    )
    messages.success(request, "교무수첩이 생성되었습니다.")
    return redirect("sheetbook:detail", pk=sheetbook.pk)


@login_required
def detail(request, pk):
    _ensure_sheetbook_enabled(request.user)
    entry_source = _sanitize_entry_source(request.GET.get("source"))
    index_search_query = str(request.GET.get("index_q") or "").strip()[:80]
    index_status_filter = _normalize_sheetbook_status_filter(request.GET.get("index_status"))
    index_page = _parse_positive_int(request.GET.get("index_page"), default=0)
    index_back_url = reverse("sheetbook:index")
    index_back_params = _build_sheetbook_index_query_params(
        search_query=index_search_query,
        entry_source=entry_source,
        status_filter=index_status_filter,
        page=index_page,
    )
    if index_back_params:
        index_back_url = f"{index_back_url}?{urlencode(index_back_params)}"
    mobile_read_only_requested = _is_sheetbook_mobile_read_only_request(request)
    sheetbook_mobile_read_only = mobile_read_only_requested
    sheetbook_mobile_read_only_message = _sheetbook_mobile_read_only_message()
    grid_limit = _parse_grid_limit(request.GET.get("grid_limit"), default=50)
    search_query = str(request.GET.get("q") or "").strip()[:80]
    focus_row_id = _parse_positive_int(request.GET.get("focus_row_id"), default=0)
    focus_col_id = _parse_positive_int(request.GET.get("focus_col_id"), default=0)
    selected_saved_view_id = _parse_positive_int(request.GET.get("view"), default=0)
    has_manual_view_filter = "view_filter" in request.GET
    has_manual_sort_col = "sort_col" in request.GET
    has_manual_sort_dir = "sort_dir" in request.GET
    has_manual_view_overrides = has_manual_view_filter or has_manual_sort_col or has_manual_sort_dir
    grid_view_filter = _normalize_saved_view_filter_text(request.GET.get("view_filter"))
    grid_view_sort_direction = _normalize_saved_view_sort_direction(request.GET.get("sort_dir"))
    grid_view_sort_column_id = 0
    onboarding_mode = str(request.GET.get("onboarding") or "").strip().lower()
    show_sample_onboarding = onboarding_mode == "sample"
    onboarding_dismiss_url = ""
    saved_views = []
    active_saved_view = None
    sheetbook = _get_owner_sheetbook_or_404(request.user, pk)
    sheetbook_archived_read_only = bool(sheetbook.is_archived)
    if sheetbook_archived_read_only:
        sheetbook_mobile_read_only = True
        sheetbook_mobile_read_only_message = _sheetbook_archive_read_only_message()
    _remember_sheetbook_entry_source(request, sheetbook.id, entry_source)
    search_results = {"tabs": [], "cells": [], "actions": []}
    if search_query:
        search_results = _build_sheetbook_search_results(sheetbook, search_query)
        _log_sheetbook_metric(
            "sheetbook_search_requested",
            user_id=request.user.id,
            sheetbook_id=sheetbook.id,
            query=search_query[:40],
            tab_result_count=len(search_results["tabs"]),
            cell_result_count=len(search_results["cells"]),
            action_result_count=len(search_results["actions"]),
        )
    tabs = list(sheetbook.tabs.all().order_by("sort_order", "id"))
    selected_tab = tabs[0] if tabs else None
    schedule_source_tab = None
    action_invocations = []
    action_history_has_more = False
    action_history_next_cursor = None
    selected_tab_id = request.GET.get("tab")
    if selected_tab_id and tabs:
        for tab in tabs:
            if str(tab.id) == str(selected_tab_id):
                selected_tab = tab
                break
    if selected_tab and selected_tab.tab_type == SheetTab.TYPE_CALENDAR:
        schedule_source_tab = _resolve_schedule_source_tab(sheetbook)
    if selected_tab and selected_tab.tab_type == SheetTab.TYPE_GRID:
        saved_views = _list_saved_views_for_tab(selected_tab)
        if selected_saved_view_id:
            active_saved_view = next(
                (saved_view for saved_view in saved_views if saved_view.id == selected_saved_view_id),
                None,
            )
        if active_saved_view and not has_manual_view_overrides:
            grid_view_filter = _normalize_saved_view_filter_text(active_saved_view.filter_text)
            grid_view_sort_column_id = active_saved_view.sort_column_id or 0
            grid_view_sort_direction = _normalize_saved_view_sort_direction(active_saved_view.sort_direction)
        elif not active_saved_view and not selected_saved_view_id and not has_manual_view_overrides:
            default_saved_view = next((saved_view for saved_view in saved_views if saved_view.is_default), None)
            if default_saved_view:
                active_saved_view = default_saved_view
                selected_saved_view_id = default_saved_view.id
                grid_view_filter = _normalize_saved_view_filter_text(default_saved_view.filter_text)
                grid_view_sort_column_id = default_saved_view.sort_column_id or 0
                grid_view_sort_direction = _normalize_saved_view_sort_direction(default_saved_view.sort_direction)
        elif has_manual_sort_col:
            grid_view_sort_column_id = _parse_positive_int(request.GET.get("sort_col"), default=0)

        grid_view_sort_column_id = _coerce_saved_view_sort_column_id(
            selected_tab,
            grid_view_sort_column_id,
            default=0,
        )
        if focus_row_id:
            grid_limit = max(grid_limit, 1000)
        initial_limit = 8
        action_invocation_batch = list(
            ActionInvocation.objects.filter(
                sheetbook=sheetbook,
                tab=selected_tab,
            )
            .select_related("actor")
            .order_by("-created_at", "-id")[: initial_limit + 1]
        )
        action_history_has_more = len(action_invocation_batch) > initial_limit
        action_invocations = action_invocation_batch[:initial_limit]
        if action_history_has_more and action_invocations:
            action_history_next_cursor = action_invocations[-1].id
    else:
        focus_row_id = 0
        focus_col_id = 0
        grid_view_filter = ""
        grid_view_sort_column_id = 0
        grid_view_sort_direction = SavedView.SORT_ASC
        selected_saved_view_id = 0
        active_saved_view = None
        saved_views = []
    _log_sheetbook_metric(
        "sheetbook_detail_opened",
        user_id=request.user.id,
        sheetbook_id=sheetbook.id,
        tab_id=selected_tab.id if selected_tab else None,
        tab_type=selected_tab.tab_type if selected_tab else "",
        entry_source=entry_source,
    )
    if mobile_read_only_requested:
        _log_sheetbook_metric(
            "sheetbook_mobile_read_mode_opened",
            user_id=request.user.id,
            sheetbook_id=sheetbook.id,
            tab_id=selected_tab.id if selected_tab else None,
        )
    if sheetbook_archived_read_only:
        _log_sheetbook_metric(
            "sheetbook_archive_read_mode_opened",
            user_id=request.user.id,
            sheetbook_id=sheetbook.id,
            tab_id=selected_tab.id if selected_tab else None,
        )
    dismiss_params = _build_sheetbook_detail_query_params(
        tab_id=selected_tab.id if selected_tab else 0,
        entry_source=entry_source,
        view_id=selected_saved_view_id if active_saved_view else 0,
        view_filter=grid_view_filter,
        sort_col=grid_view_sort_column_id,
        sort_dir=grid_view_sort_direction,
    )
    if search_query:
        dismiss_params["q"] = search_query
    if focus_row_id:
        dismiss_params["focus_row_id"] = focus_row_id
    if focus_col_id:
        dismiss_params["focus_col_id"] = focus_col_id
    onboarding_dismiss_url = reverse("sheetbook:detail", kwargs={"pk": sheetbook.pk})
    if dismiss_params:
        onboarding_dismiss_url = f"{onboarding_dismiss_url}?{urlencode(dismiss_params)}"
    tab_create_form = SheetTabCreateForm()
    return render(
        request,
        "sheetbook/detail.html",
        {
            "sheetbook": sheetbook,
            "tabs": tabs,
            "selected_tab": selected_tab,
            "entry_source": entry_source,
            "schedule_source_tab": schedule_source_tab,
            "action_invocations": action_invocations,
            "action_history_has_more": action_history_has_more,
            "action_history_next_cursor": action_history_next_cursor,
            "grid_limit": grid_limit,
            "tab_create_form": tab_create_form,
            "search_query": search_query,
            "search_results": search_results,
            "focus_row_id": focus_row_id,
            "focus_col_id": focus_col_id,
            "saved_views": saved_views,
            "active_saved_view": active_saved_view,
            "active_saved_view_id": selected_saved_view_id if active_saved_view else 0,
            "grid_view_filter": grid_view_filter,
            "grid_view_sort_column_id": grid_view_sort_column_id,
            "grid_view_sort_direction": grid_view_sort_direction,
            "show_sample_onboarding": show_sample_onboarding,
            "onboarding_dismiss_url": onboarding_dismiss_url,
            "sheetbook_mobile_read_only": sheetbook_mobile_read_only,
            "sheetbook_mobile_read_only_message": sheetbook_mobile_read_only_message,
            "sheetbook_archived_read_only": sheetbook_archived_read_only,
            "index_back_url": index_back_url,
        },
    )


@login_required
def consent_seed_review(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    if tab.tab_type != SheetTab.TYPE_GRID:
        messages.error(request, "이 화면은 표(그리드) 탭에서만 열 수 있어요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)
    if request.method == "POST":
        blocked = _maybe_block_mobile_read_only_edit(
            request,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            blocked_action="consent_seed_review_submit",
        )
        if blocked:
            return blocked

    seed_token = (request.POST.get("sb_seed") or request.GET.get("sb_seed") or "").strip()
    seed = _peek_sheetbook_action_seed(
        request,
        seed_token,
        expected_action=ActionInvocation.ACTION_CONSENT,
    )
    if not isinstance(seed, dict):
        _log_sheetbook_metric(
            "consent_review_missing_seed",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
        )
        messages.error(request, "동의서 준비 정보를 찾지 못했어요. 교무수첩에서 다시 실행해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    seed_data = seed.get("data", {}) if isinstance(seed.get("data"), dict) else {}
    base_title = str(seed_data.get("title") or "").strip()[:200]
    base_message = str(seed_data.get("message") or "").strip()[:4000]
    base_document_title = str(seed_data.get("document_title") or "").strip()[:200]
    base_recipients_text = str(seed_data.get("recipients_text") or "").strip()
    recipients, recipients_meta = _parse_recipients_text_block_with_meta(base_recipients_text)

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()[:200] or base_title
        message = (request.POST.get("message") or "").strip()[:4000] or base_message
        document_title = (request.POST.get("document_title") or "").strip()[:200] or base_document_title
        recipients_text_raw = str(request.POST.get("recipients_text") or "").strip()
        parsed_recipients, parsed_recipients_meta = _parse_recipients_text_block_with_meta(recipients_text_raw)
        normalized_recipients_text = _build_recipients_text_block(parsed_recipients)
        cleanup_applied = bool(_parse_bool_or_none(request.POST.get("recipients_cleanup_applied")))
        cleanup_removed_count = _parse_positive_int(request.POST.get("recipients_cleanup_removed_count"), default=0)
        cleanup_undo_used = bool(_parse_bool_or_none(request.POST.get("recipients_cleanup_undo_used")))
        issue_copy_used = bool(_parse_bool_or_none(request.POST.get("recipients_issue_copy_used")))
        issue_jump_count = _parse_positive_int(request.POST.get("recipients_issue_jump_count"), default=0)
        issue_jump_count = min(issue_jump_count, 999)
        expected_removed_count = max(
            0,
            parsed_recipients_meta["duplicate_count"] + parsed_recipients_meta["skipped_line_count"],
        )
        if not cleanup_applied:
            cleanup_removed_count = 0
        cleanup_removed_count = min(cleanup_removed_count, expected_removed_count)

        updated_data = dict(seed_data)
        updated_data.update(
            {
                "title": title,
                "message": message,
                "document_title": document_title,
                "recipients_text": normalized_recipients_text,
            }
        )
        updated = _update_sheetbook_action_seed_data(
            request,
            seed_token,
            expected_action=ActionInvocation.ACTION_CONSENT,
            data=updated_data,
        )
        if not updated:
            _log_sheetbook_metric(
                "consent_review_update_failed",
                user_id=request.user.id,
                sheetbook_id=tab.sheetbook_id,
                tab_id=tab.id,
            )
            messages.error(request, "동의서 확인 내용을 저장하지 못했어요. 다시 시도해 주세요.")
            return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

        _log_sheetbook_metric(
            "consent_review_submitted",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            recipients_count=len(parsed_recipients),
            recipients_input_line_count=parsed_recipients_meta["input_line_count"],
            recipients_duplicate_count=parsed_recipients_meta["duplicate_count"],
            recipients_skipped_count=parsed_recipients_meta["skipped_line_count"],
            recipients_cleanup_applied=cleanup_applied,
            recipients_cleanup_removed_count=cleanup_removed_count,
            recipients_cleanup_undo_used=cleanup_undo_used,
            recipients_issue_copy_used=issue_copy_used,
            recipients_issue_jump_count=issue_jump_count,
        )
        if parsed_recipients:
            excluded_count = (
                parsed_recipients_meta["duplicate_count"] + parsed_recipients_meta["skipped_line_count"]
            )
            if excluded_count:
                messages.info(
                    request,
                    (
                        f"입력 {parsed_recipients_meta['input_line_count']}줄 중 "
                        f"{len(parsed_recipients)}명을 반영했어요. "
                        f"중복/형식 오류 {excluded_count}줄은 제외했어요."
                    ),
                )
            else:
                messages.info(request, f"수신자 {len(parsed_recipients)}명을 확인했어요. 동의서 작성 화면으로 이동합니다.")
        else:
            messages.info(request, "수신자가 비어 있어요. 다음 단계에서 직접 추가해 주세요.")
        consent_url = f"{reverse('consent:create_step1')}?{urlencode({'sb_seed': seed_token, 'from': 'sheetbook'})}"
        return redirect(consent_url)

    _log_sheetbook_metric(
        "consent_review_opened",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        recipients_count=len(recipients),
        recipients_input_line_count=recipients_meta["input_line_count"],
        recipients_duplicate_count=recipients_meta["duplicate_count"],
        recipients_skipped_count=recipients_meta["skipped_line_count"],
    )
    preview_recipients = recipients[:CONSENT_REVIEW_PREVIEW_LIMIT]
    preview_lines = []
    for rec in preview_recipients:
        phone = str(rec.get("phone_number") or "").strip()
        base_line = f"{rec.get('student_name', '')} - {rec.get('parent_name', '')}"
        if phone:
            base_line = f"{base_line} ({phone})"
        preview_lines.append(base_line)
    preview_text = "\n".join(preview_lines)
    issue_line_numbers = list(recipients_meta.get("issue_line_numbers") or [])
    issue_line_preview = issue_line_numbers[:CONSENT_REVIEW_ISSUE_LINE_LIMIT]
    issue_line_overflow_count = max(0, len(issue_line_numbers) - len(issue_line_preview))
    issue_preview_items = []
    for item in recipients_meta.get("duplicate_issue_items", []):
        issue_preview_items.append(
            {
                "line_no": int(item.get("line_no") or 0),
                "text": str(item.get("text") or ""),
                "issue_type": "duplicate",
                "issue_type_label": "중복",
            }
        )
    for item in recipients_meta.get("skipped_issue_items", []):
        issue_preview_items.append(
            {
                "line_no": int(item.get("line_no") or 0),
                "text": str(item.get("text") or ""),
                "issue_type": "format",
                "issue_type_label": "형식 확인",
            }
        )
    issue_preview_items = sorted(
        [item for item in issue_preview_items if item["line_no"] > 0],
        key=lambda item: (item["line_no"], item["issue_type"]),
    )[:CONSENT_REVIEW_ISSUE_LINE_LIMIT]
    return render(
        request,
        "sheetbook/consent_review.html",
        {
            "sheetbook": tab.sheetbook,
            "tab": tab,
            "seed_token": seed_token,
            "range_label": str(seed_data.get("range_label") or "").strip(),
            "initial_title": base_title,
            "initial_message": base_message,
            "initial_document_title": base_document_title,
            "recipients_text": base_recipients_text,
            "recipients_count": len(recipients),
            "recipients_preview_text": preview_text,
            "recipients_preview_limit": CONSENT_REVIEW_PREVIEW_LIMIT,
            "recipients_overflow_count": max(0, len(recipients) - CONSENT_REVIEW_PREVIEW_LIMIT),
            "recipients_input_line_count": recipients_meta["input_line_count"],
            "recipients_duplicate_count": recipients_meta["duplicate_count"],
            "recipients_skipped_count": recipients_meta["skipped_line_count"],
            "recipients_duplicate_samples": recipients_meta["duplicate_samples"],
            "recipients_skipped_samples": recipients_meta["skipped_samples"],
            "recipients_issue_line_numbers": issue_line_preview,
            "recipients_issue_line_overflow_count": issue_line_overflow_count,
            "recipients_issue_preview_items": issue_preview_items,
            "back_url": f"{reverse('sheetbook:detail', kwargs={'pk': tab.sheetbook_id})}?{urlencode({'tab': tab.id})}",
        },
    )


@login_required
@require_POST
def create_tab(request, pk):
    _ensure_sheetbook_enabled(request.user)
    sheetbook = _get_owner_sheetbook_or_404(request.user, pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=sheetbook.id,
        blocked_action="create_tab",
    )
    if blocked:
        return blocked
    form = SheetTabCreateForm(request.POST)
    if form.is_valid():
        tab = form.save(commit=False)
        tab.sheetbook = sheetbook
        tab.sort_order = _next_tab_order(sheetbook)
        tab.save()
        if tab.tab_type == SheetTab.TYPE_GRID:
            _seed_default_columns(tab)
        if _is_htmx_request(request):
            return _render_tab_list_partial(request, sheetbook)
        messages.success(request, "탭이 추가되었습니다.")
    else:
        if _is_htmx_request(request):
            return _render_tab_list_partial(request, sheetbook, tab_error="탭 생성 정보를 다시 확인해 주세요.", status=400)
        messages.error(request, "탭 생성 정보를 다시 확인해 주세요.")
    return redirect("sheetbook:detail", pk=sheetbook.pk)


@login_required
@require_POST
def rename_tab(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="rename_tab",
    )
    if blocked:
        return blocked
    form = SheetTabRenameForm(request.POST)
    if form.is_valid():
        tab.name = form.cleaned_data["name"].strip()
        tab.save(update_fields=["name", "updated_at"])
        if _is_htmx_request(request):
            return _render_tab_list_partial(request, tab.sheetbook)
        messages.success(request, "탭 이름이 변경되었습니다.")
    else:
        if _is_htmx_request(request):
            return _render_tab_list_partial(request, tab.sheetbook, tab_error="탭 이름을 100자 이내로 입력해 주세요.", status=400)
        messages.error(request, "탭 이름을 100자 이내로 입력해 주세요.")
    return redirect("sheetbook:detail", pk=pk)


@login_required
@require_POST
def delete_tab(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="delete_tab",
    )
    if blocked:
        return blocked
    sheetbook = tab.sheetbook
    tab.delete()
    _normalize_tab_sort_orders(sheetbook)
    if _is_htmx_request(request):
        return _render_tab_list_partial(request, sheetbook)
    messages.success(request, "탭이 삭제되었습니다.")
    return redirect("sheetbook:detail", pk=pk)


@login_required
@require_POST
def move_tab_up(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="move_tab_up",
    )
    if blocked:
        return blocked
    previous_tab = (
        tab.sheetbook.tabs.filter(sort_order__lt=tab.sort_order)
        .order_by("-sort_order", "-id")
        .first()
    )
    if previous_tab:
        previous_order = previous_tab.sort_order
        previous_tab.sort_order = tab.sort_order
        tab.sort_order = previous_order
        previous_tab.save(update_fields=["sort_order", "updated_at"])
        tab.save(update_fields=["sort_order", "updated_at"])
    if _is_htmx_request(request):
        return _render_tab_list_partial(request, tab.sheetbook)
    if previous_tab:
        messages.success(request, "탭 순서를 변경했습니다.")
    return redirect("sheetbook:detail", pk=pk)


@login_required
@require_POST
def move_tab_down(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="move_tab_down",
    )
    if blocked:
        return blocked
    next_tab = (
        tab.sheetbook.tabs.filter(sort_order__gt=tab.sort_order)
        .order_by("sort_order", "id")
        .first()
    )
    if next_tab:
        next_order = next_tab.sort_order
        next_tab.sort_order = tab.sort_order
        tab.sort_order = next_order
        next_tab.save(update_fields=["sort_order", "updated_at"])
        tab.save(update_fields=["sort_order", "updated_at"])
    if _is_htmx_request(request):
        return _render_tab_list_partial(request, tab.sheetbook)
    if next_tab:
        messages.success(request, "탭 순서를 변경했습니다.")
    return redirect("sheetbook:detail", pk=pk)


@login_required
@require_POST
def create_grid_row(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=True,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="create_grid_row",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        return JsonResponse({"ok": False, "error": "그리드 탭에서만 행을 추가할 수 있습니다."}, status=400)

    row = SheetRow.objects.create(
        tab=tab,
        sort_order=_next_row_order(tab),
        created_by=request.user,
        updated_by=request.user,
    )
    if _is_htmx_request(request):
        return _render_grid_editor_partial(request, tab.sheetbook, tab)
    return JsonResponse({"ok": True, "row_id": row.id})


@login_required
@require_POST
def create_grid_column(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=True,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="create_grid_column",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        return JsonResponse({"ok": False, "error": "그리드 탭에서만 열을 추가할 수 있습니다."}, status=400)

    payload = _parse_request_payload(request)
    label = (payload.get("label") or "").strip() or "새 열"
    column_type = payload.get("column_type") or SheetColumn.TYPE_TEXT
    valid_types = {choice[0] for choice in SheetColumn.TYPE_CHOICES}
    if column_type not in valid_types:
        column_type = SheetColumn.TYPE_TEXT

    column = SheetColumn.objects.create(
        tab=tab,
        key=_build_unique_column_key(tab, label),
        label=label,
        column_type=column_type,
        sort_order=_next_column_order(tab),
    )
    if _is_htmx_request(request):
        return _render_grid_editor_partial(request, tab.sheetbook, tab)
    return JsonResponse({"ok": True, "column_id": column.id})


def _saved_view_redirect(request, tab, *, view_id=0, view_filter="", sort_col=0, sort_dir=SavedView.SORT_ASC):
    source = ""
    if request.method == "POST":
        source = request.POST.get("source")
    if not source:
        source = request.GET.get("source")
    return _redirect_sheetbook_tab_detail(
        tab.sheetbook_id,
        tab.id,
        entry_source=source,
        view_id=view_id,
        view_filter=view_filter,
        sort_col=sort_col,
        sort_dir=sort_dir,
    )


@login_required
@require_POST
def create_saved_view(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    wants_json = _request_prefers_json(request)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=wants_json,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="create_saved_view",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        return JsonResponse({"ok": False, "error": "그리드 탭에서만 보기를 저장할 수 있어요."}, status=400)

    payload = _parse_request_payload(request)
    name = str(payload.get("name") or "").strip()[:80]
    if not name:
        if wants_json:
            return JsonResponse({"ok": False, "error": "보기 이름을 입력해 주세요."}, status=400)
        messages.error(request, "보기 이름을 입력해 주세요.")
        return _saved_view_redirect(
            request,
            tab,
            view_filter=payload.get("view_filter"),
            sort_col=payload.get("sort_col"),
            sort_dir=payload.get("sort_dir"),
        )

    filter_text = _normalize_saved_view_filter_text(payload.get("view_filter"))
    sort_column_id = _coerce_saved_view_sort_column_id(tab, payload.get("sort_col"), default=0)
    sort_direction = _normalize_saved_view_sort_direction(payload.get("sort_dir"))
    is_favorite = bool(_parse_bool_or_none(payload.get("is_favorite")))
    is_default = bool(_parse_bool_or_none(payload.get("is_default")))
    sort_column = tab.columns.filter(id=sort_column_id).only("id").first() if sort_column_id else None

    with transaction.atomic():
        if is_default:
            SavedView.objects.filter(tab=tab, is_default=True).update(is_default=False)
        saved_view = SavedView.objects.create(
            tab=tab,
            name=name,
            filter_text=filter_text,
            sort_column=sort_column,
            sort_direction=sort_direction,
            is_favorite=is_favorite,
            is_default=is_default,
            created_by=request.user,
        )

    _log_sheetbook_metric(
        "sheetbook_saved_view_created",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        saved_view_id=saved_view.id,
        is_default=is_default,
        is_favorite=is_favorite,
        has_filter=bool(filter_text),
        has_sort=bool(sort_column_id),
    )

    if wants_json:
        saved_views = _list_saved_views_for_tab(tab)
        return JsonResponse(
            {
                "ok": True,
                "saved_view": _serialize_saved_view(saved_view),
                "saved_views": [_serialize_saved_view(item) for item in saved_views],
            }
        )

    messages.success(request, f"보기 '{saved_view.name}'를 저장했어요.")
    return _saved_view_redirect(request, tab, view_id=saved_view.id)


@login_required
@require_POST
def delete_saved_view(request, pk, tab_pk, view_pk):
    _ensure_sheetbook_enabled(request.user)
    saved_view = _get_owner_saved_view_or_404(request.user, pk, tab_pk, view_pk)
    wants_json = _request_prefers_json(request)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=wants_json,
        sheetbook_id=saved_view.tab.sheetbook_id,
        tab_id=saved_view.tab_id,
        blocked_action="delete_saved_view",
    )
    if blocked:
        return blocked
    tab = saved_view.tab
    saved_view_name = saved_view.name
    saved_view.delete()

    _log_sheetbook_metric(
        "sheetbook_saved_view_deleted",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        saved_view_name=saved_view_name,
    )

    if wants_json:
        saved_views = _list_saved_views_for_tab(tab)
        return JsonResponse(
            {
                "ok": True,
                "saved_views": [_serialize_saved_view(item) for item in saved_views],
            }
        )

    messages.success(request, f"보기 '{saved_view_name}'를 삭제했어요.")
    return _saved_view_redirect(
        request,
        tab,
        view_filter=request.POST.get("view_filter"),
        sort_col=request.POST.get("sort_col"),
        sort_dir=request.POST.get("sort_dir"),
    )


@login_required
@require_POST
def toggle_saved_view_favorite(request, pk, tab_pk, view_pk):
    _ensure_sheetbook_enabled(request.user)
    saved_view = _get_owner_saved_view_or_404(request.user, pk, tab_pk, view_pk)
    wants_json = _request_prefers_json(request)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=wants_json,
        sheetbook_id=saved_view.tab.sheetbook_id,
        tab_id=saved_view.tab_id,
        blocked_action="toggle_saved_view_favorite",
    )
    if blocked:
        return blocked
    saved_view.is_favorite = not saved_view.is_favorite
    saved_view.save(update_fields=["is_favorite", "updated_at"])

    _log_sheetbook_metric(
        "sheetbook_saved_view_favorite_toggled",
        user_id=request.user.id,
        sheetbook_id=saved_view.tab.sheetbook_id,
        tab_id=saved_view.tab_id,
        saved_view_id=saved_view.id,
        is_favorite=saved_view.is_favorite,
    )

    if wants_json:
        return JsonResponse({"ok": True, "saved_view": _serialize_saved_view(saved_view)})

    messages.success(
        request,
        f"보기 '{saved_view.name}'를 {'즐겨찾기에 추가' if saved_view.is_favorite else '즐겨찾기에서 해제'}했어요.",
    )
    return _saved_view_redirect(request, saved_view.tab, view_id=saved_view.id)


@login_required
@require_POST
def set_saved_view_default(request, pk, tab_pk, view_pk):
    _ensure_sheetbook_enabled(request.user)
    saved_view = _get_owner_saved_view_or_404(request.user, pk, tab_pk, view_pk)
    wants_json = _request_prefers_json(request)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=wants_json,
        sheetbook_id=saved_view.tab.sheetbook_id,
        tab_id=saved_view.tab_id,
        blocked_action="set_saved_view_default",
    )
    if blocked:
        return blocked
    with transaction.atomic():
        make_default = not saved_view.is_default
        SavedView.objects.filter(tab=saved_view.tab, is_default=True).exclude(pk=saved_view.pk).update(is_default=False)
        saved_view.is_default = make_default
        saved_view.save(update_fields=["is_default", "updated_at"])

    _log_sheetbook_metric(
        "sheetbook_saved_view_default_toggled",
        user_id=request.user.id,
        sheetbook_id=saved_view.tab.sheetbook_id,
        tab_id=saved_view.tab_id,
        saved_view_id=saved_view.id,
        is_default=saved_view.is_default,
    )

    if wants_json:
        saved_views = _list_saved_views_for_tab(saved_view.tab)
        return JsonResponse(
            {
                "ok": True,
                "saved_view": _serialize_saved_view(saved_view),
                "saved_views": [_serialize_saved_view(item) for item in saved_views],
            }
        )

    if saved_view.is_default:
        messages.success(request, f"보기 '{saved_view.name}'를 기본 보기로 지정했어요.")
    else:
        messages.success(request, f"보기 '{saved_view.name}' 기본 지정을 해제했어요.")
    return _saved_view_redirect(request, saved_view.tab, view_id=saved_view.id if saved_view.is_default else 0)


def _build_grid_rows_from_values(rows, columns, cell_map):
    serialized_rows = []
    for row in rows:
        row_id = row["id"]
        values = {}
        for column in columns:
            cell_data = cell_map.get((row_id, column.id))
            values[str(column.id)] = _serialize_cell_value_from_data(cell_data, column.column_type)
        serialized_rows.append(
            {
                "id": row_id,
                "sort_order": row["sort_order"],
                "values": values,
            }
        )
    return serialized_rows


def _row_matches_grid_filter(row, columns, normalized_filter):
    if not normalized_filter:
        return True
    values = row.get("values", {})
    for column in columns:
        raw_value = values.get(str(column.id))
        if raw_value in (None, ""):
            continue
        if isinstance(raw_value, list):
            token = " ".join(str(item).strip() for item in raw_value if str(item).strip())
        else:
            token = str(raw_value).strip()
        if token and normalized_filter in token.lower():
            return True
    return False


def _normalize_grid_sort_value(value, column_type):
    if column_type == SheetColumn.TYPE_NUMBER:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0
    if column_type == SheetColumn.TYPE_CHECKBOX:
        return 1 if value else 0
    if column_type == SheetColumn.TYPE_MULTI_SELECT and isinstance(value, list):
        return ", ".join(str(item).strip() for item in value if str(item).strip()).lower()
    return str(value).strip().lower()


def _sort_grid_rows(rows, columns, sort_column_id, sort_direction):
    if not rows or not sort_column_id:
        return rows

    sort_column = next((column for column in columns if column.id == sort_column_id), None)
    if not sort_column:
        return rows

    non_blank = []
    blank = []
    for idx, row in enumerate(rows):
        value = row.get("values", {}).get(str(sort_column_id))
        is_blank = value in (None, "") or (isinstance(value, list) and not value)
        if is_blank:
            blank.append((idx, row))
            continue
        non_blank.append((_normalize_grid_sort_value(value, sort_column.column_type), idx, row))

    reverse = sort_direction == SavedView.SORT_DESC
    non_blank.sort(key=lambda item: (item[0], item[1]), reverse=reverse)
    return [item[2] for item in non_blank] + [item[1] for item in blank]


def _build_grid_data_payload(
    tab,
    offset=0,
    limit=50,
    view_filter_text="",
    view_sort_column_id=0,
    view_sort_direction=SavedView.SORT_ASC,
):
    offset = max(int(offset), 0)
    limit = min(max(int(limit), 1), 1000)

    columns = list(tab.columns.order_by("sort_order", "id").only("id", "key", "label", "column_type", "sort_order"))
    rows_qs = tab.rows.order_by("sort_order", "id")
    total_rows = rows_qs.count()

    normalized_filter = _normalize_saved_view_filter_text(view_filter_text).lower()
    normalized_sort_column_id = _coerce_saved_view_sort_column_id(tab, view_sort_column_id, default=0)
    normalized_sort_direction = _normalize_saved_view_sort_direction(view_sort_direction)
    use_view_processing = bool(normalized_filter or normalized_sort_column_id)

    if use_view_processing:
        scoped_rows = list(rows_qs[:1000].values("id", "sort_order"))
    else:
        scoped_rows = list(rows_qs[offset : offset + limit].values("id", "sort_order"))

    row_ids = [row["id"] for row in scoped_rows]
    column_ids = [column.id for column in columns]
    cell_map = {}
    if row_ids and column_ids:
        cells = SheetCell.objects.filter(row_id__in=row_ids, column_id__in=column_ids).values(
            "row_id",
            "column_id",
            "value_text",
            "value_number",
            "value_bool",
            "value_date",
            "value_json",
        )
        cell_map = {(cell["row_id"], cell["column_id"]): cell for cell in cells}

    serialized_rows = _build_grid_rows_from_values(scoped_rows, columns, cell_map)
    if use_view_processing:
        if normalized_filter:
            serialized_rows = [
                row for row in serialized_rows if _row_matches_grid_filter(row, columns, normalized_filter)
            ]
        serialized_rows = _sort_grid_rows(
            serialized_rows,
            columns,
            normalized_sort_column_id,
            normalized_sort_direction,
        )
        total_rows = len(serialized_rows)
        serialized_rows = serialized_rows[offset : offset + limit]

    return {
        "tab": {
            "id": tab.id,
            "name": tab.name,
            "tab_type": tab.tab_type,
        },
        "columns": [
            {
                "id": column.id,
                "key": column.key,
                "label": column.label,
                "type": column.column_type,
                "sort_order": column.sort_order,
            }
            for column in columns
        ],
        "rows": serialized_rows,
        "offset": offset,
        "limit": limit,
        "total_rows": total_rows,
        "count": len(serialized_rows),
        "view_filter": normalized_filter,
        "view_sort_column_id": normalized_sort_column_id,
        "view_sort_direction": normalized_sort_direction,
    }


@login_required
def grid_data(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)

    try:
        offset = request.GET.get("offset", 0)
        limit = request.GET.get("limit", 50)
        payload = _build_grid_data_payload(
            tab,
            offset=offset,
            limit=limit,
            view_filter_text=request.GET.get("view_filter"),
            view_sort_column_id=request.GET.get("sort_col"),
            view_sort_direction=request.GET.get("sort_dir"),
        )
    except (TypeError, ValueError):
        payload = _build_grid_data_payload(tab, offset=0, limit=50)
    return JsonResponse(payload)


@login_required
def search_suggest(request):
    _ensure_sheetbook_enabled(request.user)
    query = str(request.GET.get("q") or "").strip()[:80]
    limit = _parse_positive_int(request.GET.get("limit"), default=6)
    limit = max(1, min(20, limit))
    if not query:
        return JsonResponse({"ok": True, "query": "", "tabs": [], "cells": [], "actions": []})

    payload = _build_global_sheetbook_search_results(
        request.user,
        query,
        limit_per_group=limit,
    )
    _log_sheetbook_metric(
        "sheetbook_global_search_requested",
        user_id=request.user.id,
        query=query[:40],
        tab_result_count=len(payload["tabs"]),
        cell_result_count=len(payload["cells"]),
        action_result_count=len(payload["actions"]),
    )
    return JsonResponse(
        {
            "ok": True,
            "query": query,
            "tabs": payload["tabs"],
            "cells": payload["cells"],
            "actions": payload["actions"],
        }
    )


@login_required
@require_POST
def update_cell(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=True,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="update_cell",
    )
    if blocked:
        return blocked
    payload = _parse_request_payload(request)

    row_id = payload.get("row_id")
    column_id = payload.get("column_id")
    raw_value = payload.get("value")
    client_original = payload.get("client_original")

    if not row_id or not column_id:
        return JsonResponse({"ok": False, "error": "row_id와 column_id가 필요합니다."}, status=400)

    row = get_object_or_404(SheetRow, id=row_id, tab=tab)
    column = get_object_or_404(SheetColumn, id=column_id, tab=tab)

    cell, _ = SheetCell.objects.get_or_create(row=row, column=column)
    server_original_text = _serialize_cell_value_for_compare(cell, column)
    if client_original is not None:
        client_original_text = str(client_original)
        if client_original_text != server_original_text:
            _log_sheetbook_metric(
                "grid_cell_conflict_detected",
                user_id=request.user.id,
                sheetbook_id=tab.sheetbook_id,
                tab_id=tab.id,
                row_id=row.id,
                column_id=column.id,
            )
            return JsonResponse(
                {
                    "ok": False,
                    "error": "다른 화면에서 먼저 수정돼 충돌이 생겼어요. 현재 값을 확인한 뒤 다시 저장해 주세요.",
                    "conflict": True,
                    "current_value": _serialize_cell_value(cell, column),
                    "server_display": server_original_text,
                },
                status=409,
            )

    ok, error_message = _apply_cell_value(cell, column, raw_value)
    if not ok:
        return JsonResponse({"ok": False, "error": error_message}, status=400)

    cell.save()
    row.updated_by = request.user
    row.save(update_fields=["updated_by", "updated_at"])

    return JsonResponse(
        {
            "ok": True,
            "row_id": row.id,
            "column_id": column.id,
            "value": _serialize_cell_value(cell, column),
        }
    )


def _ensure_rows_for_paste(tab, required_rows, actor, now_ts, batch_size):
    existing_count = tab.rows.count()
    if existing_count >= required_rows:
        return 0

    missing_count = required_rows - existing_count
    last_row = tab.rows.order_by("-sort_order", "-id").only("sort_order").first()
    base_sort_order = last_row.sort_order if last_row else 0
    new_rows = [
        SheetRow(
            tab=tab,
            sort_order=base_sort_order + index,
            created_by=actor,
            updated_by=actor,
            created_at=now_ts,
            updated_at=now_ts,
        )
        for index in range(1, missing_count + 1)
    ]
    SheetRow.objects.bulk_create(new_rows, batch_size=batch_size)
    return missing_count


def _paste_matrix_into_grid_tab(tab, matrix, start_row_index, start_col_index, actor, batch_size=None):
    if batch_size is None:
        batch_size = _get_sheetbook_grid_bulk_batch_size()

    columns = list(tab.columns.order_by("sort_order", "id"))
    if not columns:
        raise ValueError("NO_COLUMNS")

    required_rows = start_row_index + len(matrix)
    now_ts = timezone.now()

    with transaction.atomic():
        rows_added = _ensure_rows_for_paste(
            tab=tab,
            required_rows=required_rows,
            actor=actor,
            now_ts=now_ts,
            batch_size=batch_size,
        )
        target_rows = list(tab.rows.order_by("sort_order", "id")[start_row_index:required_rows])
        row_ids = [row.id for row in target_rows]
        candidate_column_ids = [column.id for column in columns[start_col_index:]]
        existing_cells = list(SheetCell.objects.filter(row_id__in=row_ids, column_id__in=candidate_column_ids))
        cell_map = {(cell.row_id, cell.column_id): cell for cell in existing_cells}

        cells_to_create = []
        cells_to_update = []
        touched_rows = {}
        updated = 0
        skipped = 0
        invalid_rows = set()
        row_errors = {}

        for r_offset, line in enumerate(matrix):
            row = target_rows[r_offset]
            touched_rows[row.id] = row
            for c_offset, raw_value in enumerate(line):
                target_col_index = start_col_index + c_offset
                if target_col_index >= len(columns):
                    skipped += 1
                    invalid_rows.add(r_offset + 1)
                    row_errors.setdefault(r_offset + 1, set()).add("열 개수를 넘는 값은 제외했어요.")
                    continue

                column = columns[target_col_index]
                cell_key = (row.id, column.id)
                cell = cell_map.get(cell_key)
                is_new_cell = cell is None
                if cell is None:
                    cell = SheetCell(row=row, column=column)
                ok, err = _apply_cell_value(cell, column, raw_value)
                if not ok:
                    skipped += 1
                    invalid_rows.add(r_offset + 1)
                    row_errors.setdefault(r_offset + 1, set()).add(err or "형식을 확인해 주세요.")
                    continue
                if is_new_cell:
                    cell.created_at = now_ts
                    cell.updated_at = now_ts
                    cells_to_create.append(cell)
                    cell_map[cell_key] = cell
                else:
                    cell.updated_at = now_ts
                    cells_to_update.append(cell)
                updated += 1

        if cells_to_create:
            SheetCell.objects.bulk_create(cells_to_create, batch_size=batch_size)
        if cells_to_update:
            SheetCell.objects.bulk_update(
                cells_to_update,
                fields=[
                    "value_text",
                    "value_number",
                    "value_bool",
                    "value_date",
                    "value_json",
                    "updated_at",
                ],
                batch_size=batch_size,
            )
        if touched_rows:
            touched_rows_list = list(touched_rows.values())
            for row in touched_rows_list:
                row.updated_by = actor
                row.updated_at = now_ts
            SheetRow.objects.bulk_update(
                touched_rows_list,
                fields=["updated_by", "updated_at"],
                batch_size=batch_size,
            )

    return {
        "updated": updated,
        "skipped": skipped,
        "invalid_rows": sorted(list(invalid_rows)),
        "row_errors": [
            {"row": row_no, "reasons": sorted(list(reasons))}
            for row_no, reasons in sorted(row_errors.items())
        ],
        "rows_added": rows_added,
        "batch_size": batch_size,
    }


@login_required
@require_POST
def paste_cells(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=True,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="paste_cells",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        return JsonResponse({"ok": False, "error": "그리드 탭에서만 붙여넣기가 가능합니다."}, status=400)

    payload = _parse_request_payload(request)
    raw_text = payload.get("raw_text")
    start_row_index = payload.get("start_row_index", 0)
    start_col_index = payload.get("start_col_index", 0)

    try:
        start_row_index = max(int(start_row_index), 0)
        start_col_index = max(int(start_col_index), 0)
    except (ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "시작 위치가 올바르지 않습니다."}, status=400)

    matrix = _parse_clipboard_matrix(raw_text)
    if not matrix:
        return JsonResponse({"ok": False, "error": "붙여넣을 데이터가 없습니다."}, status=400)

    if not tab.columns.exists():
        return JsonResponse({"ok": False, "error": "먼저 열을 추가해 주세요."}, status=400)
    result = _paste_matrix_into_grid_tab(
        tab=tab,
        matrix=matrix,
        start_row_index=start_row_index,
        start_col_index=start_col_index,
        actor=request.user,
    )

    return JsonResponse(
        {
            "ok": True,
            "updated": result["updated"],
            "skipped": result["skipped"],
            "invalid_rows": result["invalid_rows"],
            "row_errors": result["row_errors"],
            "rows_added": result["rows_added"],
        }
    )


@login_required
@require_POST
def import_grid_tab_file(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="import_grid_tab_file",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        messages.error(request, "이 기능은 표(그리드) 탭에서만 쓸 수 있어요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        messages.error(request, "가져올 파일을 먼저 선택해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    has_header = str(request.POST.get("has_header") or "").lower() in {"1", "true", "on", "yes", "y"}
    replace_existing = str(request.POST.get("replace_existing") or "").lower() in {"1", "true", "on", "yes", "y"}
    auto_add_columns = str(request.POST.get("auto_add_columns") or "").lower() in {"1", "true", "on", "yes", "y"}

    try:
        matrix, source_format = _parse_uploaded_grid_file(uploaded_file)
    except ValueError as exc:
        _log_sheetbook_metric(
            "grid_import_failed",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            reason=str(exc),
        )
        messages.error(request, str(exc))
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    if not matrix:
        messages.error(request, "파일 안에 가져올 데이터가 없어요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    created_columns = 0
    if has_header:
        headers = _normalize_import_header_labels(matrix[0])
        if auto_add_columns:
            created_columns = _ensure_grid_columns_for_import(tab, headers)
        matrix = matrix[1:]

    if not matrix:
        messages.error(request, "제목 줄은 읽었지만 실제 데이터 줄이 없어요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    if not tab.columns.exists():
        messages.error(request, "먼저 열을 추가한 뒤 다시 시도해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    if replace_existing:
        tab.rows.all().delete()

    result = _paste_matrix_into_grid_tab(
        tab=tab,
        matrix=matrix,
        start_row_index=0,
        start_col_index=0,
        actor=request.user,
    )
    _log_sheetbook_metric(
        "grid_import_completed",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        source_format=source_format,
        has_header=has_header,
        replace_existing=replace_existing,
        auto_add_columns=auto_add_columns,
        created_columns=created_columns,
        imported_rows=len(matrix),
        updated_cells=result["updated"],
        skipped_cells=result["skipped"],
        invalid_row_count=len(result["invalid_rows"]),
    )

    summary_text = (
        f"가져오기 완료: {len(matrix)}줄 반영, {result['updated']}칸 저장"
        f"{', 새 열 ' + str(created_columns) + '개 추가' if created_columns else ''}"
    )
    messages.success(request, summary_text)

    if result["skipped"]:
        invalid_row_text = ", ".join(str(item) for item in result["invalid_rows"][:10])
        if len(result["invalid_rows"]) > 10:
            invalid_row_text = f"{invalid_row_text} 외 {len(result['invalid_rows']) - 10}줄"
        messages.warning(
            request,
            f"형식 또는 열 범위를 벗어난 값 {result['skipped']}칸은 제외했어요. 확인 줄: {invalid_row_text}",
        )
    return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)


@login_required
def export_grid_tab_csv(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    if tab.tab_type != SheetTab.TYPE_GRID:
        raise Http404("Grid tab only")

    matrix = _build_grid_export_matrix(tab)
    if not matrix:
        messages.error(request, "내보낼 열이 없어요. 먼저 열을 추가해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    output = io.StringIO()
    writer = csv.writer(output)
    for row in matrix:
        writer.writerow(row)
    csv_text = "\ufeff" + output.getvalue()

    _log_sheetbook_metric(
        "grid_export_downloaded",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        export_format="csv",
        row_count=max(0, len(matrix) - 1),
        column_count=len(matrix[0]) if matrix else 0,
    )

    response = HttpResponse(csv_text, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{_build_grid_export_filename(tab, "csv")}"'
    return response


@login_required
def export_grid_tab_xlsx(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    if tab.tab_type != SheetTab.TYPE_GRID:
        raise Http404("Grid tab only")
    if Workbook is None:
        messages.error(request, "엑셀 내보내기 기능이 준비되지 않았어요. 관리자에게 문의해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    matrix = _build_grid_export_matrix(tab)
    if not matrix:
        messages.error(request, "내보낼 열이 없어요. 먼저 열을 추가해 주세요.")
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (tab.name or "Sheet1")[:31]
    for row in matrix:
        worksheet.append(row)

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    _log_sheetbook_metric(
        "grid_export_downloaded",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        export_format="xlsx",
        row_count=max(0, len(matrix) - 1),
        column_count=len(matrix[0]) if matrix else 0,
    )

    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_build_grid_export_filename(tab, "xlsx")}"'
    return response


@login_required
@require_POST
def execute_grid_action(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    wants_json = _request_prefers_json(request)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=wants_json,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        blocked_action="execute_grid_action",
    )
    if blocked:
        return blocked
    if tab.tab_type != SheetTab.TYPE_GRID:
        error_message = "이 기능은 표(그리드) 탭에서만 쓸 수 있어요."
        if wants_json:
            return JsonResponse({"ok": False, "error": error_message}, status=400)
        messages.error(request, error_message)
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    payload = _parse_request_payload(request)
    entry_source = _resolve_sheetbook_entry_source(
        request,
        tab.sheetbook_id,
        payload.get("entry_source"),
    )
    action_type = str(payload.get("action") or "").strip()
    supported_actions = {
        ActionInvocation.ACTION_CALENDAR,
        ActionInvocation.ACTION_COLLECT,
        ActionInvocation.ACTION_SIGNATURE,
        ActionInvocation.ACTION_HANDOFF,
        ActionInvocation.ACTION_CONSENT,
        ActionInvocation.ACTION_NOTICE,
    }
    recommendation_primary = str(payload.get("recommendation_primary") or "").strip()
    if recommendation_primary not in supported_actions:
        recommendation_primary = ""
    raw_recommendation_signals = payload.get("recommendation_signals")
    if not isinstance(raw_recommendation_signals, dict):
        raw_recommendation_signals = {}
    recommendation_token_count = _parse_positive_int(raw_recommendation_signals.get("token_count"), default=0)
    recommendation_date_ratio = _parse_ratio_or_default(raw_recommendation_signals.get("date_ratio"), default=0.0)
    recommendation_phone_ratio = _parse_ratio_or_default(raw_recommendation_signals.get("phone_ratio"), default=0.0)
    recommendation_name_ratio = _parse_ratio_or_default(raw_recommendation_signals.get("name_ratio"), default=0.0)
    recommendation_meta = {}
    if recommendation_primary:
        recommendation_meta["primary_action"] = recommendation_primary
    if recommendation_token_count:
        recommendation_meta["token_count"] = recommendation_token_count
    if recommendation_date_ratio:
        recommendation_meta["date_ratio"] = recommendation_date_ratio
    if recommendation_phone_ratio:
        recommendation_meta["phone_ratio"] = recommendation_phone_ratio
    if recommendation_name_ratio:
        recommendation_meta["name_ratio"] = recommendation_name_ratio

    if action_type not in supported_actions:
        error_message = "알 수 없는 실행 버튼입니다. 다시 시도해 주세요."
        _log_sheetbook_metric(
            "action_execute_invalid_action",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            action=action_type,
            entry_source=entry_source,
        )
        if wants_json:
            return JsonResponse({"ok": False, "error": error_message}, status=400)
        messages.error(request, error_message)
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    row_count = tab.rows.count()
    col_count = tab.columns.count()
    bounds, bounds_error = _normalize_selection_bounds(payload, row_count, col_count)
    if bounds_error:
        _log_sheetbook_metric(
            "action_execute_invalid_bounds",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            action=action_type,
            reason=bounds_error,
            entry_source=entry_source,
        )
        if wants_json:
            return JsonResponse({"ok": False, "error": bounds_error}, status=400)
        messages.error(request, bounds_error)
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    rows, columns, cell_map = _collect_selection_snapshot(tab, bounds)
    if not rows or not columns:
        error_message = "선택한 칸에 데이터가 없어요."
        _log_sheetbook_metric(
            "action_execute_empty_selection",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            action=action_type,
            entry_source=entry_source,
        )
        if wants_json:
            return JsonResponse({"ok": False, "error": error_message}, status=400)
        messages.error(request, error_message)
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    _log_sheetbook_metric(
        "action_execute_requested",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        action=action_type,
        selected_cell_count=bounds["count"],
        entry_source=entry_source,
        recommended_action=recommendation_primary,
        recommendation_token_count=recommendation_token_count,
        recommendation_date_ratio=recommendation_date_ratio,
        recommendation_phone_ratio=recommendation_phone_ratio,
        recommendation_name_ratio=recommendation_name_ratio,
    )

    invocation_payload = {}
    if recommendation_meta:
        invocation_payload["recommendation"] = recommendation_meta
    invocation = ActionInvocation.objects.create(
        sheetbook=tab.sheetbook,
        tab=tab,
        actor=request.user,
        action_type=action_type,
        status=ActionInvocation.STATUS_FAILED,
        selection_start_row=bounds["min_row"],
        selection_start_col=bounds["min_col"],
        selection_end_row=bounds["max_row"],
        selection_end_col=bounds["max_col"],
        selected_cell_count=bounds["count"],
        payload=invocation_payload,
    )

    try:
        if action_type == ActionInvocation.ACTION_CALENDAR:
            outcome = _execute_calendar_registration(
                user=request.user,
                sheetbook=tab.sheetbook,
                tab=tab,
                rows=rows,
                columns=columns,
                cell_map=cell_map,
                bounds=bounds,
                invocation=invocation,
            )
        elif action_type == ActionInvocation.ACTION_COLLECT:
            outcome = _execute_collect_creation(
                user=request.user,
                sheetbook=tab.sheetbook,
                tab=tab,
                rows=rows,
                columns=columns,
                cell_map=cell_map,
                bounds=bounds,
            )
        elif action_type == ActionInvocation.ACTION_HANDOFF:
            outcome = _execute_handoff_creation(
                user=request.user,
                sheetbook=tab.sheetbook,
                tab=tab,
                rows=rows,
                columns=columns,
                cell_map=cell_map,
                bounds=bounds,
            )
        else:
            outcome = _execute_guide_action(
                request=request,
                action_type=action_type,
                sheetbook=tab.sheetbook,
                tab=tab,
                rows=rows,
                columns=columns,
                cell_map=cell_map,
                bounds=bounds,
            )
    except ValueError as exc:
        invocation.summary = str(exc)
        error_payload = {"error": str(exc)}
        if recommendation_meta:
            error_payload["recommendation"] = recommendation_meta
        invocation.payload = error_payload
        invocation.save(update_fields=["summary", "payload"])
        _log_sheetbook_metric(
            "action_execute_failed",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            action=action_type,
            error=str(exc),
            error_type="validation",
            entry_source=entry_source,
            recommended_action=recommendation_primary,
            recommendation_token_count=recommendation_token_count,
            recommendation_date_ratio=recommendation_date_ratio,
            recommendation_phone_ratio=recommendation_phone_ratio,
            recommendation_name_ratio=recommendation_name_ratio,
        )
        if wants_json:
            return JsonResponse(
                {
                    "ok": False,
                    "error": str(exc),
                    "invocation": _serialize_action_invocation(invocation),
                },
                status=400,
            )
        messages.error(request, str(exc))
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)
    except Exception:
        fallback_message = _action_failure_fallback_message(action_type)
        invocation.summary = fallback_message
        error_payload = {"error": "unexpected_error"}
        if recommendation_meta:
            error_payload["recommendation"] = recommendation_meta
        invocation.payload = error_payload
        invocation.save(update_fields=["summary", "payload"])
        _log_sheetbook_metric(
            "action_execute_failed",
            user_id=request.user.id,
            sheetbook_id=tab.sheetbook_id,
            tab_id=tab.id,
            action=action_type,
            error="unexpected_error",
            error_type="unexpected",
            entry_source=entry_source,
            recommended_action=recommendation_primary,
            recommendation_token_count=recommendation_token_count,
            recommendation_date_ratio=recommendation_date_ratio,
            recommendation_phone_ratio=recommendation_phone_ratio,
            recommendation_name_ratio=recommendation_name_ratio,
        )
        if wants_json:
            return JsonResponse(
                {
                    "ok": False,
                    "error": fallback_message,
                    "invocation": _serialize_action_invocation(invocation),
                },
                status=500,
            )
        messages.error(request, fallback_message)
        return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)

    invocation.status = ActionInvocation.STATUS_SUCCESS
    invocation.summary = (outcome.get("summary") or "").strip()
    invocation.result_label = (outcome.get("result_label") or "").strip()
    invocation.result_url = (outcome.get("result_url") or "").strip()
    invocation_payload = outcome.get("payload") or {}
    if recommendation_meta:
        invocation_payload["recommendation"] = recommendation_meta
    invocation.payload = invocation_payload
    invocation.save(
        update_fields=[
            "status",
            "summary",
            "result_label",
            "result_url",
            "payload",
        ]
    )

    success_message = invocation.summary or f"{invocation.get_action_type_display()} 실행 완료"
    _log_sheetbook_metric(
        "action_execute_succeeded",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        action=action_type,
        invocation_id=invocation.id,
        selected_cell_count=bounds["count"],
        entry_source=entry_source,
        recommended_action=recommendation_primary,
        recommendation_token_count=recommendation_token_count,
        recommendation_date_ratio=recommendation_date_ratio,
        recommendation_phone_ratio=recommendation_phone_ratio,
        recommendation_name_ratio=recommendation_name_ratio,
    )
    if wants_json:
        return JsonResponse(
            {
                "ok": True,
                "message": success_message,
                "invocation": _serialize_action_invocation(invocation),
            }
        )
    messages.success(request, success_message)
    return _redirect_sheetbook_tab_detail(tab.sheetbook_id, tab.id)


@login_required
def action_history(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    if tab.tab_type != SheetTab.TYPE_GRID:
        return JsonResponse({"ok": False, "error": "표(그리드) 탭에서만 기록을 볼 수 있어요."}, status=400)

    limit = _parse_positive_int(request.GET.get("limit"), default=8)
    limit = min(limit, 20)
    cursor_id = _parse_positive_int(request.GET.get("cursor_id"), default=0)

    queryset = ActionInvocation.objects.filter(
        sheetbook=tab.sheetbook,
        tab=tab,
    ).select_related("actor")
    if cursor_id:
        queryset = queryset.filter(id__lt=cursor_id)

    items = list(
        queryset.order_by("-created_at", "-id")[: limit + 1]
    )
    has_more = len(items) > limit
    items = items[:limit]
    next_cursor = items[-1].id if has_more and items else None
    _log_sheetbook_metric(
        "action_history_loaded",
        user_id=request.user.id,
        sheetbook_id=tab.sheetbook_id,
        tab_id=tab.id,
        requested_limit=limit,
        cursor_id=cursor_id,
        loaded_count=len(items),
        has_more=has_more,
    )

    return JsonResponse(
        {
            "ok": True,
            "items": [_serialize_action_invocation(item) for item in items],
            "has_more": has_more,
            "next_cursor": next_cursor,
        }
    )


@login_required
@require_POST
def sync_calendar_from_schedule(request, pk, tab_pk):
    _ensure_sheetbook_enabled(request.user)
    calendar_tab = _get_owner_tab_or_404(request.user, pk, tab_pk)
    blocked = _maybe_block_mobile_read_only_edit(
        request,
        wants_json=True,
        sheetbook_id=calendar_tab.sheetbook_id,
        tab_id=calendar_tab.id,
        blocked_action="sync_calendar_from_schedule",
    )
    if blocked:
        return blocked
    if calendar_tab.tab_type != SheetTab.TYPE_CALENDAR:
        return JsonResponse({"ok": False, "error": "달력 탭에서만 실행할 수 있습니다."}, status=400)

    sheetbook = calendar_tab.sheetbook
    schedule_tab = _resolve_schedule_source_tab(sheetbook)
    if not schedule_tab:
        return JsonResponse({"ok": False, "error": "연동할 일정 탭이 없습니다."}, status=400)

    rows = list(schedule_tab.rows.order_by("sort_order", "id"))
    _, cell_map, key_columns = _extract_schedule_cells(schedule_tab, rows)
    date_column = key_columns.get("date")
    if not date_column:
        return JsonResponse({"ok": False, "error": "일정 탭에 날짜 열이 없어 연동할 수 없습니다."}, status=400)

    title_column = key_columns.get("title")
    note_column = key_columns.get("note")
    start_time_column = key_columns.get("start_time")
    end_time_column = key_columns.get("end_time")
    prefix = f"{sheetbook.id}:{schedule_tab.id}:"
    tz = timezone.get_current_timezone()

    existing_events = list(
        CalendarEvent.objects.filter(
            author=request.user,
            integration_source=SHEETBOOK_CALENDAR_SYNC_SOURCE,
            integration_key__startswith=prefix,
        ).prefetch_related("blocks")
    )
    existing_map = {event.integration_key: event for event in existing_events}
    desired_keys = set()

    created = 0
    updated = 0
    skipped = 0
    timed_synced = 0
    all_day_synced = 0
    default_duration_minutes = _sheetbook_schedule_default_duration_minutes()

    for row in rows:
        date_cell = cell_map.get((row.id, date_column.id))
        if not date_cell or not date_cell.value_date:
            continue

        row_date = date_cell.value_date
        start_time_value = ""
        end_time_value = ""
        if start_time_column:
            start_time_value = _cell_value_as_text(cell_map.get((row.id, start_time_column.id)), start_time_column)
        if end_time_column:
            end_time_value = _cell_value_as_text(cell_map.get((row.id, end_time_column.id)), end_time_column)

        parsed_start_time = _parse_time_or_none(start_time_value)
        parsed_end_time = _parse_time_or_none(end_time_value)

        if parsed_start_time:
            start_dt = timezone.make_aware(datetime.combine(row_date, parsed_start_time), tz)
            if parsed_end_time:
                end_dt = timezone.make_aware(datetime.combine(row_date, parsed_end_time), tz)
                if end_dt <= start_dt:
                    end_dt = start_dt + timedelta(minutes=default_duration_minutes)
            else:
                end_dt = start_dt + timedelta(minutes=default_duration_minutes)
            is_all_day = False
            timed_synced += 1
        else:
            start_dt = timezone.make_aware(datetime.combine(row_date, time.min), tz)
            end_dt = start_dt + timedelta(days=1)
            is_all_day = True
            all_day_synced += 1

        title = "일정"
        if title_column:
            title_cell = cell_map.get((row.id, title_column.id))
            if title_cell and (title_cell.value_text or "").strip():
                title = title_cell.value_text.strip()

        note = ""
        if note_column:
            note_cell = cell_map.get((row.id, note_column.id))
            if note_cell and (note_cell.value_text or "").strip():
                note = note_cell.value_text.strip()

        integration_key = f"{sheetbook.id}:{schedule_tab.id}:{row.id}"
        desired_keys.add(integration_key)
        event = existing_map.get(integration_key)
        if event is None:
            event = CalendarEvent.objects.create(
                title=title,
                author=request.user,
                start_time=start_dt,
                end_time=end_dt,
                is_all_day=is_all_day,
                visibility=CalendarEvent.VISIBILITY_TEACHER,
                source=CalendarEvent.SOURCE_LOCAL,
                color="indigo",
                integration_source=SHEETBOOK_CALENDAR_SYNC_SOURCE,
                integration_key=integration_key,
                is_locked=False,
            )
            _persist_calendar_note(event, note)
            created += 1
            continue

        changed_fields = []
        if event.title != title:
            event.title = title
            changed_fields.append("title")
        if event.start_time != start_dt:
            event.start_time = start_dt
            changed_fields.append("start_time")
        if event.end_time != end_dt:
            event.end_time = end_dt
            changed_fields.append("end_time")
        if event.is_all_day != is_all_day:
            event.is_all_day = is_all_day
            changed_fields.append("is_all_day")
        if event.visibility != CalendarEvent.VISIBILITY_TEACHER:
            event.visibility = CalendarEvent.VISIBILITY_TEACHER
            changed_fields.append("visibility")
        if event.source != CalendarEvent.SOURCE_LOCAL:
            event.source = CalendarEvent.SOURCE_LOCAL
            changed_fields.append("source")
        if event.integration_source != SHEETBOOK_CALENDAR_SYNC_SOURCE:
            event.integration_source = SHEETBOOK_CALENDAR_SYNC_SOURCE
            changed_fields.append("integration_source")
        if event.integration_key != integration_key:
            event.integration_key = integration_key
            changed_fields.append("integration_key")
        if event.is_locked:
            event.is_locked = False
            changed_fields.append("is_locked")

        if changed_fields:
            changed_fields.append("updated_at")
            event.save(update_fields=changed_fields)
            updated += 1
        _persist_calendar_note(event, note)

    stale_ids = [event.id for event in existing_events if event.integration_key not in desired_keys]
    deleted = 0
    if stale_ids:
        deleted, _ = CalendarEvent.objects.filter(id__in=stale_ids).delete()
    if not desired_keys:
        skipped = len(rows)

    return JsonResponse(
        {
            "ok": True,
            "schedule_tab_id": schedule_tab.id,
            "schedule_tab_name": schedule_tab.name,
            "created": created,
            "updated": updated,
            "deleted": deleted,
            "skipped": skipped,
            "synced": created + updated,
            "timed_synced": timed_synced,
            "all_day_synced": all_day_synced,
        }
    )


@login_required
def metrics_dashboard(request):
    _ensure_sheetbook_enabled(request.user)
    if not request.user.is_superuser:
        raise Http404("Not found")

    allowed_days = {7, 14, 30}
    requested_days = _parse_positive_int(request.GET.get("days"), default=7)
    days = requested_days if requested_days in allowed_days else 7
    since = timezone.now() - timedelta(days=days)

    event_qs = SheetbookMetricEvent.objects.filter(created_at__gte=since)
    summary = {
        "event_total": event_qs.count(),
        "active_user_count": event_qs.exclude(user_id__isnull=True).values("user_id").distinct().count(),
        "sheetbook_created_count": event_qs.filter(event_name="sheetbook_created").count(),
        "sheetbook_archived_count": event_qs.filter(event_name="sheetbook_archived").count(),
        "sheetbook_unarchived_count": event_qs.filter(event_name="sheetbook_unarchived").count(),
        "sheetbook_archive_bulk_updated_count": event_qs.filter(event_name="sheetbook_archive_bulk_updated").count(),
        "sheetbook_archive_read_mode_opened_count": event_qs.filter(event_name="sheetbook_archive_read_mode_opened").count(),
        "workspace_home_opened_count": event_qs.filter(event_name="workspace_home_opened").count(),
        "creator_user_count": event_qs.filter(event_name="sheetbook_created").exclude(user_id__isnull=True).values("user_id").distinct().count(),
        "action_requested_count": event_qs.filter(event_name="action_execute_requested").count(),
        "action_success_count": event_qs.filter(event_name="action_execute_succeeded").count(),
        "action_failed_count": event_qs.filter(event_name="action_execute_failed").count(),
    }
    bulk_archive_rows = list(
        event_qs.filter(event_name="sheetbook_archive_bulk_updated").values("metadata")
    )
    bulk_archive_changed_count = 0
    bulk_unarchive_changed_count = 0
    bulk_ignored_count = 0
    bulk_unchanged_count = 0
    for row in bulk_archive_rows:
        metadata = row.get("metadata") or {}
        changed_count = _parse_positive_int(metadata.get("changed_count"), default=0)
        ignored_count = _parse_positive_int(metadata.get("ignored_count"), default=0)
        unchanged_count = _parse_positive_int(metadata.get("unchanged_count"), default=0)
        archive_action = str(metadata.get("archive_action") or "").strip().lower()
        bulk_ignored_count += ignored_count
        bulk_unchanged_count += unchanged_count
        if archive_action == "unarchive":
            bulk_unarchive_changed_count += changed_count
        else:
            bulk_archive_changed_count += changed_count
    summary["sheetbook_bulk_archive_changed_count"] = bulk_archive_changed_count
    summary["sheetbook_bulk_unarchive_changed_count"] = bulk_unarchive_changed_count
    summary["sheetbook_bulk_ignored_count"] = bulk_ignored_count
    summary["sheetbook_bulk_unchanged_count"] = bulk_unchanged_count

    success = summary["action_success_count"]
    failed = summary["action_failed_count"]
    summary["action_total_count"] = success + failed
    summary["action_success_rate"] = round((success / (success + failed)) * 100, 1) if (success + failed) else 0.0
    consent_review_rows = list(
        event_qs.filter(event_name="consent_review_submitted").values("metadata")
    )
    consent_review_submitted_count = len(consent_review_rows)
    consent_cleanup_applied_count = 0
    consent_cleanup_removed_total = 0
    consent_cleanup_undo_used_count = 0
    consent_issue_copy_used_count = 0
    consent_issue_jump_total = 0
    consent_issue_jump_used_count = 0
    for row in consent_review_rows:
        metadata = row.get("metadata") or {}
        cleanup_applied = bool(_parse_bool_or_none(metadata.get("recipients_cleanup_applied")))
        cleanup_undo_used = bool(_parse_bool_or_none(metadata.get("recipients_cleanup_undo_used")))
        issue_copy_used = bool(_parse_bool_or_none(metadata.get("recipients_issue_copy_used")))
        issue_jump_count = _parse_positive_int(
            metadata.get("recipients_issue_jump_count"),
            default=0,
        )
        if cleanup_undo_used:
            consent_cleanup_undo_used_count += 1
        if issue_copy_used:
            consent_issue_copy_used_count += 1
        if issue_jump_count > 0:
            consent_issue_jump_used_count += 1
        consent_issue_jump_total += issue_jump_count
        if not cleanup_applied:
            continue
        consent_cleanup_applied_count += 1
        consent_cleanup_removed_total += _parse_positive_int(
            metadata.get("recipients_cleanup_removed_count"),
            default=0,
        )
    summary["consent_review_submitted_count"] = consent_review_submitted_count
    summary["consent_cleanup_applied_count"] = consent_cleanup_applied_count
    summary["consent_cleanup_apply_rate"] = (
        round((consent_cleanup_applied_count / consent_review_submitted_count) * 100, 1)
        if consent_review_submitted_count
        else 0.0
    )
    summary["consent_cleanup_removed_avg"] = (
        round((consent_cleanup_removed_total / consent_cleanup_applied_count), 1)
        if consent_cleanup_applied_count
        else 0.0
    )
    summary["consent_cleanup_undo_used_count"] = consent_cleanup_undo_used_count
    summary["consent_cleanup_undo_use_rate"] = (
        round((consent_cleanup_undo_used_count / consent_review_submitted_count) * 100, 1)
        if consent_review_submitted_count
        else 0.0
    )
    summary["consent_issue_copy_used_count"] = consent_issue_copy_used_count
    summary["consent_issue_copy_use_rate"] = (
        round((consent_issue_copy_used_count / consent_review_submitted_count) * 100, 1)
        if consent_review_submitted_count
        else 0.0
    )
    summary["consent_issue_jump_total"] = consent_issue_jump_total
    summary["consent_issue_jump_used_count"] = consent_issue_jump_used_count
    summary["consent_issue_jump_use_rate"] = (
        round((consent_issue_jump_used_count / consent_review_submitted_count) * 100, 1)
        if consent_review_submitted_count
        else 0.0
    )
    summary["consent_issue_jump_avg"] = (
        round((consent_issue_jump_total / consent_review_submitted_count), 1)
        if consent_review_submitted_count
        else 0.0
    )
    consent_cleanup_target_rate = _parse_percentage_or_default(
        getattr(settings, "SHEETBOOK_CONSENT_CLEANUP_TARGET_RATE", 30.0),
        default=30.0,
    )
    consent_cleanup_min_sample = _parse_positive_int(
        getattr(settings, "SHEETBOOK_CONSENT_CLEANUP_MIN_SAMPLE", 5),
        default=5,
    )
    consent_cleanup_undo_alert_rate = _parse_percentage_or_default(
        getattr(settings, "SHEETBOOK_CONSENT_CLEANUP_UNDO_ALERT_RATE", 20.0),
        default=20.0,
    )
    consent_cleanup_undo_min_sample = _parse_positive_int(
        getattr(settings, "SHEETBOOK_CONSENT_CLEANUP_UNDO_MIN_SAMPLE", 5),
        default=5,
    )
    summary["consent_cleanup_target_rate"] = consent_cleanup_target_rate
    summary["consent_cleanup_min_sample"] = consent_cleanup_min_sample
    summary["consent_cleanup_undo_alert_rate"] = consent_cleanup_undo_alert_rate
    summary["consent_cleanup_undo_min_sample"] = consent_cleanup_undo_min_sample
    summary["consent_cleanup_gap"] = round(
        summary["consent_cleanup_apply_rate"] - summary["consent_cleanup_target_rate"],
        1,
    )
    summary["consent_cleanup_needs_attention"] = (
        consent_review_submitted_count >= summary["consent_cleanup_min_sample"]
        and summary["consent_cleanup_apply_rate"] < summary["consent_cleanup_target_rate"]
    )
    summary["consent_cleanup_undo_gap"] = round(
        summary["consent_cleanup_undo_use_rate"] - summary["consent_cleanup_undo_alert_rate"],
        1,
    )
    summary["consent_cleanup_undo_needs_attention"] = (
        consent_review_submitted_count >= summary["consent_cleanup_undo_min_sample"]
        and summary["consent_cleanup_undo_use_rate"] > summary["consent_cleanup_undo_alert_rate"]
    )
    revisit_by_user = (
        event_qs.filter(event_name="sheetbook_detail_opened")
        .exclude(user_id__isnull=True)
        .annotate(day=TruncDate("created_at"))
        .values("user_id")
        .annotate(day_count=Count("day", distinct=True))
    )
    summary["revisitor_user_count"] = revisit_by_user.filter(day_count__gte=2).count()
    summary["revisit_rate"] = (
        round((summary["revisitor_user_count"] / summary["active_user_count"]) * 100, 1)
        if summary["active_user_count"]
        else 0.0
    )

    index_first_map = {
        row["user_id"]: row["first_at"]
        for row in (
            event_qs.filter(event_name__in=["sheetbook_index_opened", "workspace_home_opened"])
            .exclude(user_id__isnull=True)
            .values("user_id")
            .annotate(first_at=Min("created_at"))
        )
    }
    create_first_map = {
        row["user_id"]: row["first_at"]
        for row in (
            event_qs.filter(event_name="sheetbook_created")
            .exclude(user_id__isnull=True)
            .values("user_id")
            .annotate(first_at=Min("created_at"))
        )
    }
    quick_create_user_count = 0
    for user_id, first_index_at in index_first_map.items():
        first_create_at = create_first_map.get(user_id)
        if not first_create_at:
            continue
        gap = first_create_at - first_index_at
        if timedelta(seconds=0) <= gap <= timedelta(minutes=10):
            quick_create_user_count += 1
    summary["quick_create_base_count"] = len(index_first_map)
    summary["quick_create_user_count"] = quick_create_user_count
    summary["quick_create_rate"] = (
        round((quick_create_user_count / summary["quick_create_base_count"]) * 100, 1)
        if summary["quick_create_base_count"]
        else 0.0
    )

    workspace_source_index_count = 0
    workspace_source_detail_count = 0
    workspace_source_create_count = 0
    workspace_source_action_requested_count = 0
    workspace_source_action_success_count = 0
    source_sample_rows = list(
        event_qs.filter(
            event_name__in=[
                "sheetbook_index_opened",
                "sheetbook_detail_opened",
                "sheetbook_created",
                "action_execute_requested",
                "action_execute_succeeded",
            ]
        )
        .values("event_name", "metadata")
    )
    for row in source_sample_rows:
        metadata = row.get("metadata") or {}
        source = str(metadata.get("entry_source") or "").strip().lower()
        if not source.startswith("workspace_home"):
            continue
        if row["event_name"] == "sheetbook_index_opened":
            workspace_source_index_count += 1
        elif row["event_name"] == "sheetbook_detail_opened":
            workspace_source_detail_count += 1
        elif row["event_name"] == "sheetbook_created":
            workspace_source_create_count += 1
        elif row["event_name"] == "action_execute_requested":
            workspace_source_action_requested_count += 1
        elif row["event_name"] == "action_execute_succeeded":
            workspace_source_action_success_count += 1
    summary["workspace_source_index_count"] = workspace_source_index_count
    summary["workspace_source_detail_count"] = workspace_source_detail_count
    summary["workspace_source_create_count"] = workspace_source_create_count
    summary["workspace_source_action_requested_count"] = workspace_source_action_requested_count
    summary["workspace_source_action_success_count"] = workspace_source_action_success_count
    workspace_home_opened_count = summary["workspace_home_opened_count"]
    summary["workspace_to_index_rate"] = (
        round((workspace_source_index_count / workspace_home_opened_count) * 100, 1)
        if workspace_home_opened_count
        else 0.0
    )
    summary["workspace_to_detail_rate"] = (
        round((workspace_source_detail_count / workspace_home_opened_count) * 100, 1)
        if workspace_home_opened_count
        else 0.0
    )
    summary["workspace_to_create_rate"] = (
        round((workspace_source_create_count / workspace_home_opened_count) * 100, 1)
        if workspace_home_opened_count
        else 0.0
    )
    summary["workspace_to_action_requested_rate"] = (
        round((workspace_source_action_requested_count / workspace_home_opened_count) * 100, 1)
        if workspace_home_opened_count
        else 0.0
    )
    summary["workspace_to_action_success_rate"] = (
        round((workspace_source_action_success_count / workspace_home_opened_count) * 100, 1)
        if workspace_home_opened_count
        else 0.0
    )
    summary["workspace_create_to_action_requested_rate"] = (
        round((workspace_source_action_requested_count / workspace_source_create_count) * 100, 1)
        if workspace_source_create_count
        else 0.0
    )
    workspace_to_create_target_rate = _parse_percentage_or_default(
        getattr(settings, "SHEETBOOK_WORKSPACE_TO_CREATE_TARGET_RATE", 60.0),
        default=60.0,
    )
    workspace_create_to_action_target_rate = _parse_percentage_or_default(
        getattr(settings, "SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_TARGET_RATE", 50.0),
        default=50.0,
    )
    workspace_to_create_min_sample = _parse_positive_int(
        getattr(settings, "SHEETBOOK_WORKSPACE_TO_CREATE_MIN_SAMPLE", 5),
        default=5,
    )
    workspace_create_to_action_min_sample = _parse_positive_int(
        getattr(settings, "SHEETBOOK_WORKSPACE_CREATE_TO_ACTION_MIN_SAMPLE", 5),
        default=5,
    )

    summary["workspace_to_create_target_rate"] = workspace_to_create_target_rate
    summary["workspace_to_create_min_sample"] = workspace_to_create_min_sample
    summary["workspace_to_create_gap"] = round(
        summary["workspace_to_create_rate"] - summary["workspace_to_create_target_rate"],
        1,
    )
    summary["workspace_to_create_needs_attention"] = (
        workspace_home_opened_count >= workspace_to_create_min_sample
        and summary["workspace_to_create_rate"] < summary["workspace_to_create_target_rate"]
    )
    summary["workspace_create_to_action_target_rate"] = workspace_create_to_action_target_rate
    summary["workspace_create_to_action_min_sample"] = workspace_create_to_action_min_sample
    summary["workspace_create_to_action_gap"] = round(
        summary["workspace_create_to_action_requested_rate"] - summary["workspace_create_to_action_target_rate"],
        1,
    )
    summary["workspace_create_to_action_needs_attention"] = (
        workspace_source_create_count >= workspace_create_to_action_min_sample
        and summary["workspace_create_to_action_requested_rate"] < summary["workspace_create_to_action_target_rate"]
    )

    event_breakdown = list(
        event_qs.values("event_name")
        .annotate(total=Count("id"))
        .order_by("-total", "event_name")[:20]
    )
    event_label_map = {
        "workspace_home_opened": "로그인 홈 열기",
        "sheetbook_index_opened": "교무수첩 목록 열기",
        "sheetbook_created": "새 교무수첩 만들기",
        "sheetbook_detail_opened": "교무수첩 상세 열기",
        "sheetbook_archived": "교무수첩 아카이브",
        "sheetbook_unarchived": "교무수첩 아카이브 해제",
        "sheetbook_archive_bulk_updated": "교무수첩 다건 아카이브/복구",
        "sheetbook_mobile_read_mode_opened": "휴대폰 읽기 모드 진입",
        "sheetbook_mobile_read_mode_blocked": "휴대폰 읽기 모드 편집 차단",
        "sheetbook_archive_read_mode_opened": "아카이브 읽기 모드 진입",
        "sheetbook_archive_read_mode_blocked": "아카이브 읽기 모드 편집 차단",
        "action_execute_requested": "기능 실행 시작",
        "action_execute_succeeded": "기능 실행 완료",
        "action_execute_failed": "기능 실행 실패",
        "action_history_loaded": "최근 기록 불러오기",
        "consent_review_opened": "동의서 확인 화면 열기",
        "consent_review_submitted": "동의서 확인 후 다음 단계 이동",
        "consent_review_missing_seed": "동의서 준비 정보 누락",
    }
    for item in event_breakdown:
        event_name = str(item.get("event_name") or "").strip()
        item["event_label"] = event_label_map.get(event_name, event_name)

    action_breakdown = list(
        event_qs.exclude(action_type="")
        .values("action_type")
        .annotate(total=Count("id"))
        .order_by("-total", "action_type")
    )
    action_label_map = dict(ActionInvocation.ACTION_CHOICES)
    for item in action_breakdown:
        action_code = str(item.get("action_type") or "").strip()
        item["action_label"] = action_label_map.get(action_code, action_code or "기타")
    daily_stats = list(
        event_qs.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total=Count("id"),
            sheetbook_created=Count("id", filter=Q(event_name="sheetbook_created")),
            action_success=Count("id", filter=Q(event_name="action_execute_succeeded")),
            action_failed=Count("id", filter=Q(event_name="action_execute_failed")),
        )
        .order_by("-day")
    )

    _log_sheetbook_metric(
        "metrics_dashboard_opened",
        user_id=request.user.id,
        days=days,
        loaded_event_count=summary["event_total"],
    )
    return render(
        request,
        "sheetbook/metrics_dashboard.html",
        {
            "days": days,
            "summary": summary,
            "event_breakdown": event_breakdown,
            "action_breakdown": action_breakdown,
            "daily_stats": daily_stats,
        },
    )
