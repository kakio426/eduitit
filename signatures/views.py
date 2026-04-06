import base64
import io
import json
import logging
import mimetypes
import re
import uuid
from collections import Counter, defaultdict
from datetime import timedelta
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import qrcode
from django.contrib.auth import get_user_model
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, Http404, JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.utils.text import get_valid_filename
from django.utils import timezone
from django_ratelimit.decorators import ratelimit
from handoff.models import HandoffRosterGroup
from .models import (
    SIGNATURE_ATTACHMENT_MAX_FILES,
    SIGNATURE_ATTACHMENT_MAX_TOTAL_BYTES,
    AffiliationCorrectionLog,
    ExpectedParticipant,
    Signature,
    SignatureAuditLog,
    TrainingSession,
    TrainingSessionAttachment,
)
from .forms import TrainingSessionForm, SignatureForm, validate_training_session_attachment_files
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

logger = logging.getLogger(__name__)
CALENDAR_INTEGRATION_SOURCE = "signatures_training"
WORKFLOW_ACTION_SEED_SESSION_KEY = "workflow_action_seeds"
SHEETBOOK_ACTION_SEED_SESSION_KEY = "sheetbook_action_seeds"
SIGNATURE_CREATE_DRAFT_SESSION_KEY = "signature_create_drafts"
SIGNATURE_PROXY_MANAGER_USERNAMES = {"kakio"}
DEFAULT_AFFILIATION_SUGGESTIONS = [
    "교사",
    "교감",
    "교장",
    "담임",
    "전담",
    "보건교사",
    "사서교사",
    "영양교사",
    "상담교사",
    "행정실",
]

User = get_user_model()


def _apply_sensitive_cache_headers(response):
    response["Cache-Control"] = "no-store, private"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def _signature_public_ratelimit_key(group, request):
    resolver_match = getattr(request, "resolver_match", None)
    session_uuid = ""
    if resolver_match is not None:
        session_uuid = resolver_match.kwargs.get("uuid", "")
    return f"{_request_client_ip(request) or 'unknown'}:{session_uuid or 'unknown'}"


def _local_datetime_display(value):
    if not value:
        return ""
    try:
        return timezone.localtime(value).strftime("%Y.%m.%d %H:%M")
    except Exception:
        return value.strftime("%Y.%m.%d %H:%M")


def _format_file_size(size):
    try:
        value = int(size or 0)
    except (TypeError, ValueError):
        value = 0
    if value <= 0:
        return ""
    if value < 1024:
        return f"{value}B"
    if value < 1024 * 1024:
        return f"{(value / 1024):.1f}KB"
    return f"{(value / (1024 * 1024)):.1f}MB"


def _clean_saved_signature_image_data(raw_value):
    value = str(raw_value or "").strip()
    if not value:
        raise ValueError("저장할 손서명이 없습니다.")
    if len(value) > 2_000_000:
        raise ValueError("서명 이미지가 너무 커서 저장할 수 없습니다. 다시 그려 주세요.")
    prefix = "data:image/png;base64,"
    if not value.startswith(prefix):
        raise ValueError("저장할 서명 이미지를 다시 확인해 주세요.")
    try:
        base64.b64decode(value[len(prefix):], validate=True)
    except Exception as exc:
        raise ValueError("저장할 서명 이미지를 다시 확인해 주세요.") from exc
    return value


def _serialize_saved_signature(signature):
    return {
        "id": signature.id,
        "image_data": signature.image_data,
        "created_at_display": timezone.localtime(signature.created_at).strftime("%Y.%m.%d %H:%M"),
    }


def _safe_attachment_download_name(raw_name, *, fallback="attachment"):
    original = str(raw_name or "").strip()
    original_ext = Path(original).suffix or ""
    original_base = Path(original).stem if original else fallback
    safe_base = get_valid_filename(original_base) or fallback
    safe_ext = re.sub(r"[^A-Za-z0-9.]", "", original_ext)[:12]
    return f"{safe_base[:180]}{safe_ext}"


def _get_session_attachments(session):
    return list(session.attachments.order_by("sort_order", "id"))


def _summarize_attachment_state(attachments):
    total_count = len(attachments)
    total_bytes = sum(attachment.file_size for attachment in attachments)
    return {
        "count": total_count,
        "total_bytes": total_bytes,
        "total_size_display": _format_file_size(total_bytes),
        "summary_label": f"첨부 {total_count}개 포함" if total_count else "",
    }


def _build_signature_attachment_rows(session, *, public=False):
    rows = []
    route_name = "signatures:sign_attachment_download" if public else "signatures:attachment_download"
    for attachment in _get_session_attachments(session):
        rows.append(
            {
                "id": attachment.id,
                "name": attachment.original_name,
                "size_display": _format_file_size(attachment.file_size),
                "download_url": reverse(
                    route_name,
                    kwargs={"uuid": session.uuid, "attachment_id": attachment.id},
                ),
            }
        )
    return rows


def _build_signature_share_package_text(session, share_link, attachment_count=0):
    lines = [
        f"[{session.title}] 서명 부탁드립니다.",
        f"일시: {_local_datetime_display(session.datetime)}",
        f"장소: {session.location}",
        "아래 링크에서 로그인 없이 바로 서명하실 수 있습니다.",
    ]
    if attachment_count:
        lines.append(f"첨부 파일 {attachment_count}개 확인 후 서명 부탁드립니다.")
    lines.append(share_link)
    return "\n".join(lines)


def _collect_attachment_upload_state(session, new_files, attachments_to_remove):
    current_attachments = _get_session_attachments(session) if session else []
    remove_ids = {attachment.id for attachment in attachments_to_remove}
    remaining_attachments = [attachment for attachment in current_attachments if attachment.id not in remove_ids]
    remaining_count = len(remaining_attachments)
    remaining_total_bytes = sum(attachment.file_size for attachment in remaining_attachments)
    new_total_bytes = sum(int(getattr(file_obj, "size", 0) or 0) for file_obj in new_files)
    return {
        "remaining_attachments": remaining_attachments,
        "remaining_count": remaining_count,
        "remaining_total_bytes": remaining_total_bytes,
        "new_total_bytes": new_total_bytes,
        "combined_count": remaining_count + len(new_files),
        "combined_total_bytes": remaining_total_bytes + new_total_bytes,
    }


def _validate_session_attachment_batch(session, new_files, attachments_to_remove):
    errors = []
    upload_state = _collect_attachment_upload_state(session, new_files, attachments_to_remove)
    if upload_state["combined_count"] > SIGNATURE_ATTACHMENT_MAX_FILES:
        errors.append(f"첨부 파일은 최대 {SIGNATURE_ATTACHMENT_MAX_FILES}개까지 넣을 수 있습니다.")
    if upload_state["combined_total_bytes"] > SIGNATURE_ATTACHMENT_MAX_TOTAL_BYTES:
        errors.append(
            f"첨부 파일 전체 용량은 {_format_file_size(SIGNATURE_ATTACHMENT_MAX_TOTAL_BYTES)} 이하로 맞춰 주세요."
        )
    return errors, upload_state


def _resolve_attachment_removals(session, raw_ids):
    attachment_ids = []
    for raw_id in raw_ids or []:
        try:
            attachment_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue
    if not session or not attachment_ids:
        return []
    return list(session.attachments.filter(id__in=attachment_ids).order_by("sort_order", "id"))


def _sync_training_session_attachments(session, *, new_files, attachments_to_remove):
    for attachment in attachments_to_remove:
        attachment.delete()
    for file_obj in new_files:
        TrainingSessionAttachment.objects.create(
            training_session=session,
            file=file_obj,
            original_name=Path(getattr(file_obj, "name", "") or "").name[:255],
        )


def _build_attachment_download_response(attachment):
    if not attachment.file or not attachment.file.name:
        raise Http404("첨부 파일을 찾을 수 없습니다.")

    download_name = _safe_attachment_download_name(
        attachment.original_name or attachment.file.name,
        fallback=f"attachment-{attachment.pk}",
    )
    content_type = mimetypes.guess_type(download_name)[0] or "application/octet-stream"
    try:
        file_obj = attachment.file.open("rb")
    except Exception as exc:
        logger.exception(
            "[signatures] attachment open failed session_uuid=%s attachment_id=%s",
            attachment.training_session.uuid,
            attachment.id,
        )
        raise Http404("첨부 파일을 찾을 수 없습니다.") from exc

    response = FileResponse(
        file_obj,
        content_type=content_type,
        as_attachment=True,
        filename=download_name,
    )
    return _apply_sensitive_cache_headers(response)


def _normalize_affiliation_text(value):
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    normalized = normalized.replace("—", "-").replace("–", "-")
    normalized = re.sub(r"\s+", " ", normalized)
    normalized = re.sub(r"\s*/\s*", "/", normalized)
    normalized = re.sub(r"\s*-\s*", "-", normalized)
    return normalized[:100]


def _natural_string_sort_key(value):
    normalized = _normalize_affiliation_text(value).lower()
    if not normalized:
        return ((1, ""),)

    parts = re.split(r"(\d+)", normalized)
    key = []
    for part in parts:
        if not part:
            continue
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def _signature_affiliation_sort_key(signature):
    affiliation = signature.display_affiliation
    return (
        1 if not affiliation else 0,
        _natural_string_sort_key(affiliation),
        _natural_string_sort_key(signature.participant_name),
        signature.created_at,
        signature.id,
    )


def _signature_submitted_sort_key(signature):
    return (
        signature.created_at,
        signature.id,
    )


def _signature_manual_sort_key(signature):
    return (
        signature.manual_sort_order or 0,
        signature.created_at,
        signature.id,
    )


def _get_signature_sort_mode(session):
    sort_mode = str(session.signature_sort_mode or "").strip()
    valid_modes = {
        TrainingSession.SIGNATURE_SORT_SUBMITTED,
        TrainingSession.SIGNATURE_SORT_AFFILIATION,
        TrainingSession.SIGNATURE_SORT_MANUAL,
    }
    if sort_mode not in valid_modes:
        return TrainingSession.SIGNATURE_SORT_AFFILIATION
    return sort_mode


def _get_participant_sort_mode(session):
    sort_mode = str(session.participant_sort_mode or "").strip()
    valid_modes = {
        TrainingSession.SIGNATURE_SORT_SUBMITTED,
        TrainingSession.SIGNATURE_SORT_AFFILIATION,
        TrainingSession.SIGNATURE_SORT_MANUAL,
    }
    if sort_mode not in valid_modes:
        return TrainingSession.SIGNATURE_SORT_SUBMITTED
    return sort_mode


def _get_attendance_sort_choices(has_expected_participants):
    if has_expected_participants:
        return [
            (TrainingSession.SIGNATURE_SORT_SUBMITTED, "명단 입력 순서"),
            (TrainingSession.SIGNATURE_SORT_AFFILIATION, "학년반/이름 정렬"),
            (TrainingSession.SIGNATURE_SORT_MANUAL, "직접 순서 정하기"),
        ]
    return TrainingSession.SIGNATURE_SORT_CHOICES


def _sort_signatures_for_display(signatures, sort_mode):
    signature_list = list(signatures)
    if sort_mode == TrainingSession.SIGNATURE_SORT_SUBMITTED:
        return sorted(signature_list, key=_signature_submitted_sort_key)
    if sort_mode == TrainingSession.SIGNATURE_SORT_MANUAL:
        return sorted(signature_list, key=_signature_manual_sort_key)
    return sorted(signature_list, key=_signature_affiliation_sort_key)


def _participant_submitted_sort_key(participant):
    return (
        participant.created_at,
        participant.id,
    )


def _expected_participant_sort_key(participant):
    affiliation = participant.display_affiliation
    return (
        1 if not affiliation else 0,
        _natural_string_sort_key(affiliation),
        _natural_string_sort_key(participant.name),
        participant.created_at,
        participant.id,
    )


def _participant_manual_sort_key(participant):
    return (
        participant.manual_sort_order or 0,
        participant.created_at,
        participant.id,
    )


def _sort_expected_participants_for_display(participants, sort_mode):
    participant_list = list(participants)
    if sort_mode == TrainingSession.SIGNATURE_SORT_SUBMITTED:
        return sorted(participant_list, key=_participant_submitted_sort_key)
    if sort_mode == TrainingSession.SIGNATURE_SORT_MANUAL:
        return sorted(participant_list, key=_participant_manual_sort_key)
    return sorted(participant_list, key=_expected_participant_sort_key)


def _resequence_manual_participant_order(session):
    ordered_participants = _sort_expected_participants_for_display(
        session.expected_participants.all(),
        TrainingSession.SIGNATURE_SORT_MANUAL,
    )
    updates = []
    for index, participant in enumerate(ordered_participants, start=1):
        if participant.manual_sort_order == index:
            continue
        participant.manual_sort_order = index
        updates.append(participant)
    if updates:
        ExpectedParticipant.objects.bulk_update(updates, ["manual_sort_order"])
    return ordered_participants


def _move_participant_to_position(session, participant, target_position):
    ordered_participants = _resequence_manual_participant_order(session)
    total_count = len(ordered_participants)
    if total_count <= 1:
        return ordered_participants

    target_position = max(1, min(int(target_position), total_count))
    current_index = next(
        (index for index, item in enumerate(ordered_participants) if item.id == participant.id),
        None,
    )
    if current_index is None:
        return ordered_participants

    participant_item = ordered_participants.pop(current_index)
    ordered_participants.insert(target_position - 1, participant_item)
    for index, item in enumerate(ordered_participants, start=1):
        item.manual_sort_order = index
    ExpectedParticipant.objects.bulk_update(ordered_participants, ["manual_sort_order"])
    return ordered_participants


def _resequence_manual_signature_order(session):
    ordered_signatures = _sort_signatures_for_display(
        session.signatures.all(),
        TrainingSession.SIGNATURE_SORT_MANUAL,
    )
    updates = []
    for index, signature in enumerate(ordered_signatures, start=1):
        if signature.manual_sort_order == index:
            continue
        signature.manual_sort_order = index
        updates.append(signature)
    if updates:
        Signature.objects.bulk_update(updates, ["manual_sort_order"])
    return ordered_signatures


def _move_signature_to_position(session, signature, target_position):
    ordered_signatures = _resequence_manual_signature_order(session)
    total_count = len(ordered_signatures)
    if total_count <= 1:
        return ordered_signatures

    target_position = max(1, min(int(target_position), total_count))
    current_index = next(
        (index for index, item in enumerate(ordered_signatures) if item.id == signature.id),
        None,
    )
    if current_index is None:
        return ordered_signatures

    signature_item = ordered_signatures.pop(current_index)
    ordered_signatures.insert(target_position - 1, signature_item)
    for index, item in enumerate(ordered_signatures, start=1):
        item.manual_sort_order = index
    Signature.objects.bulk_update(ordered_signatures, ["manual_sort_order"])
    return ordered_signatures


def _request_client_ip(request):
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if x_forwarded_for:
        return x_forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def _request_user_agent(request):
    return (request.META.get("HTTP_USER_AGENT", "") or "")[:1000]


def _is_signature_proxy_manager(user):
    return bool(
        getattr(user, "is_authenticated", False)
        and user.is_superuser
        and user.username in SIGNATURE_PROXY_MANAGER_USERNAMES
    )


def _get_signature_accessible_sessions(user):
    if not getattr(user, "is_authenticated", False):
        return TrainingSession.objects.none()

    queryset = TrainingSession.objects.filter(created_by=user)
    if _is_signature_proxy_manager(user):
        queryset = TrainingSession.objects.filter(
            Q(created_by=user) | Q(proxy_created_by=user)
        )
    return queryset.select_related("created_by", "proxy_created_by", "shared_roster_group")


def _get_signature_session_or_404(user, uuid):
    return get_object_or_404(_get_signature_accessible_sessions(user), uuid=uuid)


def _get_signature_proxy_target_user(current_user, raw_user_id):
    if not _is_signature_proxy_manager(current_user):
        return None

    raw_user_id = str(raw_user_id or "").strip()
    if not raw_user_id:
        return None
    try:
        return User.objects.filter(
            pk=int(raw_user_id),
            is_active=True,
        ).exclude(pk=current_user.pk).exclude(is_staff=True).exclude(is_superuser=True).first()
    except (TypeError, ValueError):
        return None


def _get_signature_user_display_name(user):
    if user is None:
        return ""

    try:
        profile = user.userprofile
    except Exception:
        profile = None
    nickname = str(getattr(profile, "nickname", "") or "").strip()
    full_name = str(getattr(user, "get_full_name", lambda: "")() or "").strip()
    return nickname or full_name or user.username


def _parse_expected_participants_text(raw_text):
    participants = []
    for line in str(raw_text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split(",")]
        name = str(parts[0] if parts else "").strip()[:100]
        affiliation = _normalize_affiliation_text(parts[1] if len(parts) > 1 else "")
        if not name:
            continue
        participants.append({
            "name": name,
            "affiliation": affiliation,
        })
    return participants


def _normalize_access_code(value):
    normalized = re.sub(r"\s+", "", str(value or "").strip())
    return normalized[:6]


def _build_access_code_state(session, *, now=None):
    current_time = now or timezone.now()
    status = session.access_code_status(now=current_time)
    expires_at = session.active_access_code_expires_at
    return {
        "enabled": session.access_code_required,
        "duration_minutes": int(session.access_code_duration_minutes or 0),
        "status": status,
        "is_active": status == "active",
        "is_pending": status == "pending",
        "is_expired": status == "expired",
        "expires_at": expires_at,
        "current_code": session.active_access_code or "",
    }


def _validate_session_access_code(session, submitted_code, *, now=None):
    current_time = now or timezone.now()
    if not session.access_code_required:
        return None

    status = session.access_code_status(now=current_time)
    if status == "pending":
        return "담당 교사가 아직 현장 코드를 열지 않았습니다. 잠시 후 다시 시도해 주세요."
    if status == "expired":
        return "현장 코드 시간이 끝났습니다. 담당 교사에게 새 코드를 요청해 주세요."
    if status != "active":
        return "현장 코드를 확인할 수 없습니다. 담당 교사에게 다시 받아 주세요."

    normalized_code = _normalize_access_code(submitted_code)
    if not re.fullmatch(r"\d{4,6}", normalized_code):
        return "현장 코드는 4~6자리 숫자로 입력해 주세요."
    if normalized_code != (session.active_access_code or ""):
        return "현장 코드를 다시 확인해 주세요."
    return None


def _build_affiliation_suggestions(session, max_items=40):
    counter = Counter()
    for raw in session.expected_participants.values_list("affiliation", flat=True):
        value = _normalize_affiliation_text(raw)
        if value:
            counter[value] += 1
    for raw in session.expected_participants.values_list("corrected_affiliation", flat=True):
        value = _normalize_affiliation_text(raw)
        if value:
            counter[value] += 1
    for raw in session.signatures.values_list("participant_affiliation", flat=True):
        value = _normalize_affiliation_text(raw)
        if value:
            counter[value] += 1
    for raw in session.signatures.values_list("corrected_affiliation", flat=True):
        value = _normalize_affiliation_text(raw)
        if value:
            counter[value] += 1

    suggestions = []
    seen = set()
    for value, _ in counter.most_common(max_items):
        if value not in seen:
            suggestions.append(value)
            seen.add(value)
    for default in DEFAULT_AFFILIATION_SUGGESTIONS:
        value = _normalize_affiliation_text(default)
        if value and value not in seen:
            suggestions.append(value)
            seen.add(value)
    for grade in range(1, 7):
        for classroom in range(1, 7):
            value = f"{grade}-{classroom}"
            if value not in seen:
                suggestions.append(value)
                seen.add(value)
            if len(suggestions) >= max_items:
                return suggestions
    return suggestions[:max_items]


def _apply_signature_affiliation_correction(signature, corrected_affiliation, reason, user):
    raw_source = str(signature.participant_affiliation or "").strip()
    normalized_source = _normalize_affiliation_text(raw_source)
    normalized_corrected = _normalize_affiliation_text(corrected_affiliation)
    reason = str(reason or "").strip()[:200]

    if not normalized_corrected:
        signature.corrected_affiliation = ""
        signature.affiliation_correction_reason = ""
        signature.affiliation_corrected_by = None
        signature.affiliation_corrected_at = None
        return
    if normalized_corrected == normalized_source and raw_source == normalized_corrected:
        signature.corrected_affiliation = ""
        signature.affiliation_correction_reason = ""
        signature.affiliation_corrected_by = None
        signature.affiliation_corrected_at = None
        return

    signature.corrected_affiliation = normalized_corrected
    signature.affiliation_correction_reason = reason
    signature.affiliation_corrected_by = user
    signature.affiliation_corrected_at = timezone.now()


def _apply_expected_participant_affiliation_correction(participant, corrected_affiliation, reason, user):
    raw_source = str(participant.affiliation or "").strip()
    normalized_source = _normalize_affiliation_text(raw_source)
    normalized_corrected = _normalize_affiliation_text(corrected_affiliation)
    reason = str(reason or "").strip()[:200]

    if not normalized_corrected:
        participant.corrected_affiliation = ""
        participant.affiliation_correction_reason = ""
        participant.affiliation_corrected_by = None
        participant.affiliation_corrected_at = None
        return
    if normalized_corrected == normalized_source and raw_source == normalized_corrected:
        participant.corrected_affiliation = ""
        participant.affiliation_correction_reason = ""
        participant.affiliation_corrected_by = None
        participant.affiliation_corrected_at = None
        return

    participant.corrected_affiliation = normalized_corrected
    participant.affiliation_correction_reason = reason
    participant.affiliation_corrected_by = user
    participant.affiliation_corrected_at = timezone.now()


def _create_affiliation_correction_log(
    *,
    session,
    target_type,
    mode,
    before_affiliation,
    after_affiliation,
    reason,
    corrected_by=None,
    signature=None,
    expected_participant=None,
):
    AffiliationCorrectionLog.objects.create(
        training_session=session,
        target_type=target_type,
        mode=mode,
        signature=signature,
        expected_participant=expected_participant,
        before_affiliation=_normalize_affiliation_text(before_affiliation),
        after_affiliation=_normalize_affiliation_text(after_affiliation),
        reason=str(reason or "").strip()[:200],
        corrected_by=corrected_by,
    )


def _build_qr_data_url(raw_text):
    if not raw_text:
        return ""

    qr_image = qrcode.make(raw_text)
    with io.BytesIO() as buffer:
        qr_image.save(buffer, format="PNG")
        encoded = base64.b64encode(buffer.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _to_aware_datetime(value):
    if timezone.is_naive(value):
        return timezone.make_aware(value, timezone.get_current_timezone())
    return timezone.localtime(value)


def _sync_calendar_event_for_training(session):
    return


def _delete_calendar_event_for_training(session):
    return


def _sync_expected_participants_from_shared_roster(session):
    """Pull active members from linked handoff roster into expected participants."""
    group = session.shared_roster_group
    if not group:
        return {"created": 0, "skipped": 0, "total": 0, "group_name": ""}
    if group.owner_id != session.created_by_id:
        return {"created": 0, "skipped": 0, "total": 0, "group_name": group.name}

    created = 0
    skipped = 0
    members = group.members.filter(is_active=True).order_by("sort_order", "id")
    for member in members:
        name = (member.display_name or "").strip()
        if not name:
            skipped += 1
            continue
        affiliation = _normalize_affiliation_text(member.affiliation or member.note)
        _, was_created = ExpectedParticipant.objects.get_or_create(
            training_session=session,
            name=name,
            affiliation=affiliation,
        )
        if was_created:
            created += 1
        else:
            skipped += 1
    return {"created": created, "skipped": skipped, "total": members.count(), "group_name": group.name}


def _peek_sheetbook_seed(request, token, *, expected_action=""):
    token = (token or "").strip()
    if not token:
        return None
    for session_key in (WORKFLOW_ACTION_SEED_SESSION_KEY, SHEETBOOK_ACTION_SEED_SESSION_KEY):
        seeds = request.session.get(session_key, {})
        if not isinstance(seeds, dict):
            continue
        seed = seeds.get(token)
        if not isinstance(seed, dict):
            continue
        if expected_action and seed.get("action") != expected_action:
            continue
        return seed
    return None


def _pop_sheetbook_seed(request, token, *, expected_action=""):
    token = (token or "").strip()
    if not token:
        return None
    found_seed = None
    for session_key in (WORKFLOW_ACTION_SEED_SESSION_KEY, SHEETBOOK_ACTION_SEED_SESSION_KEY):
        seeds = request.session.get(session_key, {})
        if not isinstance(seeds, dict):
            continue
        seed = seeds.get(token)
        if not isinstance(seed, dict):
            continue
        if expected_action and seed.get("action") != expected_action:
            continue
        if found_seed is None:
            found_seed = seed
        seeds.pop(token, None)
        request.session[session_key] = seeds
    if found_seed is not None:
        request.session.modified = True
    return found_seed


def _parse_sheetbook_signature_participants(text, max_count=300):
    participants = []
    seen = set()
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split(",")]
        name = (parts[0] if parts else "").strip()[:100]
        affiliation = _normalize_affiliation_text(parts[1] if len(parts) >= 2 else "")
        if not name:
            continue
        dedupe_key = (name, affiliation)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        participants.append({"name": name, "affiliation": affiliation})
        if len(participants) >= max_count:
            break
    return participants


def _is_truthy_flag(value, *, default=True):
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if not normalized:
        return False
    return normalized not in {"0", "false", "off", "no", "n"}


def _to_datetime_local_input_value(value):
    if not value:
        return ""
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime("%Y-%m-%dT%H:%M")


def _append_query_params(url, **params):
    split_result = urlsplit(url)
    query = dict(parse_qsl(split_result.query, keep_blank_values=True))
    for key, value in params.items():
        if value in (None, ""):
            query.pop(key, None)
            continue
        query[str(key)] = str(value)
    return urlunsplit(
        (
            split_result.scheme,
            split_result.netloc,
            split_result.path,
            urlencode(query),
            split_result.fragment,
        )
    )


def _get_signature_create_drafts(request):
    drafts = request.session.get(SIGNATURE_CREATE_DRAFT_SESSION_KEY, {})
    if not isinstance(drafts, dict):
        drafts = {}
    return drafts


def _peek_signature_create_draft(request, token):
    token = str(token or "").strip()
    if not token:
        return {}
    draft = _get_signature_create_drafts(request).get(token, {})
    return draft if isinstance(draft, dict) else {}


def _store_signature_create_draft(request, draft_payload, *, token=""):
    token = str(token or "").strip() or uuid.uuid4().hex
    drafts = _get_signature_create_drafts(request)
    drafts[token] = draft_payload
    request.session[SIGNATURE_CREATE_DRAFT_SESSION_KEY] = drafts
    request.session.modified = True
    return token


def _pop_signature_create_draft(request, token):
    token = str(token or "").strip()
    if not token:
        return {}
    drafts = _get_signature_create_drafts(request)
    draft = drafts.pop(token, {})
    request.session[SIGNATURE_CREATE_DRAFT_SESSION_KEY] = drafts
    request.session.modified = True
    return draft if isinstance(draft, dict) else {}


def _build_signature_create_draft_payload(data):
    payload = {
        "title": str(data.get("title") or "").strip()[:200],
        "print_title": str(data.get("print_title") or "").strip()[:200],
        "instructor": str(data.get("instructor") or "").strip()[:100],
        "datetime": str(data.get("datetime") or "").strip()[:16],
        "location": str(data.get("location") or "").strip()[:200],
        "description": str(data.get("description") or "").strip()[:2000],
        "shared_roster_group": str(data.get("shared_roster_group") or "").strip(),
        "expected_count": str(data.get("expected_count") or "").strip()[:10],
        "acting_for_user": str(data.get("acting_for_user") or "").strip(),
        "proxy_participants_text": str(data.get("proxy_participants_text") or "").strip()[:5000],
        "copy_from_uuid": str(data.get("copy_from_uuid") or "").strip(),
        "sheetbook_seed_token": str(data.get("sheetbook_seed_token") or "").strip(),
        "draft_token": str(data.get("draft_token") or "").strip(),
        "apply_sheetbook_participants": _is_truthy_flag(
            data.get("apply_sheetbook_participants"),
            default=True,
        ),
        "is_active": _is_truthy_flag(data.get("is_active"), default=False),
    }
    return payload


def _build_signature_create_initial_from_draft(draft_payload):
    if not draft_payload:
        return {}

    initial = {
        "title": draft_payload.get("title", ""),
        "print_title": draft_payload.get("print_title", ""),
        "instructor": draft_payload.get("instructor", ""),
        "datetime": draft_payload.get("datetime", ""),
        "location": draft_payload.get("location", ""),
        "description": draft_payload.get("description", ""),
        "is_active": bool(draft_payload.get("is_active", False)),
    }
    expected_count = str(draft_payload.get("expected_count") or "").strip()
    if expected_count.isdigit() and int(expected_count) > 0:
        initial["expected_count"] = int(expected_count)
    acting_for_user = str(draft_payload.get("acting_for_user") or "").strip()
    if acting_for_user:
        initial["acting_for_user"] = acting_for_user
    proxy_participants_text = str(draft_payload.get("proxy_participants_text") or "").strip()
    if proxy_participants_text:
        initial["proxy_participants_text"] = proxy_participants_text
    shared_roster_group = str(draft_payload.get("shared_roster_group") or "").strip()
    if shared_roster_group:
        initial["shared_roster_group"] = shared_roster_group
    return initial


def _get_owned_roster_group(owner, group_id):
    group_id = str(group_id or "").strip()
    if not group_id:
        return None
    try:
        return HandoffRosterGroup.objects.filter(owner=owner, id=group_id).first()
    except Exception:
        return None


def _build_copy_initial_from_session(session):
    expected_count = session.expected_count or session.expected_participants.count() or ""
    initial = {
        "title": session.title,
        "print_title": session.print_title,
        "instructor": session.instructor,
        "location": session.location,
        "description": session.description,
        "datetime": _to_datetime_local_input_value(session.datetime),
        "expected_count": expected_count,
        "is_active": True,
    }
    if session.shared_roster_group_id:
        initial["shared_roster_group"] = session.shared_roster_group_id
    return initial


def _build_session_stage_payload(session, *, signature_count, pending_count=None, can_show_absentees=False):
    if pending_count is None:
        if can_show_absentees:
            pending_count = session.expected_participants.filter(matched_signature__isnull=True).count()
        elif session.expected_count:
            pending_count = max(0, int(session.expected_count) - int(signature_count))

    detail_href = reverse("signatures:detail", kwargs={"uuid": session.uuid})
    copy_href = f"{reverse('signatures:create')}?copy_from={session.uuid}"
    share_href = reverse("signatures:sign", kwargs={"uuid": session.uuid})
    print_href = reverse("signatures:print", kwargs={"uuid": session.uuid})

    if not session.is_active:
        stage = "closed"
        stage_badge = "마감됨"
        hero_title = "결과를 확인하세요"
        if can_show_absentees and pending_count:
            hero_description = f"미참여 {pending_count}명을 확인하고, 결과를 저장한 뒤 다시 열 수도 있습니다."
        elif pending_count:
            hero_description = f"예상 인원 기준으로 아직 {pending_count}명이 남았습니다. 결과를 저장하거나 새 요청으로 복제하세요."
        else:
            hero_description = "출력 또는 PDF 저장 후, 같은 형식으로 새 요청을 빠르게 만들 수 있습니다."
        list_primary_action = {
            "label": "결과 보기",
            "href": detail_href,
        }
        list_secondary_actions = [
            {
                "label": "출력/PDF",
                "href": print_href,
            },
            {
                "label": "복제하기",
                "href": copy_href,
            },
        ]
        primary_actions = [
            {
                "label": "결과 보기",
                "kind": "anchor",
                "href": "#attendanceSummary",
                "variant": "primary",
            },
            {
                "label": "출력/PDF 저장",
                "kind": "link",
                "href": print_href,
                "variant": "secondary",
            },
            {
                "label": "다시 열기",
                "kind": "button",
                "onclick": "toggleActive()",
                "variant": "secondary",
            },
            {
                "label": "복제해서 새로 만들기",
                "kind": "link",
                "href": copy_href,
                "variant": "ghost",
            },
        ]
    elif signature_count == 0:
        stage = "ready"
        stage_badge = "공유 준비"
        hero_title = "이제 참여 링크를 보내세요"
        hero_description = "참석자는 로그인 없이 참여하고, 휴대폰에서 바로 서명할 수 있습니다."
        list_primary_action = {
            "label": "공유 시작",
            "href": detail_href,
        }
        list_secondary_actions = [
            {
                "label": "참여 링크 열기",
                "href": share_href,
            },
            {
                "label": "출력/PDF",
                "href": print_href,
            },
        ]
        primary_actions = [
            {
                "label": "공유 패키지 복사",
                "kind": "button",
                "onclick": "copySharePackage()",
                "variant": "primary",
            },
            {
                "label": "참여 현황 보기",
                "kind": "anchor",
                "href": "#attendanceSummary",
                "variant": "secondary",
            },
        ]
    else:
        stage = "collecting"
        stage_badge = "진행 중"
        hero_title = "서명이 들어오고 있어요"
        if can_show_absentees and pending_count:
            hero_description = f"미참여 {pending_count}명을 이름으로 바로 확인할 수 있습니다."
        elif pending_count:
            hero_description = f"예상 인원 기준으로 아직 {pending_count}명이 남았습니다."
        else:
            hero_description = "현재 참여 현황을 보면서 필요할 때 바로 마감할 수 있습니다."
        list_primary_action = {
            "label": "참여 현황",
            "href": detail_href,
        }
        list_secondary_actions = [
            {
                "label": "참여 링크 열기",
                "href": share_href,
            },
            {
                "label": "출력/PDF",
                "href": print_href,
            },
        ]
        pending_action_label = "미참여 보기" if can_show_absentees else "남은 인원 보기"
        primary_actions = [
            {
                "label": "공유 패키지 복사",
                "kind": "button",
                "onclick": "copySharePackage()",
                "variant": "primary",
            },
            {
                "label": pending_action_label,
                "kind": "anchor",
                "href": "#pendingSummary" if pending_count is not None else "#attendanceSummary",
                "variant": "secondary",
            },
            {
                "label": "참여 현황",
                "kind": "anchor",
                "href": "#attendanceSummary",
                "variant": "ghost",
            },
            {
                "label": "마감하기",
                "kind": "button",
                "onclick": "toggleActive()",
                "variant": "ghost",
            },
        ]

    return {
        "stage": stage,
        "stage_badge": stage_badge,
        "hero_title": hero_title,
        "hero_description": hero_description,
        "pending_count": pending_count,
        "can_show_absentees": can_show_absentees,
        "list_primary_action": list_primary_action,
        "list_secondary_actions": list_secondary_actions,
        "primary_actions": primary_actions,
    }


@login_required
def session_list(request):
    """내가 만든 연수 목록"""
    can_proxy_create = _is_signature_proxy_manager(request.user)
    sessions = _get_signature_accessible_sessions(request.user)
    session_cards = []
    for session in sessions:
        signature_count = session.signatures.count()
        can_show_absentees = session.expected_participants.exists()
        pending_count = (
            session.expected_participants.filter(matched_signature__isnull=True).count()
            if can_show_absentees
            else max(0, int(session.expected_count or 0) - signature_count)
            if session.expected_count
            else None
        )
        stage_payload = _build_session_stage_payload(
            session,
            signature_count=signature_count,
            pending_count=pending_count,
            can_show_absentees=can_show_absentees,
        )
        session_cards.append(
            {
                "session": session,
                "signature_count": signature_count,
                "pending_count": pending_count,
                "can_show_absentees": can_show_absentees,
                "show_owner_label": bool(can_proxy_create and session.created_by_id != request.user.id),
                "owner_label": _get_signature_user_display_name(session.created_by),
                **stage_payload,
            }
        )

    return render(
        request,
        'signatures/list.html',
        {
            'session_cards': session_cards,
            'can_proxy_create': can_proxy_create,
        },
    )


@login_required
@require_POST
def prepare_roster_return(request):
    """서명 생성 초안을 저장한 뒤 명단 서비스로 이동."""
    draft_payload = _build_signature_create_draft_payload(request.POST)
    draft_token = _store_signature_create_draft(
        request,
        draft_payload,
        token=draft_payload.get("draft_token", ""),
    )
    return_to = _append_query_params(
        reverse("signatures:create"),
        draft_token=draft_token,
    )
    return redirect(
        _append_query_params(
            reverse("handoff:dashboard"),
            return_to=return_to,
        )
    )


@login_required
def session_create(request):
    """연수 생성"""
    can_proxy_create = _is_signature_proxy_manager(request.user)
    draft_token = (
        request.POST.get("draft_token")
        or request.GET.get("draft_token")
        or ""
    ).strip()
    draft_payload = _peek_signature_create_draft(request, draft_token)
    proxy_target_user_id = (
        request.POST.get("acting_for_user")
        or draft_payload.get("acting_for_user", "")
        or request.GET.get("acting_for_user")
        or ""
    ).strip()
    proxy_target_user = _get_signature_proxy_target_user(request.user, proxy_target_user_id)
    session_owner = proxy_target_user or request.user
    is_proxy_mode = proxy_target_user is not None
    copy_from_uuid = (
        request.POST.get("copy_from_uuid")
        or draft_payload.get("copy_from_uuid", "")
        or request.GET.get("copy_from")
        or ""
    ).strip()
    sheetbook_seed_token = (
        request.POST.get("sheetbook_seed_token")
        or draft_payload.get("sheetbook_seed_token", "")
        or request.GET.get("sb_seed")
        or ""
    ).strip()
    sheetbook_seed = _peek_sheetbook_seed(
        request,
        sheetbook_seed_token,
        expected_action="signature",
    )
    seed_data = sheetbook_seed.get("data", {}) if isinstance(sheetbook_seed, dict) else {}
    seed_data = seed_data if isinstance(seed_data, dict) else {}
    seed_participants = _parse_sheetbook_signature_participants(seed_data.get("participants_text", ""))
    prefill_source_label = (str(seed_data.get("source_label") or "").strip() if seed_data else "") or "교무수첩에서 가져온 내용으로 먼저 채워두었어요."
    prefill_origin_label = str(seed_data.get("origin_label") or "").strip() if seed_data else ""
    prefill_origin_url = str(seed_data.get("origin_url") or "").strip() if seed_data else ""
    copy_source_session = None
    copy_source_participants = []
    if copy_from_uuid and not seed_data:
        copy_source_session = get_object_or_404(
            TrainingSession,
            uuid=copy_from_uuid,
            created_by=session_owner,
        )
        copy_source_participants = list(
            copy_source_session.expected_participants.order_by("manual_sort_order", "id")
        )

    prefill_initial = {}
    if draft_payload:
        prefill_initial = _build_signature_create_initial_from_draft(draft_payload)
    elif seed_data:
        prefill_initial = {
            "title": str(seed_data.get("title") or "").strip()[:200],
            "print_title": str(seed_data.get("print_title") or "").strip()[:200],
            "instructor": str(seed_data.get("instructor") or "").strip()[:100],
            "location": str(seed_data.get("location") or "").strip()[:200],
            "description": str(seed_data.get("description") or "").strip()[:2000],
            "datetime": str(seed_data.get("datetime") or "").strip()[:16],
        }
        expected_count = seed_data.get("expected_count")
        if isinstance(expected_count, int) and expected_count > 0:
            prefill_initial["expected_count"] = expected_count
        elif str(expected_count).isdigit():
            prefill_initial["expected_count"] = int(str(expected_count))
        elif seed_participants:
            prefill_initial["expected_count"] = len(seed_participants)
    elif copy_source_session:
        prefill_initial = _build_copy_initial_from_session(copy_source_session)
    if is_proxy_mode:
        prefill_initial["acting_for_user"] = str(proxy_target_user.id)

    selected_roster_group_id = (
        request.GET.get("shared_roster_group")
        or draft_payload.get("shared_roster_group", "")
        or prefill_initial.get("shared_roster_group", "")
    )
    selected_roster_group = _get_owned_roster_group(session_owner, selected_roster_group_id)
    restored_roster_group = _get_owned_roster_group(
        session_owner,
        request.GET.get("shared_roster_group"),
    )
    if selected_roster_group is not None:
        prefill_initial["shared_roster_group"] = str(selected_roster_group.id)

    apply_sheetbook_participants = _is_truthy_flag(
        request.POST.get("apply_sheetbook_participants")
        if request.method == "POST"
        else draft_payload.get("apply_sheetbook_participants"),
        default=True,
    )
    existing_attachments = []
    selected_remove_attachment_ids = []

    if request.method == 'POST':
        form = TrainingSessionForm(
            request.POST,
            request.FILES,
            owner=session_owner,
            can_delegate=can_proxy_create,
            delegate_user=request.user,
        )
        form_valid = form.is_valid()
        attachment_files, attachment_file_errors = validate_training_session_attachment_files(
            request.FILES.getlist("attachments")
        )
        attachment_batch_errors, _ = _validate_session_attachment_batch(
            None,
            attachment_files,
            [],
        )
        for error in attachment_file_errors + attachment_batch_errors:
            form.add_error("attachments", error)

        if form_valid and not form.errors:
            session = form.save(commit=False)
            session.created_by = session_owner
            session.proxy_created_by = request.user if is_proxy_mode else None
            session.save()
            roster_result = _sync_expected_participants_from_shared_roster(session)
            seed_created_count = 0
            seed_skipped_count = 0
            proxy_created_count = 0
            proxy_skipped_count = 0
            proxy_participants = _parse_expected_participants_text(
                form.cleaned_data.get("proxy_participants_text", "")
            )
            if seed_participants and apply_sheetbook_participants:
                for participant in seed_participants:
                    _, was_created = ExpectedParticipant.objects.get_or_create(
                        training_session=session,
                        name=participant["name"],
                        affiliation=participant["affiliation"],
                    )
                    if was_created:
                        seed_created_count += 1
                    else:
                        seed_skipped_count += 1
            if proxy_participants:
                for participant in proxy_participants:
                    _, was_created = ExpectedParticipant.objects.get_or_create(
                        training_session=session,
                        name=participant["name"],
                        affiliation=participant["affiliation"],
                    )
                    if was_created:
                        proxy_created_count += 1
                    else:
                        proxy_skipped_count += 1
            copied_participant_count = 0
            if copy_source_session and not seed_participants:
                for participant in copy_source_participants:
                    _, was_created = ExpectedParticipant.objects.get_or_create(
                        training_session=session,
                        name=participant.name,
                        affiliation=participant.affiliation,
                    )
                    if was_created:
                        copied_participant_count += 1
            _sync_training_session_attachments(
                session,
                new_files=attachment_files,
                attachments_to_remove=[],
            )
            if sheetbook_seed:
                _pop_sheetbook_seed(
                    request,
                    sheetbook_seed_token,
                    expected_action="signature",
                )
            if draft_token:
                _pop_signature_create_draft(request, draft_token)
            _sync_calendar_event_for_training(session)
            message_parts = []
            if is_proxy_mode:
                message_parts.append(
                    f"{_get_signature_user_display_name(session_owner)} 선생님 대신 요청을 만들었습니다."
                )
            if roster_result["total"] > 0:
                message_parts.append(
                    f"연수가 생성되었습니다. 공유 명단 '{roster_result['group_name']}'에서 {roster_result['created']}명 가져왔습니다."
                )
            else:
                message_parts.append("연수가 생성되었습니다.")
            if seed_created_count > 0:
                message_parts.append(f"연결된 서비스에서 참석자 후보 {seed_created_count}명을 반영했습니다.")
            elif seed_participants and apply_sheetbook_participants and seed_skipped_count > 0:
                message_parts.append("연결된 서비스 참석자 후보는 이미 모두 포함되어 있었어요.")
            if proxy_created_count > 0:
                message_parts.append(f"교사가 보낸 명단 {proxy_created_count}명도 바로 넣었습니다.")
            elif proxy_participants and proxy_skipped_count > 0:
                message_parts.append("교사가 보낸 명단은 이미 모두 포함되어 있었어요.")
            if copied_participant_count > 0:
                message_parts.append(f"이전 요청 명단 {copied_participant_count}명도 복사했습니다.")
            if attachment_files:
                message_parts.append(f"첨부 파일 {len(attachment_files)}개도 함께 넣었습니다.")
            messages.success(request, " ".join(message_parts))
            return redirect('signatures:detail', uuid=session.uuid)
    else:
        form = TrainingSessionForm(
            owner=session_owner,
            initial=prefill_initial,
            can_delegate=can_proxy_create,
            delegate_user=request.user,
        )

    participant_preview = []
    preview_source = seed_participants
    if not preview_source and copy_source_participants:
        preview_source = copy_source_participants

    for participant in preview_source[:5]:
        if isinstance(participant, dict):
            participant_name = participant["name"]
            participant_affiliation = participant["affiliation"]
        else:
            participant_name = participant.name
            participant_affiliation = participant.display_affiliation
        if participant_affiliation:
            participant_preview.append(f"{participant_name} ({participant_affiliation})")
        else:
            participant_preview.append(participant_name)

    return render(
        request,
        'signatures/create.html',
        {
            'form': form,
            'draft_token': draft_token,
            'copy_from_uuid': copy_from_uuid if copy_source_session else "",
            'copy_source_session': copy_source_session,
            'sheetbook_seed_token': sheetbook_seed_token if seed_data else "",
            'sheetbook_prefill_active': bool(seed_data),
            'sheetbook_prefill_source_label': prefill_source_label if seed_data else "",
            'sheetbook_prefill_origin_label': prefill_origin_label,
            'sheetbook_prefill_origin_url': prefill_origin_url,
            'sheetbook_prefill_participants_count': len(seed_participants) or len(copy_source_participants),
            'sheetbook_prefill_participants_preview': participant_preview,
            'apply_sheetbook_participants': apply_sheetbook_participants,
            'can_proxy_create': can_proxy_create,
            'is_proxy_mode': is_proxy_mode,
            'proxy_target_user': proxy_target_user,
            'proxy_target_user_label': _get_signature_user_display_name(proxy_target_user),
            'has_roster_groups': form.fields["shared_roster_group"].queryset.exists(),
            'restored_roster_group': restored_roster_group,
            'existing_attachments': existing_attachments,
            'selected_remove_attachment_ids': selected_remove_attachment_ids,
            'attachment_limits': form.attachment_limits,
        },
    )


@login_required
def session_detail(request, uuid):
    """연수 상세 (관리자용) - 미매칭과 중복 점검 포함"""
    from django.http import HttpResponse
    import traceback

    try:
        session = _get_signature_session_or_404(request.user, uuid)
        signatures = session.signatures.all()
        expected = session.expected_participants.all()
        has_expected_participants = expected.exists()
        signature_sort_mode = _get_signature_sort_mode(session)
        participant_sort_mode = _get_participant_sort_mode(session)
        attendance_sort_mode = participant_sort_mode if has_expected_participants else signature_sort_mode
        attendance_sort_choices = _get_attendance_sort_choices(has_expected_participants)
        signature_rows = (
            _sort_signatures_for_display(signatures, signature_sort_mode)
            if not has_expected_participants
            else list(signatures.order_by("created_at", "id"))
        )

        suggestions = []
        if has_expected_participants:
            matched_sig_ids = expected.filter(
                matched_signature__isnull=False
            ).values_list('matched_signature_id', flat=True)

            unmatched_signatures = signatures.exclude(id__in=matched_sig_ids)

            for sig in unmatched_signatures:
                exact_matches = expected.filter(
                    name=sig.participant_name,
                    matched_signature__isnull=True
                )

                suggestions.append({
                    'signature': sig,
                    'exact_matches': list(exact_matches),
                    'has_matches': exact_matches.exists(),
                })

        sig_dict = defaultdict(list)
        for sig in signatures:
            key = (sig.participant_name, sig.display_affiliation or '')
            sig_dict[key].append(sig)

        duplicates = [sigs for sigs in sig_dict.values() if len(sigs) > 1]

        share_link = request.build_absolute_uri(
            reverse("signatures:sign", kwargs={"uuid": session.uuid})
        )
        share_qr_data_url = _build_qr_data_url(share_link)
        attachment_rows = _build_signature_attachment_rows(session)
        attachment_state = _summarize_attachment_state(_get_session_attachments(session))
        share_package_text = _build_signature_share_package_text(
            session,
            share_link,
            attachment_count=attachment_state["count"],
        )
        correction_logs = session.affiliation_correction_logs.select_related(
            "corrected_by",
            "signature",
            "expected_participant",
        )[:20]
        access_code_state = _build_access_code_state(session)
        unmatched_expected_participants = (
            _sort_expected_participants_for_display(
                expected.filter(matched_signature__isnull=True),
                participant_sort_mode,
            )
            if has_expected_participants
            else []
        )
        signature_count = signatures.count()
        pending_count = (
            len(unmatched_expected_participants)
            if has_expected_participants
            else max(0, int(session.expected_count or 0) - signature_count)
            if session.expected_count
            else None
        )
        stage_payload = _build_session_stage_payload(
            session,
            signature_count=signature_count,
            pending_count=pending_count,
            can_show_absentees=has_expected_participants,
        )

        return render(request, 'signatures/detail.html', {
            'session': session,
            'signatures': signatures,
            'signature_count': signature_count,
            'signature_rows': signature_rows,
            'attendance_sort_mode': attendance_sort_mode,
            'attendance_sort_choices': attendance_sort_choices,
            'participant_sort_mode': participant_sort_mode,
            'signature_sort_mode': signature_sort_mode,
            'can_customize_signature_sort': not has_expected_participants,
            'can_customize_participant_sort': has_expected_participants,
            'expected_participants': expected,
            'unmatched_expected_participants': unmatched_expected_participants,
            'unmatched_suggestions': suggestions,
            'duplicates': duplicates,
            'has_unmatched': len(suggestions) > 0,
            'has_duplicates': len(duplicates) > 0,
            'share_link': share_link,
            'share_qr_data_url': share_qr_data_url,
            'share_package_text': share_package_text,
            'attachment_rows': attachment_rows,
            'attachment_state': attachment_state,
            'access_code_state': access_code_state,
            'affiliation_suggestions': _build_affiliation_suggestions(session),
            'affiliation_correction_logs': correction_logs,
            'pending_participants_preview': unmatched_expected_participants[:5],
            **stage_payload,
        })
    except Exception as e:
        traceback.print_exc()
        return HttpResponse(f"Server Error in session_detail: {str(e)}<br><pre>{traceback.format_exc()}</pre>", status=500)


@login_required
def session_edit(request, uuid):
    """연수 수정"""
    session = _get_signature_session_or_404(request.user, uuid)
    existing_attachments = _get_session_attachments(session)
    selected_remove_attachment_ids = []
    if request.method == 'POST':
        selected_remove_attachment_ids = request.POST.getlist("remove_attachment_ids")
        attachments_to_remove = _resolve_attachment_removals(session, selected_remove_attachment_ids)
        form = TrainingSessionForm(request.POST, request.FILES, instance=session, owner=session.created_by)
        form_valid = form.is_valid()
        attachment_files, attachment_file_errors = validate_training_session_attachment_files(
            request.FILES.getlist("attachments")
        )
        attachment_batch_errors, _ = _validate_session_attachment_batch(
            session,
            attachment_files,
            attachments_to_remove,
        )
        for error in attachment_file_errors + attachment_batch_errors:
            form.add_error("attachments", error)

        if form_valid and not form.errors:
            session = form.save()
            _sync_training_session_attachments(
                session,
                new_files=attachment_files,
                attachments_to_remove=attachments_to_remove,
            )
            roster_result = _sync_expected_participants_from_shared_roster(session)
            _sync_calendar_event_for_training(session)
            message_parts = []
            if roster_result["total"] > 0 and roster_result["created"] > 0:
                message_parts.append(
                    f"연수 정보가 수정되었습니다. 공유 명단에서 {roster_result['created']}명을 추가 반영했습니다."
                )
            else:
                message_parts.append("연수 정보가 수정되었습니다.")
            if attachment_files:
                message_parts.append(f"첨부 파일 {len(attachment_files)}개를 추가했습니다.")
            if attachments_to_remove:
                message_parts.append(f"첨부 파일 {len(attachments_to_remove)}개를 제거했습니다.")
            messages.success(request, " ".join(message_parts))
            return redirect('signatures:detail', uuid=session.uuid)
    else:
        form = TrainingSessionForm(instance=session, owner=session.created_by)
    return render(
        request,
        'signatures/edit.html',
        {
            'form': form,
            'session': session,
            'existing_attachments': existing_attachments,
            'selected_remove_attachment_ids': [str(item) for item in selected_remove_attachment_ids],
            'attachment_limits': form.attachment_limits,
        },
    )


@login_required
@require_POST
def sync_expected_participants_from_roster(request, uuid):
    """연결된 공유 명단을 예상 참석자 목록으로 다시 가져오기."""
    session = _get_signature_session_or_404(request.user, uuid)
    if not session.shared_roster_group:
        messages.error(request, "먼저 연수 수정에서 공유 명단을 선택해 주세요.")
        return redirect("signatures:detail", uuid=session.uuid)

    result = _sync_expected_participants_from_shared_roster(session)
    if result["total"] == 0:
        messages.warning(request, "공유 명단에 가져올 활성 멤버가 없습니다.")
    else:
        messages.success(
            request,
            f"공유 명단 '{result['group_name']}' 동기화 완료: {result['created']}명 추가, {result['skipped']}명 중복/제외",
        )
    return redirect("signatures:detail", uuid=session.uuid)


@login_required
def session_delete(request, uuid):
    """연수 삭제"""
    session = _get_signature_session_or_404(request.user, uuid)
    if request.method == 'POST':
        _delete_calendar_event_for_training(session)
        session.delete()
        messages.success(request, '연수가 삭제되었습니다.')
        return redirect('signatures:list')
    return render(request, 'signatures/delete_confirm.html', {'session': session})


@ratelimit(key=_signature_public_ratelimit_key, rate="120/10m", method="POST", block=True, group="signatures_public_sign")
def sign(request, uuid):
    """서명 페이지 (공개 - 로그인 불필요)"""
    session = get_object_or_404(TrainingSession, uuid=uuid)
    affiliation_suggestions = _build_affiliation_suggestions(session)
    request_time = timezone.now()
    public_attachments = _build_signature_attachment_rows(session, public=True)
    roster_participant_options = _sort_expected_participants_for_display(
        session.expected_participants.all(),
        _get_participant_sort_mode(session),
    )
    show_roster_selection = bool(roster_participant_options)
    access_code_state = _build_access_code_state(session, now=request_time)
    walk_in_mode = (
        _is_truthy_flag(request.POST.get("walk_in_mode"), default=False)
        if request.method == "POST" and show_roster_selection
        else False
    )
    selected_expected_participant_id = (
        request.POST.get("expected_participant_id")
        if request.method == "POST"
        else ""
    ) or ""

    if not session.is_active:
        response = render(request, 'signatures/closed.html', {'session': session})
        return _apply_sensitive_cache_headers(response)

    if request.method == 'POST':
        form = SignatureForm(
            request.POST,
            use_roster_selection=show_roster_selection,
            use_access_code=access_code_state["is_active"],
        )
        if form.is_valid():
            selected_participant = None
            selected_participant_id = form.cleaned_data.get("expected_participant_id")
            access_code_error = _validate_session_access_code(
                session,
                form.cleaned_data.get("access_code"),
                now=request_time,
            )
            if access_code_error:
                if access_code_state["is_active"]:
                    form.add_error("access_code", access_code_error)
                else:
                    form.add_error(None, access_code_error)

            if not form.errors and show_roster_selection and selected_participant_id and not walk_in_mode:
                selected_participant = (
                    ExpectedParticipant.objects
                    .filter(training_session=session, id=selected_participant_id)
                    .first()
                )
                if selected_participant is None:
                    form.add_error(None, "선택한 이름을 다시 확인해 주세요.")
                elif selected_participant.matched_signature_id:
                    form.add_error(None, "이미 서명이 완료된 이름입니다. 다른 이름을 선택해 주세요.")

            if not form.errors:
                with transaction.atomic():
                    if show_roster_selection and selected_participant_id and not walk_in_mode:
                        selected_participant = get_object_or_404(
                            ExpectedParticipant.objects.select_for_update(),
                            training_session=session,
                            id=selected_participant_id,
                        )
                        if selected_participant.matched_signature_id:
                            form.add_error(None, "이미 서명이 완료된 이름입니다. 다른 이름을 선택해 주세요.")

                    if not form.errors:
                        signature = form.save(commit=False)
                        signature.training_session = session
                        if selected_participant is not None:
                            signature.participant_name = selected_participant.name
                            signature.participant_affiliation = _normalize_affiliation_text(
                                selected_participant.display_affiliation or selected_participant.affiliation
                            )
                        else:
                            signature.participant_affiliation = _normalize_affiliation_text(signature.participant_affiliation)
                        signature.submission_mode = Signature.SUBMISSION_MODE_OPEN
                        signature.ip_address = _request_client_ip(request)
                        signature.user_agent = _request_user_agent(request)
                        signature.save()

                        if selected_participant is not None:
                            selected_participant.matched_signature = signature
                            selected_participant.is_confirmed = True
                            selected_participant.save(update_fields=["matched_signature", "is_confirmed"])

                        SignatureAuditLog.objects.create(
                            training_session=session,
                            signature=signature,
                            expected_participant=selected_participant,
                            event_type=SignatureAuditLog.EVENT_SIGN_SUBMITTED,
                            event_meta={
                                'participant_name': signature.participant_name,
                                'participant_affiliation': signature.display_affiliation,
                                'submission_mode': signature.submission_mode,
                                'expected_participant_id': selected_participant.id if selected_participant else None,
                                'walk_in_mode': bool(selected_participant is None),
                                'access_code_required': access_code_state["enabled"],
                                'access_code_verified': access_code_state["enabled"],
                                'access_code_duration_minutes': access_code_state["duration_minutes"] or None,
                            },
                            ip_address=signature.ip_address,
                            user_agent=signature.user_agent,
                        )
                        response = render(request, 'signatures/sign_success.html', {'session': session})
                        return _apply_sensitive_cache_headers(response)
    else:
        form = SignatureForm(
            use_roster_selection=show_roster_selection,
            use_access_code=access_code_state["is_active"],
        )

    response = render(request, 'signatures/sign.html', {
        'session': session,
        'form': form,
        'affiliation_suggestions': affiliation_suggestions,
        'access_code_state': access_code_state,
        'show_roster_selection': show_roster_selection,
        'roster_participant_options': roster_participant_options,
        'walk_in_mode': walk_in_mode,
        'selected_expected_participant_id': str(selected_expected_participant_id),
        'public_attachments': public_attachments,
    })
    return _apply_sensitive_cache_headers(response)


@login_required
def session_attachment_download(request, uuid, attachment_id):
    session = _get_signature_session_or_404(request.user, uuid)
    attachment = get_object_or_404(
        TrainingSessionAttachment,
        training_session=session,
        id=attachment_id,
    )
    return _build_attachment_download_response(attachment)


@ratelimit(
    key=_signature_public_ratelimit_key,
    rate="60/10m",
    method="GET",
    block=True,
    group="signatures_public_attachment_download",
)
def sign_attachment_download(request, uuid, attachment_id):
    """공개 서명 페이지에서 첨부 파일 다운로드."""
    session = get_object_or_404(TrainingSession, uuid=uuid)
    if not session.is_active:
        raise Http404("첨부 파일을 찾을 수 없습니다.")
    attachment = get_object_or_404(
        TrainingSessionAttachment,
        training_session=session,
        id=attachment_id,
    )
    return _build_attachment_download_response(attachment)


@login_required
def print_view(request, uuid):
    """출석부 인쇄 페이지 - 명단 유무에 따라 동작 변경"""
    session = _get_signature_session_or_404(request.user, uuid)
    signature_sort_mode = _get_signature_sort_mode(session)
    participant_sort_mode = _get_participant_sort_mode(session)
    
    # 데이터 준비
    print_items = []
    signed_count = 0
    
    if session.expected_participants.exists():
        # Case A: 명단이 있는 경우 (Phase 2) -> 명단 기준 + 미매칭 서명
        participants = _sort_expected_participants_for_display(
            session.expected_participants.all(),
            participant_sort_mode,
        )
        
        # 1. 예상 참석자 추가
        for p in participants:
            item = {
                'name': p.name,
                'affiliation': p.display_affiliation,
                'original_affiliation': p.affiliation,
                'is_affiliation_corrected': bool(p.corrected_affiliation),
                'signature_data': p.matched_signature.signature_data if p.matched_signature else None,
            }
            print_items.append(item)
            if item['signature_data']:
                signed_count += 1
                
        # 2. 명단에 없는 추가 서명(Walk-ins) 추가
        matched_sig_ids = [p.matched_signature.id for p in participants if p.matched_signature]
        unmatched_sigs = _sort_signatures_for_display(
            session.signatures.exclude(id__in=matched_sig_ids),
            signature_sort_mode,
        )
        
        for sig in unmatched_sigs:
            print_items.append({
                'name': sig.participant_name,
                'affiliation': sig.display_affiliation,
                'original_affiliation': sig.participant_affiliation,
                'is_affiliation_corrected': bool(sig.corrected_affiliation),
                'signature_data': sig.signature_data,
            })
            signed_count += 1
            
        total_expected = session.expected_count or len(participants)
        
    else:
        # Case B: 명단이 없는 경우 (Phase 1) -> 서명 기준
        signatures = _sort_signatures_for_display(session.signatures.all(), signature_sort_mode)
        for sig in signatures:
            print_items.append({
                'name': sig.participant_name,
                'affiliation': sig.display_affiliation,
                'original_affiliation': sig.participant_affiliation,
                'is_affiliation_corrected': bool(sig.corrected_affiliation),
                'signature_data': sig.signature_data,
            })
        signed_count = len(print_items)
        total_expected = session.expected_count or signed_count
    
    # 페이지네이션 처리
    total_items = len(print_items)
    SIGS_PER_PAGE = 60
    pages = []
    
    for page_num in range(0, total_items, SIGS_PER_PAGE):
        # 이번 페이지의 아이템들 (최대 60개)
        page_items = print_items[page_num:page_num + SIGS_PER_PAGE]
        
        # 좌우 분할 (30개씩)
        left_items = page_items[:30]
        right_items = page_items[30:60]
        
        # 빈 줄 채우기 (항상 30줄이 되도록)
        # left_rows/right_rows는 순번만 계산
        current_base_idx = page_num
        
        pages.append({
            'page_number': (page_num // SIGS_PER_PAGE) + 1,
            'left_items': left_items,
            'right_items': right_items,
            'left_start_index': current_base_idx + 1,
            'right_start_index': current_base_idx + 31,
            'left_padding': range(30 - len(left_items)),
            'right_padding': range(30 - len(right_items)),
        })
    
    # 페이지가 하나도 없으면 빈 페이지 하나 생성
    if not pages:
        pages.append({
            'page_number': 1,
            'left_items': [], 'right_items': [],
            'left_start_index': 1, 'right_start_index': 31,
            'left_padding': range(30), 'right_padding': range(30)
        })

    response = render(request, 'signatures/print_view.html', {
        'session': session,
        'pages': pages,
        'total_count': total_expected,
        'signed_count': signed_count,
        'unsigned_count': max(0, total_expected - signed_count),
        'total_pages': len(pages),
        'signature_sort_mode': signature_sort_mode,
    })
    return _apply_sensitive_cache_headers(response)


@login_required
@require_POST
def update_signature_sort_mode(request, uuid):
    """명단 유무에 따라 출석부 정렬 방식을 변경."""
    session = _get_signature_session_or_404(request.user, uuid)
    has_expected_participants = session.expected_participants.exists()

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    sort_mode = str(payload.get("sort_mode") or "").strip()
    valid_modes = {choice[0] for choice in TrainingSession.SIGNATURE_SORT_CHOICES}
    if sort_mode not in valid_modes:
        return JsonResponse({'success': False, 'error': '정렬 방식을 다시 확인해 주세요.'}, status=400)

    if has_expected_participants:
        if sort_mode == TrainingSession.SIGNATURE_SORT_MANUAL:
            _resequence_manual_participant_order(session)
        if session.participant_sort_mode != sort_mode:
            session.participant_sort_mode = sort_mode
            session.save(update_fields=["participant_sort_mode"])
        ordered_items = _sort_expected_participants_for_display(
            session.expected_participants.all(),
            sort_mode,
        )
        ordered_ids = [participant.id for participant in ordered_items]
    else:
        if sort_mode == TrainingSession.SIGNATURE_SORT_MANUAL:
            _resequence_manual_signature_order(session)
        if session.signature_sort_mode != sort_mode:
            session.signature_sort_mode = sort_mode
            session.save(update_fields=["signature_sort_mode"])
        ordered_items = _sort_signatures_for_display(session.signatures.all(), sort_mode)
        ordered_ids = [signature.id for signature in ordered_items]

    return JsonResponse({
        'success': True,
        'sort_mode': sort_mode,
        'ordered_ids': ordered_ids,
        'uses_participants': has_expected_participants,
    })


@login_required
@require_POST
def update_signature_manual_order(request, uuid, signature_id):
    """명단 없는 연수에서 수동 순서를 저장."""
    session = _get_signature_session_or_404(request.user, uuid)
    if session.expected_participants.exists():
        return JsonResponse(
            {'success': False, 'error': '명단이 있는 연수는 명단 기준으로 출력됩니다.'},
            status=400,
        )

    signature = get_object_or_404(Signature, id=signature_id, training_session=session)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    raw_position = payload.get("position")
    try:
        target_position = int(raw_position)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': '이동할 순번을 숫자로 입력해 주세요.'}, status=400)

    if target_position < 1:
        return JsonResponse({'success': False, 'error': '순번은 1 이상이어야 합니다.'}, status=400)

    with transaction.atomic():
        ordered_signatures = _move_signature_to_position(session, signature, target_position)
        if session.signature_sort_mode != TrainingSession.SIGNATURE_SORT_MANUAL:
            session.signature_sort_mode = TrainingSession.SIGNATURE_SORT_MANUAL
            session.save(update_fields=["signature_sort_mode"])

    return JsonResponse({
        'success': True,
        'sort_mode': TrainingSession.SIGNATURE_SORT_MANUAL,
        'ordered_signature_ids': [item.id for item in ordered_signatures],
        'ordered_signature_names': [item.participant_name for item in ordered_signatures],
    })


@login_required
@require_POST
def update_expected_participant_manual_order(request, uuid, participant_id):
    """명단이 있는 연수에서 참석자 수동 순서를 저장."""
    session = _get_signature_session_or_404(request.user, uuid)
    participant = get_object_or_404(ExpectedParticipant, id=participant_id, training_session=session)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    raw_position = payload.get("position")
    try:
        target_position = int(raw_position)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': '이동할 순번을 숫자로 입력해 주세요.'}, status=400)

    if target_position < 1:
        return JsonResponse({'success': False, 'error': '순번은 1 이상이어야 합니다.'}, status=400)

    with transaction.atomic():
        ordered_participants = _move_participant_to_position(session, participant, target_position)
        if session.participant_sort_mode != TrainingSession.SIGNATURE_SORT_MANUAL:
            session.participant_sort_mode = TrainingSession.SIGNATURE_SORT_MANUAL
            session.save(update_fields=["participant_sort_mode"])

    return JsonResponse({
        'success': True,
        'sort_mode': TrainingSession.SIGNATURE_SORT_MANUAL,
        'ordered_participant_ids': [item.id for item in ordered_participants],
        'ordered_participant_names': [item.name for item in ordered_participants],
    })


@login_required
@require_POST
def toggle_active(request, uuid):
    """서명 받기 활성화/비활성화 토글 (AJAX)"""
    session = _get_signature_session_or_404(request.user, uuid)
    session.is_active = not session.is_active
    session.save()
    return JsonResponse({
        'success': True,
        'is_active': session.is_active,
    })


@login_required
@require_POST
def update_access_code(request, uuid):
    """현장 코드 사용 여부와 현재 코드를 업데이트."""
    session = _get_signature_session_or_404(request.user, uuid)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    raw_duration = payload.get("duration_minutes", TrainingSession.ACCESS_CODE_DISABLED)
    try:
        duration_minutes = int(raw_duration or 0)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': '코드 시간을 다시 선택해 주세요.'}, status=400)

    if duration_minutes not in {
        TrainingSession.ACCESS_CODE_DISABLED,
        TrainingSession.ACCESS_CODE_5_MINUTES,
        TrainingSession.ACCESS_CODE_10_MINUTES,
    }:
        return JsonResponse({'success': False, 'error': '코드 시간은 5분 또는 10분만 사용할 수 있습니다.'}, status=400)

    if duration_minutes == TrainingSession.ACCESS_CODE_DISABLED:
        session.access_code_duration_minutes = TrainingSession.ACCESS_CODE_DISABLED
        session.active_access_code = ""
        session.active_access_code_expires_at = None
        session.save(
            update_fields=[
                "access_code_duration_minutes",
                "active_access_code",
                "active_access_code_expires_at",
            ]
        )
        return JsonResponse({
            'success': True,
            'message': '현장 코드를 사용하지 않도록 변경했습니다.',
            'access_code_state': _build_access_code_state(session),
        })

    access_code = _normalize_access_code(payload.get("access_code"))
    if not re.fullmatch(r"\d{4,6}", access_code):
        return JsonResponse({'success': False, 'error': '현장 코드는 4~6자리 숫자로 입력해 주세요.'}, status=400)

    session.access_code_duration_minutes = duration_minutes
    session.active_access_code = access_code
    session.active_access_code_expires_at = timezone.now() + timedelta(minutes=duration_minutes)
    session.save(
        update_fields=[
            "access_code_duration_minutes",
            "active_access_code",
            "active_access_code_expires_at",
        ]
    )
    return JsonResponse({
        'success': True,
        'message': f'현장 코드를 {duration_minutes}분 동안 적용했습니다.',
        'access_code_state': _build_access_code_state(session),
    })


@login_required
@require_POST
def delete_signature(request, pk):
    """개별 서명 삭제 (AJAX)"""
    signature = get_object_or_404(
        Signature,
        pk=pk,
        training_session__in=_get_signature_accessible_sessions(request.user),
    )
    session = signature.training_session
    signature.delete()
    if not session.expected_participants.exists():
        _resequence_manual_signature_order(session)
    return JsonResponse({'success': True})


@login_required
def style_list(request):
    """내 서명 보관함"""
    from .models import SavedSignature, SignatureStyle

    saved_signatures = SavedSignature.objects.filter(user=request.user).order_by("-created_at")[:12]
    styles = SignatureStyle.objects.filter(user=request.user)
    return render(
        request,
        'signatures/style_list.html',
        {
            'saved_signatures': saved_signatures,
            'styles': styles,
        },
    )


@login_required
@require_POST
def save_style_api(request):
    """스타일 즐겨찾기 저장 API"""
    try:
        data = json.loads(request.body)
        from .models import SignatureStyle

        SignatureStyle.objects.create(
            user=request.user,
            name=data.get('name', '내 서명 스타일'),
            font_family=data.get('font_family'),
            color=data.get('color'),
            background_color=data.get('background_color')
        )

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_POST
def save_signature_image_api(request):
    """서명 이미지 저장 API (스타일 없이 이미지만)"""
    try:
        data = json.loads(request.body)
        from .models import SavedSignature

        image_data = _clean_saved_signature_image_data(data.get('image_data'))
        saved_signature = SavedSignature.objects.filter(user=request.user, image_data=image_data).first()
        created = False
        if saved_signature is None:
            saved_signature = SavedSignature.objects.create(
                user=request.user,
                image_data=image_data,
            )
            created = True

        return JsonResponse(
            {
                'success': True,
                'created': created,
                'message': '손으로 쓴 서명을 보관함에 저장했습니다.' if created else '같은 손서명이 이미 보관함에 있습니다.',
                'signature': _serialize_saved_signature(saved_signature),
            }
        )
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def get_my_signatures_api(request):
    """내 저장된 서명 이미지 목록 가져오기"""
    from .models import SavedSignature

    signatures = SavedSignature.objects.filter(user=request.user).order_by('-created_at')[:8]
    data = [_serialize_saved_signature(sig) for sig in signatures]
    return JsonResponse({'signatures': data})


@login_required
@require_POST
def delete_saved_signature_api(request, pk):
    """손서명 보관함 항목 삭제"""
    from .models import SavedSignature

    signature = get_object_or_404(SavedSignature, pk=pk, user=request.user)
    signature.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def delete_style_api(request, pk):
    """스타일 삭제"""
    from .models import SignatureStyle
    style = get_object_or_404(SignatureStyle, pk=pk, user=request.user)
    style.delete()
    return JsonResponse({'success': True})


def signature_maker(request):
    """전자 서명 제작 도구 (비회원 개방)"""
    # 추천 폰트 리스트
    fonts = [
        'Nanum Brush Script', 'Nanum Pen Script', 'Cafe24 Ssurround Air', 
        'Gowun Batang', 'Gamja Flower', 'Poor Story'
    ]
    selected_font = request.GET.get("font", "").strip()
    if selected_font not in fonts:
        selected_font = fonts[0]

    selected_color = request.GET.get("color", "").strip()
    if not re.fullmatch(r"#[0-9a-fA-F]{6}", selected_color):
        selected_color = "#000000"

    initial_name = (
        request.user.first_name.strip()
        if request.user.is_authenticated and request.user.first_name
        else "스쿨잇"
    )
    return render(request, 'signatures/maker.html', {
        'fonts': fonts,
        'is_guest': not request.user.is_authenticated,
        'selected_font': selected_font,
        'selected_color': selected_color,
        'initial_name': initial_name,
    })


# ===== Phase 2: Expected Participants Management =====

@login_required
@require_POST
def add_expected_participants(request, uuid):
    """예상 참석자 명단 일괄 등록"""
    from .models import ExpectedParticipant
    
    session = _get_signature_session_or_404(request.user, uuid)
    
    try:
        data = json.loads(request.body)
        participants_text = data.get('participants', '')
        
        if not participants_text.strip():
            return JsonResponse({'success': False, 'error': '명단이 비어있습니다.'})
        
        # Parse input (format: "이름, 소속" or "이름")
        lines = participants_text.strip().split('\n')
        created_count = 0
        skipped_count = 0
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            parts = [p.strip() for p in line.split(',')]
            name = parts[0] if parts else ''
            affiliation = _normalize_affiliation_text(parts[1] if len(parts) > 1 else '')
            
            if not name:
                skipped_count += 1
                continue
            
            # Create or skip if duplicate
            _, created = ExpectedParticipant.objects.get_or_create(
                training_session=session,
                name=name,
                affiliation=affiliation
            )
            
            if created:
                created_count += 1
            else:
                skipped_count += 1
        
        return JsonResponse({
            'success': True,
            'created': created_count,
            'skipped': skipped_count,
            'message': f'{created_count}명 등록 완료 (중복 {skipped_count}명 제외)'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def upload_participants_file(request, uuid):
    """파일(CSV, XLSX)을 통한 명단 등록"""
    from .models import ExpectedParticipant
    import csv
    import io
    
    session = _get_signature_session_or_404(request.user, uuid)
    file_obj = request.FILES.get('file')
    
    if not file_obj:
        return JsonResponse({'success': False, 'error': '파일이 없습니다.'})
    
    file_name = file_obj.name.lower()
    participants = []
    
    try:
        if file_name.endswith('.csv'):
            # CSV 처리
            decoded_file = file_obj.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file)
            for row in reader:
                if row:
                    name = row[0].strip()
                    affiliation = _normalize_affiliation_text(row[1] if len(row) > 1 else '')
                    if name: participants.append((name, affiliation))
                    
        elif file_name.endswith('.xlsx'):
            # Excel 처리
            import openpyxl
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row and row[0]:
                    name = str(row[0]).strip()
                    affiliation = _normalize_affiliation_text(row[1] if len(row) > 1 and row[1] else '')
                    if name: participants.append((name, affiliation))
        else:
            return JsonResponse({'success': False, 'error': '참석자 명단 파일(.csv, .xlsx)만 업로드 가능합니다.'})
        
        # 데이터 저장
        created_count = 0
        skipped_count = 0
        for name, affiliation in participants:
            _, created = ExpectedParticipant.objects.get_or_create(
                training_session=session,
                name=name,
                affiliation=affiliation
            )
            if created: created_count += 1
            else: skipped_count += 1
            
        return JsonResponse({
            'success': True,
            'created': created_count,
            'skipped': skipped_count,
            'message': f'{created_count}명 등록 완료 (중복 {skipped_count}명 제외)'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'파일 처리 중 오류: {str(e)}'})


@login_required
def get_expected_participants(request, uuid):
    """예상 참석자 목록 조회 (JSON)"""
    from .models import ExpectedParticipant
    
    session = _get_signature_session_or_404(request.user, uuid)
    participant_sort_mode = _get_participant_sort_mode(session)
    participants = _sort_expected_participants_for_display(
        session.expected_participants.all(),
        participant_sort_mode,
    )
    
    data = []
    for index, p in enumerate(participants, start=1):
        data.append({
            'id': p.id,
            'name': p.name,
            'affiliation': p.display_affiliation,
            'display_affiliation': p.display_affiliation,
            'original_affiliation': p.affiliation,
            'corrected_affiliation': p.corrected_affiliation,
            'has_signed': p.has_signed,
            'signature_id': p.matched_signature.id if p.matched_signature else None,
            'match_note': p.match_note,
            'position': index,
        })
    
    return JsonResponse({
        'participants': data,
        'sort_mode': participant_sort_mode,
    })


@login_required
@require_POST
def delete_expected_participant(request, uuid, participant_id):
    """예상 참석자 삭제"""
    from .models import ExpectedParticipant
    
    session = _get_signature_session_or_404(request.user, uuid)
    participant = get_object_or_404(
        ExpectedParticipant,
        id=participant_id,
        training_session=session
    )
    participant.delete()
    _resequence_manual_participant_order(session)
    
    return JsonResponse({'success': True})


@login_required
@require_POST
def match_signature(request, uuid, signature_id):
    """서명을 예상 참석자와 수동으로 연결"""
    from .models import ExpectedParticipant
    import json
    
    session = _get_signature_session_or_404(request.user, uuid)
    signature = get_object_or_404(Signature, id=signature_id, training_session=session)
    
    try:
        data = json.loads(request.body)
        participant_id = data.get('participant_id')
        
        if not participant_id:
            return JsonResponse({'success': False, 'error': '참석자 ID가 필요합니다.'})
        
        participant = get_object_or_404(
            ExpectedParticipant,
            id=participant_id,
            training_session=session
        )
        
        # 기존 매칭 해제 (다른 서명과 연결되어 있었다면)
        if participant.matched_signature:
            return JsonResponse({
                'success': False,
                'error': f'{participant.name}은(는) 이미 다른 서명과 연결되어 있습니다.'
            })
        
        # 매칭 설정
        participant.matched_signature = signature
        participant.is_confirmed = True
        participant.save()

        return JsonResponse({
            'success': True,
            'message': f'{signature.participant_name} → {participant.name} 연결 완료'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def correct_signature_affiliation(request, uuid, signature_id):
    """개별 서명의 직위/학년반 정정."""
    session = _get_signature_session_or_404(request.user, uuid)
    signature = get_object_or_404(Signature, id=signature_id, training_session=session)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    corrected_affiliation = payload.get('corrected_affiliation', '')
    reason = payload.get('reason', '')
    before_display_affiliation = signature.display_affiliation
    before_state = (
        signature.corrected_affiliation,
        signature.affiliation_correction_reason,
        signature.affiliation_corrected_by_id,
        signature.affiliation_corrected_at,
    )
    _apply_signature_affiliation_correction(signature, corrected_affiliation, reason, request.user)
    after_state = (
        signature.corrected_affiliation,
        signature.affiliation_correction_reason,
        signature.affiliation_corrected_by_id,
        signature.affiliation_corrected_at,
    )
    if before_state != after_state:
        signature.save(
            update_fields=[
                'corrected_affiliation',
                'affiliation_correction_reason',
                'affiliation_corrected_by',
                'affiliation_corrected_at',
            ]
        )
        _create_affiliation_correction_log(
            session=session,
            target_type=AffiliationCorrectionLog.TARGET_SIGNATURE,
            mode=AffiliationCorrectionLog.MODE_SINGLE,
            before_affiliation=before_display_affiliation,
            after_affiliation=signature.display_affiliation,
            reason=reason,
            corrected_by=request.user,
            signature=signature,
        )
    return JsonResponse({
        'success': True,
        'display_affiliation': signature.display_affiliation,
        'original_affiliation': signature.participant_affiliation,
        'corrected_affiliation': signature.corrected_affiliation,
    })


@login_required
@require_POST
def correct_expected_participant_affiliation(request, uuid, participant_id):
    """개별 예상 참석자의 직위/학년반 정정."""
    session = _get_signature_session_or_404(request.user, uuid)
    participant = get_object_or_404(ExpectedParticipant, id=participant_id, training_session=session)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    corrected_affiliation = payload.get('corrected_affiliation', '')
    reason = payload.get('reason', '')
    before_display_affiliation = participant.display_affiliation
    before_state = (
        participant.corrected_affiliation,
        participant.affiliation_correction_reason,
        participant.affiliation_corrected_by_id,
        participant.affiliation_corrected_at,
    )
    _apply_expected_participant_affiliation_correction(participant, corrected_affiliation, reason, request.user)
    after_state = (
        participant.corrected_affiliation,
        participant.affiliation_correction_reason,
        participant.affiliation_corrected_by_id,
        participant.affiliation_corrected_at,
    )
    if before_state != after_state:
        participant.save(
            update_fields=[
                'corrected_affiliation',
                'affiliation_correction_reason',
                'affiliation_corrected_by',
                'affiliation_corrected_at',
            ]
        )
        _create_affiliation_correction_log(
            session=session,
            target_type=AffiliationCorrectionLog.TARGET_PARTICIPANT,
            mode=AffiliationCorrectionLog.MODE_SINGLE,
            before_affiliation=before_display_affiliation,
            after_affiliation=participant.display_affiliation,
            reason=reason,
            corrected_by=request.user,
            expected_participant=participant,
        )
    return JsonResponse({
        'success': True,
        'display_affiliation': participant.display_affiliation,
        'original_affiliation': participant.affiliation,
        'corrected_affiliation': participant.corrected_affiliation,
    })


@login_required
@require_POST
def bulk_correct_affiliation(request, uuid):
    """직위/학년반 일괄 정정."""
    session = _get_signature_session_or_404(request.user, uuid)

    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '요청 형식이 올바르지 않습니다.'}, status=400)

    source_affiliation = _normalize_affiliation_text(payload.get('source_affiliation', ''))
    corrected_affiliation = _normalize_affiliation_text(payload.get('corrected_affiliation', ''))
    reason = str(payload.get('reason') or '').strip()[:200] or '일괄 정정'
    target = str(payload.get('target') or 'all').strip()

    if not source_affiliation:
        return JsonResponse({'success': False, 'error': '원본 직위/학년반을 입력해 주세요.'}, status=400)
    if not corrected_affiliation:
        return JsonResponse({'success': False, 'error': '정정할 직위/학년반을 입력해 주세요.'}, status=400)
    if target not in {'all', 'participants', 'signatures'}:
        return JsonResponse({'success': False, 'error': '정정 대상을 확인해 주세요.'}, status=400)

    updated_signatures = 0
    updated_participants = 0

    if target in {'all', 'signatures'}:
        signatures = session.signatures.all()
        for signature in signatures:
            if _normalize_affiliation_text(signature.display_affiliation) != source_affiliation:
                continue
            before_display_affiliation = signature.display_affiliation
            before_state = (
                signature.corrected_affiliation,
                signature.affiliation_correction_reason,
                signature.affiliation_corrected_by_id,
                signature.affiliation_corrected_at,
            )
            _apply_signature_affiliation_correction(signature, corrected_affiliation, reason, request.user)
            after_state = (
                signature.corrected_affiliation,
                signature.affiliation_correction_reason,
                signature.affiliation_corrected_by_id,
                signature.affiliation_corrected_at,
            )
            if before_state == after_state:
                continue
            signature.save(
                update_fields=[
                    'corrected_affiliation',
                    'affiliation_correction_reason',
                    'affiliation_corrected_by',
                    'affiliation_corrected_at',
                ]
            )
            _create_affiliation_correction_log(
                session=session,
                target_type=AffiliationCorrectionLog.TARGET_SIGNATURE,
                mode=AffiliationCorrectionLog.MODE_BULK,
                before_affiliation=before_display_affiliation,
                after_affiliation=signature.display_affiliation,
                reason=reason,
                corrected_by=request.user,
                signature=signature,
            )
            updated_signatures += 1

    if target in {'all', 'participants'}:
        participants = session.expected_participants.all()
        for participant in participants:
            if _normalize_affiliation_text(participant.display_affiliation) != source_affiliation:
                continue
            before_display_affiliation = participant.display_affiliation
            before_state = (
                participant.corrected_affiliation,
                participant.affiliation_correction_reason,
                participant.affiliation_corrected_by_id,
                participant.affiliation_corrected_at,
            )
            _apply_expected_participant_affiliation_correction(participant, corrected_affiliation, reason, request.user)
            after_state = (
                participant.corrected_affiliation,
                participant.affiliation_correction_reason,
                participant.affiliation_corrected_by_id,
                participant.affiliation_corrected_at,
            )
            if before_state == after_state:
                continue
            participant.save(
                update_fields=[
                    'corrected_affiliation',
                    'affiliation_correction_reason',
                    'affiliation_corrected_by',
                    'affiliation_corrected_at',
                ]
            )
            _create_affiliation_correction_log(
                session=session,
                target_type=AffiliationCorrectionLog.TARGET_PARTICIPANT,
                mode=AffiliationCorrectionLog.MODE_BULK,
                before_affiliation=before_display_affiliation,
                after_affiliation=participant.display_affiliation,
                reason=reason,
                corrected_by=request.user,
                expected_participant=participant,
            )
            updated_participants += 1

    return JsonResponse({
        'success': True,
        'updated_signatures': updated_signatures,
        'updated_participants': updated_participants,
        'updated_total': updated_signatures + updated_participants,
    })


def download_participant_template(request, format='csv'):
    """예상 참석자 명단 양식 다운로드 (CSV 또는 Excel)"""

    if format == 'csv':
        # CSV 파일 생성
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="참석자명단_양식.csv"'

        writer = csv.writer(response)
        writer.writerow(['이름', '소속/학년반'])
        writer.writerow(['홍길동', '1-1'])
        writer.writerow(['김철수', '1-2'])
        writer.writerow(['박영희', '교사'])
        writer.writerow(['이순신', '2-1'])
        writer.writerow(['최영', '3-1'])

        return response

    elif format == 'excel':
        # Excel 파일 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "참석자 명단"

        # 헤더 스타일
        header_fill = PatternFill(start_color="7B68EE", end_color="7B68EE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 작성
        ws['A1'] = '이름'
        ws['B1'] = '소속/학년반'

        for cell in ['A1', 'B1']:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = header_alignment

        # 예시 데이터
        example_data = [
            ['홍길동', '1-1'],
            ['김철수', '1-2'],
            ['박영희', '교사'],
            ['이순신', '2-1'],
            ['최영', '3-1'],
        ]

        for idx, row in enumerate(example_data, start=2):
            ws[f'A{idx}'] = row[0]
            ws[f'B{idx}'] = row[1]
            # 텍스트 형식으로 명시 (날짜 자동 변환 방지)
            ws[f'A{idx}'].number_format = '@'
            ws[f'B{idx}'].number_format = '@'

        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20

        # 안내 시트 추가
        ws_guide = wb.create_sheet("사용 안내")
        ws_guide['A1'] = "📋 참석자 명단 작성 안내"
        ws_guide['A1'].font = Font(bold=True, size=14, color="7B68EE")

        ws_guide['A3'] = "1. 첫 번째 열에 참석자 이름을 입력하세요."
        ws_guide['A4'] = "2. 두 번째 열에 소속이나 학년반을 입력하세요."
        ws_guide['A5'] = "3. 헤더(첫 번째 행)는 삭제하지 마세요."
        ws_guide['A6'] = "4. 예시 데이터는 삭제하고 실제 데이터를 입력하세요."
        ws_guide['A7'] = "5. 완성 후 파일을 저장하고 업로드하세요."

        ws_guide.column_dimensions['A'].width = 60

        # 파일 저장
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="참석자명단_양식.xlsx"'

        return response

    else:
        return HttpResponse("Invalid format", status=400)
